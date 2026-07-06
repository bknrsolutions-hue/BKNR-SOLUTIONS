from app.utils.timezone import ist_now
# app/routers/bills/electricity_log.py

from fastapi import APIRouter, Request, Depends, Body, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, and_
from datetime import date, datetime
import datetime as dt
import calendar
import logging
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from app.database import get_db
from app.database.models.bills import ElectricityLog
from app.database.models.processing import AuditLog  # మాస్టర్ ఆడిట్ ట్రాక్ మోడల్ సింక్
from app.database.models.criteria import production_at

router = APIRouter(
    prefix="/electricity",
    tags=["Electricity"]
)

templates = Jinja2Templates(directory="app/templates")
logger = logging.getLogger(__name__)


# ============================================================
# 📋 PYDANTIC SCHEMAS FOR JSON PAYLOADS (AJAX COMPATIBLE)
# ============================================================
class ElectricitySchema(BaseModel):
    unit_id: int
    reading_date: date
    opening_kwh: float
    closing_kwh: float
    unit_rate: float


# ============================================================
# 🔌 1. MAIN ENTRY PAGE (GET) - DEFAULT EMPTY STATE (FY LOCKED)
# ============================================================
@router.get("/entry", response_class=HTMLResponse)
def electricity_entry_page(
    request: Request,
    fy: str = Query(None), # Financial Year Query Param
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    company_code = request.session.get("company_code")

    if not email or not company_code:
        return RedirectResponse("/", status_code=302)

    # 🔹 Fetching production units for the dropdown
    units = (
        db.query(production_at)
        .filter(production_at.company_id == company_code, ElectricityLog.is_cancelled != True)
        .order_by(production_at.production_at)
        .all()
    )

    # 🔹 Electricity History Filtered by Financial Year (April to March)
    history = []
    if fy:
        selected_year = int(fy)
        start_date = dt.date(selected_year, 4, 1)
        end_date = dt.date(selected_year + 1, 3, 31)

        history = (
            db.query(
                ElectricityLog.id,
                ElectricityLog.reading_date,
                ElectricityLog.opening_kwh,
                ElectricityLog.closing_kwh,
                ElectricityLog.unit_rate,
                ElectricityLog.total_cost,
                production_at.production_at.label("location_name")
            )
            .join(production_at, ElectricityLog.unit_id == production_at.id)
            .filter(
                production_at.company_id == company_code, ElectricityLog.is_cancelled != True,
                ElectricityLog.is_cancelled != True,
                ElectricityLog.reading_date >= start_date,
                ElectricityLog.reading_date <= end_date
            )
            .order_by(desc(ElectricityLog.reading_date), desc(ElectricityLog.id))
            .all()
        )

    return templates.TemplateResponse(
        request=request,
        name="bills/electricity_entry.html",
        context={
            "units": units,
            "history": history,
            "email": email,
            "company_id": company_code,
            "selected_fy": fy
        }
    )


# ============================================================
# 🔍 2. LOOKUP LAST READING & AUTO-FILL RATE (GET AJAX)
# ============================================================
@router.get("/lookup/{unit_id}")
def lookup_last_reading(unit_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("email"):
        return JSONResponse({"error": "Session expired"}, status_code=401)

    # 1. Master Info for Default Rate
    unit_info = db.query(production_at).filter(production_at.id == unit_id).first()

    # 2. Last Log for Closing Reading & Current Rate
    last_log = (
        db.query(ElectricityLog)
        .filter(ElectricityLog.unit_id == unit_id)
        .order_by(desc(ElectricityLog.reading_date), desc(ElectricityLog.id))
        .first()
    )

    current_rate = 0.0
    if last_log and last_log.unit_rate:
        current_rate = last_log.unit_rate
    elif unit_info and getattr(unit_info, 'unit_rate', None):
        current_rate = unit_info.unit_rate

    return JSONResponse({
        "last_closing": float(last_log.closing_kwh) if last_log else 0.0,
        "unit_rate": float(current_rate),
        "meter_number": getattr(unit_info, 'meter_number', 'N/A') if unit_info else "N/A"
    })


# ============================================================
# 💾 3. SAVE ELECTRICITY ENTRY (POST JSON Payload Target)
# ============================================================
@router.post("/save")
async def save_electricity_entry(
    request: Request,
    payload: ElectricitySchema,
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    company_code = request.session.get("company_code")
    if not email or not company_code:
        return JSONResponse({"success": False, "message": "Session expired"}, status_code=401)

    # Avoid duplicate entries for same unit and date
    exists = db.query(ElectricityLog).filter(
        ElectricityLog.unit_id == payload.unit_id,
        ElectricityLog.reading_date == payload.reading_date
    ).first()
    
    if exists:
        return JSONResponse(
            {"success": False, "status": "error", "message": "Entry already exists for this unit & date"},
            status_code=400
        )

    # Calculation
    units_consumed = payload.closing_kwh - payload.opening_kwh
    if units_consumed < 0:
        return JSONResponse(
            {"success": False, "status": "error", "message": "Closing reading cannot be less than opening"},
            status_code=400
        )

    total_cost = round(units_consumed * payload.unit_rate, 2)

    entry = ElectricityLog(
        unit_id=payload.unit_id,
        reading_date=payload.reading_date,
        opening_kwh=payload.opening_kwh,
        closing_kwh=payload.closing_kwh,
        unit_rate=payload.unit_rate,
        total_cost=total_cost
    )

    try:
        db.add(entry)
        db.flush()

        # 📜 Add Initial Operational Audit Entry
        db.add(AuditLog(
            table_name="electricity_logs", record_id=entry.id, company_id=company_code,
            field_name="CREATE", old_value="NONE", new_value=f"Units Consumed: {units_consumed} KWH (Cost: ₹{total_cost})",
            edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
        ))

        db.commit()
        return JSONResponse({
            "success": True,
            "status": "success", 
            "message": "Data saved successfully",
            "units_consumed": units_consumed,
            "total_cost": total_cost
        })
    except Exception as e:
        db.rollback()
        logger.error(f"ELECTRICITY SAVE ERROR: {str(e)}")
        return JSONResponse({"success": False, "status": "error", "message": str(e)}, status_code=500)


# ============================================================
# 📋 4. MASTER AUDIT HISTORY LOG ENGINE (GET)
# ============================================================
@router.get("/audit_all")
async def get_all_electricity_audit(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    logs = (
        db.query(AuditLog, production_at.production_at)
        .join(ElectricityLog, AuditLog.record_id == ElectricityLog.id)
        .join(production_at, ElectricityLog.unit_id == production_at.id)
        .filter(AuditLog.table_name == "electricity_logs", production_at.company_id == comp_code)
        .order_by(AuditLog.edited_at.desc()).limit(100).all()
    )

    return [{
        "timestamp": l.AuditLog.edited_at.strftime("%d-%m-%Y %H:%M:%S"),
        "user": l.AuditLog.edited_by.split('@')[0] if l.AuditLog.edited_by else "System",
        "email": l.AuditLog.edited_by if l.AuditLog.edited_by else "System",
        "batch": f"Unit: {l.production_at}" if l.production_at else f"ID Ref: {l.AuditLog.record_id}",
        "action": l.AuditLog.field_name,
        "details": l.AuditLog.new_value if l.AuditLog.old_value == "NONE" else f"{l.AuditLog.old_value} ➔ {l.AuditLog.new_value}"
    } for l in logs]


# ============================================================
# 🗑️ 5. DELETE ACTION WITH TRACE AUDIT (POST JSON)
# ============================================================
@router.post("/delete/{log_id}")
def delete_electricity_log(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")

    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    entry = db.query(ElectricityLog).join(production_at, ElectricityLog.unit_id == production_at.id).filter(
        ElectricityLog.id == log_id,
        production_at.company_id == comp_code
    ).first()

    if entry:
        try:
            db.add(AuditLog(
                table_name="electricity_logs", record_id=entry.id, company_id=comp_code,
                field_name="is_cancelled", old_value="False", new_value="True",
                edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
            ))
            entry.is_cancelled = True
            db.commit()
            return {"success": True, "message": "Electricity log entry cancelled successfully"}
        except Exception as e:
            db.rollback()
            return JSONResponse({"success": False, "message": str(e)}, status_code=500)

    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# 📊 6. GLOBAL MASTER EXCEL EXPORT LEDGER (GET)
# ============================================================
@router.get("/export/excel")
def export_electricity_excel(request: Request, db: Session = Depends(get_db)):
    company_code = request.session.get("company_code")
    if not company_code:
        return RedirectResponse("/", status_code=302)

    history = (
        db.query(ElectricityLog, production_at.production_at.label("location_name"))
        .join(production_at, ElectricityLog.unit_id == production_at.id)
        .filter(production_at.company_id == company_code, ElectricityLog.is_cancelled != True)
        .order_by(desc(ElectricityLog.reading_date), desc(ElectricityLog.id))
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Electricity Ledger"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="143465", end_color="143465", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    headers = ["Sl No", "Reading Date", "Location Unit", "Opening KWH", "Closing KWH", "Consumed Units", "Unit Rate", "Total Cost"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, log in enumerate(history, 1):
        consumed = (log.ElectricityLog.closing_kwh or 0) - (log.ElectricityLog.opening_kwh or 0)
        row_data = [
            idx,
            log.ElectricityLog.reading_date.strftime("%Y-%m-%d") if log.ElectricityLog.reading_date else "",
            log.location_name,
            log.ElectricityLog.opening_kwh,
            log.ElectricityLog.closing_kwh,
            consumed,
            log.ElectricityLog.unit_rate,
            log.ElectricityLog.total_cost
        ]
        ws.append(row_data)
        
        curr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in [4, 5, 6, 7, 8]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [1, 2]:
                cell.alignment = Alignment(horizontal="center")

    last_row = ws.max_row
    total_row_idx = last_row + 1
    ws.cell(row=total_row_idx, column=1, value="Total Summary").font = total_font
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=5)
    ws.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="right")

    ws.cell(row=total_row_idx, column=6, value=f"=SUM(F2:F{last_row})").number_format = '#,##0.00'
    ws.cell(row=total_row_idx, column=8, value=f"=SUM(H2:H{last_row})").number_format = '#,##0.00'

    for col in [6, 8]:
        c = ws.cell(row=total_row_idx, column=col)
        c.font = total_font
        c.alignment = Alignment(horizontal="right")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"Electricity_Consumption_Ledger_{ist_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================
# 📊 7. MONTHLY SUMMARY (API FOR REPORTS / UNCHANGED BUSINESS LOGIC)
# ============================================================
@router.get("/summary/{year}/{month}")
def electricity_monthly_summary(year: int, month: int, request: Request, db: Session = Depends(get_db)):
    company_code = request.session.get("company_code")
    if not company_code:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    last_day = calendar.monthrange(year, month)[1]
    start_date = date(year, month, 1)
    end_date = date(year, month, last_day)

    rows = (
        db.query(ElectricityLog)
        .join(production_at, ElectricityLog.unit_id == production_at.id)
        .filter(
            production_at.company_id == company_code, ElectricityLog.is_cancelled != True,
                ElectricityLog.is_cancelled != True,
            ElectricityLog.reading_date.between(start_date, end_date)
        )
        .all()
    )

    summary = {}
    for r in rows:
        u_id = r.unit_id
        if u_id not in summary:
            summary[u_id] = {"units": 0, "cost": 0}
        summary[u_id]["units"] += (r.closing_kwh - r.opening_kwh)
        summary[u_id]["cost"] += r.total_cost

    return JSONResponse(summary)