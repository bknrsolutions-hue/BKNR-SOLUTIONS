from fastapi import APIRouter, Request, Depends, Body, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, and_
from datetime import date, datetime
import datetime as dt
import logging
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from app.database import get_db
from app.database.models.bills import OtherExpense
from app.database.models.processing import AuditLog
from app.database.models.criteria import production_at

router = APIRouter(
    prefix="/expenses",
    tags=["Other Expenses"]
)

templates = Jinja2Templates(directory="app/templates")
logger = logging.getLogger(__name__)


# ============================================================
# 📋 PYDANTIC SCHEMAS FOR JSON PAYLOADS
# ============================================================
class ExpenseSchema(BaseModel):
    production_at_id: int
    expense_date: date
    category: str
    paid_to: str
    remarks: str = ""
    voucher_no: str = ""
    amount: float
    gst_per: float = 0.0
    grand_total: float


# ============================================================
# 🧾 1. EXPENSE ENTRY PAGE (GET) - COMPLIANT WITH MODEL 'date'
# ============================================================
@router.get("/entry", response_class=HTMLResponse)
def expenses_entry_page(
    request: Request,
    fy: str = Query(None), # Financial Year Query Param
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    company_code = request.session.get("company_code")

    if not email or not company_code:
        return RedirectResponse("/", status_code=302)

    # 🔹 Units / Production Locations filter by company
    locations = (
        db.query(production_at)
        .filter(production_at.company_id == company_code)
        .order_by(production_at.production_at)
        .all()
    )

    # 🔹 Expenses History Filtered by Financial Year
    expense_history = []
    if fy:
        selected_year = int(fy)
        start_date = dt.date(selected_year, 4, 1)
        end_date = dt.date(selected_year + 1, 3, 31)

        try:
            # 🌟 మోడల్‌లోని 'date' కాలమ్ బేస్డ్ ఫిల్టరింగ్
            expense_history = (
                db.query(OtherExpense, production_at.production_at.label("location_name"))
                .join(production_at, OtherExpense.unit_id == production_at.id)
                .filter(
                    production_at.company_id == company_code,
                    OtherExpense.date >= start_date,
                    OtherExpense.date <= end_date
                )
                .order_by(OtherExpense.id.desc())
                .all()
            )

            # ఒకవేళ ఫాల్‌బ్యాక్ కింద పాత డేటా (remarks లో స్టోర్ అయిన పాత ఫార్మాట్) లోడ్ అవ్వడానికి:
            if not expense_history:
                fy_filters = [OtherExpense.remarks.like(f"Date: {selected_year}-%")]
                for m in [1, 2, 3]:
                    fy_filters.append(OtherExpense.remarks.like(f"Date: {selected_year + 1}-{m:02d}-%"))
                
                from sqlalchemy import or_
                expense_history = (
                    db.query(OtherExpense, production_at.production_at.label("location_name"))
                    .join(production_at, OtherExpense.unit_id == production_at.id)
                    .filter(
                        production_at.company_id == company_code,
                        or_(*fy_filters)
                    )
                    .order_by(OtherExpense.id.desc())
                    .all()
                )
        except Exception as e:
            logger.error(f"EXPENSE HISTORY FETCH ERROR: {str(e)}")
            expense_history = []

    return templates.TemplateResponse(
        request=request,
        name="bills/expenses_entry.html",
        context={
            "locations": locations,
            "expense_history": expense_history,
            "email": email,
            "company_id": company_code,
            "selected_fy": fy
        }
    )


# ============================================================
# 💾 2. SAVE EXPENSE (POST JSON - AUTOMATIC DATE MAPPED TO 'date')
# ============================================================
@router.post("/save")
async def save_expense(
    request: Request,
    payload: ExpenseSchema,
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    company_code = request.session.get("company_code")
    if not email or not company_code:
        return JSONResponse({"success": False, "message": "Session expired"}, status_code=401)

    # 📝 Combine extra info into remarks exactly matching structure
    full_remarks = (
        f"Date: {payload.expense_date} | "
        f"Paid To: {payload.paid_to} | "
        f"Voucher: {payload.voucher_no} | "
        f"GST: {payload.gst_per}% | "
        f"Notes: {payload.remarks}"
    )

    # 🌟 ఇక్కడ మోడల్ కాలమ్ 'date' కు payload.expense_date ని మ్యాప్ చేసాము 📅
    new_entry = OtherExpense(
        unit_id=payload.production_at_id,
        category=payload.category.upper().strip(),
        amount=payload.grand_total,   
        remarks=full_remarks,
        date=payload.expense_date if payload.expense_date else dt.date.today()
    )

    try:
        db.add(new_entry)
        db.flush()

        # 📜 Add Initial Operational Audit Entry
        db.add(AuditLog(
            table_name="other_expenses", record_id=new_entry.id, company_id=company_code,
            field_name="CREATE", old_value="NONE", new_value=f"Category: {new_entry.category} (Amount: ₹{payload.grand_total})",
            edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
        ))

        db.commit()
        return JSONResponse({"success": True, "message": "Expense transaction logged successfully!"})
    except Exception as e:
        db.rollback()
        logger.error(f"EXPENSE SAVE ERROR: {str(e)}")
        return JSONResponse({"success": False, "message": f"Database Error: {str(e)}"}, status_code=500)


# ============================================================
# 📋 3. MASTER AUDIT HISTORY LOG ENGINE (GET)
# ============================================================
@router.get("/audit_all")
async def get_all_expenses_audit(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    logs = (
        db.query(AuditLog, production_at.production_at)
        .join(OtherExpense, AuditLog.record_id == OtherExpense.id)
        .join(production_at, OtherExpense.unit_id == production_at.id)
        .filter(AuditLog.table_name == "other_expenses", production_at.company_id == comp_code)
        .order_by(AuditLog.edited_at.desc()).limit(100).all()
    )

    return [{
        "timestamp": l.AuditLog.edited_at.strftime("%d-%m-%Y %H:%M:%S"),
        "user": l.AuditLog.edited_by.split('@')[0],
        "invoice_no": l.production_at,
        "action": l.AuditLog.field_name,
        "details": l.AuditLog.new_value if l.AuditLog.old_value == "NONE" else f"{l.AuditLog.old_value} ➔ {l.AuditLog.new_value}"
    } for l in logs]


# ============================================================
# 🗑️ 4. SECURE DELETE ACTION WITH TRACE AUDIT (POST)
# ============================================================
@router.post("/delete/{expense_id}")
def delete_expense(expense_id: int, request: Request, db: Session = Depends(get_db)):
    company_code = request.session.get("company_code")
    email = request.session.get("email")
    if not company_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    
    entry = db.query(OtherExpense).join(
        production_at, OtherExpense.unit_id == production_at.id
    ).filter(
        OtherExpense.id == expense_id,
        production_at.company_id == company_code
    ).first()

    if entry:
        try:
            db.add(AuditLog(
                table_name="other_expenses", record_id=entry.id, company_id=company_code,
                field_name="DELETE", old_value=f"Category: {entry.category}", new_value="DELETED",
                edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
            ))
            db.delete(entry)
            db.commit()
            return {"success": True, "message": "Expense record dropped successfully!"}
        except Exception as e:
            db.rollback()
            return JSONResponse({"success": False, "message": str(e)}, status_code=500)
    
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# 📊 5. GLOBAL MASTER EXCEL EXPORT LEDGER (GET)
# ============================================================
@router.get("/export/excel")
def export_expenses_excel(request: Request, db: Session = Depends(get_db)):
    company_code = request.session.get("company_code")
    if not company_code:
        return RedirectResponse("/", status_code=302)

    history = (
        db.query(OtherExpense, production_at.production_at.label("location_name"))
        .join(production_at, OtherExpense.unit_id == production_at.id)
        .filter(production_at.company_id == company_code)
        .order_by(OtherExpense.id.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Other Expenses"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="143465", end_color="143465", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    headers = ["Sl No", "Location Unit", "Expense Category", "Total Amount (Inclusive)", "Remarks / Metadata History Log"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, log in enumerate(history, 1):
        row_data = [
            idx,
            log.location_name,
            log.OtherExpense.category,
            log.OtherExpense.amount,
            log.OtherExpense.remarks
        ]
        ws.append(row_data)
        
        curr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 4:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [1, 3]:
                cell.alignment = Alignment(horizontal="center")

    last_row = ws.max_row
    total_row_idx = last_row + 1
    ws.cell(row=total_row_idx, column=1, value="Total Summary").font = total_font
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=3)
    ws.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="right")

    ws.cell(row=total_row_idx, column=4, value=f"=SUM(D2:D{last_row})").number_format = '#,##0.00'
    ws.cell(row=total_row_idx, column=4).font = total_font
    ws.cell(row=total_row_idx, column=4).alignment = Alignment(horizontal="right")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"Other_Expenses_Ledger_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )