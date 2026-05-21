# app/routers/bills/diesel_log.py

from fastapi import APIRouter, Request, Depends, Body, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, and_
from datetime import date, datetime
import datetime as dt
import logging
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.database.models.bills import DieselLog
from app.database.models.processing import AuditLog  # మాస్టర్ ఆడిట్ ట్రాక్ మోడల్ సింక్
from app.database.models.criteria import production_at

router = APIRouter(
    prefix="/diesel",
    tags=["Diesel"]
)

templates = Jinja2Templates(directory="app/templates")
logger = logging.getLogger(__name__)


# ============================================================
# 📋 PYDANTIC SCHEMAS FOR JSON PAYLOADS (AJAX COMPATIBLE)
# ============================================================
class DieselInSchema(BaseModel):
    entry_date: date
    in_unit_id: int
    bill_date: date
    grn_no: str
    bill_no: str
    vendor: str
    received_qty: float
    rate: float
    tax_per: float
    net_amount: float
    closing_stock: float

class DieselOutSchema(BaseModel):
    out_date: date
    unit_id: int
    out_qty: float
    out_closing: float


# ============================================================
# 1. DIESEL ENTRY PAGE (GET) - DEFAULT EMPTY STATE (FY LOCKED)
# ============================================================
@router.get("/entry", response_class=HTMLResponse)
def diesel_entry_page(
    request: Request,
    fy: str = Query(None), # Financial Year Query Param
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    company_code = request.session.get("company_code")

    if not email or not company_code:
        return RedirectResponse("/", status_code=302)

    # 🔹 Production Locations List
    locations = (
        db.query(production_at)
        .filter(production_at.company_id == company_code)
        .order_by(production_at.production_at)
        .all()
    )

    # 🔹 Diesel History Filtered by Financial Year (April to March)
    diesel_history = []
    if fy:
        selected_year = int(fy)
        start_date = dt.date(selected_year, 4, 1)
        end_date = dt.date(selected_year + 1, 3, 31)

        diesel_history = (
            db.query(DieselLog, production_at.production_at.label("location_name"))
            .join(production_at, DieselLog.unit_id == production_at.id)
            .filter(
                production_at.company_id == company_code,
                DieselLog.log_date >= start_date,
                DieselLog.log_date <= end_date
            )
            .order_by(desc(DieselLog.log_date), desc(DieselLog.id))
            .all()
        )

    return templates.TemplateResponse(
        request=request,
        name="bills/diesel_entry.html",
        context={
            "locations": locations,
            "diesel_history": diesel_history,
            "email": email,
            "selected_fy": fy
        }
    )


# ============================================================
# 2. LOOKUP LAST ACTIVE STOCK (GET AJAX)
# ============================================================
@router.get("/lookup/{unit_id}")
def lookup_diesel_status(unit_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("email"):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    last = (
        db.query(DieselLog)
        .filter(DieselLog.unit_id == unit_id)
        .order_by(desc(DieselLog.id))
        .first()
    )

    return {
        "last_closing": float(last.closing_stock) if last else 0.0,
        "last_rate": float(last.avg_price) if last else 0.0
    }


# ============================================================
# 3. SAVE DIESEL IN / GRN STOCK (POST JSON Payload Target)
# ============================================================
@router.post("/save_in")
async def save_diesel_in(
    request: Request,
    payload: DieselInSchema,
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    company_code = request.session.get("company_code")
    if not email or not company_code:
        return JSONResponse({"success": False, "message": "Session Expired"}, status_code=401)

    try:
        last = db.query(DieselLog).filter(DieselLog.unit_id == payload.in_unit_id).order_by(desc(DieselLog.id)).first()
        opening = last.closing_stock if last else 0.0

        new_log = DieselLog(
            unit_id=payload.in_unit_id,
            log_date=payload.entry_date,
            bill_date=payload.bill_date,
            type="IN",
            grn_no=payload.grn_no.upper().strip(),
            bill_no=payload.bill_no.upper().strip(),
            vendor=payload.vendor,
            opening_stock=opening,
            purchase_qty=payload.received_qty,
            consumption=0.0,
            closing_stock=payload.closing_stock,
            avg_price=payload.rate,
            tax_per=payload.tax_per,
            net_val=payload.net_amount,
            email=email
        )

        db.add(new_log)
        db.flush()

        # 📜 Add Audit Log Trace
        db.add(AuditLog(
            table_name="diesel_logs", record_id=new_log.id, company_id=company_code,
            field_name="CREATE_IN", old_value="NONE", new_value=f"GRN: {new_log.grn_no} ({payload.received_qty} Ltr)",
            edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
        ))

        db.commit()
        return JSONResponse({"success": True, "message": "Diesel Stock IN record committed successfully!"})
    except Exception as e:
        db.rollback()
        logger.error(f"DIESEL IN SAVE ERROR: {str(e)}")
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


# ============================================================
# 4. SAVE DIESEL OUT / CONSUMPTION (POST JSON Payload Target)
# ============================================================
@router.post("/save_out")
async def save_diesel_out(
    request: Request,
    payload: DieselOutSchema,
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    company_code = request.session.get("company_code")
    if not email or not company_code:
        return JSONResponse({"success": False, "message": "Session Expired"}, status_code=401)

    try:
        last = db.query(DieselLog).filter(DieselLog.unit_id == payload.unit_id).order_by(desc(DieselLog.id)).first()
        opening = float(last.closing_stock) if last else 0.0
        rate = float(last.avg_price) if last else 0.0

        new_log = DieselLog(
            unit_id=payload.unit_id,
            log_date=payload.out_date,
            type="OUT",
            grn_no="", bill_no="", vendor="Internal Consumption",
            opening_stock=opening,
            purchase_qty=0.0,
            consumption=payload.out_qty,
            closing_stock=payload.out_closing,
            avg_price=rate,
            tax_per=0.0,
            net_val=round(payload.out_qty * rate, 2),
            email=email
        )

        db.add(new_log)
        db.flush()

        # 📜 Add Audit Log Trace
        db.add(AuditLog(
            table_name="diesel_logs", record_id=new_log.id, company_id=company_code,
            field_name="CREATE_OUT", old_value="NONE", new_value=f"Consumed: {payload.out_qty} Ltr",
            edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
        ))

        db.commit()
        return JSONResponse({"success": True, "message": "Diesel Consumption record committed successfully!"})
    except Exception as e:
        db.rollback()
        logger.error(f"DIESEL OUT SAVE ERROR: {str(e)}")
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


# ============================================================
# 5. MASTER AUDIT HISTORY LOG ENGINE (GET)
# ============================================================
@router.get("/audit_all")
async def get_all_diesel_audit(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    logs = (
        db.query(AuditLog, production_at.production_at)
        .join(DieselLog, AuditLog.record_id == DieselLog.id)
        .join(production_at, DieselLog.unit_id == production_at.id)
        .filter(AuditLog.table_name == "diesel_logs", production_at.company_id == comp_code)
        .order_by(AuditLog.edited_at.desc()).limit(100).all()
    )

    return [{
        "timestamp": l.AuditLog.edited_at.strftime("%d-%m-%Y %H:%M:%S"),
        "user": l.AuditLog.edited_by.split('@')[0],
        "invoice_no": l.production_at, # Location Unit string token redirection
        "action": l.AuditLog.field_name,
        "details": l.AuditLog.new_value if l.AuditLog.old_value == "NONE" else f"{l.AuditLog.old_value} ➔ {l.AuditLog.new_value}"
    } for l in logs]


# ============================================================
# 6. DELETE ACTION WITH TRACE AUDIT (POST)
# ============================================================
@router.post("/delete/{log_id}")
def delete_diesel_log(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")

    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    log_entry = db.query(DieselLog).join(production_at, DieselLog.unit_id == production_at.id).filter(
        DieselLog.id == log_id,
        production_at.company_id == comp_code
    ).first()

    if log_entry:
        try:
            db.add(AuditLog(
                table_name="diesel_logs", record_id=log_entry.id, company_id=comp_code,
                field_name="DELETE", old_value=f"Type: {log_entry.type}", new_value="DELETED",
                edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
            ))
            db.delete(log_entry)
            db.commit()
            return {"success": True, "message": "Diesel record destroyed successfully"}
        except Exception as e:
            db.rollback()
            return JSONResponse({"success": False, "message": str(e)}, status_code=500)

    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# 7. GLOBAL MASTER EXCEL EXPORT LEDGER (GET)
# ============================================================
@router.get("/export/excel")
def export_diesel_excel(request: Request, db: Session = Depends(get_db)):
    company_code = request.session.get("company_code")
    if not company_code:
        return RedirectResponse("/", status_code=302)

    history = (
        db.query(DieselLog, production_at.production_at.label("location_name"))
        .join(production_at, DieselLog.unit_id == production_at.id)
        .filter(production_at.company_id == company_code)
        .order_by(desc(DieselLog.log_date), desc(DieselLog.id))
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Diesel Stock Ledger"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="143465", end_color="143465", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    headers = [
        "Sl No", "Log Date", "Location Unit", "Type", "GRN Ref", "Bill Number", 
        "Vendor Name", "Opening (L)", "Stock In (L)", "Consume (L)", "Closing (L)", "Avg Rate", "Net Value"
    ]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, log in enumerate(history, 1):
        row_data = [
            idx,
            log.DieselLog.log_date.strftime("%Y-%m-%d") if log.DieselLog.log_date else "",
            log.location_name,
            "STOCK IN" if log.DieselLog.type == "IN" else "CONSUMPTION",
            log.DieselLog.grn_no or "-",
            log.DieselLog.bill_no or "-",
            log.DieselLog.vendor or "Internal",
            log.DieselLog.opening_stock,
            log.DieselLog.purchase_qty,
            log.DieselLog.consumption,
            log.DieselLog.closing_stock,
            log.DieselLog.avg_price,
            log.DieselLog.net_val
        ]
        ws.append(row_data)
        
        curr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in [8, 9, 10, 11, 12, 13]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [1, 2, 4, 5, 6]:
                cell.alignment = Alignment(horizontal="center")

    last_row = ws.max_row
    total_row_idx = last_row + 1
    ws.cell(row=total_row_idx, column=1, value="Total Summary").font = total_font
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=7)
    ws.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="right")

    ws.cell(row=total_row_idx, column=9, value=f"=SUM(I2:I{last_row})").number_format = '#,##0.00'
    ws.cell(row=total_row_idx, column=10, value=f"=SUM(J2:J{last_row})").number_format = '#,##0.00'
    ws.cell(row=total_row_idx, column=13, value=f"=SUM(M2:M{last_row})").number_format = '#,##0.00'

    for col in [9, 10, 13]:
        c = ws.cell(row=total_row_idx, column=col)
        c.font = total_font
        c.alignment = Alignment(horizontal="right")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"Diesel_Inventory_Ledger_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )