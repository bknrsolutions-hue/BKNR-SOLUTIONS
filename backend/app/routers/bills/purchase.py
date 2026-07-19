from app.utils.timezone import ist_now
from fastapi import APIRouter, Request, Depends, Body, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, and_
from datetime import date, datetime
import datetime as dt
import logging
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services.pdf_renderer import render_pdf_from_html

from app.database import get_db
from app.database.models.bills import PurchaseInvoice
from app.database.models.processing import AuditLog 
from app.database.models.inventory_management import pending_orders, sales_dispatch
from app.database.models.criteria import (
    production_at,
    vendors,
    hsn_codes
)
from app.services.bill_accounting import (
    cancel_linked_bill_voucher,
    ensure_bill_accounting_schema,
    list_posting_ledgers,
    post_vendor_bill,
    resolve_posting_ledger,
)

router = APIRouter(
    prefix="/purchase",
    tags=["Purchase"]
)

templates = Jinja2Templates(directory="app/templates")
logger = logging.getLogger(__name__)


def purchase_stock_account(product_name: str):
    clean = (product_name or "").strip().lower()
    if "sticker" in clean or "label" in clean:
        return {
            "ledger_name": "Stickers Expense A/c",
            "group_name": "Direct Expenses",
            "group_type": "EXPENSE",
            "parent_group_name": None,
        }
    if any(word in clean for word in ["chemical", "chem", "salt", "stpp", "soda", "powder", "dry", "wet"]):
        return {
            "ledger_name": "Chemicals Stock A/c",
            "group_name": "Current Assets",
            "group_type": "ASSET",
            "parent_group_name": None,
        }
    return {
        "ledger_name": "Packing Material Stock A/c",
        "group_name": "Current Assets",
        "group_type": "ASSET",
        "parent_group_name": None,
    }


# ============================================================
# 📋 PYDANTIC SCHEMAS FOR JSON PAYLOAD
# ============================================================
class PurchaseInvoiceSchema(BaseModel):
    unit_id: int
    vendor_id: int
    invoice_date: date
    invoice_no: str
    product_name: str
    hsn_code: str
    qty: float
    base_price: float
    gst_percent: float
    po_number: str = None
    accounting_ledger_id: int | None = None


# --- Helper: Get Financial Year Year (April to March) ---
def get_fin_year(date_val):
    if not date_val: return None
    return date_val.year if date_val.month >= 4 else date_val.year - 1


# ============================================================
# 1. MAIN ENTRY PAGE (GET) - DEFAULT EMPTY STATE (FY LOCKED)
# ============================================================
@router.get("/entry", response_class=HTMLResponse)
def purchase_entry_page(
    request: Request, 
    fy: str = Query(None), # Financial Year Query Param
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    company_id = request.session.get("company_code")

    if not email or not company_id:
        return RedirectResponse("/", status_code=302)

    ensure_bill_accounting_schema(db)
    if fy is None:
        today = ist_now().date()
        fy = str(today.year if today.month >= 4 else today.year - 1)

    # 🔹 Production Units / Locations
    locations = db.query(production_at).filter(
        production_at.company_id == company_id
    ).order_by(production_at.production_at).all()

    # 🔹 Vendors
    all_vendors = db.query(vendors).filter(
        vendors.company_id == company_id
    ).order_by(vendors.name).all()

    # 🔹 HSN & GST % Combo Data
    hsn_list = db.query(hsn_codes).filter(
        hsn_codes.company_id == company_id
    ).order_by(hsn_codes.hsn_code).all()

    posting_ledgers = list_posting_ledgers(
        db,
        company_id,
        group_types={"EXPENSE", "ASSET"},
        group_names={"Purchase Accounts", "Direct Expenses", "Indirect Expenses", "Current Assets"},
    )

    # 🔹 ⚡ Financial Year Query Logic
    invoice_history = []
    if fy:
        selected_year = int(fy)
        start_date = dt.date(selected_year, 4, 1)
        end_date = dt.date(selected_year + 1, 3, 31)
        
        invoice_history = (
            db.query(PurchaseInvoice)
            .filter(
                PurchaseInvoice.company_id == company_id,
                PurchaseInvoice.invoice_date >= start_date,
                PurchaseInvoice.invoice_date <= end_date
            )
            .order_by(desc(PurchaseInvoice.invoice_date), desc(PurchaseInvoice.id))
            .all()
        )

    vendor_map = {v.id: v.name for v in all_vendors}
    location_code_map = {loc.id: loc.production_at for loc in locations}

    # 📋 PO NUMBER DROPDOWN EXTRACTOR (PERFORMANCE OPTIMIZED)
    po_from_pending = db.query(pending_orders.po_number).filter(
        pending_orders.company_id == company_id,
        pending_orders.po_number.isnot(None),
        pending_orders.po_number != ""
    ).distinct().limit(500).all()

    po_from_sales = db.query(sales_dispatch.po_number).filter(
        sales_dispatch.company_id == company_id,
        sales_dispatch.po_number.isnot(None),
        sales_dispatch.po_number != ""
    ).distinct().limit(500).all()

    po_set = set()
    for p in po_from_pending: po_set.add(p[0].strip())
    for p in po_from_sales: po_set.add(p[0].strip())

    po_list = sorted(list(po_set))
    po_list.insert(0, "N/A")

    return templates.TemplateResponse(
        request=request,
        name="bills/purchase_entry.html",
        context={
            "locations": locations,
            "vendors": all_vendors,
            "location_code_map": location_code_map,
            "vendor_map": vendor_map,
            "hsn_list": hsn_list,
            "invoice_history": invoice_history,
            "po_list": po_list,
            "posting_ledgers": posting_ledgers,
            "email": email,
            "company_id": company_id,
            "selected_fy": fy
        }
    )


# ============================================================
# 2. SAVE/CREATE ACTION (POST) - BULLETPROOF IST PROTECTED
# ============================================================
@router.post("/save")
async def save_purchase_invoice(
    request: Request,
    payload: PurchaseInvoiceSchema,
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    company_id = request.session.get("company_code")

    if not email or not company_id:
        return JSONResponse({"success": False, "message": "Session Expired"}, status_code=401)

    try:
        ensure_bill_accounting_schema(db)
        duplicate = db.query(PurchaseInvoice).filter(
            PurchaseInvoice.invoice_no == payload.invoice_no.strip(),
            PurchaseInvoice.company_id == company_id, PurchaseInvoice.is_cancelled != True
        ).first()

        if duplicate:
            return JSONResponse(
                {"success": False, "message": f"Invoice {payload.invoice_no} already exists in records"},
                status_code=400
            )

        # 🧮 Calculations
        taxable_value = round(payload.qty * payload.base_price, 2)
        tax_amount = round((taxable_value * payload.gst_percent) / 100, 2)
        grand_total = round(taxable_value + tax_amount, 2)

        po_number = payload.po_number
        if po_number:
            po_number = po_number.strip()
            if po_number.upper() in ["N/A", "", "-"]:
                po_number = None

        # 🛠️ FIX: Explicit IST deployment timestamp tracking
        current_ist = ist_now()
        vendor = db.query(vendors).filter(vendors.id == payload.vendor_id, vendors.company_id == company_id).first()
        vendor_name = vendor.name if vendor else f"Vendor {payload.vendor_id}"

        new_invoice = PurchaseInvoice(
            unit_id=payload.unit_id,
            production_at_id=payload.unit_id,
            vendor_id=payload.vendor_id,
            invoice_date=payload.invoice_date,
            invoice_no=payload.invoice_no.upper().strip(),
            product_name=payload.product_name.upper().strip(),
            hsn_code=payload.hsn_code.strip(),
            qty=payload.qty,
            base_price=payload.base_price,
            gst_percent=payload.gst_percent,
            tax_amount=tax_amount,
            grand_total=grand_total,
            po_number=po_number,
            company_id=company_id,
            email=email,
            date=current_ist.strftime("%Y-%m-%d"), # 🟢 Fixed Midnight Date Drift
            time=current_ist.strftime("%H:%M:%S"),  # 🟢 Fixed dt.ist_now() AttributeError Crash
            status="DRAFT"
        )

        db.add(new_invoice)
        db.flush()

        # 📜 Master Audit Track Creation (IST Calibrated)
        db.add(AuditLog(
            table_name="purchase_invoice", record_id=new_invoice.id, company_id=company_id,
            field_name="CREATE", old_value="NONE", new_value=new_invoice.invoice_no,
            edited_by=email, edited_at=current_ist # 🟢 Synced onto common IST timeline
        ))

        default_posting = purchase_stock_account(new_invoice.product_name)
        posting_ledger = resolve_posting_ledger(
            db,
            company_id,
            payload.accounting_ledger_id,
            default_posting["ledger_name"],
            default_posting["group_name"],
            default_posting["group_type"],
            default_posting["parent_group_name"],
        )
        voucher = post_vendor_bill(
            db,
            company_id,
            payload.invoice_date,
            new_invoice.invoice_no,
            vendor_name,
            posting_ledger["ledger_name"],
            taxable_value,
            tax_amount,
            grand_total,
            f"Purchase invoice {new_invoice.invoice_no}",
            email,
            expense_group_name=posting_ledger["group_name"],
            expense_group_type=posting_ledger["group_type"],
            expense_parent_group_name=posting_ledger["parent_group_name"],
            voucher_type="Purchase",
        )
        new_invoice.journal_id = voucher.id
        new_invoice.status = "POSTED"
        
        db.commit()
        return JSONResponse({"success": True, "message": f"Invoice {payload.invoice_no} saved and posted: {voucher.voucher_no}"})

    except Exception as e:
        db.rollback()
        logger.error(f"PURCHASE SAVE ERROR: {str(e)}")
        return JSONResponse({"success": False, "message": f"Internal Server Error: {str(e)}"}, status_code=500)

# ============================================================
# 3. UPDATE ACTION (PUT) - WITH DYNAMIC FIELD LOG TRACKING
# ============================================================
@router.put("/update/{inv_id}")
async def update_purchase_invoice(
    inv_id: int,
    request: Request,
    payload: PurchaseInvoiceSchema,
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    company_id = request.session.get("company_code")

    if not email or not company_id:
        return JSONResponse({"success": False, "message": "Session Expired"}, status_code=401)

    try:
        invoice = db.query(PurchaseInvoice).filter(
            PurchaseInvoice.id == inv_id,
            PurchaseInvoice.company_id == company_id, PurchaseInvoice.is_cancelled != True
        ).first()

        if not invoice:
            return JSONResponse({"success": False, "message": "Invoice record not found"}, status_code=404)
        if invoice.journal_id and invoice.status == "POSTED":
            return JSONResponse({"success": False, "message": "Posted invoice cannot be edited. Cancel and re-enter to keep accounts correct."}, status_code=400)

        current_ist = ist_now()

        # 📜 Field Level Tracking and Auditing
        tracked_fields = {
            "unit_id": payload.unit_id,
            "vendor_id": payload.vendor_id,
            "invoice_date": payload.invoice_date,
            "invoice_no": payload.invoice_no.upper().strip(),
            "product_name": payload.product_name.upper().strip(),
            "hsn_code": payload.hsn_code.strip(),
            "qty": payload.qty,
            "base_price": payload.base_price,
            "gst_percent": payload.gst_percent,
            "po_number": payload.po_number.strip() if payload.po_number else None
        }

        for key, new_val in tracked_fields.items():
            old_val = str(getattr(invoice, key) or "")
            check_new = str(new_val or "")
            if old_val != check_new:
                db.add(AuditLog(
                    table_name="purchase_invoice", record_id=invoice.id, company_id=company_id,
                    field_name=key, old_value=old_val, new_value=check_new,
                    edited_by=email, edited_at=current_ist # 🟢 Synced onto IST timeline
                ))
                setattr(invoice, key, new_val)

        # 🧮 Re-calculations
        taxable_value = round(invoice.qty * invoice.base_price, 2)
        invoice.tax_amount = round((taxable_value * invoice.gst_percent) / 100, 2)
        invoice.grand_total = round(taxable_value + invoice.tax_amount, 2)
        invoice.production_at_id = invoice.unit_id
        
        invoice.updated_by = email
        invoice.updated_date = current_ist.strftime("%Y-%m-%d") # 🟢 Synced onto IST date

        db.commit()
        return JSONResponse({"success": True, "message": f"Invoice {invoice.invoice_no} updated successfully!"})

    except Exception as e:
        db.rollback()
        logger.error(f"PURCHASE UPDATE ERROR: {str(e)}")
        return JSONResponse({"success": False, "message": f"Internal Server Error: {str(e)}"}, status_code=500)


# ============================================================
# 4. MASTER AUDIT HISTORY LOG ENGINE (GET) - ALL RECORDS MATCH
# ============================================================
@router.get("/audit_all")
async def get_all_purchase_audit(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    logs = (
        db.query(AuditLog, PurchaseInvoice.invoice_no)
        .join(PurchaseInvoice, AuditLog.record_id == PurchaseInvoice.id)
        .filter(AuditLog.table_name == "purchase_invoice", AuditLog.company_id == comp_code)
        .order_by(AuditLog.edited_at.desc()).limit(100).all()
    )
    
    return [{
        "record_id": l.AuditLog.record_id,
        "timestamp": l.AuditLog.edited_at.strftime("%d-%m-%Y %H:%M:%S"),
        "user": l.AuditLog.edited_by.split('@')[0] if l.AuditLog.edited_by else "System",
        "email": l.AuditLog.edited_by if l.AuditLog.edited_by else "System",
        "batch": f"Row ID #{l.AuditLog.record_id} • Invoice: {l.invoice_no}" if l.invoice_no else f"Row ID #{l.AuditLog.record_id}",
        "action": f"Changed {l.AuditLog.field_name.replace('_', ' ').title()}" if l.AuditLog.field_name not in ["CREATE", "DELETE"] else l.AuditLog.field_name,
        "details": f"{l.AuditLog.old_value} ➔ {l.AuditLog.new_value}" if l.AuditLog.old_value != "NONE" else f"Created Invoice Entry: {l.AuditLog.new_value}"
    } for l in logs]


# ============================================================
# 5. DELETE ACTION (POST) - WITH AUDIT DESTROY TRACE
# ============================================================
@router.post("/delete/{inv_id}")
def delete_invoice(inv_id: int, request: Request, db: Session = Depends(get_db)):
    company_id = request.session.get("company_code")
    email = request.session.get("email")
    if not company_id:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    invoice = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.id == inv_id,
        PurchaseInvoice.company_id == company_id, PurchaseInvoice.is_cancelled != True
    ).first()

    if invoice:
        try:
            db.add(AuditLog(
                table_name="purchase_invoice", record_id=invoice.id, company_id=company_id,
                field_name="is_cancelled", old_value="False", new_value="True",
                edited_by=email, edited_at=ist_now() # 🟢 Synced onto IST timeline
            ))
            cancel_linked_bill_voucher(db, company_id, invoice.journal_id, email)
            invoice.is_cancelled = True
            invoice.status = "CANCELLED"
            db.commit()
            return {"success": True, "message": "Record cancelled successfully"}
        except Exception as e:
            db.rollback()
            return JSONResponse({"success": False, "message": str(e)}, status_code=500)

    return JSONResponse({"success": False, "message": "Invoice not found"}, status_code=404)


# ============================================================
# 6. GLOBAL EXCEL EXPORT
# ============================================================
@router.get("/export/excel")
def export_purchase_excel(request: Request, db: Session = Depends(get_db)):
    company_id = request.session.get("company_code")
    if not company_id:
        return RedirectResponse("/", status_code=302)

    invoices = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.company_id == company_id, PurchaseInvoice.is_cancelled != True
    ).order_by(desc(PurchaseInvoice.invoice_date)).all()

    all_vendors = db.query(vendors).filter(vendors.company_id == company_id).all()
    vendor_map = {v.id: v.name for v in all_vendors}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Ledger"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="143465", end_color="143465", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    headers = [
        "Sl No", "Date", "Invoice No", "PO Number", 
        "Vendor Name", "Product Description", "HSN Code", 
        "Qty", "Rate", "Taxable Value", "GST %", "Tax Amount", "Grand Total"
    ]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, inv in enumerate(invoices, 1):
        taxable_val = round(inv.qty * inv.base_price, 2)
        v_name = vendor_map.get(inv.vendor_id, f"ID: {inv.vendor_id}")

        row_data = [
            idx, inv.invoice_date.strftime("%Y-%m-%d") if inv.invoice_date else "",
            inv.invoice_no, inv.po_number or "N/A", v_name, inv.product_name, inv.hsn_code,
            inv.qty, inv.base_price, taxable_val, f"{inv.gst_percent}%", inv.tax_amount, inv.grand_total
        ]
        ws.append(row_data)
        
        curr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in [8, 9, 10, 12, 13]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [1, 2, 3, 4, 7, 11]:
                cell.alignment = Alignment(horizontal="center")

    last_row = ws.max_row
    total_row_idx = last_row + 1
    ws.cell(row=total_row_idx, column=1, value="Total Summary").font = total_font
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=7)
    ws.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="right")

    ws.cell(row=total_row_idx, column=8, value=f"=SUM(H2:H{last_row})").number_format = '#,##0.00'
    ws.cell(row=total_row_idx, column=10, value=f"=SUM(J2:J{last_row})").number_format = '#,##0.00'
    ws.cell(row=total_row_idx, column=12, value=f"=SUM(L2:L{last_row})").number_format = '#,##0.00'
    ws.cell(row=total_row_idx, column=13, value=f"=SUM(M2:M{last_row})").number_format = '#,##0.00'

    for col in [8, 10, 12, 13]:
        c = ws.cell(row=total_row_idx, column=col)
        c.font = total_font
        c.alignment = Alignment(horizontal="right")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"Purchase_Ledger_{ist_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================
# 7. WEASYPRINT PDF EXPORT FROM 3-DOTS
# ============================================================
@router.get("/export/pdf/{inv_id}")
def export_invoice_pdf(inv_id: int, request: Request, db: Session = Depends(get_db)):
    company_id = request.session.get("company_code")
    if not company_id:
        return RedirectResponse("/", status_code=302)
        
    invoice = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.id == inv_id, PurchaseInvoice.company_id == company_id, PurchaseInvoice.is_cancelled != True
    ).first()
    
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice record not found")
        
    all_vendors = db.query(vendors).filter(vendors.company_id == company_id).all()
    vendor_map = {v.id: v.name for v in all_vendors}
    
    html_content = templates.get_template("reports/purchase_print.html").render({
        "request": request,
        "invoice": invoice,
        "vendor_name": vendor_map.get(invoice.vendor_id, "Unknown Vendor"),
        "printed_on": ist_now()
    })
    
    pdf = render_pdf_from_html(html_content)
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Invoice_{invoice.invoice_no}.pdf"}
    )


# ============================================================
# 8. PRINT INDIVIDUAL INVOICE FROM 3-DOTS
# ============================================================
@router.get("/print/{inv_id}", response_class=HTMLResponse)
def print_individual_invoice(inv_id: int, request: Request, db: Session = Depends(get_db)):
    company_id = request.session.get("company_code")
    if not company_id:
        return RedirectResponse("/", status_code=302)

    invoice = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.id == inv_id, PurchaseInvoice.company_id == company_id, PurchaseInvoice.is_cancelled != True
    ).first()

    if not invoice:
        return HTMLResponse(content="<h3>Invoice Record Not Found</h3>", status_code=404)

    all_vendors = db.query(vendors).filter(vendors.company_id == company_id).all()
    vendor_map = {v.id: v.name for v in all_vendors}

    return templates.TemplateResponse(
        request=request,
        name="reports/purchase_print.html",
        context={
            "invoice": invoice, 
            "vendor_name": vendor_map.get(invoice.vendor_id, "Unknown Vendor"),
            "company_id": company_id
        }
    )
