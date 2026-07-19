from app.utils.timezone import ist_now
from fastapi import APIRouter, Request, Depends, Body, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, and_
from datetime import date, datetime
import datetime as dt
import logging
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.database.models.bills import ContainerLog, PurchaseInvoice  
from app.database.models.processing import AuditLog 
from app.database.models.criteria import vendors, production_at # 🟢 ADDED: production_at
from app.database.models.inventory_management import pending_orders, sales_dispatch
from app.utils.global_filters import get_global_filters # 🟢 ADDED: Global Filters
from app.services.bill_accounting import (
    cancel_linked_bill_voucher,
    ensure_bill_accounting_schema,
    list_posting_ledgers,
    post_vendor_bill,
    resolve_posting_ledger,
)

router = APIRouter(
    prefix="/container",
    tags=["Container Logistics"]
)

templates = Jinja2Templates(directory="app/templates")
logger = logging.getLogger(__name__)


# ============================================================
# 📋 PYDANTIC SCHEMAS FOR JSON PAYLOAD
# ============================================================
class ContainerLogisticsSchema(BaseModel):
    po_number: str
    production_at: str  # 🟢 ADDED: Location Field
    container_no: str
    size: str
    shipping_line_id: int
    ocean_cost: float
    local_cost: float
    handling: float
    detention: float
    gst_percent: float
    accounting_ledger_id: int | None = None


# ============================================================
# 🚢 1. MAIN ENTRY PAGE (GET) - FIXED FOR NATIVE DATE & LOCATION FILTER
# ============================================================
@router.get("/entry", response_class=HTMLResponse)
def container_entry_page(
    request: Request,
    fy: str = Query(None), # Financial Year Query Param
    location: str = Query(None), # 🟢 ADDED: Location Query Param
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    comp_code = request.session.get("company_code")

    if not email or not comp_code:
        return RedirectResponse("/", status_code=302)

    ensure_bill_accounting_schema(db)
    if fy is None:
        today = ist_now().date()
        fy = str(today.year if today.month >= 4 else today.year - 1)

    # 🟢 1. Location Filters Setup
    _, cookie_loc = get_global_filters(request)
    global_location = location if location is not None else cookie_loc
    g_loc_clean = global_location.strip().upper() if global_location else None

    session_locations = request.session.get("allowed_locations", [])
    user_allowed_locations = [loc.strip().upper() for loc in session_locations.split(",") if loc.strip()] if isinstance(session_locations, str) else [str(loc).strip().upper() for loc in session_locations if str(loc).strip()]

    # 🟢 2. Production At List for Dropdown
    pl_q = db.query(production_at.production_at).filter(production_at.company_id == comp_code)
    if user_allowed_locations:
        pl_q = pl_q.filter(func.upper(func.trim(production_at.production_at)).in_(user_allowed_locations))
    production_at_list = [p[0] for p in pl_q.order_by(production_at.production_at).all() if p[0]]

    # 🔹 Vendors (Shipping Lines)
    vendor_list = db.query(vendors).filter(
        vendors.company_id == comp_code
    ).order_by(vendors.name).all()

    posting_ledgers = list_posting_ledgers(
        db,
        comp_code,
        group_types={"EXPENSE"},
        group_names={"Direct Expenses", "Indirect Expenses"},
    )

    # 🔹 PO LIST DROP-DOWN (PENDING + SALES)
    pending_po = db.query(pending_orders.po_number).filter(
        pending_orders.company_id == comp_code,
        pending_orders.po_number.isnot(None),
        pending_orders.po_number != ""
    ).distinct().limit(500).all()

    sales_po = db.query(sales_dispatch.po_number).filter(
        sales_dispatch.company_id == comp_code,
        sales_dispatch.po_number.isnot(None),
        sales_dispatch.po_number != ""
    ).distinct().limit(500).all()

    po_set = set()
    for p in pending_po:
        if p[0] and str(p[0]).strip().upper() not in ["N/A", "-", "NONE"]:
            po_set.add(str(p[0]).strip())

    for s in sales_po:
        if s[0] and str(s[0]).strip().upper() not in ["N/A", "-", "NONE"]:
            po_set.add(str(s[0]).strip())

    po_list = sorted(list(po_set))

    # 🔹 ⚡ HISTORY WITH NEW NATIVE DATE & LOCATION OVERSIGHT
    container_history = []
    if fy:
        selected_year = int(fy)
        start_date = dt.date(selected_year, 4, 1)
        end_date = dt.date(selected_year + 1, 3, 31)

        try:
            # Base Query
            query = db.query(ContainerLog, vendors.name.label("v_name")).join(
                vendors, ContainerLog.vendor_id == vendors.id
            ).filter(
                ContainerLog.company_id == comp_code,
                ContainerLog.date >= start_date,
                ContainerLog.date <= end_date
            )

            # 🟢 Apply Location Filter
            if g_loc_clean and g_loc_clean != "ALL":
                query = query.filter(func.upper(func.trim(ContainerLog.production_at)) == g_loc_clean)
            elif user_allowed_locations:
                query = query.filter(func.upper(func.trim(ContainerLog.production_at)).in_(user_allowed_locations))

            container_history = query.order_by(desc(ContainerLog.id)).all()
                
        except Exception as e:
            logger.error(f"FETCH CONTAINER LOGS ERROR: {e}")
            container_history = []

    return templates.TemplateResponse(
        request=request,
        name="bills/container_entry.html",
        context={
            "shipping_vendors": vendor_list,
            "container_history": container_history,
            "po_list": po_list,
            "comp_code": comp_code,
            "email": email,
            "selected_fy": fy,
            "production_at_list": production_at_list,   # 🟢 ADDED
            "selected_location": global_location or "",  # 🟢 ADDED
            "posting_ledgers": posting_ledgers
        }
    )


# ============================================================
# 💾 2. SAVE/CREATE ACTION (POST JSON - DEFAULT DATE INCLUDED)
# ============================================================
@router.post("/save")
async def save_container_log(
    request: Request,
    payload: ContainerLogisticsSchema,
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    comp_code = request.session.get("company_code")

    if not email or not comp_code:
        return JSONResponse({"success": False, "message": "Session Expired"}, status_code=401)

    try:
        ensure_bill_accounting_schema(db)
        subtotal = payload.ocean_cost + payload.local_cost + payload.handling + payload.detention
        tax_calculated = round((subtotal * payload.gst_percent) / 100, 2)
        grand_total = round(subtotal + tax_calculated, 2)
        shipping_vendor = db.query(vendors).filter(vendors.id == payload.shipping_line_id, vendors.company_id == comp_code).first()
        vendor_name = shipping_vendor.name if shipping_vendor else f"Shipping Vendor {payload.shipping_line_id}"

        new_entry = ContainerLog(
            company_id=comp_code,
            unit_id=request.session.get("unit_id", 0),
            po_number=(payload.po_number or "N/A").upper().strip(),
            production_at=(payload.production_at or "").upper().strip(), # 🟢 ADDED
            container_no=payload.container_no.upper().strip(),
            size=payload.size,
            vendor_id=payload.shipping_line_id,
            ocean_cost=payload.ocean_cost,
            local_cost=payload.local_cost,
            handling=payload.handling,
            detention=payload.detention,
            lended_total=grand_total,
            vessel_name="",
            date=dt.date.today(),
            status="DRAFT"
        )

        db.add(new_entry)
        db.flush()

        db.add(AuditLog(
            table_name="container_logistics", record_id=new_entry.id, company_id=comp_code,
            field_name="CREATE", old_value="NONE", new_value=new_entry.container_no,
            edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
        ))

        posting_ledger = resolve_posting_ledger(
            db,
            comp_code,
            payload.accounting_ledger_id,
            "Freight & Logistics Expense A/c",
            "Direct Expenses",
            "EXPENSE",
        )
        voucher = post_vendor_bill(
            db,
            comp_code,
            new_entry.date,
            new_entry.container_no,
            vendor_name,
            posting_ledger["ledger_name"],
            subtotal,
            tax_calculated,
            grand_total,
            f"Logistics invoice {new_entry.container_no}",
            email,
            expense_group_name=posting_ledger["group_name"],
            voucher_type="Purchase",
        )
        new_entry.journal_id = voucher.id
        new_entry.status = "POSTED"

        db.commit()
        return JSONResponse({"success": True, "message": f"Logistics container {payload.container_no} saved and posted: {voucher.voucher_no}"})

    except Exception as e:
        db.rollback()
        logger.error(f"CONTAINER SAVE ERROR: {str(e)}")
        return JSONResponse({"success": False, "message": f"Internal Server Error: {str(e)}"}, status_code=500)


# ============================================================
# 🔄 3. UPDATE ACTION (PUT JSON - TRACKED MODIFICATIONS)
# ============================================================
@router.put("/update/{log_id}")
async def update_container_log(
    log_id: int,
    request: Request,
    payload: ContainerLogisticsSchema,
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    comp_code = request.session.get("company_code")

    if not email or not comp_code:
        return JSONResponse({"success": False, "message": "Session Expired"}, status_code=401)

    try:
        entry = db.query(ContainerLog).filter(
            ContainerLog.id == log_id,
            ContainerLog.company_id == comp_code, ContainerLog.is_cancelled != True
        ).first()

        if not entry:
            return JSONResponse({"success": False, "message": "Logistics entry record not found"}, status_code=404)
        if entry.journal_id and entry.status == "POSTED":
            return JSONResponse({"success": False, "message": "Posted logistics bill cannot be edited. Cancel and re-enter to keep accounts correct."}, status_code=400)

        # 📜 Field Level Tracking
        tracked_fields = {
            "po_number": (payload.po_number or "N/A").upper().strip(),
            "production_at": (payload.production_at or "").upper().strip(), # 🟢 ADDED
            "container_no": payload.container_no.upper().strip(),
            "size": payload.size,
            "vendor_id": payload.shipping_line_id,
            "ocean_cost": payload.ocean_cost,
            "local_cost": payload.local_cost,
            "handling": payload.handling,
            "detention": payload.detention
        }

        for key, new_val in tracked_fields.items():
            old_val = str(getattr(entry, key) or "")
            check_new = str(new_val or "")
            if old_val != check_new:
                db.add(AuditLog(
                    table_name="container_logistics", record_id=entry.id, company_id=comp_code,
                    field_name=key, old_value=old_val, new_value=check_new,
                    edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
                ))
                setattr(entry, key, new_val)

        subtotal = entry.ocean_cost + entry.local_cost + entry.handling + entry.detention
        tax_calculated = round((subtotal * payload.gst_percent) / 100, 2)
        entry.lended_total = round(subtotal + tax_calculated, 2)

        db.commit()
        return JSONResponse({"success": True, "message": f"Container Logistics {entry.container_no} updated successfully!"})

    except Exception as e:
        db.rollback()
        logger.error(f"CONTAINER UPDATE ERROR: {str(e)}")
        return JSONResponse({"success": False, "message": f"Internal Server Error: {str(e)}"}, status_code=500)


# ============================================================
# 📋 4. MASTER AUDIT HISTORY LOG ENGINE (GET)
# ============================================================
@router.get("/audit_all")
async def get_all_container_logistics_audit(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    logs = (
        db.query(AuditLog, ContainerLog.container_no)
        .join(ContainerLog, AuditLog.record_id == ContainerLog.id)
        .filter(AuditLog.table_name == "container_logistics", AuditLog.company_id == comp_code)
        .order_by(AuditLog.edited_at.desc()).limit(100).all()
    )

    return [{
        "record_id": l.AuditLog.record_id,
        "timestamp": l.AuditLog.edited_at.strftime("%d-%m-%Y %H:%M:%S"),
        "user": l.AuditLog.edited_by.split('@')[0] if l.AuditLog.edited_by else "System",
        "email": l.AuditLog.edited_by if l.AuditLog.edited_by else "System",
        "batch": f"Row ID #{l.AuditLog.record_id} • Container: {l.container_no}" if l.container_no else f"Row ID #{l.AuditLog.record_id}",
        "action": f"Changed {l.AuditLog.field_name.replace('_', ' ').title()}" if l.AuditLog.field_name not in ["CREATE", "DELETE"] else l.AuditLog.field_name,
        "details": f"{l.AuditLog.old_value} ➔ {l.AuditLog.new_value}" if l.AuditLog.old_value != "NONE" else f"Created Entry Trace: {l.AuditLog.new_value}"
    } for l in logs]


# ============================================================
# 🗑️ 5. DELETE ACTION WITH TRACE AUDIT (POST JSON)
# ============================================================
@router.post("/delete/{log_id}")
def delete_container_log(
    log_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")

    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    log_entry = db.query(ContainerLog).filter(
        ContainerLog.id == log_id,
        ContainerLog.company_id == comp_code, ContainerLog.is_cancelled != True
    ).first()

    if log_entry:
        try:
            db.add(AuditLog(
                table_name="container_logistics", record_id=log_entry.id, company_id=comp_code,
                field_name="is_cancelled", old_value="False", new_value="True",
                edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
            ))
            cancel_linked_bill_voucher(db, comp_code, log_entry.journal_id, email)
            log_entry.is_cancelled = True
            log_entry.status = "CANCELLED"
            db.commit()
            return {"status": "success", "success": True, "message": "Record cancelled successfully"}
        except Exception as e:
            db.rollback()
            return JSONResponse({"success": False, "message": str(e)}, status_code=500)

    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# 📊 6. EXCEL REPORT GENERATOR
# ============================================================
@router.get("/export/excel")
def export_container_excel(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return RedirectResponse("/", status_code=302)

    logs = db.query(ContainerLog, vendors.name.label("v_name")).join(
        vendors, ContainerLog.vendor_id == vendors.id
    ).filter(ContainerLog.company_id == comp_code, ContainerLog.is_cancelled != True).order_by(ContainerLog.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logistics Ledger"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="143465", end_color="143465", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    # 🟢 ADDED "Production At" in Excel Headers
    headers = ["Sl No", "PO Number", "Production At", "Container No", "Size", "Vendor Line", "Ocean Freight", "Local Trans", "Handling", "Detention", "Grand Total"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, (log, v_name) in enumerate(logs, 1):
        row_data = [
            idx, log.po_number, log.production_at, log.container_no, log.size, v_name,
            log.ocean_cost, log.local_cost, log.handling, log.detention, log.lended_total
        ]
        ws.append(row_data)
        
        curr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            # 🟢 Adjusted Column Indices for Currency Format due to new column
            if col_idx in [7, 8, 9, 10, 11]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [1, 2, 3, 4, 5]:
                cell.alignment = Alignment(horizontal="center")

    last_row = ws.max_row
    total_row_idx = last_row + 1
    ws.cell(row=total_row_idx, column=1, value="Total Summary").font = total_font
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=6)
    ws.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="right")

    # 🟢 Adjusted Summation Formula Target Columns
    ws.cell(row=total_row_idx, column=7, value=f"=SUM(G2:G{last_row})").number_format = '#,##0.00'
    ws.cell(row=total_row_idx, column=11, value=f"=SUM(K2:K{last_row})").number_format = '#,##0.00'

    for col in [7, 11]:
        c = ws.cell(row=total_row_idx, column=col)
        c.font = total_font
        c.alignment = Alignment(horizontal="right")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"Logistics_Ledger_{ist_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
