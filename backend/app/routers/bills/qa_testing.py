from app.utils.timezone import ist_now
from fastapi import APIRouter, Request, Depends, Body, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text, desc, and_
from datetime import date, datetime
import datetime as dt
import logging
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from app.database import get_db
from app.database.models.bills import QATestingLog
from app.database.models.processing import AuditLog  
from app.database.models.criteria import production_at, varieties, vendors
from app.services.bill_accounting import (
    cancel_linked_bill_voucher,
    ensure_bill_accounting_schema,
    list_posting_ledgers,
    post_vendor_bill,
    resolve_posting_ledger,
)

router = APIRouter(
    prefix="/qa",
    tags=["QA Testing"]
)

templates = Jinja2Templates(directory="app/templates")
logger = logging.getLogger(__name__)


# ============================================================
# 📋 PYDANTIC SCHEMAS FOR JSON PAYLOADS (AJAX COMPATIBLE)
# ============================================================
class QaTestingSchema(BaseModel):
    test_date: date
    production_at_id: int
    product_name: str
    batch_no: str
    lab_id: int
    parameters: str = ""
    report_ref: str
    base_cost: float
    gst_percent: float
    grand_total: float
    accounting_ledger_id: int | None = None


# ============================================================
# 🧪 1. MAIN ENTRY PAGE (GET) - FIXED NATIVE DATE FILTERS
# ============================================================
@router.get("/entry", response_class=HTMLResponse)
def qa_entry_page(
    request: Request,
    fy: str = Query(None), # Financial Year Query Param
    db: Session = Depends(get_db)
):
    email = request.session.get("email")
    comp_code = request.session.get("company_code")

    if not email or not comp_code:
        return RedirectResponse("/", status_code=302)

    ensure_bill_accounting_schema(db)
    today = ist_now().date()
    default_fy = today.year if today.month >= 4 else today.year - 1
    try:
        selected_year = int(fy) if fy is not None and fy != "" else default_fy
    except (TypeError, ValueError):
        selected_year = default_fy
    selected_fy = "" if fy == "" else str(selected_year)

    # 1. Products and production locations fetch safely
    products_list = []
    production_at_list = []
    try:
        products_list = db.query(varieties).filter(varieties.company_id == comp_code).order_by(varieties.variety_name).all()
    except Exception as e:
        logger.error(f"Products Table Error: {e}")
    try:
        production_at_list = db.query(production_at).filter(production_at.company_id == comp_code).order_by(production_at.production_at).all()
    except Exception as e:
        logger.error(f"Production At Table Error: {e}")

    # 2. External Lab Facility lookup: Inhouse + vendors whose Service For contains "Lab".
    labs_list = [{"id": 0, "name": "Inhouse"}]
    existing_lab_names = {"inhouse"}
    try:
        vendor_lab_rows = (
            db.query(vendors)
            .filter(
                vendors.company_id == comp_code,
                vendors.service_for.ilike("%Lab%"),
            )
            .order_by(vendors.name)
            .all()
        )
        for vendor in vendor_lab_rows:
            lab_name = (vendor.name or "").strip()
            if lab_name and lab_name.lower() not in existing_lab_names:
                labs_list.append({"id": -int(vendor.id), "name": lab_name})
                existing_lab_names.add(lab_name.lower())
    except Exception as e:
        logger.error(f"Vendor lab lookup error: {e}")

    posting_ledgers = list_posting_ledgers(
        db,
        comp_code,
        group_types={"EXPENSE"},
        group_names={"Direct Expenses", "Indirect Expenses"},
    )

    # 3. History Data (Filtered by Financial Year April to March)
    qa_history = []
    start_date = dt.date(selected_year, 4, 1)
    end_date = dt.date(selected_year + 1, 3, 31)

    try:
        if fy == "":
            raw_history = []
        else:
            raw_history = (
                db.query(QATestingLog, production_at.production_at.label("production_at_name"))
                .join(production_at, QATestingLog.unit_id == production_at.id)
                .filter(
                    production_at.company_id == comp_code,
                    QATestingLog.test_date >= start_date,
                    QATestingLog.test_date <= end_date
                )
                .order_by(desc(QATestingLog.id))
                .all()
            )

        qa_history = [{
            "id": row.QATestingLog.id,
            "batch_no": row.QATestingLog.batch_no,
            "lab_name": row.QATestingLog.lab_name,
            "parameters": getattr(row.QATestingLog, "parameters", "") or "",
            "total": row.QATestingLog.test_cost,
            "report_ref": row.QATestingLog.report_ref,
            "production_at": row.production_at_name,
            "product": row.QATestingLog.product_name or "-",
            "date": row.QATestingLog.test_date.strftime("%Y-%m-%d") if row.QATestingLog.test_date else today.strftime("%Y-%m-%d"),
            "base_cost": row.QATestingLog.test_cost,
            "is_cancelled": bool(row.QATestingLog.is_cancelled),
            "gst_amt": 0.0,
            "status": row.QATestingLog.status,
            "journal_id": row.QATestingLog.journal_id
        } for row in raw_history]
    except Exception as e:
        logger.error(f"History Fetch Error: {e}")
        qa_history = []

    return templates.TemplateResponse(
        request=request,
        name="bills/qa_testing_entry.html",
        context={
            "products": products_list,
            "production_at_list": production_at_list,
            "labs": labs_list,
            "qa_history": qa_history,
            "comp_code": comp_code,
            "email": email,
            "selected_fy": selected_fy,
            "posting_ledgers": posting_ledgers
        }
    )


# ============================================================
# 💾 2. SAVE QA TESTING RECORD (POST JSON - AUTOMATIC DATE)
# ============================================================
@router.post("/save")
async def save_qa_testing(
    request: Request,
    payload: QaTestingSchema,
    db: Session = Depends(get_db)
):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code or not email:
        return JSONResponse({"success": False, "message": "Session Expired"}, status_code=401)

    try:
        ensure_bill_accounting_schema(db)
        try:
            db.execute(text("ALTER TABLE qa_testing_logs ADD COLUMN IF NOT EXISTS parameters TEXT"))
            db.execute(text("ALTER TABLE qa_testing_logs ADD COLUMN IF NOT EXISTS product_name VARCHAR(150)"))
            db.flush()
        except Exception as exc:
            db.rollback()
            logger.warning("QA parameters schema check failed: %s", exc)
            ensure_bill_accounting_schema(db)

        # lookup lab name from labs table or vendor master service-lab lookup.
        lab_name = "Inhouse"
        if payload.lab_id and payload.lab_id > 0:
            try:
                lab_row = db.execute(
                    text("SELECT name FROM labs WHERE id = :l_id"),
                    {"l_id": payload.lab_id}
                ).fetchone()
                if lab_row:
                    lab_name = lab_row[0]
            except Exception as e:
                logger.error(f"Lab lookup implicit warning: {e}")
        elif payload.lab_id and payload.lab_id < 0:
            try:
                vendor_lab = (
                    db.query(vendors)
                    .filter(
                        vendors.id == abs(payload.lab_id),
                        vendors.company_id == comp_code,
                    )
                    .first()
                )
                if vendor_lab and vendor_lab.name:
                    lab_name = vendor_lab.name
            except Exception as e:
                logger.error(f"Vendor lab lookup implicit warning: {e}")

        clean_product = ", ".join(
            item.strip().upper()
            for item in (payload.product_name or "").split(",")
            if item.strip()
        )
        if not clean_product:
            return JSONResponse({"success": False, "message": "Select at least one product"}, status_code=400)
        for product_item in [item.strip() for item in clean_product.split(",") if item.strip()]:
            existing_product = db.query(varieties).filter(
                varieties.company_id == comp_code,
                varieties.variety_name == product_item,
            ).first()
            if not existing_product:
                db.add(varieties(
                    company_id=comp_code,
                    variety_name=product_item,
                    email=email,
                    date=payload.test_date,
                    time=dt.datetime.now().strftime("%H:%M:%S"),
                ))
                db.flush()

        #  payload   'test_date' ‌   📅
        new_entry = QATestingLog(
            unit_id=payload.production_at_id,
            product_name=clean_product,
            batch_no=payload.batch_no.upper().strip(),
            po_number=payload.batch_no.upper().strip(),
            lab_name=lab_name,
            test_cost=payload.grand_total,
            report_ref=payload.report_ref.upper().strip(),
            test_date=payload.test_date if payload.test_date else dt.date.today(),
            status="DRAFT"
        )
        setattr(new_entry, "parameters", payload.parameters.strip())
        
        db.add(new_entry)
        db.flush()

        # 📜 Add Initial Master Operational Audit Entry Trace
        db.add(AuditLog(
            table_name="qa_testing_logs", record_id=new_entry.id, company_id=comp_code,
            field_name="CREATE", old_value="NONE", new_value=f"Batch: {new_entry.batch_no} (Cost: ₹{payload.grand_total})",
            edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
        ))

        taxable_value = round(float(payload.base_cost or 0.0), 2)
        gst_amount = round(float(payload.grand_total or 0.0) - taxable_value, 2)
        posting_ledger = resolve_posting_ledger(
            db,
            comp_code,
            payload.accounting_ledger_id,
            "QA Testing Expense A/c",
            "Direct Expenses",
            "EXPENSE",
        )
        voucher = post_vendor_bill(
            db,
            comp_code,
            new_entry.test_date,
            new_entry.report_ref or new_entry.batch_no,
            lab_name,
            posting_ledger["ledger_name"],
            taxable_value,
            gst_amount,
            payload.grand_total,
            f"QA testing bill {new_entry.report_ref or new_entry.batch_no}",
            email,
            expense_group_name=posting_ledger["group_name"],
            voucher_type="Purchase",
        )
        new_entry.journal_id = voucher.id
        new_entry.status = "POSTED"

        db.commit()
        return JSONResponse({"success": True, "message": f"QA Lab entry {payload.batch_no} saved and posted: {voucher.voucher_no}"})

    except Exception as e:
        db.rollback()
        logger.error(f"QA Save Error: {e}")
        return JSONResponse({"success": False, "message": f"Database Error: {str(e)}"}, status_code=500)


# ============================================================
# 📋 3. MASTER AUDIT HISTORY LOG ENGINE (GET)
# ============================================================
@router.get("/audit_all")
async def get_all_qa_audit(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    logs = (
        db.query(AuditLog, QATestingLog.batch_no)
        .join(QATestingLog, AuditLog.record_id == QATestingLog.id)
        .join(production_at, QATestingLog.unit_id == production_at.id)
        .filter(AuditLog.table_name == "qa_testing_logs", production_at.company_id == comp_code)
        .order_by(AuditLog.edited_at.desc()).limit(100).all()
    )

    return [{
        "record_id": l.AuditLog.record_id,
        "timestamp": l.AuditLog.edited_at.strftime("%d-%m-%Y %H:%M:%S"),
        "user": l.AuditLog.edited_by.split('@')[0] if l.AuditLog.edited_by else "System",
        "email": l.AuditLog.edited_by if l.AuditLog.edited_by else "System",
        "batch": f"Row ID #{l.AuditLog.record_id} • Batch: {l.batch_no}" if l.batch_no else f"Row ID #{l.AuditLog.record_id}",
        "action": l.AuditLog.field_name,
        "details": l.AuditLog.new_value if l.AuditLog.old_value == "NONE" else f"{l.AuditLog.old_value} ➔ {l.AuditLog.new_value}"
    } for l in logs]


# ============================================================
# 🗑️ 4. SECURE DELETE ACTION WITH TRACE AUDIT (POST)
# ============================================================
@router.post("/delete/{expense_id}")
def delete_qa_log(expense_id: int, request: Request, db: Session = Depends(get_db)):
    company_code = request.session.get("company_code")
    email = request.session.get("email")
    if not company_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    
    entry = db.query(QATestingLog).join(
        production_at, QATestingLog.unit_id == production_at.id
    ).filter(
        QATestingLog.id == expense_id,
        production_at.company_id == company_code
    ).first()

    if entry:
        try:
            db.add(AuditLog(
                table_name="qa_testing_logs", record_id=entry.id, company_id=company_code,
                field_name="is_cancelled", old_value="False", new_value="True",
                edited_by=email, edited_at=dt.datetime.now(dt.timezone.utc)
            ))
            cancel_linked_bill_voucher(db, company_code, entry.journal_id, email)
            entry.is_cancelled = True
            entry.status = "CANCELLED"
            db.commit()
            return {"success": True, "message": "QA testing charges cancelled successfully!"}
        except Exception as e:
            db.rollback()
            return JSONResponse({"success": False, "message": str(e)}, status_code=500)
    
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# 📊 5. GLOBAL MASTER EXCEL EXPORT LEDGER (GET)
# ============================================================
@router.get("/export/excel")
def export_qa_excel(request: Request, db: Session = Depends(get_db)):
    company_code = request.session.get("company_code")
    if not company_code:
        return RedirectResponse("/", status_code=302)

    history = (
        db.query(QATestingLog, production_at.production_at.label("production_at_name"))
        .join(production_at, QATestingLog.unit_id == production_at.id)
        .filter(production_at.company_id == company_code)
        .order_by(QATestingLog.id.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QA Testing Ledger"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="143465", end_color="143465", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    headers = ["Sl No", "Batch / PO Number", "Product", "Production At", "Parameters", "Lab Name Facility", "Report Ref", "Grand Total Cost"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, log in enumerate(history, 1):
        row_data = [
            idx, log.QATestingLog.batch_no, log.QATestingLog.product_name or "-", log.production_at_name,
            getattr(log.QATestingLog, "parameters", "") or "-", log.QATestingLog.lab_name,
            log.QATestingLog.report_ref, log.QATestingLog.test_cost
        ]
        ws.append(row_data)
        
        curr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 8:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [1, 2, 7]:
                cell.alignment = Alignment(horizontal="center")

    last_row = ws.max_row
    total_row_idx = last_row + 1
    ws.cell(row=total_row_idx, column=1, value="Total Summary").font = total_font
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=5)
    ws.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="right")

    ws.cell(row=total_row_idx, column=6, value=f"=SUM(F2:F{last_row})").number_format = '#,##0.00'
    ws.cell(row=total_row_idx, column=6).font = total_font
    ws.cell(row=total_row_idx, column=6).alignment = Alignment(horizontal="right")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"QA_Testing_Ledger_{ist_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
