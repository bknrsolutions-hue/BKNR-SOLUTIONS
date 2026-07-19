# ============================================================================
# STOCK REPORT ROUTER – FINAL (WITH FY FILTERS & EXPORTS)
# ============================================================================

from fastapi import APIRouter, Request, Depends, Body, HTTPException, Query
from fastapi.encoders import jsonable_encoder
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func, case, and_
from datetime import datetime, date, timezone
from zoneinfo import ZoneInfo
from app.utils.timezone import ist_now
from app.utils.global_filters import get_global_filters
from app.services.cache import cache_get_or_set, invalidate_company_cache

import openpyxl
from io import BytesIO
from app.services.pdf_renderer import render_pdf_from_html
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.database import get_db
from app.database.models.inventory_management import stock_entry
from app.database.models.users import Company
from app.database.models.processing import AuditLog, GateEntry
from app.database.models.criteria import (
    production_for as production_for_model, production_at, freezers, packing_styles, production_types,
    glazes, varieties, grades, brands, species as species_model
)

router = APIRouter(prefix="/stock_report", tags=["STOCK REPORT"])


def format_audit_timestamp(value):
    if not value:
        return "-"
    utc_value = value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value
    return utc_value.astimezone(ZoneInfo("Asia/Kolkata")).strftime("%d-%m-%Y %H:%M:%S")


def signed_stock_movement(column):
    movement = case((stock_entry.cargo_movement_type == "IN", column), else_=-column)
    return case((stock_entry.is_cancelled == True, -movement), else_=movement)

# ------------------------------------------------------------
# HELPER: GET COMPANY INFO
# ------------------------------------------------------------
def get_company_info(db: Session, comp_code: str):
    c = db.query(Company).filter(Company.company_code == comp_code).first()
    return (c.company_name or "", c.address or "") if c else ("", "")


def get_company_mpeda_code(db: Session, comp_code: str):
    c = db.query(Company).filter(Company.company_code == comp_code).first()
    return c.mpeda_registration_code if c and c.mpeda_registration_code else ""

# ------------------------------------------------------------
# HELPER: CHECK STOCK AVAILABILITY
# ------------------------------------------------------------
def check_stock_availability(db: Session, comp_code: str, batch: str, location: str, grade: str, packing: str, exclude_id: int = None):
    q = db.query(
        func.sum(signed_stock_movement(stock_entry.no_of_mc)).label("bal_mc"),
        func.sum(signed_stock_movement(stock_entry.loose)).label("bal_ls")
    ).filter(
        stock_entry.company_id == comp_code,
        stock_entry.batch_number == batch,
        stock_entry.location == location,
        stock_entry.grade == grade,
        stock_entry.packing_style == packing
    )
    if exclude_id:
        q = q.filter(stock_entry.id != exclude_id)
    res = q.first()
    return int(res.bal_mc or 0), int(res.bal_ls or 0)

# ------------------------------------------------------------
# STOCK REPORT PAGE (GET) - WITH UNIVERSAL FILTERS LAYER
# ------------------------------------------------------------
@router.get("", response_class=HTMLResponse)
async def stock_report_page(
    request: Request,
    db: Session = Depends(get_db),
    from_date: str = "",
    to_date: str = "",
    fy: str = Query(None),
    production_for: str | None = Query(None),
    location: str | None = Query(None),
):
    # 🟢 FETCH ACTIVE UNIVERSAL FILTERS FROM CONTEXT LAYER
    cookie_production_for, cookie_location = get_global_filters(request)
    global_production_for = production_for if production_for is not None else cookie_production_for
    global_location = location if location is not None else cookie_location

    email = request.session.get("email")
    comp_code = request.session.get("company_code")
    role = request.session.get("role")

    if not email or not comp_code:
        return RedirectResponse("/auth/login", status_code=302)

    def build_stock_report_masters():
        all_dates = db.query(GateEntry.date).filter(GateEntry.company_id == comp_code, GateEntry.date != None).all()
        fy_set = set()
        for d_tuple in all_dates:
            d = d_tuple[0]
            current_year = d.year
            fy_set.add(f"{current_year}" if d.month >= 4 else f"{current_year - 1}")

        def get_list(model, attr):
            return [getattr(x, attr) for x in db.query(model).filter(model.company_id == comp_code).all()]

        try:
            p_types_data = db.query(production_types).filter(production_types.company_id == comp_code).all()
        except Exception:
            p_types_data = db.query(production_types).all()

        prod_types_list = []
        for x in p_types_data:
            val = getattr(x, "type_name", None) or getattr(x, "production_type", None) or getattr(x, "type_of_production", None)
            if val:
                prod_types_list.append(val)
        prod_types_list = sorted(list(set(prod_types_list))) or ["PROCESSED", "SEMIPROCESSED", "RAW"]

        company_name, company_address = get_company_info(db, comp_code)
        return {
            "financial_years": sorted(list(fy_set), reverse=True),
            "species_list": get_list(species_model, "species_name"),
            "brands_list": get_list(brands, "brand_name"),
            "production_for_list": sorted({x.production_for for x in db.query(production_for_model).filter(production_for_model.company_id == comp_code).all() if x.production_for}),
            "type_of_production_list": prod_types_list,
            "production_at_list": get_list(production_at, "production_at"),
            "freezers_list": get_list(freezers, "freezer_name"),
            "packing_styles_list": get_list(packing_styles, "packing_style"),
            "glazes_list": get_list(glazes, "glaze_name"),
            "varieties_list": get_list(varieties, "variety_name"),
            "grades_list": get_list(grades, "grade_name"),
            "company_name": company_name,
            "company_address": company_address,
        }

    master_context = cache_get_or_set(f"bknr:inventory_report:{comp_code}:stock_report_masters", build_stock_report_masters, ttl=300)
    financial_years = master_context["financial_years"]

    # --- Financial Year Logic ---
    selected_fy = fy if fy is not None else ""

    #       LEFT OUTER JOIN
    q = db.query(stock_entry).outerjoin(
        GateEntry,
        and_(
            stock_entry.batch_number == GateEntry.batch_number,
            stock_entry.company_id == GateEntry.company_id,
        ),
    ).filter(
        stock_entry.company_id == comp_code
    )

    # 🟢 INJECT ACTIVE UNIVERSAL FILTERS ON BASE QUERY POOL
    if global_production_for:
        q = q.filter(func.trim(stock_entry.production_for) == func.trim(global_production_for))
    if global_location:
        q = q.filter(func.trim(stock_entry.production_at) == func.trim(global_location))

    if selected_fy:
        start_year = int(selected_fy)
        fy_start = date(start_year, 4, 1)
        fy_end = date(start_year + 1, 3, 31)

        #      ,
        q = q.filter(
            case(
                (GateEntry.date != None, and_(GateEntry.date >= fy_start, GateEntry.date <= fy_end)),
                else_=and_(stock_entry.date >= fy_start, stock_entry.date <= fy_end)
            )
        )
        if not from_date: from_date = fy_start.isoformat()
        if not to_date: to_date = fy_end.isoformat()
    else:
        if from_date: q = q.filter(stock_entry.date >= date.fromisoformat(from_date))
        if to_date: q = q.filter(stock_entry.date <= date.fromisoformat(to_date))

    # If no FY and no dates, load all records by default
    rows = q.order_by(stock_entry.date.desc(), stock_entry.time.desc()).all()

    context = {
        "request": request,
        "rows": rows,
        "from_date": from_date,
        "to_date": to_date,
        "financial_years": financial_years,
        "selected_fy": selected_fy,
        "selected_production_for": global_production_for, # 🟢 For Dropdown State Memory Lock
        "selected_location": global_location,             # 🟢 For Dropdown State Memory Lock
        "is_admin": role in {"admin", "Admin", "super_admin", "Super Admin"} or email == "bknr.solutions@gmail.com"
    }
    context.update(master_context)

    if request.query_params.get("format") == "json":
        json_context = {key: value for key, value in context.items() if key != "request"}
        json_context["rows"] = [
            {column.name: getattr(row, column.name) for column in row.__table__.columns}
            for row in rows
        ]
        return JSONResponse(content=jsonable_encoder(json_context))

    return request.app.state.templates.TemplateResponse(
        request=request,
        name="inventory_management/stock_report.html",
        context={k: v for k, v in context.items()}
    )

# ------------------------------------------------------------
# UPDATE STOCK (WITH VALIDATION & AUDIT) - TRANSACTIONAL
# ------------------------------------------------------------
@router.post("/update")
async def update_stock(request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    user_email = request.session.get("email")
    role = request.session.get("role")
    email = request.session.get("email")
    if role not in {"admin", "Admin", "super_admin", "Super Admin"} and email != "bknr.solutions@gmail.com":
        raise HTTPException(status_code=403)

    row = db.query(stock_entry).filter(stock_entry.id == payload.get("id"), stock_entry.company_id == comp_code).first()
    if not row: raise HTTPException(status_code=404)

    new_mc = int(payload.get("no_of_mc", row.no_of_mc))
    new_ls = int(payload.get("loose", row.loose))
    new_batch = payload.get("batch_number", row.batch_number)
    new_loc = payload.get("location", row.location)
    new_grade = payload.get("grade", row.grade)
    new_pack_name = payload.get("packing_style", row.packing_style)

    if row.cargo_movement_type == "OUT":
        bal_mc, bal_ls = check_stock_availability(db, comp_code, new_batch, new_loc, new_grade, new_pack_name, exclude_id=row.id)
        if new_mc > bal_mc or new_ls > bal_ls:
            raise HTTPException(status_code=400, detail=f"Insufficient Stock! Available: {bal_mc} MC, {bal_ls} Lse")

    fields = ["batch_number", "location", "brand", "freezer", "glaze", "species", "packing_style", "variety", "grade", "no_of_mc", "loose", "product_kg_value", "sales_reference_rate", "hlso_count", "hoso_count", "purpose", "type_of_production"]
    for f in fields:
        if f in payload:
            old_v, new_v = str(getattr(row, f) or "").strip(), str(payload[f] or "").strip()
            if old_v != new_v:
                db.add(AuditLog(
                    table_name="stock_entry", record_id=row.id, company_id=comp_code,
                    field_name=f, old_value=old_v, new_value=new_v,
                    edited_by=user_email, edited_at=datetime.utcnow()
                ))
                setattr(row, f, payload[f])

    pack = db.query(packing_styles).filter(packing_styles.company_id == comp_code, packing_styles.packing_style == row.packing_style).first()
    if pack:
        row.quantity = (float(row.no_of_mc or 0) * float(pack.mc_weight or 0)) + (float(row.loose or 0) * float(pack.slab_weight or 0))
        row.inventory_value = round(row.quantity * float(row.product_kg_value or 0), 2)

    db.commit()
    invalidate_company_cache(comp_code, "inventory_report")
    invalidate_company_cache(comp_code, "inventory_dashboard")
    invalidate_company_cache(comp_code, "costing_dashboard")
    return {"status": "success"}

# ------------------------------------------------------------
# EXPORT EXCEL (WITH FULL FILTERS & GLOBAL LAYERS)
# ------------------------------------------------------------
@router.get("/export_xlsx")
def export_xlsx(
    request: Request, db: Session = Depends(get_db),
    from_date: str = "", to_date: str = "", type: str = "",
    batch: str = "", brand: str = "", species: str = "",
    variety: str = "", location: str = ""
):
    comp_code = request.session.get("company_code")

    # 🟢 FIX: Extract with unique names to bypass explicit route variable clash
    global_production_for, global_location = get_global_filters(request)

    q = db.query(stock_entry).filter(
        stock_entry.company_id == comp_code
    )

    # 🟢 Layer universal bindings safely onto workbook generation parameters pool
    if global_production_for:
        q = q.filter(func.trim(stock_entry.production_for) == func.trim(global_production_for))
    if global_location:
        q = q.filter(func.trim(stock_entry.production_at) == func.trim(global_location))

    if from_date: q = q.filter(stock_entry.date >= date.fromisoformat(from_date))
    if to_date: q = q.filter(stock_entry.date <= date.fromisoformat(to_date))
    if type: q = q.filter(stock_entry.cargo_movement_type == type)
    if batch: q = q.filter(stock_entry.batch_number == batch)
    if brand: q = q.filter(stock_entry.brand == brand)
    if species: q = q.filter(stock_entry.species == species)
    if variety: q = q.filter(stock_entry.variety == variety)
    if location: q = q.filter(stock_entry.location == location)

    rows = q.order_by(stock_entry.date.asc(), stock_entry.time.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Ledger"

    headers = ["Date", "Batch #", "Type", "Brand", "Species", "Variety", "Grade", "Glaze", "Freezer", "Pack Style", "Location", "PO #", "Prod For", "Prod At", "HLSO", "HOSO", "MC", "Lse", "Qty", "Cost", "Value", "Sales Rate", "User"]
    ws.append(headers)

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        s = -1 if r.cargo_movement_type == "OUT" else 1
        if r.is_cancelled:
            s *= -1
        ws.append([
            str(r.date), r.batch_number, r.cargo_movement_type, r.brand, r.species, r.variety, r.grade, r.glaze, r.freezer, r.packing_style, r.location, r.po_number or "", r.production_for or "", r.production_at or "", r.hlso_count or 0, r.hoso_count or 0,
            s * int(r.no_of_mc or 0), s * int(r.loose or 0), s * float(r.quantity or 0), float(r.product_kg_value or 0), float(r.inventory_value or 0), float(r.sales_reference_rate or 0), r.email.split('@')[0] if r.email else ""
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=Stock_Ledger.xlsx"})

# ------------------------------------------------------------
# EXPORT PDF (WITH FULL FILTERS & GLOBAL LAYERS)
# ------------------------------------------------------------
@router.get("/export_pdf")
def export_pdf(
    request: Request, db: Session = Depends(get_db),
    from_date: str = "", to_date: str = "", type: str = "",
    batch: str = "", brand: str = "", species: str = "",
    variety: str = "", location: str = "",
    download: bool = Query(False)
):
    comp_code = request.session.get("company_code")

    # 🟢 FIX: Extract with unique names to bypass explicit route variable clash
    global_production_for, global_location = get_global_filters(request)

    q = db.query(stock_entry).filter(
        stock_entry.company_id == comp_code
    )

    # 🟢 Layer universal bindings safely onto PDF render parameter pipeline
    if global_production_for:
        q = q.filter(func.trim(stock_entry.production_for) == func.trim(global_production_for))
    if global_location:
        q = q.filter(func.trim(stock_entry.production_at) == func.trim(global_location))

    if from_date: q = q.filter(stock_entry.date >= date.fromisoformat(from_date))
    if to_date: q = q.filter(stock_entry.date <= date.fromisoformat(to_date))
    if type: q = q.filter(stock_entry.cargo_movement_type == type)
    if batch: q = q.filter(stock_entry.batch_number == batch)
    if brand: q = q.filter(stock_entry.brand == brand)
    if species: q = q.filter(stock_entry.species == species)
    if variety: q = q.filter(stock_entry.variety == variety)
    if location: q = q.filter(stock_entry.location == location)

    rows = q.order_by(stock_entry.date.asc(), stock_entry.time.asc()).all()
    company_name, company_address = get_company_info(db, comp_code)

    html = request.app.state.templates.get_template("inventory_management/stock_report_print.html").render({
        "request": request, "rows": rows, "company_name": company_name,
        "company_address": company_address,
        "mpeda_registration_code": get_company_mpeda_code(db, comp_code),
        "printed_on": ist_now().strftime("%d-%m-%Y %H:%M:%S")
    })

    pdf_file = render_pdf_from_html(html)
    disposition = "attachment" if download else "inline"

    return StreamingResponse(
        BytesIO(pdf_file),
        media_type="application/pdf",
        headers={"Content-Disposition": f"{disposition}; filename=Stock_Report.pdf"}
    )

# ------------------------------------------------------------
# AUDIT LOGS FETCH
# ------------------------------------------------------------
@router.get("/audit_all")
async def get_stock_audits(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    logs = (db.query(AuditLog, stock_entry.batch_number).join(stock_entry, AuditLog.record_id == stock_entry.id)
            .filter(AuditLog.table_name == "stock_entry", AuditLog.company_id == comp_code)
            .order_by(AuditLog.edited_at.desc()).limit(100).all())
    return JSONResponse([{
        "record_id": l.AuditLog.record_id,
        "timestamp": format_audit_timestamp(l.AuditLog.edited_at),
        "user": l.AuditLog.edited_by.split('@')[0] if l.AuditLog.edited_by else "System",
        "email": l.AuditLog.edited_by if l.AuditLog.edited_by else "System",
        "batch": f"Row ID #{l.AuditLog.record_id} • Batch: {l.batch_number}" if l.batch_number else f"Row ID #{l.AuditLog.record_id}",
        "action": f"Changed {l.AuditLog.field_name.replace('_', ' ').title()}" if l.AuditLog.field_name != "DELETE" else "Deleted Record",
        "details": f"{l.AuditLog.old_value} ➔ {l.AuditLog.new_value}"
    } for l in logs])

# ------------------------------------------------------------
# DELETE RECORD - TRANSACTIONAL
# ------------------------------------------------------------
@router.post("/delete")
async def delete_stock(request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if request.session.get("role") != "admin": raise HTTPException(status_code=403)
    row = db.query(stock_entry).filter(stock_entry.id == payload["id"], stock_entry.company_id == comp_code).first()
    if row:
        db.add(AuditLog(table_name="stock_entry", record_id=row.id, company_id=comp_code, field_name="DELETE", old_value="Exist", new_value="Deleted", edited_by=request.session.get("email"), edited_at=datetime.utcnow()))
        db.delete(row)
        db.commit()
        invalidate_company_cache(comp_code, "inventory_report")
        invalidate_company_cache(comp_code, "inventory_dashboard")
        invalidate_company_cache(comp_code, "costing_dashboard")
    return {"status": "success"}
