# app/routers/dashboard/hr_dashboard.py

from fastapi import APIRouter, Request, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_
from datetime import datetime, date, timedelta
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from app.database import get_db
from app.database.models.attendance import DailyAttendance, EmployeeRegistration, EmployeeIncrement, EmployeeSalaryAdvance
from app.database.models.processing import AuditLog

router = APIRouter(
    prefix="/hr-dashboard",
    tags=["HR DASHBOARD ANALYTICS"]
)

templates = Jinja2Templates(directory="app/templates")

# ============================================================
# 📅 1. DASHBOARD MAIN ROUTE (GET)
# ============================================================
@router.get("", response_class=HTMLResponse)
def hr_dashboard_page(request: Request, db: Session = Depends(get_db)):
    email = request.session.get("email")
    company_code = request.session.get("company_code")

    if not email or not company_code:
        return RedirectResponse("/", status_code=302)

    today = date.today()

    # 1️⃣ 🧮 TOP LEVEL METRICS
    total_staff = db.query(EmployeeRegistration).filter(
        EmployeeRegistration.company_id == company_code,
        EmployeeRegistration.status == "ACTIVE"
    ).count()

    today_present = db.query(DailyAttendance.employee_id).distinct().join(
        EmployeeRegistration, 
        and_(
            DailyAttendance.employee_id == EmployeeRegistration.employee_id,
            DailyAttendance.company_id == EmployeeRegistration.company_id
        )
    ).filter(
        EmployeeRegistration.company_id == company_code,
        DailyAttendance.duty_date == today,
        DailyAttendance.status != "ABSENT"
    ).count()

    inside_gate = db.query(DailyAttendance.employee_id).distinct().join(
        EmployeeRegistration,
        and_(
            DailyAttendance.employee_id == EmployeeRegistration.employee_id,
            DailyAttendance.company_id == EmployeeRegistration.company_id
        )
    ).filter(
        EmployeeRegistration.company_id == company_code,
        DailyAttendance.duty_date == today,
        DailyAttendance.status == "OPEN"
    ).count()

    double_duty_alerts = db.query(DailyAttendance).join(
        EmployeeRegistration,
        and_(
            DailyAttendance.employee_id == EmployeeRegistration.employee_id,
            DailyAttendance.company_id == EmployeeRegistration.company_id
        )
    ).filter(
        EmployeeRegistration.company_id == company_code,
        DailyAttendance.duty_date == today,
        DailyAttendance.working_hours >= 14.0
    ).count()

    # 2️⃣ 💵 FINANCIAL & OPERATIONAL COUNTERS
    total_payroll_gross = db.query(func.sum(EmployeeRegistration.current_salary)).filter(
        EmployeeRegistration.company_id == company_code,
        EmployeeRegistration.status == "ACTIVE"
    ).scalar() or 0.0

    total_advance_running = db.query(func.sum(EmployeeSalaryAdvance.remaining_balance)).filter(
        EmployeeSalaryAdvance.company_id == company_code,
        EmployeeSalaryAdvance.status == "APPROVED"
    ).scalar() or 0.0

    # 3️⃣ 📊 DATA MATRICES & PANELS
    dept_rows = db.query(
        EmployeeRegistration.department,
        func.count(DailyAttendance.id).label("on_floor_count")
    ).join(
        DailyAttendance, 
        and_(
            EmployeeRegistration.employee_id == DailyAttendance.employee_id,
            EmployeeRegistration.company_id == DailyAttendance.company_id
        )
    ).filter(
        EmployeeRegistration.company_id == company_code,
        DailyAttendance.duty_date == today,
        DailyAttendance.status == "OPEN"
    ).group_by(EmployeeRegistration.department).all()

    live_feed_rows = db.query(DailyAttendance).join(
        EmployeeRegistration,
        and_(
            DailyAttendance.employee_id == EmployeeRegistration.employee_id,
            DailyAttendance.company_id == EmployeeRegistration.company_id
        )
    ).filter(
        EmployeeRegistration.company_id == company_code,
        DailyAttendance.duty_date == today
    ).order_by(DailyAttendance.first_in.desc()).limit(10).all()

    audit_logs_rows = db.query(AuditLog).filter(
        AuditLog.table_name == "daily_attendance",
        AuditLog.company_id == company_code
    ).order_by(AuditLog.edited_at.desc()).limit(10).all()

    audit_list = [{
        "timestamp": l.edited_at.strftime("%d-%b %H:%M"),
        "invoice_no": l.field_name,
        "details": l.new_value
    } for l in audit_logs_rows]

    recent_increments = db.query(EmployeeIncrement).join(
        EmployeeRegistration, EmployeeIncrement.employee_id == EmployeeRegistration.employee_id
    ).filter(
        EmployeeRegistration.company_id == company_code
    ).order_by(EmployeeIncrement.id.desc()).limit(5).all()

    active_loans = db.query(EmployeeSalaryAdvance).filter(
        EmployeeSalaryAdvance.company_id == company_code,
        EmployeeSalaryAdvance.remaining_balance > 0
    ).order_by(EmployeeSalaryAdvance.id.desc()).limit(5).all()

    all_employees = db.query(EmployeeRegistration).filter(
        EmployeeRegistration.company_id == company_code,
        EmployeeRegistration.status == "ACTIVE"
    ).order_by(EmployeeRegistration.employee_id.asc()).all()

    # 4️⃣ 📈 CHART DATA COMPUTATION (Last 6 Days Attendance Trend)
    trend_labels = []
    trend_values = []
    for i in range(5, -1, -1):
        target_date = today - timedelta(days=i)
        trend_labels.append(target_date.strftime("%d-%b"))
        
        count = db.query(DailyAttendance.employee_id).distinct().join(
            EmployeeRegistration,
            and_(
                DailyAttendance.employee_id == EmployeeRegistration.employee_id,
                DailyAttendance.company_id == EmployeeRegistration.company_id
            )
        ).filter(
            EmployeeRegistration.company_id == company_code,
            DailyAttendance.duty_date == target_date,
            DailyAttendance.status != "ABSENT"
        ).count()
        trend_values.append(count)

    # 📊 CHART DATA COMPUTATION (Department Distribution)
    dept_labels = [d.department if d.department else "GENERAL" for d in dept_rows]
    dept_values = [d.on_floor_count for d in dept_rows]

    return templates.TemplateResponse(
        request=request,
        name="attendance/hr_dashboard.html",
        context={
            "email": email,
            "company_id": company_code,
            "total_staff": total_staff,
            "today_present": today_present,
            "inside_gate": inside_gate,
            "double_duty_alerts": double_duty_alerts,
            "total_payroll": total_payroll_gross,
            "total_advance": total_advance_running,
            "dept_matrix": dept_rows,
            "live_feed": live_feed_rows,
            "audit_logs": audit_list,
            "increments": recent_increments,
            "loans": active_loans,
            "employees": all_employees,
            "dept_labels": dept_labels,
            "dept_values": dept_values,
            "attendance_labels": trend_labels,
            "attendance_values": trend_values
        }
    )


# ============================================================
# 📈 2. DASHBOARD FINANCIAL EXCEL EXPORT LEDGER (GET)
# ============================================================
@router.get("/export/financial-excel")
def export_dashboard_financial_excel(request: Request, db: Session = Depends(get_db)):
    company_id = request.session.get("company_code")
    if not company_id:
        return RedirectResponse("/", status_code=302)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HR Operational Summary"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="143465", end_color="143465", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    headers = ["Sl No", "Financial Metric Context", "Ledger Gross Balance Amount"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_payroll_gross = db.query(func.sum(EmployeeRegistration.current_salary)).filter(
        EmployeeRegistration.company_id == company_id,
        EmployeeRegistration.status == "ACTIVE"
    ).scalar() or 0.0

    total_advance_running = db.query(func.sum(EmployeeSalaryAdvance.remaining_balance)).filter(
        EmployeeSalaryAdvance.company_id == company_id,
        EmployeeSalaryAdvance.status == "APPROVED"
    ).scalar() or 0.0

    metrics_data = [
        (1, "Estimated Monthly Gross Payroll Run", total_payroll_gross),
        (2, "Outstanding Salary Advances / Loans", total_advance_running)
    ]

    for idx, title, amt in metrics_data:
        ws.append([idx, title, amt])
        curr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx == 3:
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"HR_Operational_Financial_Ledger_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )