# app/routers/attendance/daily_attendance.py

from fastapi import APIRouter, Request, Depends, Body, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_, desc, func
from sqlalchemy.orm.attributes import flag_modified
from datetime import datetime, date, timedelta
from app.utils.timezone import ist_now
import datetime as dt
import logging
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from app.database import get_db
from app.database.models.attendance import DailyAttendance, EmployeeRegistration, Shift
from app.database.models.criteria import contractors
from app.database.models.processing import AuditLog
from app.services.bill_accounting import ensure_bill_accounting_schema, post_contractor_source_charge

# 🌐 UNIVERSAL GLOBAL FILTERS HELPER
from app.utils.global_filters import get_global_filters

router = APIRouter(tags=["ATTENDANCE"])
templates = Jinja2Templates(directory="app/templates")
logger = logging.getLogger(__name__)


# ============================================================
# 📋 PYDANTIC SCHEMAS FOR JSON PAYLOADS
# ============================================================
class AttendanceEntrySchema(BaseModel):
    employee_id: str
    action: str  
    shift_name: str = "GENERAL"  
    location: str = None  # 👈 🟢 FETCH DROP FIX: Added Location to Payload


# ============================================================
# 🟢 🔴 HELPER: STRICT LOCATION CHECK
# ============================================================
def get_strict_location(request: Request):
    _, global_location = get_global_filters(request)
    session_locations = request.session.get("allowed_locations", [])
    user_allowed_locations = [loc.strip().upper() for loc in session_locations.split(",") if loc.strip()] if isinstance(session_locations, str) else [str(loc).strip().upper() for loc in session_locations if str(loc).strip()]

    actual_location = global_location.strip().upper() if global_location and global_location.upper() != "ALL" else None
    
    if not actual_location and user_allowed_locations and len(user_allowed_locations) == 1:
        actual_location = user_allowed_locations[0]

    return actual_location, user_allowed_locations


def get_shift_required_hours(db: Session, company_id: str, shift_name: str) -> float:
    shift = db.query(Shift).filter(
        Shift.company_id == company_id,
        Shift.shift_name == (shift_name or "GENERAL"),
    ).first()
    if not shift or not shift.start_time or not shift.end_time:
        return 8.0
    start_dt = datetime.combine(date.today(), shift.start_time)
    end_dt = datetime.combine(date.today(), shift.end_time)
    if end_dt < start_dt:
        end_dt += timedelta(days=1)
    hours = (end_dt - start_dt).total_seconds() / 3600.0
    break_hours = (shift.break_minutes or 0) / 60.0
    return max(1.0, hours - break_hours)


def attendance_payable_credit(working_hours: float, required_hours: float) -> float:
    raw_credit = float(working_hours or 0.0) / required_hours if required_hours > 0 else 0.0
    if raw_credit < 0.5:
        return 0.0
    if raw_credit < 1.0:
        return 0.5
    if raw_credit < 1.5:
        return 1.0
    if raw_credit < 2.0:
        return 1.5
    if raw_credit < 2.5:
        return 2.0
    if raw_credit < 3.0:
        return 2.5
    return 3.0


def calculate_duty_type_and_ot(working_hours: float) -> tuple[str, float]:
    wh = float(working_hours or 0.0)
    if wh >= 14:
        return "DOUBLE", round(wh - 16, 2) if wh > 16 else 0.0
    if wh >= 8:
        return "SINGLE", round(wh - 8, 2) if wh > 8 else 0.0
    if wh > 0:
        return "HALF", 0.0
    return "ABSENT", 0.0


def close_attendance_duty(
    duty: DailyAttendance,
    close_time: datetime,
    movement_type: str,
    pending_approval: bool = False,
) -> float:
    safe_close = close_time.replace(tzinfo=None)
    safe_first_in = duty.first_in.replace(tzinfo=None) if duty.first_in else safe_close
    wh = round(min(24.0, max(0.0, (safe_close - safe_first_in).total_seconds() / 3600)), 2)
    duty.working_hours = max(0.0, wh)
    duty.exit_time = close_time
    duty.status = "CLOSED"
    duty.duty_type, duty.calculated_ot_hours = calculate_duty_type_and_ot(duty.working_hours)
    duty.ot_status = "PENDING"
    duty.approved_ot_hours = 0.0
    if pending_approval:
        duty.duty_status = "PENDING"
        duty.duty_approved_by = None
    elif not duty.duty_status or duty.duty_status in {"OPEN", "PENDING"}:
        duty.duty_status = "APPROVED"
        duty.approved_duty_credit = duty.approved_duty_credit or 1.0

    movements = list(duty.movements) if duty.movements else []
    movements.append({
        "type": movement_type,
        "time": close_time.strftime("%H:%M"),
        "date": close_time.strftime("%Y-%m-%d"),
    })
    duty.movements = movements
    flag_modified(duty, "movements")
    return duty.working_hours


def auto_close_stale_attendance(
    db: Session,
    company_id: str,
    email: str,
    employee_id: str | None = None,
    location: str | None = None,
    allowed_locations: list[str] | None = None,
) -> list[DailyAttendance]:
    now = ist_now()
    cutoff = now.replace(tzinfo=None) - timedelta(hours=24)
    query = db.query(DailyAttendance).filter(
        DailyAttendance.company_id == company_id,
        DailyAttendance.status != "CLOSED",
        DailyAttendance.first_in != None,
        DailyAttendance.first_in <= cutoff,
    )
    if employee_id:
        query = query.filter(DailyAttendance.employee_id == employee_id)
    if location and location != "ALL":
        query = query.filter(func.upper(func.trim(DailyAttendance.production_at)) == location)
    elif allowed_locations:
        query = query.filter(func.upper(func.trim(DailyAttendance.production_at)).in_(allowed_locations))

    closed_rows = []
    for duty in query.all():
        close_time = duty.first_in + timedelta(hours=24)
        wh = close_attendance_duty(
            duty,
            close_time,
            "AUTO OUT",
            pending_approval=True,
        )
        db.add(AuditLog(
            table_name="daily_attendance",
            record_id=duty.id,
            company_id=company_id,
            field_name="AUTO_OUT_24H",
            old_value="OPEN",
            new_value=f"Emp: {duty.employee_name} ({duty.employee_id}) | Auto closed after 24 hours | {wh} Hrs | Duty approval pending",
            edited_by=email or "SYSTEM",
            edited_at=datetime.now(dt.timezone.utc),
        ))
        closed_rows.append(duty)
    if closed_rows:
        db.flush()
    return closed_rows


def contractor_gst_percent(db: Session, company_id: str, contractor_name: str) -> float:
    row = db.query(contractors).filter(
        contractors.company_id == company_id,
        contractors.contractor_name == contractor_name,
    ).first()
    return float(row.gst_percent or 0) if row else 0.0


# ============================================================
# 📅 1. PAGE LOAD (GET)
# ============================================================
@router.get("/daily", response_class=HTMLResponse)
def daily_attendance_page(
    request: Request,
    format: str = Query(default="html"),
    db: Session = Depends(get_db),
):
    email = request.session.get("email")
    company_code = request.session.get("company_code")

    if not email or not company_code:
        return RedirectResponse("/", status_code=302)

    ensure_bill_accounting_schema(db)
    actual_location, _ = get_strict_location(request)
    auto_close_stale_attendance(db, company_code, email, location=actual_location)
    db.commit()

    plant_shifts = []
    if actual_location and actual_location != "UNASSIGNED":
        plant_shifts = db.query(Shift).filter(
            Shift.company_id == company_code, 
            Shift.is_active == True,
            func.upper(func.trim(Shift.production_at)) == actual_location
        ).all()

    if format.lower() == "json":
        return JSONResponse({
            "status": "success",
            "email": email,
            "company_id": company_code,
            "actual_location": actual_location or "",
            "location_required": not bool(actual_location),
            "shifts": [
                {
                    "id": shift.id,
                    "shift_name": shift.shift_name,
                    "production_at": shift.production_at,
                    "start_time": shift.start_time.strftime("%H:%M") if shift.start_time else "",
                    "end_time": shift.end_time.strftime("%H:%M") if shift.end_time else "",
                    "break_minutes": shift.break_minutes or 0,
                    "is_night_shift": bool(shift.is_night_shift),
                }
                for shift in plant_shifts
            ],
        })

    return templates.TemplateResponse(
        request=request,
        name="attendance/daily_attendance.html",
        context={
            "email": email,
            "company_id": company_code,
            "shifts": plant_shifts,
            "actual_location": actual_location  
        }
    )


# ============================================================
# ⏱️ 2. ATTENDANCE ENTRY (IN / OUT / EXIT - POST JSON)
# ============================================================
@router.post("/entry")
async def attendance_entry(
    request: Request, 
    payload: AttendanceEntrySchema,
    db: Session = Depends(get_db)
):
    company_id = request.session.get("company_code")
    email = request.session.get("email")
    
    if not company_id or not email:
        return JSONResponse({"success": False, "error": "INVALID_SESSION"}, status_code=401)
    ensure_bill_accounting_schema(db)

    # 🟢 🔴 Explicit Location Fallback Fix
    frontend_location = payload.location.strip().upper() if payload.location else None
    backend_location, _ = get_strict_location(request)
    actual_location = frontend_location or backend_location
    
    if not actual_location:
        return JSONResponse({"success": False, "error": "GLOBAL_FILTER_REQUIRED"}, status_code=400)

    input_id = payload.employee_id.strip()
    action = payload.action.upper().strip()
    shift_name = payload.shift_name.strip() 

    now = ist_now()
    time_str = now.strftime("%H:%M")

    search_pattern = f"%{input_id.zfill(5)}"

    emp = db.query(EmployeeRegistration).filter(
        EmployeeRegistration.employee_id.like(search_pattern),
        EmployeeRegistration.company_id == company_id,
        EmployeeRegistration.status == "ACTIVE"
    ).first()

    if not emp:
        return JSONResponse({"success": False, "error": f"ID {input_id} Not Found"}, status_code=404)

    full_employee_id = emp.employee_id
    auto_closed = auto_close_stale_attendance(
        db,
        company_id,
        email,
        employee_id=full_employee_id,
        location=actual_location,
    )

    duty = db.query(DailyAttendance).filter(
        DailyAttendance.employee_id == full_employee_id,
        DailyAttendance.company_id == company_id,
        DailyAttendance.status != "CLOSED"
    ).first()

    if auto_closed and action in {"OUT", "EXIT"} and not duty:
        db.commit()
        return {
            "success": True,
            "employee_name": emp.employee_name,
            "message": "Previous duty auto-closed after 24 hours and sent for duty approval."
        }

    duty_count = db.query(DailyAttendance).filter(
        DailyAttendance.employee_id == full_employee_id,
        DailyAttendance.company_id == company_id,
        DailyAttendance.duty_date == now.date(),
        DailyAttendance.status == "CLOSED"
    ).count()

    try:
        audit_details = ""
        
        if action == "IN":
            if duty and duty.status == "OPEN":
                return JSONResponse({"success": False, "error": "ALREADY_INSIDE"}, status_code=400)
            
            if duty_count >= 2 and not duty:
                return JSONResponse({"success": False, "error": "DAILY_DUTY_LIMIT_REACHED"}, status_code=403)

            if duty: 
                movements = list(duty.movements) if duty.movements else []
                movements.append({"type": "IN", "time": time_str, "date": now.strftime("%Y-%m-%d"), "shift": shift_name})
                duty.movements = movements
                duty.status = "OPEN"
                flag_modified(duty, "movements")
                audit_details = f"Re-entry In at {time_str} [{shift_name}]"
            else: 
                new_duty = DailyAttendance(
                    company_id=company_id,
                    employee_id=full_employee_id,
                    employee_name=emp.employee_name,
                    designation=emp.designation, 
                    employee_type=emp.employee_type, 
                    production_at=actual_location,   
                    duty_date=now.date(),
                    first_in=now,
                    shift_name=shift_name,
                    movements=[{"type": "IN", "time": time_str, "date": now.strftime("%Y-%m-%d"), "shift": shift_name}],
                    status="OPEN",
                    duty_status="OPEN",
                )
                db.add(new_duty)
                audit_details = f"Fresh Shift Punch In at {time_str} (Loc: {actual_location}) [{shift_name}]"

        elif action == "OUT": 
            if not duty: 
                return JSONResponse({"success": False, "error": "NO_ACTIVE_DUTY"}, status_code=400)
            if duty.status == "AWAY": 
                return JSONResponse({"success": False, "error": "ALREADY_ON_BREAK"}, status_code=400)
            
            movements = list(duty.movements) if duty.movements else []
            movements.append({"type": "OUT", "time": time_str, "date": now.strftime("%Y-%m-%d")})
            duty.movements = movements
            duty.status = "AWAY"
            flag_modified(duty, "movements")
            audit_details = f"Break Out at {time_str}"

        elif action == "EXIT": 
            if not duty: 
                return JSONResponse({"success": False, "error": "NO_ACTIVE_DUTY"}, status_code=400)
            
            wh = close_attendance_duty(duty, now, "EXIT", pending_approval=False)
            required_hours = get_shift_required_hours(db, company_id, duty.shift_name)
            suggested_credit = attendance_payable_credit(wh, required_hours)
            duty.approved_duty_credit = suggested_credit if suggested_credit <= 1.0 else 0.0
            if suggested_credit > 1.0:
                duty.duty_status = "PENDING"
                duty.duty_approved_by = None
            else:
                duty.duty_status = "APPROVED"
                duty.duty_approved_by = "SYSTEM"
            audit_details = f"Final Shift Close at {time_str} ({wh} Hrs Worked - Duty: {duty.duty_type})"

            if (
                str(emp.employee_type or "").strip().upper() in {"CONTRACT", "CONTRACTOR"}
                and emp.contractor_name
                and not duty.journal_id
                and duty.duty_status == "APPROVED"
            ):
                required_hours = get_shift_required_hours(db, company_id, duty.shift_name)
                payable_days = attendance_payable_credit(wh, required_hours)
                per_day_rate = float(emp.current_salary or 0.0) / 26.0 if emp.current_salary else 0.0
                payable_amount = round(payable_days * per_day_rate, 2)
                if payable_amount > 0:
                    db.flush()
                    voucher = post_contractor_source_charge(
                        db=db,
                        company_id=company_id,
                        voucher_date=duty.duty_date,
                        reference_no=f"ATT-{duty.id}",
                        contractor_name=emp.contractor_name,
                        charge_type="Processing",
                        taxable_amount=payable_amount,
                        gst_percent=contractor_gst_percent(db, company_id, emp.contractor_name),
                        created_by=email,
                        quantity=payable_days,
                        rate=per_day_rate,
                    )
                    duty.journal_id = voucher.id

        db.add(AuditLog(
            table_name="daily_attendance", record_id=emp.id, company_id=company_id,
            field_name=f"PUNCH_{action}", old_value="ATTENDANCE_STREAM", 
            new_value=f"Emp: {emp.employee_name} ({full_employee_id}) | {audit_details}",
            edited_by=email, edited_at=datetime.now(dt.timezone.utc)
        ))

        db.commit()
        return {"success": True, "employee_name": emp.employee_name, "message": "Punch committed successfully"}

    except Exception as e:
        db.rollback()
        logger.error(f"ATTENDANCE POST ERROR: {str(e)}")
        return JSONResponse({"success": False, "error": f"Database processing fault: {str(e)}"}, status_code=500)


# ============================================================
# 📊 3. FETCH TODAY'S DATA
# ============================================================
@router.get("/today_all")
def today_attendance_list(request: Request, location: str = None, db: Session = Depends(get_db)):
    company_id = request.session.get("company_code")
    if not company_id: 
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    ensure_bill_accounting_schema(db)
    today = ist_now().date()

    # 🟢 🔴 Support query param location if fetch drops cookie
    backend_location, user_allowed_locations = get_strict_location(request)
    actual_location = location.strip().upper() if location else backend_location
    auto_close_stale_attendance(
        db,
        company_id,
        request.session.get("email") or "SYSTEM",
        location=actual_location,
        allowed_locations=user_allowed_locations,
    )
    db.commit()

    query = db.query(
        DailyAttendance, 
        EmployeeRegistration.department,
        EmployeeRegistration.designation
    ).join(
        EmployeeRegistration, 
        and_(
            DailyAttendance.employee_id == EmployeeRegistration.employee_id,
            DailyAttendance.company_id == EmployeeRegistration.company_id
        )
    ).filter(
        DailyAttendance.company_id == company_id,
        or_(
            DailyAttendance.duty_date == today,
            DailyAttendance.status != "CLOSED"
        )
    )

    if actual_location and actual_location != "ALL":
        query = query.filter(func.upper(func.trim(DailyAttendance.production_at)) == actual_location)
    elif user_allowed_locations:
        query = query.filter(func.upper(func.trim(DailyAttendance.production_at)).in_(user_allowed_locations))

    rows = query.order_by(DailyAttendance.first_in.desc()).all()

    results = []
    for da, dept, desg in rows:
        wh = float(da.working_hours or 0)
        duty_type = da.duty_type if da.status == "CLOSED" else "ON-DUTY"
        
        results.append({
            "id": da.id,
            "employee_id": da.employee_id,
            "employee_name": da.employee_name,
            "department": dept or "GENERAL", 
            "designation": desg or "STAFF", 
            "working_hours": wh,
            "duty_type": duty_type,
            "status": da.status,
            "shift_name": da.shift_name,
            "calculated_ot_hours": da.calculated_ot_hours or 0.0,
            "movements": da.movements or []
        })
    return results


# ============================================================
# 📜 4. AUDIT & EXPORT
# ============================================================
@router.get("/audit_all")
async def get_all_attendance_audit(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    logs = (
        db.query(AuditLog)
        .filter(AuditLog.table_name == "daily_attendance", AuditLog.company_id == comp_code)
        .order_by(AuditLog.edited_at.desc()).limit(100).all()
    )

    return [{
        "record_id": l.record_id,
        "timestamp": l.edited_at.strftime("%d-%m-%Y %H:%M:%S"),
        "user": l.edited_by.split('@')[0] if l.edited_by else "System",
        "email": l.edited_by if l.edited_by else "System",
        "batch": f"Row ID #{l.record_id} • Field: {l.field_name}" if l.field_name else f"Row ID #{l.record_id}",
        "action": "PUNCH TRANSACTION",
        "details": l.new_value
    } for l in logs]


@router.get("/export/excel")
def export_attendance_excel(request: Request, location: str = None, db: Session = Depends(get_db)):
    company_id = request.session.get("company_code")
    if not company_id:
        return RedirectResponse("/", status_code=302)

    backend_location, user_allowed_locations = get_strict_location(request)
    actual_location = location.strip().upper() if location else backend_location
    today = ist_now().date()

    query = db.query(
        DailyAttendance, 
        EmployeeRegistration.department
    ).join(
        EmployeeRegistration, 
        and_(
            DailyAttendance.employee_id == EmployeeRegistration.employee_id,
            DailyAttendance.company_id == EmployeeRegistration.company_id
        )
    ).filter(
        DailyAttendance.company_id == company_id,
        DailyAttendance.duty_date == today
    )

    if actual_location and actual_location != "ALL":
        query = query.filter(func.upper(func.trim(DailyAttendance.production_at)) == actual_location)
    elif user_allowed_locations:
        query = query.filter(func.upper(func.trim(DailyAttendance.production_at)).in_(user_allowed_locations))

    rows = query.order_by(DailyAttendance.first_in.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Today Attendance"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="143465", end_color="143465", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    headers = ["Sl No", "Employee ID", "Employee Name", "Department", "Designation", "Shift", "First In", "Status", "Working Hours", "Duty Allocation", "Calculated OT", "OT Status"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, (da, dept) in enumerate(rows, 1):
        wh = float(da.working_hours or 0)
        duty_type = da.duty_type if da.status == "CLOSED" else "ON-DUTY"
        
        row_data = [
            idx,
            da.employee_id,
            da.employee_name,
            dept or "GENERAL",
            da.designation or "STAFF",
            da.shift_name or "GENERAL",
            da.first_in.strftime("%H:%M:%S") if da.first_in else "-",
            da.status,
            wh,
            duty_type,
            da.calculated_ot_hours,
            da.ot_status
        ]
        ws.append(row_data)
        
        curr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in [9, 11]:  
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [1, 2, 6, 7, 8, 10, 12]:  
                cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    loc_str = actual_location if (actual_location and actual_location != "ALL") else "ALL_UNITS"
    filename = f"{loc_str}_Attendance_Ledger_{ist_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============================================================
# 🖥️ 5. HTML REPORT ROUTERS (IFRAME Drilldowns)
# ============================================================
@router.get("/today_report", response_class=HTMLResponse)
def today_attendance_report(request: Request):
    email = request.session.get("email")
    company_code = request.session.get("company_code")

    if not email or not company_code:
        return RedirectResponse("/", status_code=302)

    actual_location, _ = get_strict_location(request)

    return templates.TemplateResponse(
        request=request,
        name="attendance/today_report.html",
        context={
            "email": email,
            "company_id": company_code,
            "actual_location": actual_location or "ALL UNITS"
        }
    )


@router.get("/audit_report", response_class=HTMLResponse)
def attendance_audit_report(request: Request):
    email = request.session.get("email")
    company_code = request.session.get("company_code")

    if not email or not company_code:
        return RedirectResponse("/", status_code=302)

    actual_location, _ = get_strict_location(request)

    return templates.TemplateResponse(
        request=request,
        name="attendance/audit_report.html",
        context={
            "email": email,
            "company_id": company_code,
            "actual_location": actual_location or "ALL UNITS"
        }
    )
