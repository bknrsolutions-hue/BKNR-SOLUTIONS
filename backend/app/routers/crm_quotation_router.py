from fastapi import APIRouter, Request, Depends, HTTPException, Query
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, or_, and_
from datetime import date, datetime
from io import BytesIO
import openpyxl
import re
import json
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.utils.download_security import require_download_grant
from app.database.models.crm_quotation import CRMQuotation, CRMQuotationLine, CRMQuotationReply
from app.database.models.invoices import ProformaInvoice
from app.database.models.users import Company, User
from app.database.models.inventory_management import stock_entry, cold_storage_holding
from app.database.models.criteria import (
    buyers, buyer_agents, countries, species, varieties, grades,
    brands, glazes, freezers, packing_styles, production_for, production_at,
    grade_to_hoso, HOSO_HLSO_Yields
)

from app.database.models.processing import AuditLog
from app.utils.timezone import ist_now

router = APIRouter(prefix="/crm/quotation", tags=["CRM Quotation"])


class LineItemPayload(BaseModel):
    item_name: str
    brand: Optional[str] = None
    packing_style: Optional[str] = None
    freezer: Optional[str] = None
    count_glaze: Optional[str] = None
    weight_glaze: Optional[str] = None
    species: Optional[str] = None
    variety: Optional[str] = None
    grade: Optional[str] = None
    no_of_pieces: Optional[str] = "0"
    no_of_mc: Optional[int] = 0
    quantity_kg: float = Field(default=0.0, ge=0)
    rate_per_kg: float = Field(default=0.0, ge=0)
    hoso_count: Optional[str] = None
    target_hoso_rate: Optional[float] = 0.0
    expenses: Optional[float] = 0.0
    target_quotation_price: Optional[float] = 0.0
    bidding_price: Optional[float] = 0.0
    amount: float = Field(default=0.0, ge=0)



class QuotationPayload(BaseModel):
    quotation_no: str
    po_number: Optional[str] = None
    company_name: Optional[str] = None
    quotation_date: Optional[date] = None
    valid_until: date
    shipment_date: Optional[str] = None
    customer_name: str
    customer_address: Optional[str] = None
    agent: Optional[str] = None
    country: Optional[str] = None
    production_at: Optional[str] = None
    currency: str = "USD"
    exchange_rate: Optional[float] = 83.5
    incoterm: Optional[str] = "FOB"
    payment_terms: Optional[str] = None
    remarks: Optional[str] = None
    status: str = "DRAFT"
    items: List[LineItemPayload] = []


class ApprovalPayload(BaseModel):
    decision: str  # ACCEPTED, SENT, REJECTED, EXPIRED
    remarks: Optional[str] = None


class AnalyzePayload(BaseModel):
    company_name: Optional[str] = None
    production_at: Optional[str] = None
    exchange_rate: Optional[float] = 83.5
    items: List[LineItemPayload] = []


class SendQuotationEmailPayload(BaseModel):
    to_email: str
    subject: str
    from_email: Optional[str] = None
    header_text: Optional[str] = None
    footer_text: Optional[str] = None
    signoff_text: Optional[str] = None
    items: Optional[List[LineItemPayload]] = None


def calculate_pieces(grade_str, manual_pcs):
    try:
        if manual_pcs and str(manual_pcs).strip() and int(manual_pcs) > 0:
            return str(int(manual_pcs))
        nums = re.findall(r'\d+', str(grade_str or ""))
        if nums:
            last_num = int(nums[-1])
            return str(round(last_num * 2.2))
    except Exception:
        pass
    return "0"


def resolve_company_code(request: Request, db: Session) -> str:
    code, _ = resolve_session_company_info(request, db)
    return code


def resolve_session_company_info(request: Request, db: Session):
    comp_code = request.session.get("company_code")
    comp_name = request.session.get("company_name") or request.session.get("company")

    if not comp_code or not comp_name:
        email = request.session.get("email")
        if email:
            usr = db.query(User).filter(User.email == email).first()
            if usr and usr.company_id:
                c = db.query(Company).filter(Company.id == usr.company_id).first()
                if c:
                    comp_code = comp_code or c.company_code
                    comp_name = comp_name or getattr(c, "company_name", None) or getattr(c, "name", None) or c.company_code
    if not comp_code or not comp_name:
        first_c = db.query(Company).first()
        if first_c:
            comp_code = comp_code or first_c.company_code
            comp_name = comp_name or getattr(first_c, "company_name", None) or getattr(first_c, "name", None) or first_c.company_code

    return comp_code or "DEFAULT", comp_name or "BKNR ERP"


def get_next_quotation_no(db: Session, company_id: str = None) -> str:
    current_year = datetime.now().year
    prefix = f"QT-{current_year}-"
    all_nos = [q[0] for q in db.query(CRMQuotation.quotation_no).filter(CRMQuotation.quotation_no.like(f"{prefix}%")).all() if q[0]]
    max_seq = 0
    for q_no in all_nos:
        try:
            num = int(q_no.split("-")[-1])
            if num > max_seq:
                max_seq = num
        except (ValueError, IndexError):
            pass
    seq = max_seq + 1
    new_no = f"{prefix}{seq:04d}"
    while db.query(CRMQuotation).filter(CRMQuotation.quotation_no == new_no).first():
        seq += 1
        new_no = f"{prefix}{seq:04d}"
    return new_no


def log_audit(db: Session, company_id: str, email: str, action: str, quotation_no: str, details: str):
    try:
        audit = AuditLog(
            company_id=company_id,
            table_name="crm_quotations",
            field_name=action,
            old_value=quotation_no,
            new_value=details,
            edited_by=email,
            edited_at=ist_now()
        )
        db.add(audit)
    except Exception:
        pass


def create_pi_from_quotation(db: Session, quotation: CRMQuotation, user_email: str) -> Optional[ProformaInvoice]:
    """
    Automatically creates a Proforma Invoice (PI) when a CRM Quotation is accepted.
    Prevents duplicate creation for the same quotation.
    """
    try:
        comp_code = quotation.company_id
        if not comp_code:
            return None

        # Check if a PI already exists for this quotation
        existing = db.query(ProformaInvoice).filter(
            ProformaInvoice.company_id == comp_code,
            ProformaInvoice.is_cancelled == False,
            or_(
                ProformaInvoice.remarks.like(f"%{quotation.quotation_no}%"),
                and_(
                    quotation.po_number != None,
                    quotation.po_number != "",
                    ProformaInvoice.po_number == quotation.po_number
                ),
                ProformaInvoice.po_number == quotation.quotation_no
            )
        ).first()

        if existing:
            return existing

        # Generate next PI number: PI-YYYY-XXXX
        current_year = date.today().year
        prefix = f"PI-{current_year}-"
        all_pis = [
            p[0] for p in db.query(ProformaInvoice.pi_no).filter(
                ProformaInvoice.company_id == comp_code,
                ProformaInvoice.pi_no.like(f"{prefix}%")
            ).all() if p[0]
        ]
        max_seq = 0
        for pi_code in all_pis:
            try:
                num = int(pi_code.split("-")[-1])
                if num > max_seq:
                    max_seq = num
            except (ValueError, IndexError):
                pass
        seq = max_seq + 1
        new_pi_no = f"{prefix}{seq:04d}"
        while db.query(ProformaInvoice).filter(ProformaInvoice.company_id == comp_code, ProformaInvoice.pi_no == new_pi_no).first():
            seq += 1
            new_pi_no = f"{prefix}{seq:04d}"

        # Fetch quotation lines
        lines = db.query(CRMQuotationLine).filter(CRMQuotationLine.quotation_id == quotation.id).all()
        first_line = lines[0] if lines else None

        total_qty = sum(float(l.quantity_kg or 0.0) for l in lines) if lines else 0.0
        total_mc = sum(int(l.no_of_mc or 0) for l in lines) if lines else 0
        total_amt = float(quotation.total_amount or 0.0)
        if total_amt <= 0 and lines:
            total_amt = sum(float(l.amount or (l.quantity_kg * (l.bidding_price or l.rate_per_kg or 0.0))) for l in lines)

        avg_unit_price = (total_amt / total_qty) if total_qty > 0 else (float(first_line.rate_per_kg or first_line.bidding_price or 0.0) if first_line else 0.0)

        # Format items JSON and product description
        desc_items = []
        items_list = []
        for l in lines:
            l_name = (l.item_name or " ".join(filter(None, [l.species, l.variety, l.grade])) or "Shrimp Item").strip()
            l_qty = float(l.quantity_kg or 0.0)
            l_rate = float(l.bidding_price or l.rate_per_kg or 0.0)
            l_mc = int(l.no_of_mc or 0)
            l_amt = float(l.amount or (l_qty * l_rate))
            desc_items.append(f"• {l_name} (Grade: {l.grade or 'N/A'}, Pack: {l.packing_style or 'N/A'}) - {l_mc} MC / {l_qty:,.2f} KG @ ${l_rate:,.2f}")
            items_list.append({
                "item_name": l_name,
                "brand": l.brand or "",
                "packing_style": l.packing_style or "",
                "freezer": l.freezer or "",
                "count_glaze": l.count_glaze or "",
                "weight_glaze": l.weight_glaze or "",
                "species": l.species or "",
                "variety": l.variety or "",
                "grade": l.grade or "",
                "no_of_pieces": l.no_of_pieces or "0",
                "no_of_mc": l_mc,
                "quantity": l_qty,
                "quantity_kg": l_qty,
                "unit_price": l_rate,
                "rate_per_kg": l_rate,
                "total_amount": l_amt,
                "amount": l_amt
            })

        product_desc = "\n".join(desc_items) if desc_items else (quotation.remarks or "Commercial Price Quotation Items")

        ship_date_display = quotation.valid_until.strftime('%d-%b-%Y') if (quotation.valid_until and hasattr(quotation.valid_until, 'strftime')) else (str(quotation.valid_until) if quotation.valid_until else 'agreed schedule')

        pi_remarks = (
            f"1. Shipment Schedule: Expected shipment on or before {ship_date_display}.\n"
            f"2. Quantity Tolerance: Shipped quantity +/- 10% acceptable at final shipment as per actual container loading.\n"
            f"3. Quality & Inspection: Pre-shipment quality, count, glaze & net weight inspection by EIA / MPEDA authorized surveyors.\n"
            f"4. Payment & Documentation: Buyer to provide swift advice within 3 banking days of LC opening.\n"
            f"5. Validity & Force Majeure: Offer valid until {ship_date_display} and subject to vessel space availability.\n"
            f"Ref: Quotation {quotation.quotation_no}"
        )

        pi = ProformaInvoice(
            company_id=comp_code,
            pi_no=new_pi_no,
            pi_date=date.today(),
            validity_date=quotation.valid_until or date.today(),
            po_number=quotation.po_number or quotation.quotation_no,
            buyer_name=quotation.customer_name or "Unknown Buyer",
            buyer_address=quotation.customer_address or "N/A",
            country=quotation.country or "N/A",
            currency=quotation.currency or "USD",
            incoterm=quotation.incoterm or "FOB",
            payment_terms=quotation.payment_terms or "100% LC at sight",
            port_of_loading="PLANT",
            port_of_discharge=quotation.country or "N/A",
            product_description=product_desc,
            quantity=round(total_qty, 3),
            unit="KG",
            unit_price=round(avg_unit_price, 4),
            total_amount=round(total_amt, 2),
            status="ACCEPTED",
            brand=first_line.brand if first_line else None,
            packing_style=first_line.packing_style if first_line else None,
            freezer=first_line.freezer if first_line else None,
            count_glaze=first_line.count_glaze if first_line else None,
            weight_glaze=first_line.weight_glaze if first_line else None,
            species=first_line.species if first_line else None,
            variety=first_line.variety if first_line else None,
            grade=first_line.grade if first_line else None,
            no_of_pieces=first_line.no_of_pieces if first_line else None,
            no_of_mc=total_mc,
            items_json=json.dumps(items_list),
            approval_status="APPROVED",
            approved_by=user_email,
            approved_at=ist_now(),
            approval_remarks=f"Auto-approved via accepted Quotation {quotation.quotation_no}",
            remarks=pi_remarks,
            created_by=user_email,
            created_at=ist_now()
        )
        db.add(pi)
        db.flush()

        log_audit(db, comp_code, user_email, "AUTO_CREATE_PI", quotation.quotation_no, f"Auto-created Proforma Invoice {pi.pi_no} for {quotation.customer_name}")
        return pi
    except Exception as exc:
        print(f"Error auto-creating PI from quotation {quotation.quotation_no}: {exc}")
        return None


@router.post("/analyze_stock")
@router.post("/export_documents/quotation/analyze_stock", include_in_schema=False)
async def analyze_quotation_stock(payload: AnalyzePayload, request: Request, db: Session = Depends(get_db)):
    comp_code = resolve_company_code(request, db)
    exch_rate = float(payload.exchange_rate or 83.5)

    # Load stock_entry table for tenant company matching pending_orders_report.py line 103
    all_stocks = db.query(stock_entry).filter(
        (stock_entry.company_id == comp_code) | (stock_entry.company_id == "BKNR") | (stock_entry.company_id.is_(None)) | (stock_entry.company_id == "")
    ).all()
    if not all_stocks:
        all_stocks = db.query(stock_entry).all()

    eval_stock = list(all_stocks)


    # Query grade_to_hoso list for NW grade lookup matching pending_orders_report
    grade_map_list = db.query(grade_to_hoso).filter(
        (grade_to_hoso.company_id == comp_code) | (grade_to_hoso.company_id == "BKNR") | (grade_to_hoso.company_id.is_(None)) | (grade_to_hoso.company_id == "")
    ).all()
    if not grade_map_list:
        grade_map_list = db.query(grade_to_hoso).all()


    def norm_clean(val):
        if not val:
            return ""
        s = str(val).upper().strip()
        s = s.replace("GRADE GLAZE%", "").replace("WEIGHT GLAZE%", "").strip()
        return re.sub(r'[^A-Z0-9]', '', s)

    # Filter strictly by selected company_name (production_for) if provided in header
    target_comp = norm_clean(payload.company_name)
    if target_comp and target_comp != "ALL":
        eval_stock = [
            s for s in eval_stock
            if norm_clean(getattr(s, "production_for", "")) == target_comp
            or target_comp in norm_clean(getattr(s, "production_for", ""))
            or norm_clean(getattr(s, "company_id", "")) == target_comp
            or norm_clean(getattr(s, "company_name", "")) == target_comp
        ]


    # Global weighted average rates helper matching inventory dashboard
    from collections import defaultdict
    kpi_rates_helper = defaultdict(lambda: {"sum_val": 0.0, "sum_qty": 0.0})
    for item in eval_stock:
        qty = float(getattr(item, "quantity", 0) or 0)
        rate = float(getattr(item, "product_kg_value", 0) or getattr(item, "rate_per_kg", 0) or 0)
        if qty > 0 and rate > 0:
            g_key = (
                norm_clean(getattr(item, "species", "")),
                norm_clean(getattr(item, "variety", "")),
                norm_clean(getattr(item, "grade", ""))
            )
            kpi_rates_helper[g_key]["sum_val"] += (qty * rate)
            kpi_rates_helper[g_key]["sum_qty"] += qty

    global_rates = {gk: (v["sum_val"] / v["sum_qty"] if v["sum_qty"] > 0.01 else 0.0) for gk, v in kpi_rates_helper.items()}

    # Group raw stock entries into net closing stock per batch/location identity
    batch_grouped_stock = defaultdict(lambda: {
        "spec": "", "var": "", "grad": "", "pack": "", "frz": "", "gl": "",
        "location": "", "batch_number": "", "raw_brand": "", "raw_packing": "",
        "raw_freezer": "", "raw_glaze": "", "raw_grade": "", "raw_species": "",
        "raw_variety": "", "raw_prod_for": "", "batch_id": "",
        "net_mc": 0, "net_kg": 0.0, "net_cost": 0.0
    })

    for item in eval_stock:
        s_spec = norm_clean(getattr(item, "species", ""))
        s_var = norm_clean(getattr(item, "variety", ""))
        s_grad = norm_clean(getattr(item, "grade", ""))
        s_pack = norm_clean(getattr(item, "packing_style", ""))
        s_frz = norm_clean(getattr(item, "freezer", ""))
        s_gl = norm_clean(getattr(item, "glaze", ""))

        loc = getattr(item, "cold_storage_name", "") or getattr(item, "location", "") or getattr(item, "production_at", "") or "PLANT"
        batch_no = getattr(item, "batch_number", "") or "N/A"
        raw_b = getattr(item, "brand", "") or "N/A"
        raw_p = getattr(item, "packing_style", "") or "N/A"
        raw_f = getattr(item, "freezer", "") or "N/A"
        raw_g = getattr(item, "glaze", "") or "N/A"
        raw_gr = getattr(item, "grade", "") or "N/A"
        raw_sp = getattr(item, "species", "") or "N/A"
        raw_v = getattr(item, "variety", "") or "N/A"
        raw_pf = getattr(item, "production_for", "") or "N/A"

        batch_id = f"{loc}|{batch_no}|{raw_b}|{raw_p}|{raw_f}|{raw_g}|{raw_gr}|{raw_sp}|{raw_v}|{raw_pf}"

        mc = int(getattr(item, "no_of_mc", 0) or 0)
        kg = float(getattr(item, "quantity", 0) or 0.0)
        rate = float(getattr(item, "product_kg_value", 0) or getattr(item, "rate_per_kg", 0) or 0.0)

        if rate <= 0:
            rate = global_rates.get((s_spec, s_var, s_grad), 0.0)

        move = str(getattr(item, "cargo_movement_type", "") or "").strip().upper()
        sign = 1 if (not move or move == "IN") else -1

        b_entry = batch_grouped_stock[batch_id]
        b_entry["spec"] = s_spec
        b_entry["var"] = s_var
        b_entry["grad"] = s_grad
        b_entry["pack"] = s_pack
        b_entry["frz"] = s_frz
        b_entry["gl"] = s_gl
        b_entry["location"] = loc
        b_entry["batch_number"] = batch_no
        b_entry["raw_brand"] = raw_b
        b_entry["raw_packing"] = raw_p
        b_entry["raw_freezer"] = raw_f
        b_entry["raw_glaze"] = raw_g
        b_entry["raw_grade"] = raw_gr
        b_entry["raw_species"] = raw_sp
        b_entry["raw_variety"] = raw_v
        b_entry["raw_prod_for"] = raw_pf
        b_entry["batch_id"] = batch_id
        b_entry["net_mc"] += (mc * sign)
        b_entry["net_kg"] += (kg * sign)
        b_entry["net_cost"] += (kg * rate * sign)

    # Active batch list with positive net closing stock
    active_batch_list = [b for b in batch_grouped_stock.values() if b["net_mc"] > 0 or b["net_kg"] > 0.01]

    cards = []
    for idx, item in enumerate(payload.items):
        i_spec = norm_clean(item.species)
        i_var = norm_clean(item.variety)
        i_grad = norm_clean(item.grade)
        i_pack = norm_clean(item.packing_style)
        i_frz = norm_clean(item.freezer)
        i_c_gl = norm_clean(item.count_glaze)
        i_w_gl = norm_clean(item.weight_glaze)
        i_gl = i_c_gl or i_w_gl

        avail_mc = 0
        avail_kg = 0.0
        avail_cost = 0.0
        matched_batch_ids = set()
        avail_stock_details = []

        def is_combo_match(s):
            if i_spec and s["spec"] and i_spec != s["spec"]:
                return False
            if i_var and s["var"] and i_var != s["var"]:
                return False
            if i_grad and s["grad"] and i_grad != s["grad"]:
                return False
            if i_pack and s["pack"]:
                if i_pack != s["pack"] and i_pack not in s["pack"] and s["pack"] not in i_pack:
                    return False
            if i_frz and s["frz"] and i_frz != s["frz"]:
                return False
            if i_gl and s["gl"]:
                if i_gl != s["gl"] and i_gl not in s["gl"] and s["gl"] not in i_gl:
                    return False
            return True


        # 1. Combo Stock Match
        for s in active_batch_list:
            if is_combo_match(s):
                avail_mc += s["net_mc"]
                avail_kg += s["net_kg"]
                avail_cost += s["net_cost"]
                matched_batch_ids.add(s["batch_id"])
                avail_stock_details.append({
                    "cold_storage_name": s["location"],
                    "batch_number": s["batch_number"],
                    "brand": s["raw_brand"],
                    "packing_style": s["raw_packing"],
                    "freezer": s["raw_freezer"],
                    "glaze": s["raw_glaze"],
                    "grade": s["raw_grade"],
                    "species": s["raw_species"],
                    "variety": s["raw_variety"],
                    "production_for": s["raw_prod_for"],
                    "no_of_mc": s["net_mc"],
                    "quantity_kg": round(s["net_kg"], 2),
                    "rate_per_kg": round(s["net_cost"] / s["net_kg"], 2) if s["net_kg"] > 0 else 0.0,
                })

        avail_mc = max(0, avail_mc)
        avail_kg = max(0.0, round(avail_kg, 2))
        avail_avg_rate = round(avail_cost / avail_kg, 2) if avail_kg > 0 else 0.0

        # 2. Referral Stock Match matching pending_orders_report.py
        referral_mc = 0
        referral_kg = 0.0
        referral_cost = 0.0
        referral_stock_details = []

        p_spec = str(item.species or "").strip().lower()
        p_var = str(item.variety or "").strip().lower()
        p_grad = str(item.grade or "").strip().lower()
        p_pack = str(item.packing_style or "").strip().lower()
        p_frz = str(item.freezer or "N/A").strip().lower()

        gl_text = str(item.count_glaze or item.weight_glaze or "").strip().upper()
        c_gl_val = float(re.search(r'(\d+)', gl_text).group(1)) if re.search(r'(\d+)', gl_text) else 0.0
        c_gl_factor = (100.0 - c_gl_val) / 100.0 if (c_gl_val > 0 and c_gl_val < 100) else 1.0
        is_order_nwnc = "NWNC" in gl_text or c_gl_val == 0

        # Calculate pieces per MC matching pending_orders formula
        pcs = float(item.no_of_pieces or 0)
        if pcs <= 0 and p_grad:
            bounds = re.findall(r'\d+', p_grad)
            if bounds:
                pcs = round(float(bounds[-1]) * 2.2)

        # Net Count (Count Glaze based) matching pending_orders_report.py line 177
        try:
            net_cnt_calc = round((pcs / 2.20462) / c_gl_factor, 2) if pcs > 0 else 0.0
        except Exception:
            net_cnt_calc = 0.0

        # NW Grade Mapping logic (Direct Lookup -> Nearest Count Lookup -> Fallback)
        nw_grade = "-"
        direct_gm = next(
            (gm for gm in grade_map_list 
             if str(getattr(gm, "species", "") or "").strip().lower() == p_spec 
             and str(getattr(gm, "grade_name", "") or "").strip().lower() == p_grad 
             and (not gl_text or norm_clean(getattr(gm, "glaze_name", "")) == norm_clean(gl_text))),
            None
        )
        if direct_gm and getattr(direct_gm, "nw_grade", None):
            nw_grade = str(direct_gm.nw_grade).strip().lower()

        if nw_grade == "-":
            rel_grades = [
                gm for gm in grade_map_list 
                if str(getattr(gm, "species", "") or "").strip().lower() == p_spec 
                and (not p_var or not getattr(gm, "variety_name", None) or str(getattr(gm, "variety_name", "")).strip().lower() == p_var)
            ]
            if not rel_grades:
                rel_grades = [gm for gm in grade_map_list if str(getattr(gm, "species", "") or "").strip().lower() == p_spec]

            if rel_grades and net_cnt_calc > 0:
                nearest_gm = min(
                    rel_grades, 
                    key=lambda x: abs(float(getattr(x, "hlso_count", 0) or getattr(x, "hoso_count", 0) or 0) - net_cnt_calc)
                )
                if getattr(nearest_gm, "nw_grade", None):
                    nw_grade = str(nearest_gm.nw_grade).strip().lower()

        grade_mapped = True
        warning_msg = ""
        if is_order_nwnc:
            target_ref_grade = p_grad
        else:
            if nw_grade != "-" and nw_grade != "":
                target_ref_grade = nw_grade
            else:
                grade_mapped = False
                target_ref_grade = None
                warning_msg = f"Grade mapping not configured in Grade to HOSO master for species '{item.species}', grade '{item.grade}', glaze '{item.count_glaze or item.weight_glaze}'."

        for s in active_batch_list:
            if s["batch_id"] in matched_batch_ids:
                continue

            s_spec = str(s["raw_species"] or "").strip().lower()
            s_var = str(s["raw_variety"] or "").strip().lower()
            s_grad = str(s["raw_grade"] or "").strip().lower()
            s_pack = str(s["raw_packing"] or "").strip().lower()
            s_frz = str(s["raw_freezer"] or "N/A").strip().lower()

            s_gl_match = re.search(r'(\d+)', str(s["raw_glaze"] or "0"))
            s_gl_num = s_gl_match.group(1) if s_gl_match else "0"

            match_ref = False
            if s_spec == p_spec and s_var == p_var and s_frz == p_frz:
                if is_order_nwnc:
                    if s_grad == p_grad and s_gl_num == "0" and s_pack != p_pack:
                        match_ref = True
                else:
                    if target_ref_grade and s_grad == target_ref_grade and s_gl_num == "0":
                        match_ref = True

            if match_ref:
                referral_mc += s["net_mc"]
                referral_kg += s["net_kg"]
                referral_cost += s["net_cost"]
                referral_stock_details.append({
                    "cold_storage_name": s["location"],
                    "batch_number": s["batch_number"],
                    "brand": s["raw_brand"],
                    "packing_style": s["raw_packing"],
                    "freezer": s["raw_freezer"],
                    "glaze": s["raw_glaze"],
                    "grade": s["raw_grade"],
                    "species": s["raw_species"],
                    "variety": s["raw_variety"],
                    "production_for": s["raw_prod_for"],
                    "no_of_mc": s["net_mc"],
                    "quantity_kg": round(s["net_kg"], 2),
                    "rate_per_kg": round(s["net_cost"] / s["net_kg"], 2) if s["net_kg"] > 0 else 0.0,
                })

        ref_kg = max(0.0, round(referral_kg, 2))
        ref_avg_rate = round(referral_cost / ref_kg, 2) if ref_kg > 0 else 0.0

        req_mc = int(item.no_of_mc or 0)
        req_kg = float(item.quantity_kg or 0.0)
        rate_per_kg = float(item.rate_per_kg or 0.0)
        quoted_price_inr = round(rate_per_kg * exch_rate, 2)

        if req_mc > 0:
            if avail_mc >= req_mc:
                status = "AVAILABLE"
            elif avail_mc > 0:
                status = "PARTIAL"
            else:
                status = "OUT_OF_STOCK"
        else:
            status = "AVAILABLE" if avail_mc > 0 else "OUT_OF_STOCK"

        deficit_mc = max(0, req_mc - avail_mc)

        cards.append({
            "line_no": idx + 1,
            "item_name": item.item_name,
            "brand": item.brand or "N/A",
            "packing_style": item.packing_style or "N/A",
            "freezer": item.freezer or "N/A",
            "count_glaze": item.count_glaze or "N/A",
            "weight_glaze": item.weight_glaze or "N/A",
            "species": item.species or "N/A",
            "variety": item.variety or "N/A",
            "grade": item.grade or "N/A",
            "no_of_pieces": item.no_of_pieces or "0",
            "required_mc": req_mc,
            "required_kg": req_kg,
            "rate_per_kg": rate_per_kg,
            "quoted_price_inr": quoted_price_inr,
            "available_stock_mc": avail_mc,
            "available_stock_kg": avail_kg,
            "avail_stock_avg_rate": avail_avg_rate,
            "avail_stock_details": avail_stock_details,
            "deficit_mc": deficit_mc,
            "referral_stock_mc": referral_mc,
            "referral_stock_kg": ref_kg,
            "referral_stock_avg_rate": ref_avg_rate,
            "referral_stock_details": referral_stock_details,
            "grade_mapped": grade_mapped,
            "nw_grade": nw_grade,
            "warning_msg": warning_msg,
            "status": status,
        })




    return JSONResponse(content={
        "success": True,
        "cards": cards,
    })


def ensure_crm_quotation_schema(db: Session):
    statements = [
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS company_id VARCHAR(50);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS company_name VARCHAR(255);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS inquiry_id INTEGER;",
        "ALTER TABLE crm_quotations ALTER COLUMN inquiry_id DROP NOT NULL;",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS quotation_no VARCHAR(50);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS po_number VARCHAR(100);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS quotation_date DATE;",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS valid_until DATE;",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS shipment_date VARCHAR(100);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS customer_name VARCHAR(255);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS customer_address TEXT;",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS agent VARCHAR(255);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS country VARCHAR(100);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS production_at VARCHAR(255);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS currency VARCHAR(20) DEFAULT 'USD';",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS exchange_rate DOUBLE PRECISION DEFAULT 83.5;",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS incoterm VARCHAR(50);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS payment_terms VARCHAR(255);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS total_amount NUMERIC(18,2) DEFAULT 0.0;",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS status VARCHAR(20) DEFAULT 'DRAFT';",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS remarks TEXT;",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS created_by VARCHAR(255);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS updated_by VARCHAR(255);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS approved_by VARCHAR(255);",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS approved_at TIMESTAMP;",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS approval_remarks TEXT;",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP;",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP;",
        "ALTER TABLE crm_quotations ADD COLUMN IF NOT EXISTS is_cancelled BOOLEAN DEFAULT FALSE;",

        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS quotation_id INTEGER;",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS item_name VARCHAR(255);",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS brand VARCHAR(150);",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS packing_style VARCHAR(150);",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS freezer VARCHAR(150);",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS count_glaze VARCHAR(100);",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS weight_glaze VARCHAR(100);",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS species VARCHAR(150);",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS variety VARCHAR(150);",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS grade VARCHAR(100);",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS no_of_pieces VARCHAR(100);",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS no_of_mc INTEGER DEFAULT 0;",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS quantity_kg DOUBLE PRECISION DEFAULT 0;",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS rate_per_kg DOUBLE PRECISION DEFAULT 0;",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS hoso_count VARCHAR(100);",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS target_hoso_rate DOUBLE PRECISION DEFAULT 0;",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS expenses DOUBLE PRECISION DEFAULT 0;",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS target_quotation_price DOUBLE PRECISION DEFAULT 0;",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS bidding_price DOUBLE PRECISION DEFAULT 0;",
        "ALTER TABLE crm_quotation_lines ADD COLUMN IF NOT EXISTS amount DOUBLE PRECISION DEFAULT 0;",
        "CREATE TABLE IF NOT EXISTS crm_quotation_replies (id SERIAL PRIMARY KEY, quotation_id INTEGER, quotation_no VARCHAR(50) NOT NULL, sender_email VARCHAR(255) NOT NULL, recipient_email VARCHAR(255) NOT NULL, subject VARCHAR(255), message_body TEXT NOT NULL, received_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, direction VARCHAR(20) DEFAULT 'INBOUND', message_id VARCHAR(255), attachments_json TEXT);",
        "ALTER TABLE crm_quotation_replies ADD COLUMN IF NOT EXISTS attachments_json TEXT;",
    ]
    try:
        from sqlalchemy import text
        for stmt in statements:
            try:
                db.execute(text(stmt))
            except Exception:
                pass
        db.commit()
    except Exception:
        db.rollback()


@router.get("/data")
@router.get("/export_documents/quotation/data", include_in_schema=False)
async def get_quotation_data(request: Request, db: Session = Depends(get_db)):
    comp_code, comp_name = resolve_session_company_info(request, db)
    ensure_crm_quotation_schema(db)

    # Fetch Quotations
    quotations = (
        db.query(CRMQuotation)
        .filter(CRMQuotation.company_id == comp_code, CRMQuotation.is_cancelled == False)
        .order_by(desc(CRMQuotation.quotation_date), desc(CRMQuotation.id))
        .all()
    )

    rows = []
    for q in quotations:
        lines = [
            {
                "id": line.id,
                "item_name": line.item_name,
                "brand": line.brand or "",
                "packing_style": line.packing_style or "",
                "freezer": line.freezer or "",
                "count_glaze": line.count_glaze or "",
                "weight_glaze": line.weight_glaze or "",
                "species": line.species or "",
                "variety": line.variety or "",
                "grade": line.grade or "",
                "no_of_pieces": line.no_of_pieces or "0",
                "no_of_mc": line.no_of_mc or 0,
                "quantity_kg": float(line.quantity_kg or 0),
                "rate_per_kg": float(line.bidding_price or line.rate_per_kg or 0),
                "hoso_count": getattr(line, "hoso_count", "") or "",
                "target_hoso_rate": float(getattr(line, "target_hoso_rate", 0) or 0),
                "expenses": float(getattr(line, "expenses", 0) or 0),
                "target_quotation_price": float(getattr(line, "target_quotation_price", 0) or 0),
                "bidding_price": float(getattr(line, "bidding_price", 0) or line.rate_per_kg or 0),
                "amount": float(line.amount or 0),
            }
            for line in q.lines
        ]

        rows.append({
            "id": q.id,
            "quotation_no": q.quotation_no,
            "po_number": q.po_number or "",
            "company_name": q.company_name or "",
            "quotation_date": q.quotation_date.isoformat() if q.quotation_date else None,
            "valid_until": q.valid_until.isoformat() if q.valid_until else None,
            "customer_name": q.customer_name,
            "customer_address": q.customer_address or "",
            "agent": q.agent or "",
            "country": q.country or "",
            "production_at": q.production_at or "",
            "currency": q.currency or "USD",
            "exchange_rate": float(q.exchange_rate or 83.5),
            "incoterm": q.incoterm or "FOB",
            "payment_terms": q.payment_terms or "",
            "total_amount": float(q.total_amount or 0),
            "status": q.status or "DRAFT",
            "remarks": q.remarks or "",
            "created_by": q.created_by or "",
            "approved_by": q.approved_by or "",
            "approval_remarks": q.approval_remarks or "",
            "items": lines,
            "pi_id": None,
            "pi_no": "",
        })

    # Attach linked PI id/no for each quotation (for chat PDF links)
    for row in rows:
        try:
            linked_pi = db.query(ProformaInvoice).filter(
                ProformaInvoice.company_id == comp_code,
                ProformaInvoice.is_cancelled == False,
                or_(
                    ProformaInvoice.remarks.like(f"%{row['quotation_no']}%"),
                    ProformaInvoice.po_number == row["quotation_no"]
                )
            ).first()
            if linked_pi:
                row["pi_id"] = linked_pi.id
                row["pi_no"] = linked_pi.pi_no or ""
        except Exception:
            pass


    # Strict Tenant Master Lookup Helper (Prioritizes tenant's own master records)
    def get_lookup(model, field_names, defaults=None):
        try:
            q = db.query(model)
            if hasattr(model, 'company_id') and comp_code:
                res = q.filter((model.company_id == comp_code) | (model.company_id.is_(None)) | (model.company_id == "")).all()
                tenant_res = [x for x in res if getattr(x, 'company_id', '') == comp_code]
                if tenant_res:
                    res = tenant_res
            else:
                res = q.all()

            found = []
            for x in res:
                for fname in field_names:
                    val = getattr(x, fname, None)
                    if val:
                        found.append(str(val).strip())
                        break
            found = sorted(list(set(f for f in found if f)))
            if found:
                return found
        except Exception:
            pass
        return sorted(list(set(defaults or [])))

    # Unique Companies (Strict Tenant Filter)
    unique_companies = []
    if comp_name:
        unique_companies.append(comp_name)
    try:
        p_q = db.query(production_for.production_for).filter(production_for.production_for != None)
        if hasattr(production_for, 'company_id') and comp_code:
            p_q = p_q.filter((production_for.company_id == comp_code) | (production_for.company_id.is_(None)) | (production_for.company_id == ""))
        p_rows = p_q.distinct().all()
        for c in p_rows:
            if c and c[0] and c[0].strip() and c[0].strip() not in unique_companies:
                unique_companies.append(c[0].strip())
    except Exception:
        pass

    if not unique_companies:
        try:
            comp_rows = db.query(Company).filter(Company.company_code == comp_code).all() if comp_code else db.query(Company).all()
            excluded = ["MAIN PLANT", "MAIN UNIT", "GENERAL STOCK", "N/A", "NONE", "NULL"]
            for c in comp_rows:
                if c.company_name and c.company_name.upper().strip() not in excluded and c.company_name.strip() not in unique_companies:
                    unique_companies.append(c.company_name.strip())
        except Exception:
            pass

    unique_companies = sorted(list(set(unique_companies))) if unique_companies else [comp_name or "BKNR SEAFOODS PRIVATE LIMITED"]

    # Processing Locations (STRICT TENANT FILTER ONLY)
    loc_set = set()
    try:
        l_q = db.query(production_at.production_at).filter(production_at.production_at != None)
        if hasattr(production_at, 'company_id') and comp_code:
            l_q = l_q.filter(production_at.company_id == comp_code)
        for l in l_q.distinct().all():
            if l and l[0] and str(l[0]).strip():
                loc_set.add(str(l[0]).strip())
    except Exception:
        pass

    try:
        from app.database.models.processing import Peeling, RawMaterialPurchasing, DeHeading, Production, Soaking
        # Peeling peeling_at
        p_q = db.query(Peeling.peeling_at).filter(Peeling.peeling_at != None)
        if comp_code:
            p_q = p_q.filter(Peeling.company_id == comp_code)
        for r in p_q.distinct().all():
            if r and r[0] and str(r[0]).strip():
                loc_set.add(str(r[0]).strip())

        # RawMaterialPurchasing peeling_at
        rm_q = db.query(RawMaterialPurchasing.peeling_at).filter(RawMaterialPurchasing.peeling_at != None)
        if comp_code:
            rm_q = rm_q.filter(RawMaterialPurchasing.company_id == comp_code)
        for r in rm_q.distinct().all():
            if r and r[0] and str(r[0]).strip():
                loc_set.add(str(r[0]).strip())

        # DeHeading peeling_at
        dh_q = db.query(DeHeading.peeling_at).filter(DeHeading.peeling_at != None)
        if comp_code:
            dh_q = dh_q.filter(DeHeading.company_id == comp_code)
        for r in dh_q.distinct().all():
            if r and r[0] and str(r[0]).strip():
                loc_set.add(str(r[0]).strip())

        # Production production_at
        pr_q = db.query(Production.production_at).filter(Production.production_at != None)
        if comp_code:
            pr_q = pr_q.filter(Production.company_id == comp_code)
        for r in pr_q.distinct().all():
            if r and r[0] and str(r[0]).strip():
                loc_set.add(str(r[0]).strip())

        # Soaking production_at
        sk_q = db.query(Soaking.production_at).filter(Soaking.production_at != None)
        if comp_code:
            sk_q = sk_q.filter(Soaking.company_id == comp_code)
        for r in sk_q.distinct().all():
            if r and r[0] and str(r[0]).strip():
                loc_set.add(str(r[0]).strip())
    except Exception:
        pass

    production_locations = sorted(list(loc_set))

    # Buyers (Strict Tenant Filter)
    b_rows = []
    try:
        b_q = db.query(buyers)
        if comp_code:
            b_res = b_q.filter((buyers.company_id == comp_code) | (buyers.company_id.is_(None)) | (buyers.company_id == "")).all()
            tenant_b = [b for b in b_res if getattr(b, 'company_id', '') == comp_code]
            b_rows = tenant_b if tenant_b else b_res
        else:
            b_rows = b_q.all()
    except Exception:
        pass

    seen_buyers = set()
    buyer_options = []
    for b in b_rows:
        bname = str(getattr(b, 'buyer_name', '') or '').strip()
        if bname and bname not in seen_buyers:
            seen_buyers.add(bname)
            buyer_options.append({
                "name": bname,
                "address": getattr(b, 'buyer_address', "") or "",
                "country": getattr(b, 'country', "") or "",
                "currency": getattr(b, 'currency_code', "USD") or "USD",
                "payment_terms": getattr(b, 'payment_terms', "") or "",
            })
    buyer_names = sorted(list(seen_buyers))
    if not buyer_names:
        buyer_names = ["Global Seafoods LLC", "Pacific Marine Trading", "Apex Imports Inc", "Eurofood Distributors"]
        buyer_options = [{"name": b, "address": "", "country": "", "currency": "USD", "payment_terms": ""} for b in buyer_names]

    agents_list = get_lookup(buyer_agents, ["agent_name", "agent", "name"], ["Direct Buyer", "Local Agent", "International Broker"])
    brands_list = get_lookup(brands, ["brand_name", "brand", "name"], ["BKNR PREMIUM", "OCEAN FRESH", "BUYER BRAND"])
    country_options = get_lookup(countries, ["country_name", "country", "name"], ["USA", "Japan", "Vietnam", "China", "UAE", "UK", "Canada", "EU"])
    species_options = get_lookup(species, ["species_name", "species", "name"], ["VANNAMEI", "BLACK TIGER", "WHITE", "SEA CAUGHT", "PINK"])
    variety_options = get_lookup(varieties, ["variety_name", "variety", "name"], ["HLSO", "HOSO", "PUD", "PD", "PDTO", "EZP", "HEADLESS"])
    grade_options = get_lookup(grades, ["grade_name", "grade", "name"], ["16/20", "21/25", "26/30", "31/40", "41/50", "51/60", "61/70", "71/90", "91/110", "110/130"])
    glazes_list = get_lookup(glazes, ["glaze_name", "glaze", "name"], ["NET WEIGHT", "5% GLAZE", "10% GLAZE", "15% GLAZE", "20% GLAZE", "25% GLAZE"])
    freezers_list = get_lookup(freezers, ["freezer_name", "freezer", "name"], ["IQF", "BLOCK", "SEMI IQF", "PLATE FREEZER"])

    # Packing styles (Strict Tenant Filter)
    packing_options = []
    try:
        ps_q = db.query(packing_styles)
        if hasattr(packing_styles, 'company_id') and comp_code:
            ps_res = ps_q.filter((packing_styles.company_id == comp_code) | (packing_styles.company_id.is_(None)) | (packing_styles.company_id == "")).all()
            tenant_ps = [p for p in ps_res if getattr(p, 'company_id', '') == comp_code]
            ps_res = tenant_ps if tenant_ps else ps_res
        else:
            ps_res = ps_q.all()

        for p in ps_res:
            if getattr(p, 'packing_style', None):
                packing_options.append({
                    "packing_style": p.packing_style,
                    "mc_weight": float(getattr(p, 'mc_weight', 0) or 0),
                    "slab_weight": float(getattr(p, 'slab_weight', 0) or 0),
                })
    except Exception:
        pass

    if not packing_options:
        packing_options = [
            {"packing_style": "10 x 1 KG", "mc_weight": 10.0, "slab_weight": 1.0},
            {"packing_style": "6 x 1.8 KG", "mc_weight": 10.8, "slab_weight": 1.8},
            {"packing_style": "1 x 10 KG BLOCK", "mc_weight": 10.0, "slab_weight": 10.0},
            {"packing_style": "10 x 2 LBS", "mc_weight": 9.072, "slab_weight": 0.9072},
            {"packing_style": "5 x 2 KG", "mc_weight": 10.0, "slab_weight": 2.0},
        ]

    audit_data = []
    try:
        audit_logs = (
            db.query(AuditLog)
            .filter(AuditLog.company_id == comp_code, AuditLog.table_name == "crm_quotations")
            .order_by(desc(AuditLog.edited_at))
            .limit(50)
            .all()
        )
        audit_data = [
            {
                "id": a.id,
                "user_email": getattr(a, "edited_by", "") or "SYSTEM",
                "action": getattr(a, "field_name", "") or "AUDIT",
                "target_id": getattr(a, "old_value", "") or "",
                "details": getattr(a, "new_value", "") or "",
                "created_at": a.edited_at.strftime("%d-%b-%Y %I:%M %p") if getattr(a, "edited_at", None) else "",
            }
            for a in audit_logs
        ]
    except Exception:
        pass

    # Variety yields mapping matching Inventory Costing
    variety_yields = {}
    try:
        v_rows = db.query(varieties).all()
        for v in v_rows:
            if v.variety_name:
                v_name = str(v.variety_name).upper().strip()
                variety_yields[v_name] = {
                    "peeling_yield": float(v.peeling_yield or 100),
                    "soaking_yield": float(v.soaking_yield or 100),
                }
    except Exception:
        pass

    # HOSO HLSO Yields list matching Inventory Costing
    hoso_hlso_yields = []
    try:
        h_rows = db.query(HOSO_HLSO_Yields).all()
        for h in h_rows:
            hoso_hlso_yields.append({
                "species": str(h.species or "").upper().strip(),
                "hoso_count": int(h.hoso_count or 0),
                "hlso_count": int(h.hlso_count or 0),
                "hlso_yield_pct": float(h.hlso_yield_pct or 100),
            })
    except Exception:
        pass

    # Country Details with production_cost_per_kg
    country_details = []
    try:
        c_rows = db.query(countries).all()
        for c in c_rows:
            if getattr(c, "country_name", None):
                country_details.append({
                    "country_name": c.country_name,
                    "production_cost_per_kg": float(getattr(c, "production_cost_per_kg", 0) or 0),
                })
    except Exception:
        pass

    # Grade to HOSO master mapping
    grade_to_hoso_list = []
    try:
        from app.services.grade_to_hoso_sync import sync_grade_to_hoso
        sync_grade_to_hoso(db, comp_code, request.session.get("email") or "SYSTEM")
    except Exception:
        pass

    try:
        gh_rows = db.query(grade_to_hoso).all()
        for g in gh_rows:
            grade_to_hoso_list.append({
                "species": str(g.species or "").upper().strip(),
                "grade_name": str(g.grade_name or "").upper().strip(),
                "variety_name": str(g.variety_name or "").upper().strip(),
                "glaze_name": str(g.glaze_name or "").upper().strip(),
                "hlso_count": g.hlso_count,
                "hoso_count": g.hoso_count,
                "nw_grade": str(g.nw_grade or "").strip(),
            })
    except Exception:
        pass

    return JSONResponse(content={
        "success": True,
        "session_company_code": comp_code,
        "session_company_name": comp_name,
        "rows": rows,
        "buyers": buyer_options,
        "buyer_names": buyer_names,
        "agents": agents_list,
        "unique_companies": unique_companies,
        "production_locations": production_locations,
        "countries": country_options,
        "country_details": country_details,
        "brands": brands_list,
        "freezers": freezers_list,
        "glazes": glazes_list,
        "species": species_options,
        "varieties": variety_options,
        "grades": grade_options,
        "packing_styles": packing_options,
        "variety_yields": variety_yields,
        "hoso_hlso_yields": hoso_hlso_yields,
        "grade_to_hoso_list": grade_to_hoso_list,
        "next_quotation_no": get_next_quotation_no(db, comp_code),
        "audit_logs": audit_data,
        "can_approve": True,
    })


def calculate_target_quotation_price_backend(db: Session, comp_code: str, item, exchange_rate: float) -> float:
    t_price = float(getattr(item, "target_quotation_price", 0) or 0)
    if t_price > 0:
        return round(t_price, 2)

    hoso_rate = float(getattr(item, "target_hoso_rate", 0) or 0)
    exp = float(getattr(item, "expenses", 0) or 0)
    exch = float(exchange_rate or 83.5)

    if hoso_rate <= 0 and exp <= 0:
        return 0.0

    glaze_factor = 1.0
    glaze_str = str(getattr(item, "weight_glaze", "") or getattr(item, "count_glaze", "") or "").upper()
    if glaze_str and "NWNC" not in glaze_str and "NET WEIGHT" not in glaze_str:
        digits = re.findall(r"\d+", glaze_str)
        if digits:
            glaze_factor = (100 - int(digits[0])) / 100

    variety_str = str(getattr(item, "variety", "") or "").upper().strip()
    is_hoso = "HOSO" in variety_str

    peeling_yield = 1.0
    soaking_yield = 1.0
    hlso_yield = 1.0

    if not is_hoso and variety_str:
        v_obj = db.query(varieties).filter(func.upper(varieties.variety_name) == variety_str).first()
        if v_obj:
            peeling_yield = float(getattr(v_obj, "peeling_yield", 100) or 100) / 100.0
            soaking_yield = float(getattr(v_obj, "soaking_yield", 100) or 100) / 100.0

    if not is_hoso and getattr(item, "grade", None):
        grade_str = str(item.grade)
        nums = re.findall(r"\d+", grade_str)
        if nums:
            raw_grade_num = int(nums[-1])
            pcs = calculate_pieces(grade_str, getattr(item, "no_of_pieces", "0")) or round(raw_grade_num * 2.2)
            net_count_calc = round((pcs / 2.20462) / (glaze_factor if glaze_factor > 0 else 1.0), 2)
            hl_count_calc = round(net_count_calc * peeling_yield * soaking_yield, 2)

            species_str = str(getattr(item, "species", "") or "").upper().strip()
            sp_yields = db.query(HOSO_HLSO_Yields).filter(
                (HOSO_HLSO_Yields.species == species_str) | (HOSO_HLSO_Yields.species.is_(None))
            ).all()
            if sp_yields and hl_count_calc > 0:
                nearest_y = min(sp_yields, key=lambda x: abs(float(x.hlso_count or 0) - hl_count_calc))
                if nearest_y:
                    hlso_yield = float(getattr(nearest_y, "hlso_yield_pct", 100) or 100) / 100.0

    denominator = peeling_yield * soaking_yield * hlso_yield
    hoso_rm_cost = (hoso_rate * glaze_factor) / denominator if denominator > 0 else (hoso_rate * glaze_factor)
    total_cost_inr = hoso_rm_cost + exp
    target_usd = total_cost_inr / (exch if exch > 0 else 83.5)
    return round(target_usd, 2)


@router.post("/save")
@router.post("/export_documents/quotation/save", include_in_schema=False)
async def save_quotation(payload: QuotationPayload, request: Request, db: Session = Depends(get_db)):
    comp_code = resolve_company_code(request, db)
    ensure_crm_quotation_schema(db)
    email = request.session.get("email") or "SYSTEM"

    q_date = payload.quotation_date or date.today()
    if payload.valid_until and payload.valid_until < q_date:
        raise HTTPException(status_code=400, detail="Valid Until date cannot be earlier than Quotation Date")

    calc_total = sum((item.quantity_kg * item.rate_per_kg) for item in payload.items) if payload.items else 0.0

    q_no = payload.quotation_no
    if not q_no or db.query(CRMQuotation).filter(CRMQuotation.quotation_no == q_no).first():
        q_no = get_next_quotation_no(db, comp_code)

    quotation = CRMQuotation(
        company_id=comp_code,
        company_name=payload.company_name,
        inquiry_id=getattr(payload, "inquiry_id", None),
        quotation_no=q_no,
        po_number=payload.po_number,
        quotation_date=q_date,
        valid_until=payload.valid_until,
        shipment_date=payload.shipment_date,
        customer_name=payload.customer_name,
        customer_address=payload.customer_address,
        agent=payload.agent,
        country=payload.country,
        production_at=payload.production_at,
        currency=payload.currency,
        exchange_rate=payload.exchange_rate or 83.5,
        incoterm=payload.incoterm,
        payment_terms=payload.payment_terms,
        total_amount=calc_total,
        status=payload.status or "DRAFT",
        remarks=payload.remarks,
        created_by=email,
        created_at=ist_now()
    )
    db.add(quotation)
    db.flush()

    for item in payload.items:
        rate_val = item.bidding_price or item.rate_per_kg or 0.0
        line_amt = item.quantity_kg * rate_val
        calc_pcs = calculate_pieces(item.grade, item.no_of_pieces)
        target_price_val = calculate_target_quotation_price_backend(db, comp_code, item, payload.exchange_rate or 83.5)
        line = CRMQuotationLine(
            quotation_id=quotation.id,
            item_name=item.item_name,
            brand=item.brand,
            packing_style=item.packing_style,
            freezer=item.freezer,
            count_glaze=item.count_glaze,
            weight_glaze=item.weight_glaze,
            species=item.species,
            variety=item.variety,
            grade=item.grade,
            no_of_pieces=calc_pcs,
            no_of_mc=item.no_of_mc or 0,
            quantity_kg=item.quantity_kg,
            rate_per_kg=rate_val,
            hoso_count=item.hoso_count or "",
            target_hoso_rate=item.target_hoso_rate or 0.0,
            expenses=item.expenses or 0.0,
            target_quotation_price=target_price_val,
            bidding_price=rate_val,
            amount=line_amt
        )
        db.add(line)

    pi_created_msg = ""
    if quotation.status == "ACCEPTED":
        created_pi = create_pi_from_quotation(db, quotation, email)
        if created_pi:
            pi_created_msg = f" Auto-created Proforma Invoice {created_pi.pi_no}."

    log_audit(db, comp_code, email, "CREATE", quotation.quotation_no, f"Created Quotation for {payload.customer_name} total {payload.currency} {calc_total:,.2f}{pi_created_msg}")
    db.commit()

    return JSONResponse(content={"success": True, "message": f"Quotation {quotation.quotation_no} created successfully.{pi_created_msg}"})


@router.put("/{quotation_id}")
@router.put("/export_documents/quotation/{quotation_id}", include_in_schema=False)
async def update_quotation(quotation_id: int, payload: QuotationPayload, request: Request, db: Session = Depends(get_db)):
    comp_code = resolve_company_code(request, db)
    email = request.session.get("email") or "SYSTEM"

    quotation = db.query(CRMQuotation).filter(CRMQuotation.id == quotation_id, CRMQuotation.company_id == comp_code).first()
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")

    q_date = payload.quotation_date or quotation.quotation_date or date.today()
    if payload.valid_until and payload.valid_until < q_date:
        raise HTTPException(status_code=400, detail="Valid Until date cannot be earlier than Quotation Date")

    calc_total = sum((item.quantity_kg * (item.bidding_price or item.rate_per_kg or 0.0)) for item in payload.items) if payload.items else 0.0

    quotation.company_name = payload.company_name
    if payload.quotation_no and payload.quotation_no != quotation.quotation_no:
        conflict = db.query(CRMQuotation).filter(
            CRMQuotation.quotation_no == payload.quotation_no,
            CRMQuotation.id != quotation_id
        ).first()
        if not conflict:
            quotation.quotation_no = payload.quotation_no
    quotation.po_number = payload.po_number
    quotation.quotation_date = q_date
    quotation.valid_until = payload.valid_until
    quotation.shipment_date = payload.shipment_date
    quotation.customer_name = payload.customer_name
    quotation.customer_address = payload.customer_address
    quotation.agent = payload.agent
    quotation.country = payload.country
    quotation.production_at = payload.production_at
    quotation.currency = payload.currency
    quotation.exchange_rate = payload.exchange_rate or 83.5
    quotation.incoterm = payload.incoterm
    quotation.payment_terms = payload.payment_terms
    quotation.total_amount = calc_total
    quotation.status = payload.status
    quotation.remarks = payload.remarks
    quotation.updated_by = email
    quotation.updated_at = ist_now()

    # Clear existing lines and re-add
    db.query(CRMQuotationLine).filter(CRMQuotationLine.quotation_id == quotation.id).delete()

    for item in payload.items:
        rate_val = item.bidding_price or item.rate_per_kg or 0.0
        line_amt = item.quantity_kg * rate_val
        calc_pcs = calculate_pieces(item.grade, item.no_of_pieces)
        line = CRMQuotationLine(
            quotation_id=quotation.id,
            item_name=item.item_name,
            brand=item.brand,
            packing_style=item.packing_style,
            freezer=item.freezer,
            count_glaze=item.count_glaze,
            weight_glaze=item.weight_glaze,
            species=item.species,
            variety=item.variety,
            grade=item.grade,
            no_of_pieces=calc_pcs,
            no_of_mc=item.no_of_mc or 0,
            quantity_kg=item.quantity_kg,
            rate_per_kg=rate_val,
            hoso_count=item.hoso_count or "",
            target_hoso_rate=item.target_hoso_rate or 0.0,
            expenses=item.expenses or 0.0,
            target_quotation_price=item.target_quotation_price or 0.0,
            bidding_price=rate_val,
            amount=line_amt
        )
        db.add(line)

    pi_created_msg = ""
    if quotation.status == "ACCEPTED":
        created_pi = create_pi_from_quotation(db, quotation, email)
        if created_pi:
            pi_created_msg = f" Auto-created Proforma Invoice {created_pi.pi_no}."

    log_audit(db, comp_code, email, "UPDATE", quotation.quotation_no, f"Updated Quotation for {payload.customer_name}{pi_created_msg}")
    db.commit()

    return JSONResponse(content={"success": True, "message": f"Quotation {quotation.quotation_no} updated successfully.{pi_created_msg}"})


@router.post("/cancel/{quotation_id}")
@router.post("/export_documents/quotation/cancel/{quotation_id}", include_in_schema=False)
async def cancel_quotation(quotation_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = resolve_company_code(request, db)
    email = request.session.get("email") or "SYSTEM"

    quotation = db.query(CRMQuotation).filter(CRMQuotation.id == quotation_id, CRMQuotation.company_id == comp_code).first()
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")

    quotation.is_cancelled = True
    quotation.status = "CANCELLED"
    quotation.updated_by = email
    quotation.updated_at = ist_now()

    log_audit(db, comp_code, email, "CANCEL", quotation.quotation_no, f"Cancelled Quotation {quotation.quotation_no}")
    db.commit()

    return JSONResponse(content={"success": True, "message": f"Quotation {quotation.quotation_no} cancelled."})


@router.post("/{quotation_id}/approval")
@router.post("/export_documents/quotation/{quotation_id}/approval", include_in_schema=False)
async def quotation_approval(quotation_id: int, payload: ApprovalPayload, request: Request, db: Session = Depends(get_db)):
    comp_code = resolve_company_code(request, db)
    email = request.session.get("email") or "SYSTEM"

    quotation = db.query(CRMQuotation).filter(CRMQuotation.id == quotation_id, CRMQuotation.company_id == comp_code).first()
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")

    quotation.status = payload.decision
    quotation.approved_by = email
    quotation.approved_at = ist_now()
    quotation.approval_remarks = payload.remarks
    quotation.updated_by = email

    pi_created_msg = ""
    if payload.decision == "ACCEPTED":
        created_pi = create_pi_from_quotation(db, quotation, email)
        if created_pi:
            pi_created_msg = f" Auto-created Proforma Invoice {created_pi.pi_no}."

    log_audit(db, comp_code, email, "APPROVAL", quotation.quotation_no, f"Status updated to {payload.decision}. Remarks: {payload.remarks or 'N/A'}{pi_created_msg}")
    db.commit()

    return JSONResponse(content={"success": True, "message": f"Quotation {quotation.quotation_no} status changed to {payload.decision}.{pi_created_msg}"})


@router.get("/register.xlsx")
@router.get("/export_documents/quotation/register.xlsx", include_in_schema=False)
async def export_quotation_register(request: Request, grant: str = Query(...), db: Session = Depends(get_db)):
    comp_code = resolve_company_code(request, db)

    require_download_grant(grant, request)

    quotations = (
        db.query(CRMQuotation)
        .filter(CRMQuotation.company_id == comp_code, CRMQuotation.is_cancelled == False)
        .order_by(desc(CRMQuotation.quotation_date), desc(CRMQuotation.id))
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation Register"

    headers = [
        "Quotation No", "Company", "Location", "Quotation Date", "Valid Until", "Customer Name",
        "Agent", "Country", "Currency", "Exchange Rate", "Incoterm", "Payment Terms", "Total Amount",
        "Status", "Remarks", "Created By", "Approved By"
    ]
    ws.append(headers)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    for q in quotations:
        ws.append([
            q.quotation_no,
            q.company_name or "",
            q.production_at or "",
            q.quotation_date.strftime("%Y-%m-%d") if q.quotation_date else "",
            q.valid_until.strftime("%Y-%m-%d") if q.valid_until else "",
            q.customer_name,
            q.agent or "",
            q.country or "",
            q.currency or "USD",
            float(q.exchange_rate or 83.5),
            q.incoterm or "",
            q.payment_terms or "",
            float(q.total_amount or 0),
            q.status or "DRAFT",
            q.remarks or "",
            q.created_by or "",
            q.approved_by or "",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Quotation_Register_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/{quotation_id}/send-email")
@router.post("/export_documents/quotation/{quotation_id}/send-email", include_in_schema=False)
async def send_quotation_email_endpoint(quotation_id: int, payload: SendQuotationEmailPayload, request: Request, db: Session = Depends(get_db)):
    comp_code = resolve_company_code(request, db)
    email = request.session.get("email") or "SYSTEM"

    quotation = db.query(CRMQuotation).filter(CRMQuotation.id == quotation_id, CRMQuotation.company_id == comp_code).first()
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")

    db_lines = db.query(CRMQuotationLine).filter(CRMQuotationLine.quotation_id == quotation.id).all()
    items_to_use = payload.items if (payload.items and len(payload.items) > 0) else db_lines

    # Global Grouping by normalized product description
    grouped_map = {}
    for item in items_to_use:
        if isinstance(item, BaseModel):
            item_name_val = item.item_name
            species_val = item.species
            variety_val = item.variety
            grade_val = item.grade
        else:
            item_name_val = item.item_name
            species_val = item.species
            variety_val = item.variety
            grade_val = item.grade

        prod_desc = (item_name_val or " ".join(filter(None, [species_val, variety_val, grade_val])) or "Shrimp Item").strip()
        key = prod_desc.lower()
        if key not in grouped_map:
            grouped_map[key] = {'desc': prod_desc, 'items': []}
        grouped_map[key]['items'].append(item)

    tables_html = ""
    line_counter = 1
    for key, group in grouped_map.items():
        desc = group['desc']
        desc_formatted = desc.replace("\n", "<br/>")

        first_l = group['items'][0] if group['items'] else None
        b_name = (first_l.brand if isinstance(first_l, BaseModel) else first_l.brand) if first_l else "—"
        p_name = (first_l.packing_style if isinstance(first_l, BaseModel) else first_l.packing_style) if first_l else "—"

        rows_html = ""
        for l in group['items']:
            idx = line_counter
            line_counter += 1

            if isinstance(l, BaseModel):
                grade_val = l.grade
                no_mc_val = l.no_of_mc
                qty_val = float(l.quantity_kg or 0.0)
                rate_val = float(l.bidding_price or l.rate_per_kg or 0.0)
                pcs_val = l.no_of_pieces
                amt_val = float(l.amount or (rate_val * qty_val))
            else:
                grade_val = l.grade
                no_mc_val = l.no_of_mc
                qty_val = float(l.quantity_kg or 0.0)
                rate_val = float(l.bidding_price or l.rate_per_kg or 0.0)
                pcs_val = l.no_of_pieces
                amt_val = float(l.amount or (rate_val * qty_val))

            rows_html += f"""
            <tr style="border-bottom:1px solid #cbd5e1;">
                <td style="padding:6px 8px;font-weight:bold;color:#475569;border-right:1px solid #cbd5e1;">#{idx}</td>
                <td style="padding:6px 8px;font-weight:bold;color:#0f172a;border-right:1px solid #cbd5e1;">{grade_val or '—'}</td>
                <td style="padding:6px 8px;text-align:center;font-weight:bold;border-right:1px solid #cbd5e1;">{no_mc_val or 0}</td>
                <td style="padding:6px 8px;text-align:right;font-weight:bold;color:#0f172a;border-right:1px solid #cbd5e1;">{qty_val:,.2f} Kg</td>
                <td style="padding:6px 8px;text-align:center;color:#475569;border-right:1px solid #cbd5e1;">{pcs_val or '—'}</td>
                <td style="padding:6px 8px;text-align:right;font-weight:bold;color:#0f172a;border-right:1px solid #cbd5e1;">${rate_val:.2f}</td>
                <td style="padding:6px 8px;text-align:right;font-weight:bold;color:#0f172a;">${amt_val:,.2f}</td>
            </tr>
            """

        tables_html += f"""
        <div style="margin-bottom:16px;border:1px solid #cbd5e1;border-radius:6px;overflow:hidden;background:#ffffff;">
          <div style="background:#f8fafc;border-bottom:1px solid #cbd5e1;padding:8px 12px;">
            <div style="font-size:12px;font-weight:bold;color:#0f172a;line-height:1.4;">{desc_formatted}</div>
            <div style="font-size:10.5px;color:#475569;margin-top:4px;font-weight:bold;">
              BRAND: <span style="color:#0f172a;">{b_name or '—'}</span> &nbsp;&bull;&nbsp; PACKING: <span style="color:#0f172a;">{p_name or '—'}</span>
            </div>
          </div>
          <div class="table-responsive" style="width:100%;overflow-x:auto;-webkit-overflow-scrolling:touch;">
            <table class="spec-table" style="width:100%;min-width:440px;border-collapse:collapse;font-size:10px;border:none;table-layout:fixed;">
              <thead>
                <tr style="background:#f1f5f9;color:#0f172a;text-align:left;border-bottom:2px solid #64748b;">
                  <th style="padding:6px 4px;width:5%;border-right:1px solid #cbd5e1;">#</th>
                  <th style="padding:6px 4px;width:17%;border-right:1px solid #cbd5e1;">Grade</th>
                  <th style="padding:6px 4px;text-align:center;width:9%;border-right:1px solid #cbd5e1;">MC</th>
                  <th style="padding:6px 4px;text-align:right;width:20%;border-right:1px solid #cbd5e1;">Qty (KG)</th>
                  <th style="padding:6px 4px;text-align:center;width:12%;border-right:1px solid #cbd5e1;">No. of Pcs</th>
                  <th style="padding:6px 4px;text-align:right;width:16%;border-right:1px solid #cbd5e1;">Price ($/Kg)</th>
                  <th style="padding:6px 4px;text-align:right;width:21%;">Total ($)</th>
                </tr>
              </thead>
              <tbody>
                {rows_html}
              </tbody>
            </table>
          </div>
        </div>
        """

    ship_date_str = str(quotation.shipment_date) if quotation.shipment_date else (str(quotation.quotation_date) if quotation.quotation_date else date.today().strftime("%Y-%m-%d"))
    default_footer = f"Terms & Conditions:\n• Shipment Date: {ship_date_str}\n• Quotation Validity: Valid until {quotation.valid_until}\n• Payment Terms: {quotation.payment_terms or '100% LC at sight'}\n• Incoterms: {quotation.incoterm or 'FOB'}"
    header_formatted = (payload.header_text or "").replace("\n", "<br/>")
    footer_formatted = (payload.footer_text or default_footer).replace("\n", "<br/>")
    signoff_formatted = (payload.signoff_text or "").replace("\n", "<br/>")

    html_content = f"""
    <!doctype html>
    <html>
    <head>
      <meta charset="utf-8"/>
      <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
      <style>
        body {{ margin: 0; padding: 8px; font-family: Arial, Helvetica, sans-serif; background-color: #ffffff; color: #0f172a; }}
        .email-card {{ max-width: 820px; margin: 0 auto; background: #ffffff; border: 1px solid #cbd5e1; border-radius: 6px; padding: 20px; box-shadow: 0 4px 12px rgba(0,0,0,0.05); }}
        .table-responsive {{ width: 100%; overflow-x: auto; -webkit-overflow-scrolling: touch; }}
        .spec-table {{ width: 100%; min-width: 480px; border-collapse: collapse; font-size: 10px; color: #0f172a; table-layout: fixed; }}
        @media only screen and (max-width: 600px) {{
          body {{ padding: 2px !important; }}
          .email-card {{ padding: 10px !important; border-radius: 4px !important; }}
          .header-flex {{ flex-direction: column !important; align-items: flex-start !important; gap: 4px !important; }}
          .header-right {{ text-align: left !important; }}
          .spec-table {{ font-size: 8px !important; min-width: 100% !important; table-layout: fixed !important; }}
          .spec-table th, .spec-table td {{ padding: 3px 2px !important; line-height: 1.2 !important; box-sizing: border-box !important; word-break: break-word !important; }}
        }}
      </style>
    </head>
    <body style="margin:0;padding:12px;background:#ffffff;font-family:Arial,Helvetica,sans-serif;">
      <div class="email-card" style="max-width:820px;margin:0 auto;background:#ffffff;border:1px solid #cbd5e1;border-radius:6px;padding:24px;">
        
        <div class="header-flex" style="border-bottom:2px solid #0f172a;padding-bottom:12px;margin-bottom:18px;display:flex;justify-content:space-between;align-items:flex-end;">
          <div>
            <h2 style="margin:0;color:#0f172a;font-size:20px;letter-spacing:-0.3px;">{quotation.company_name or 'BKNR ERP'}</h2>
            <div style="font-size:12px;color:#475569;margin-top:3px;font-weight:bold;text-transform:uppercase;letter-spacing:0.5px;">Commercial Price Quotation</div>
          </div>
          <div class="header-right" style="text-align:right;">
            <div style="font-size:16px;font-weight:bold;color:#0f172a;">Ref: {quotation.quotation_no}</div>
            <div style="font-size:12px;color:#475569;">Date: {quotation.quotation_date}</div>
          </div>
        </div>

        <div style="font-size:13px;line-height:1.6;margin-bottom:18px;color:#1e293b;">
          {header_formatted}
        </div>

        {tables_html}

        <div style="background:#f8fafc;border:1px solid #0f172a;border-radius:4px;padding:10px 14px;margin-bottom:18px;display:flex;justify-content:space-between;align-items:center;font-weight:bold;">
          <span style="font-size:12.5px;color:#0f172a;text-transform:uppercase;">TOTAL OFFERED VALUE ({quotation.currency}):</span>
          <span style="font-size:15px;color:#0f172a;">${quotation.total_amount:,.2f}</span>
        </div>
        <div style="font-size:12px;line-height:1.6;color:#475569;background:#f8fafc;padding:12px;border-radius:6px;border:1px solid #e2e8f0;margin-bottom:18px;">
          {footer_formatted}
        </div>

        <div style="font-size:13px;line-height:1.6;color:#1e293b;">
          {signoff_formatted}
        </div>

      </div>
    </body>
    </html>
    """

    from app.utils.email_service import send_email
    from_addr = payload.from_email or email or "noreply@bknr.in"
    send_email(payload.to_email, payload.subject, html_content, from_email=from_addr, reply_to=from_addr)

    outbound = CRMQuotationReply(
        quotation_id=quotation.id,
        quotation_no=quotation.quotation_no,
        sender_email=from_addr,
        recipient_email=payload.to_email,
        subject=payload.subject,
        message_body=(payload.header_text or "") + "\n\n[Table specifications included]\n\n" + (payload.footer_text or "") + "\n\n" + (payload.signoff_text or ""),
        direction="OUTBOUND",
        received_at=ist_now()
    )
    db.add(outbound)

    quotation.status = "SENT"
    log_audit(db, comp_code, email, "SEND_EMAIL", quotation.quotation_no, f"Sent Quotation email to {payload.to_email}")
    db.commit()

    return JSONResponse(content={"success": True, "message": f"Quotation {quotation.quotation_no} email sent successfully to {payload.to_email}."})


@router.get("/{quotation_id}/replies")
@router.get("/export_documents/quotation/{quotation_id}/replies", include_in_schema=False)
async def get_quotation_replies(quotation_id: int, request: Request, db: Session = Depends(get_db)):
    ensure_crm_quotation_schema(db)
    comp_code = resolve_company_code(request, db)
    quotation = db.query(CRMQuotation).filter(CRMQuotation.id == quotation_id, CRMQuotation.company_id == comp_code).first()
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")

    replies = db.query(CRMQuotationReply).filter(
        CRMQuotationReply.quotation_id == quotation_id
    ).order_by(CRMQuotationReply.received_at.asc()).all()

    data = [{
        "id": r.id,
        "sender_email": r.sender_email,
        "recipient_email": r.recipient_email,
        "subject": r.subject,
        "message_body": r.message_body,
        "received_at": r.received_at.strftime("%Y-%m-%d %H:%M:%S") if r.received_at else "",
        "direction": r.direction,
        "attachments_json": getattr(r, "attachments_json", None)
    } for r in replies]

    return JSONResponse(content={"success": True, "replies": data, "quotation_no": quotation.quotation_no, "status": quotation.status})


class InboundReplyPayload(BaseModel):
    message_body: str
    sender_email: Optional[str] = None


class OutboundChatbotPayload(BaseModel):
    message_body: str
    to_email: str
    subject: Optional[str] = None
    attachments: Optional[List[Dict[str, Any]]] = None


@router.post("/{quotation_id}/send-chatbot-reply")
@router.post("/export_documents/quotation/{quotation_id}/send-chatbot-reply", include_in_schema=False)
async def send_chatbot_reply(quotation_id: int, payload: OutboundChatbotPayload, request: Request, db: Session = Depends(get_db)):
    """Send AI chatbot reply as email to customer with attachments and log it as OUTBOUND."""
    comp_code = resolve_company_code(request, db)
    from_email_addr = request.session.get("email") or ""

    quotation = db.query(CRMQuotation).filter(
        CRMQuotation.id == quotation_id, CRMQuotation.company_id == comp_code
    ).first()
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")

    subject = payload.subject or f"Re: Price Quotation #{quotation.quotation_no} — {quotation.customer_name}"

    # Plain-text → HTML (preserve line breaks)
    html_body = "<div style='font-family:Arial,sans-serif;font-size:14px;line-height:1.7;color:#202124;'>" + \
        payload.message_body.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br/>") + \
        "</div>"

    # Process attachments
    email_attachments = []
    db_att_list = []
    if payload.attachments:
        import base64
        for att in payload.attachments:
            b64_str = att.get("b64_data") or att.get("data_url") or ""
            mime_type = att.get("mime_type") or "application/pdf"
            filename = att.get("filename") or "attachment.pdf"

            if "," in b64_str:
                b64_str = b64_str.split(",", 1)[1]

            try:
                raw_bytes = base64.b64decode(b64_str)
                email_attachments.append((filename, raw_bytes, mime_type))
                db_att_list.append({
                    "filename": filename,
                    "mime_type": mime_type,
                    "data_url": f"data:{mime_type};base64,{b64_str}"
                })
            except Exception as e:
                logger.error(f"Failed to decode base64 attachment: {e}")

    import json
    saved_att_json = json.dumps(db_att_list) if db_att_list else None

    from app.utils.email_service import send_email
    try:
        send_email(
            payload.to_email,
            subject,
            html_body,
            text=payload.message_body,
            from_email=from_email_addr,
            reply_to=from_email_addr,
            attachments=email_attachments if email_attachments else None
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Email delivery failed: {exc}")

    # Log as OUTBOUND in DB
    outbound = CRMQuotationReply(
        quotation_id=quotation.id,
        quotation_no=quotation.quotation_no,
        sender_email=from_email_addr or "system@bknr.in",
        recipient_email=payload.to_email,
        subject=subject,
        message_body=payload.message_body,
        direction="OUTBOUND",
        attachments_json=saved_att_json,
        received_at=ist_now(),
    )
    db.add(outbound)
    quotation.status = "SENT"
    log_audit(db, comp_code, from_email_addr, "CHATBOT_EMAIL_SENT", quotation.quotation_no,
              f"AI chatbot reply sent to {payload.to_email}")
    db.commit()

    return JSONResponse(content={"success": True, "message": f"Reply sent to {payload.to_email} successfully."})


@router.post("/{quotation_id}/post-reply")
@router.post("/export_documents/quotation/{quotation_id}/post-reply", include_in_schema=False)
async def post_customer_reply(quotation_id: int, payload: InboundReplyPayload, request: Request, db: Session = Depends(get_db)):
    comp_code = resolve_company_code(request, db)
    email = request.session.get("email") or "SYSTEM"

    quotation = db.query(CRMQuotation).filter(CRMQuotation.id == quotation_id, CRMQuotation.company_id == comp_code).first()
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")

    inbound = CRMQuotationReply(
        quotation_id=quotation.id,
        quotation_no=quotation.quotation_no,
        sender_email=payload.sender_email or quotation.customer_name or "customer@buyer.com",
        recipient_email=email,
        subject=f"Re: Price Quotation #{quotation.quotation_no}",
        message_body=payload.message_body,
        direction="INBOUND",
        received_at=ist_now()
    )
    db.add(inbound)
    quotation.status = "CUSTOMER REPLIED"
    log_audit(db, comp_code, email, "CUSTOMER_REPLY", quotation.quotation_no, f"Logged customer reply for {quotation.quotation_no}")
    db.commit()

    return JSONResponse(content={"success": True, "message": "Customer reply logged successfully."})


@router.post("/sync-inbound-emails")
@router.post("/export_documents/quotation/sync-inbound-emails", include_in_schema=False)
async def sync_inbound_emails_endpoint(request: Request, db: Session = Depends(get_db)):
    from app.services.email_poller import poll_inbound_emails
    result = poll_inbound_emails(db)
    return JSONResponse(content=result)


@router.post("/{quotation_id}/ai-chatbot")
@router.post("/export_documents/quotation/{quotation_id}/ai-chatbot", include_in_schema=False)
async def generate_ai_chatbot_proposal(quotation_id: int, payload: InboundReplyPayload, request: Request, db: Session = Depends(get_db)):
    comp_code = resolve_company_code(request, db)
    quotation = db.query(CRMQuotation).filter(CRMQuotation.id == quotation_id, CRMQuotation.company_id == comp_code).first()
    if not quotation:
        raise HTTPException(status_code=404, detail="Quotation not found")

    from app.services.quotation_ai_bot import generate_ai_bot_reply

    q_items = []
    for l in quotation.lines:
        q_items.append({
            "item_name": l.item_name,
            "grade": l.grade,
            "quantity_kg": l.quantity_kg,
            "rate_per_kg": l.rate_per_kg,
            "bidding_price": l.bidding_price,
            "packing_style": l.packing_style,
            "brand": l.brand
        })

    _, comp_name = resolve_session_company_info(request, db)

    q_dict = {
        "quotation_no": quotation.quotation_no,
        "customer_name": quotation.customer_name,
        "total_amount": float(quotation.total_amount or 0.0),
        "currency": quotation.currency or "USD",
        "shipment_date": str(quotation.shipment_date) if quotation.shipment_date else "Prompt",
        "valid_until": str(quotation.valid_until) if quotation.valid_until else "30 Days",
        "company_name": comp_name or quotation.company_name or "BKNR ERP Solutions",
        "items": q_items
    }

    ai_res = generate_ai_bot_reply(q_dict, payload.message_body)
    return JSONResponse(content=ai_res)


