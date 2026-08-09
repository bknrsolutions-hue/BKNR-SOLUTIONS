import json
import logging
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

from fastapi import APIRouter, Request, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc

from app.database import get_db
from app.database.models.invoices import (
    ExportShipment, PackingList, ContainerStuffing, ShippingBill,
    BillOfLading, HealthCertificate, CommercialInvoice, ExportDocumentFile, ExportComplianceTracker
)
from app.utils.download_security import require_download_grant
from app.utils.data_management_audit import log_data_management_action
from app.services.cache import invalidate_company_cache

from app.routers.export_documents.common import (
    templates,
    get_export_company_profile,
    get_export_record_or_404,
    export_doc_config,
    build_document_payload,
    process_items_with_spans,
    render_document_pdf,
    store_export_pdf,
    set_document_path,
    write_audit,
    invalidate_export_cache,
    safe_filename,
    document_register_workbook,
    get_invoice_packing_rows,
    ExportShipmentSchema,
    PackingListSchema,
    PackingListBulkSchema,
    ContainerStuffingSchema,
    ShippingBillSchema,
    BillOfLadingSchema,
    HealthCertificateSchema,
)

logger = logging.getLogger(__name__)
router = APIRouter()


def require_company_invoice(db: Session, company_id: str, invoice_no: str) -> CommercialInvoice:
    invoice = db.query(CommercialInvoice).filter(
        CommercialInvoice.company_id == company_id,
        CommercialInvoice.invoice_no == invoice_no,
        CommercialInvoice.is_cancelled != True,
    ).first()
    if not invoice:
        raise ValueError("Select a valid commercial invoice for this company")
    return invoice


def refresh_compliance(db: Session, company_id: str, shipment_no: str) -> None:
    tracker = db.query(ExportComplianceTracker).filter(
        ExportComplianceTracker.company_id == company_id,
        ExportComplianceTracker.shipment_no == shipment_no,
    ).first()
    if not tracker:
        return
    invoice = db.query(CommercialInvoice).filter(
        CommercialInvoice.company_id == company_id,
        CommercialInvoice.shipment_no == shipment_no,
        CommercialInvoice.is_cancelled != True,
    ).first()
    tracker.invoice_pending = invoice is None
    if not invoice:
        tracker.packing_list_pending = True
        tracker.health_cert_pending = True
        tracker.shipping_bill_pending = True
        tracker.bl_pending = True
        return
    common = (lambda model: db.query(model).filter(
        model.company_id == company_id,
        model.invoice_no == invoice.invoice_no,
        model.is_cancelled != True,
    ).first() is None)
    tracker.packing_list_pending = common(PackingList)
    tracker.health_cert_pending = common(HealthCertificate)
    tracker.shipping_bill_pending = common(ShippingBill)
    tracker.bl_pending = common(BillOfLading)


def apply_invoice_container_defaults(db: Session, company_id: str, invoices: list[CommercialInvoice]) -> list[CommercialInvoice]:
    for invoice in invoices:
        if invoice.container_no:
            continue
        stuffing = db.query(ContainerStuffing).filter(
            ContainerStuffing.company_id == company_id,
            ContainerStuffing.invoice_no == invoice.invoice_no,
            ContainerStuffing.is_cancelled != True,
        ).order_by(desc(ContainerStuffing.id)).first()
        if stuffing:
            invoice.container_no = stuffing.container_no
            continue
        shipment = db.query(ExportShipment).filter(
            ExportShipment.company_id == company_id,
            ExportShipment.shipment_no == invoice.shipment_no,
            ExportShipment.is_cancelled != True,
        ).first()
        if shipment and shipment.container_no:
            invoice.container_no = shipment.container_no
    return invoices


# ============================================================
# EXPORT SHIPMENTS
# ============================================================
@router.get("/export_shipment/entry", response_class=HTMLResponse)
def export_shipment_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return RedirectResponse("/", status_code=302)
    history = db.query(ExportShipment).filter(ExportShipment.company_id == comp_code, ExportShipment.is_cancelled != True).order_by(desc(ExportShipment.created_at)).all()
    return templates.TemplateResponse(request=request, name="export_documents/export_shipment.html", context={"history": history, "company_id": comp_code})


@router.post("/export_shipment/save")
def export_shipment_save(request: Request, payload: ExportShipmentSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    
    exists = db.query(ExportShipment).filter(ExportShipment.company_id == comp_code, ExportShipment.shipment_no == payload.shipment_no).first()
    if exists:
        return JSONResponse({"success": False, "message": "Shipment Number already registered"}, status_code=400)
    
    entry = ExportShipment(company_id=comp_code, created_by=email, **payload.model_dump())
    db.add(entry)
    db.flush()
    
    compliance = ExportComplianceTracker(
        company_id=comp_code,
        shipment_no=entry.shipment_no,
        invoice_pending=True,
        packing_list_pending=True,
        health_cert_pending=True,
        shipping_bill_pending=True,
        bl_pending=True,
        payment_pending=True
    )
    db.add(compliance)
    
    write_audit(db, "export_shipments", entry.id, comp_code, "CREATE", "NONE", f"Shipment Registered: {payload.shipment_no}", email)
    db.commit()
    return {"success": True, "message": "Export shipment registered successfully"}


@router.post("/export_shipment/delete/{log_id}")
def export_shipment_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(ExportShipment).filter(ExportShipment.id == log_id, ExportShipment.company_id == comp_code).first()
    if entry:
        if entry.invoice_no:
            return JSONResponse({"success": False, "message": "Cancel linked export documents before cancelling this shipment"}, status_code=400)
        write_audit(db, "export_shipments", entry.id, comp_code, "CANCEL", entry.status, "CANCELLED", email)
        entry.is_cancelled = True
        entry.status = "CANCELLED"
        db.commit()
        return {"success": True, "message": "Export shipment deleted successfully"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# PACKING LIST
# ============================================================
@router.get("/packing_list/entry", response_class=HTMLResponse)
def packing_list_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return RedirectResponse("/", status_code=302)
    history = db.query(PackingList).filter(PackingList.company_id == comp_code, PackingList.is_cancelled != True).order_by(desc(PackingList.created_at)).all()
    invoices = apply_invoice_container_defaults(db, comp_code, db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.is_cancelled != True).all())
    return templates.TemplateResponse(request=request, name="export_documents/packing_list.html", context={"history": history, "invoices": invoices, "company_id": comp_code})


@router.post("/packing_list/save")
def packing_list_save(request: Request, payload: PackingListSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    try:
        invoice = require_company_invoice(db, comp_code, payload.invoice_no)
    except ValueError as exc:
        return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
    from app.routers.export_documents.common import repost_invoice_cogs
    entry = PackingList(company_id=comp_code, created_by=email, **payload.model_dump())
    db.add(entry)
    db.flush()
    cogs_value = repost_invoice_cogs(db, comp_code, invoice, email)
    refresh_compliance(db, comp_code, invoice.shipment_no)
    write_audit(db, "packing_lists", entry.id, comp_code, "CREATE", "NONE", f"Packing Item: {payload.packing_no}", email)
    db.commit()
    return {"success": True, "message": "Packing list and COGS accounting recorded successfully", "cogs_value": cogs_value}


@router.post("/packing_list/save-bulk")
def packing_list_save_bulk(request: Request, payload: PackingListBulkSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    try:
        invoice = require_company_invoice(db, comp_code, payload.invoice_no)
        from app.routers.export_documents.common import repost_invoice_cogs
        created_entries = []
        common = {
            "company_id": comp_code,
            "packing_no": payload.packing_no.strip(),
            "invoice_no": payload.invoice_no,
            "po_number": payload.po_number,
            "container_no": payload.container_no,
            "buyer_name": payload.buyer_name,
            "created_by": email,
        }
        if not common["packing_no"]:
            raise ValueError("Packing document number is required")
        for index, item in enumerate(payload.items, start=1):
            entry = PackingList(**common, invoice_item_no=index, **item.model_dump())
            db.add(entry)
            db.flush()
            created_entries.append(entry)
            write_audit(db, "packing_lists", entry.id, comp_code, "CREATE", "NONE", f"Packing {payload.packing_no} · Line {index}", email)
        cogs_value = repost_invoice_cogs(db, comp_code, invoice, email)
        refresh_compliance(db, comp_code, invoice.shipment_no)
        db.commit()
        invalidate_export_cache(comp_code)
        return {
            "success": True,
            "message": f"Packing list saved successfully with {len(created_entries)} line items",
            "record_ids": [entry.id for entry in created_entries],
            "cogs_value": cogs_value,
        }
    except Exception as exc:
        db.rollback()
        logger.exception("Bulk packing list save failed")
        return JSONResponse({"success": False, "message": str(exc)}, status_code=400)


@router.post("/packing_list/delete/{log_id}")
def packing_list_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(PackingList).filter(PackingList.id == log_id, PackingList.company_id == comp_code).first()
    if entry:
        invoice = require_company_invoice(db, comp_code, entry.invoice_no)
        from app.routers.export_documents.common import repost_invoice_cogs
        write_audit(db, "packing_lists", entry.id, comp_code, "CANCEL", "ACTIVE", "CANCELLED", email)
        entry.is_cancelled = True
        repost_invoice_cogs(db, comp_code, invoice, email)
        refresh_compliance(db, comp_code, invoice.shipment_no)
        db.commit()
        return {"success": True, "message": "Packing list entry deleted successfully"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# CONTAINER STUFFING
# ============================================================
@router.get("/container_stuffing/entry", response_class=HTMLResponse)
def container_stuffing_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return RedirectResponse("/", status_code=302)
    history = db.query(ContainerStuffing).filter(ContainerStuffing.company_id == comp_code, ContainerStuffing.is_cancelled != True).order_by(desc(ContainerStuffing.stuffing_date)).all()
    invoices = apply_invoice_container_defaults(db, comp_code, db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.is_cancelled != True).all())
    return templates.TemplateResponse(request=request, name="export_documents/container_stuffing.html", context={"history": history, "invoices": invoices, "company_id": comp_code})


@router.post("/container_stuffing/save")
def container_stuffing_save(request: Request, payload: ContainerStuffingSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    invoice = None
    if payload.invoice_no:
        try:
            invoice = require_company_invoice(db, comp_code, payload.invoice_no)
        except ValueError as exc:
            return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
    
    exists = db.query(ContainerStuffing).filter(ContainerStuffing.company_id == comp_code, ContainerStuffing.container_no == payload.container_no).first()
    if exists:
        return JSONResponse({"success": False, "message": "Container stuffing already logged"}, status_code=400)
    
    entry = ContainerStuffing(company_id=comp_code, created_by=email, **payload.model_dump())
    db.add(entry)
    db.flush()
    
    if payload.invoice_no and invoice:
        invoice.container_no = entry.container_no
        shipment = db.query(ExportShipment).filter(ExportShipment.company_id == comp_code, ExportShipment.shipment_no == invoice.shipment_no).first()
        if shipment:
            shipment.container_no = entry.container_no
                
    write_audit(db, "container_stuffing", entry.id, comp_code, "CREATE", "NONE", f"Container Stuffing: {payload.container_no}", email)
    db.commit()
    return {"success": True, "message": "Container stuffing log recorded successfully"}


@router.post("/container_stuffing/delete/{log_id}")
def container_stuffing_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(ContainerStuffing).filter(ContainerStuffing.id == log_id, ContainerStuffing.company_id == comp_code).first()
    if entry:
        write_audit(db, "container_stuffing", entry.id, comp_code, "CANCEL", "ACTIVE", "CANCELLED", email)
        entry.is_cancelled = True
        db.commit()
        return {"success": True, "message": "Container stuffing record deleted successfully"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# SHIPPING BILL
# ============================================================
@router.get("/shipping_bill/entry", response_class=HTMLResponse)
def shipping_bill_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return RedirectResponse("/", status_code=302)
    history = db.query(ShippingBill).filter(ShippingBill.company_id == comp_code, ShippingBill.is_cancelled != True).order_by(desc(ShippingBill.shipping_bill_date)).all()
    invoices = apply_invoice_container_defaults(db, comp_code, db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.is_cancelled != True).all())
    return templates.TemplateResponse(request=request, name="export_documents/shipping_bill.html", context={"history": history, "invoices": invoices, "company_id": comp_code})


@router.post("/shipping_bill/save")
def shipping_bill_save(request: Request, payload: ShippingBillSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    try:
        invoice = require_company_invoice(db, comp_code, payload.invoice_no)
    except ValueError as exc:
        return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
    
    exists = db.query(ShippingBill).filter(ShippingBill.company_id == comp_code, ShippingBill.shipping_bill_no == payload.shipping_bill_no).first()
    if exists:
        return JSONResponse({"success": False, "message": "Shipping Bill Number already registered"}, status_code=400)
    
    entry = ShippingBill(company_id=comp_code, created_by=email, **payload.model_dump())
    db.add(entry)
    db.flush()
    
    refresh_compliance(db, comp_code, invoice.shipment_no)
    write_audit(db, "shipping_bills", entry.id, comp_code, "CREATE", "NONE", f"Shipping Bill: {payload.shipping_bill_no}", email)
    db.commit()
    return {"success": True, "message": "Shipping Bill successfully registered"}


@router.post("/shipping_bill/delete/{log_id}")
def shipping_bill_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(ShippingBill).filter(ShippingBill.id == log_id, ShippingBill.company_id == comp_code).first()
    if entry:
        invoice = require_company_invoice(db, comp_code, entry.invoice_no)
        write_audit(db, "shipping_bills", entry.id, comp_code, "CANCEL", "ACTIVE", "CANCELLED", email)
        entry.is_cancelled = True
        refresh_compliance(db, comp_code, invoice.shipment_no)
        db.commit()
        return {"success": True, "message": "Shipping bill record removed successfully"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# BILL OF LADING
# ============================================================
@router.get("/bill_of_lading/entry", response_class=HTMLResponse)
def bill_of_lading_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return RedirectResponse("/", status_code=302)
    history = db.query(BillOfLading).filter(BillOfLading.company_id == comp_code, BillOfLading.is_cancelled != True).order_by(desc(BillOfLading.bl_date)).all()
    invoices = apply_invoice_container_defaults(db, comp_code, db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.is_cancelled != True).all())
    return templates.TemplateResponse(request=request, name="export_documents/bill_of_lading.html", context={"history": history, "invoices": invoices, "company_id": comp_code})


@router.post("/bill_of_lading/save")
def bill_of_lading_save(request: Request, payload: BillOfLadingSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    try:
        invoice = require_company_invoice(db, comp_code, payload.invoice_no)
    except ValueError as exc:
        return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
    
    exists = db.query(BillOfLading).filter(BillOfLading.company_id == comp_code, BillOfLading.bl_no == payload.bl_no).first()
    if exists:
        return JSONResponse({"success": False, "message": "BL Number already registered"}, status_code=400)
    
    entry = BillOfLading(company_id=comp_code, created_by=email, **payload.model_dump())
    db.add(entry)
    db.flush()
    
    refresh_compliance(db, comp_code, invoice.shipment_no)
    write_audit(db, "bill_of_ladings", entry.id, comp_code, "CREATE", "NONE", f"BL Entry: {payload.bl_no}", email)
    db.commit()
    return {"success": True, "message": "Bill of Lading recorded successfully"}


@router.post("/bill_of_lading/delete/{log_id}")
def bill_of_lading_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(BillOfLading).filter(BillOfLading.id == log_id, BillOfLading.company_id == comp_code).first()
    if entry:
        invoice = require_company_invoice(db, comp_code, entry.invoice_no)
        write_audit(db, "bill_of_ladings", entry.id, comp_code, "CANCEL", "ACTIVE", "CANCELLED", email)
        entry.is_cancelled = True
        refresh_compliance(db, comp_code, invoice.shipment_no)
        db.commit()
        return {"success": True, "message": "Bill of lading entry removed"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# HEALTH CERTIFICATE
# ============================================================
@router.get("/health_certificate/entry", response_class=HTMLResponse)
def health_certificate_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return RedirectResponse("/", status_code=302)
    history = db.query(HealthCertificate).filter(HealthCertificate.company_id == comp_code, HealthCertificate.is_cancelled != True).order_by(desc(HealthCertificate.issue_date)).all()
    invoices = apply_invoice_container_defaults(db, comp_code, db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.is_cancelled != True).all())
    return templates.TemplateResponse(request=request, name="export_documents/health_certificate.html", context={"history": history, "invoices": invoices, "company_id": comp_code})


@router.post("/health_certificate/save")
def health_certificate_save(request: Request, payload: HealthCertificateSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    try:
        invoice = require_company_invoice(db, comp_code, payload.invoice_no)
    except ValueError as exc:
        return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
    
    exists = db.query(HealthCertificate).filter(HealthCertificate.company_id == comp_code, HealthCertificate.certificate_no == payload.certificate_no).first()
    if exists:
        return JSONResponse({"success": False, "message": "Certificate Number already exists"}, status_code=400)
    
    entry = HealthCertificate(company_id=comp_code, created_by=email, **payload.model_dump())
    db.add(entry)
    db.flush()
    
    refresh_compliance(db, comp_code, invoice.shipment_no)
    write_audit(db, "health_certificates", entry.id, comp_code, "CREATE", "NONE", f"Health Cert: {payload.certificate_no}", email)
    db.commit()
    return {"success": True, "message": "Health Certificate recorded successfully"}


@router.post("/health_certificate/delete/{log_id}")
def health_certificate_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(HealthCertificate).filter(HealthCertificate.id == log_id, HealthCertificate.company_id == comp_code).first()
    if entry:
        invoice = require_company_invoice(db, comp_code, entry.invoice_no)
        write_audit(db, "health_certificates", entry.id, comp_code, "CANCEL", "ACTIVE", "CANCELLED", email)
        entry.is_cancelled = True
        refresh_compliance(db, comp_code, invoice.shipment_no)
        db.commit()
        return {"success": True, "message": "Health certificate deleted successfully"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# EXCEL REGISTERS & DOSSIER EXPORT
# ============================================================
@router.get("/registers.xlsx")
def export_all_document_registers(request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Unauthorized")
    content = document_register_workbook(db, comp_code)
    log_data_management_action(comp_code, "REGISTER", "All Export Registers", "Success", "Downloaded complete export-document workbook")
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Export_Document_Registers_{safe_filename(comp_code)}.xlsx"'},
    )


@router.get("/{doc_type}/register.xlsx")
def export_document_register(doc_type: str, request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Unauthorized")
    if doc_type not in export_doc_config():
        raise HTTPException(status_code=404, detail="Unsupported document type")
    content = document_register_workbook(db, comp_code, doc_type)
    label = doc_type.replace("_", " ").title()
    log_data_management_action(comp_code, "REGISTER", f"{label} Register", "Success", f"Downloaded {label} tenant register")
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_filename(doc_type)}_Register.xlsx"'},
    )


@router.get("/shipment/{shipment_id}/dossier.zip")
def export_shipment_dossier(shipment_id: int, request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Unauthorized")
    shipment = db.query(ExportShipment).filter(
        ExportShipment.id == shipment_id,
        ExportShipment.company_id == comp_code,
    ).first()
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")

    records = [("export_shipment", shipment)]
    invoice = None
    if shipment.invoice_no:
        invoice = db.query(CommercialInvoice).filter(
            CommercialInvoice.company_id == comp_code,
            CommercialInvoice.invoice_no == shipment.invoice_no,
        ).first()
    if invoice:
        records.append(("commercial_invoice", invoice))
        linked_models = (
            ("packing_list", PackingList),
            ("container_stuffing", ContainerStuffing),
            ("shipping_bill", ShippingBill),
            ("bill_of_lading", BillOfLading),
            ("health_certificate", HealthCertificate),
        )
        for doc_type, model in linked_models:
            linked_rows = db.query(model).filter(
                model.company_id == comp_code,
                model.invoice_no == invoice.invoice_no,
            ).order_by(model.id).all()
            records.extend((doc_type, row) for row in linked_rows)

    output = BytesIO()
    company_profile = get_export_company_profile(db, comp_code)
    manifest_rows = []
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for index, (doc_type, row) in enumerate(records, start=1):
            cfg = export_doc_config()[doc_type]
            document_no = str(getattr(row, cfg["no"], row.id))
            file_name = f"{index:02d}_{safe_filename(cfg['title'])}_{safe_filename(document_no)}.pdf"
            archive.writestr(
                file_name,
                render_document_pdf(
                    cfg,
                    row,
                    comp_code,
                    doc_type,
                    company_profile,
                    get_invoice_packing_rows(db, row),
                ),
            )
            manifest_rows.append((cfg["title"], document_no, "SYSTEM GENERATED", file_name))

        supporting = db.query(ExportDocumentFile).filter(
            ExportDocumentFile.company_id == comp_code,
            ExportDocumentFile.module_name == "export_supporting",
            ExportDocumentFile.record_id == shipment.id,
            ExportDocumentFile.is_current == True,
        ).order_by(ExportDocumentFile.document_kind).all()
        for file_row in supporting:
            file_name = f"Supporting/{safe_filename(file_row.document_kind)}_{safe_filename(file_row.file_name)}"
            archive.writestr(file_name, file_row.file_bytes)
            manifest_rows.append((file_row.document_kind, file_row.document_no or "", "UPLOADED COPY", file_name))

        present_types = {doc_type for doc_type, _ in records}
        for required_type in (
            "export_shipment", "commercial_invoice", "packing_list", "container_stuffing",
            "shipping_bill", "bill_of_lading", "health_certificate",
        ):
            if required_type not in present_types:
                manifest_rows.append((export_doc_config()[required_type]["title"], "", "MISSING", ""))

        manifest = openpyxl.Workbook()
        sheet = manifest.active
        sheet.title = "Shipment Dossier"
        sheet.append(["BKNR EXPORT SHIPMENT DOSSIER"])
        sheet.append(["Shipment No", shipment.shipment_no])
        sheet.append(["Company", comp_code])
        sheet.append(["Generated UTC", datetime.utcnow().strftime("%d-%b-%Y %H:%M")])
        sheet.append([])
        sheet.append(["Document Type", "Document No", "Source", "File"])
        for row in manifest_rows:
            sheet.append(row)
        for cell in sheet[6]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="176B87")
        for column, width in zip("ABCD", (28, 24, 20, 64)):
            sheet.column_dimensions[column].width = width
        manifest_output = BytesIO()
        manifest.save(manifest_output)
        archive.writestr("00_Dossier_Manifest.xlsx", manifest_output.getvalue())

    write_audit(
        db, "export_shipments", shipment.id, comp_code, "DOSSIER_EXPORT", "NONE",
        f"Exported {len(manifest_rows)} documents", request.session.get("email"),
    )
    db.commit()
    log_data_management_action(comp_code, "DOSSIER", f"Shipment {shipment.shipment_no}", "Success", "Downloaded complete shipment dossier")
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="Shipment_{safe_filename(shipment.shipment_no)}_Dossier.zip"'},
    )


@router.get("/{doc_type}/print/{record_id}", response_class=HTMLResponse)
def export_document_print(doc_type: str, record_id: int, request: Request, db: Session = Depends(get_db)):
    cfg, row, comp_code = get_export_record_or_404(db, request, doc_type, record_id)
    payload = build_document_payload(cfg, row, get_invoice_packing_rows(db, row))
    company = get_export_company_profile(db, comp_code)
    raw_items = []
    if hasattr(row, "items_json") and row.items_json:
        try:
            raw_items = json.loads(row.items_json)
        except Exception:
            raw_items = []
    if not raw_items and hasattr(row, "product_description"):
        raw_items = [{
            "product_description": getattr(row, "product_description", ""),
            "brand": getattr(row, "brand", ""),
            "packing_style": getattr(row, "packing_style", ""),
            "grade": getattr(row, "grade", ""),
            "no_of_mc": getattr(row, "no_of_mc", 0),
            "quantity": getattr(row, "quantity", 0),
            "unit_price": getattr(row, "unit_price", 0),
            "total_amount": getattr(row, "total_amount", 0),
        }]

    processed_items = process_items_with_spans(
        raw_items,
        getattr(row, "product_description", ""),
        getattr(row, "brand", ""),
        getattr(row, "packing_style", ""),
    )

    return templates.TemplateResponse(
        request=request,
        name=cfg["template"],
        context={
            **payload,
            "company_id": comp_code,
            "company": company,
            "record": row,
            "items": processed_items,
            "doc_type": doc_type,
            "generated_at": datetime.utcnow(),
        },
    )


@router.get("/{doc_type}/pdf/{record_id}")
def export_document_pdf(doc_type: str, record_id: int, request: Request, db: Session = Depends(get_db)):
    cfg, row, comp_code = get_export_record_or_404(db, request, doc_type, record_id)
    pdf_bytes = render_document_pdf(
        cfg,
        row,
        comp_code,
        doc_type,
        get_export_company_profile(db, comp_code),
        get_invoice_packing_rows(db, row),
    )
    document_no = str(getattr(row, cfg["no"], record_id))
    file_row = store_export_pdf(
        db=db,
        company_id=comp_code,
        module_name=doc_type,
        record_id=row.id,
        document_no=document_no,
        document_kind="GENERATED_PDF",
        file_name=f"{safe_filename(document_no)}.pdf",
        content=pdf_bytes,
        uploaded_by=request.session.get("email"),
        remarks="System generated international format PDF",
    )
    file_row.approval_status = "DRAFT"
    set_document_path(row, file_row.file_path)
    write_audit(db, doc_type, row.id, comp_code, "PDF_GENERATE", "NONE", file_row.file_path, request.session.get("email"))
    db.commit()
    is_download = request.query_params.get("download") in ("1", "true", "yes", "True")
    disposition = "attachment" if is_download else "inline"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"{disposition}; filename={safe_filename(document_no)}.pdf"},
    )


@router.post("/{doc_type}/upload_pdf/{record_id}")
async def export_document_upload_pdf(
    doc_type: str,
    record_id: int,
    request: Request,
    file: UploadFile = File(...),
    document_kind: str = Form("SIGNED_COPY"),
    remarks: str = Form(None),
    db: Session = Depends(get_db),
):
    cfg, row, comp_code = get_export_record_or_404(db, request, doc_type, record_id)
    if file.content_type != "application/pdf" and not file.filename.lower().endswith(".pdf"):
        return JSONResponse({"success": False, "message": "Only PDF files are allowed"}, status_code=400)
    content = await file.read()
    if not content:
        return JSONResponse({"success": False, "message": "Empty PDF file"}, status_code=400)
    if len(content) > 25 * 1024 * 1024:
        return JSONResponse({"success": False, "message": "PDF size cannot exceed 25 MB"}, status_code=400)
    if not content.startswith(b"%PDF-"):
        return JSONResponse({"success": False, "message": "Invalid PDF file"}, status_code=400)
    document_no = str(getattr(row, cfg["no"], record_id))
    file_row = store_export_pdf(
        db=db,
        company_id=comp_code,
        module_name=doc_type,
        record_id=row.id,
        document_no=document_no,
        document_kind=document_kind,
        file_name=file.filename or f"{document_no}.pdf",
        content=content,
        uploaded_by=request.session.get("email"),
        remarks=remarks,
    )
    file_row.approval_status = "DRAFT"
    set_document_path(row, file_row.file_path)
    write_audit(db, doc_type, row.id, comp_code, "PDF_UPLOAD", "NONE", file_row.file_path, request.session.get("email"))
    db.commit()
    return {
        "success": True,
        "message": "Draft PDF saved. Import the final copy in Document Center for selected-email approval.",
        "file_id": file_row.id,
        "file_path": file_row.file_path,
        "approval_status": "DRAFT",
    }


@router.get("/files/{file_id}/download")
def export_document_file_download(file_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Unauthorized")
    file_row = db.query(ExportDocumentFile).filter(ExportDocumentFile.id == file_id, ExportDocumentFile.company_id == comp_code).first()
    if not file_row:
        raise HTTPException(status_code=404, detail="File not found")
    return StreamingResponse(
        BytesIO(file_row.file_bytes),
        media_type=file_row.content_type or "application/pdf",
        headers={"Content-Disposition": f"inline; filename={safe_filename(file_row.file_name)}"},
    )


@router.post("/{doc_type}/send-email/{record_id}")
async def export_document_send_email(
    doc_type: str,
    record_id: int,
    request: Request,
    to_email: str = Form(...),
    subject: str | None = Form(None),
    body: str | None = Form(None),
    file: UploadFile | None = File(None),
    db: Session = Depends(get_db)
):
    cfg, row, comp_code = get_export_record_or_404(db, request, doc_type, record_id)
    doc_no = getattr(row, cfg["no"], str(record_id))
    company_profile = get_export_company_profile(db, comp_code)
    
    if file and file.filename:
        pdf_bytes = await file.read()
        filename = file.filename
    else:
        pdf_bytes = render_document_pdf(cfg, row, comp_code, doc_type, company_profile=company_profile)
        filename = f"{safe_filename(doc_type.upper())}_{safe_filename(doc_no)}.pdf"
    
    subject_str = subject or f"{cfg['title']} - {doc_no} from {company_profile.get('name', 'BHAGAVATHI KRISHNA EXPORTS')}"
    body_text = body or f"Dear Partner,\n\nPlease find attached the official {cfg['title']} ({doc_no}) from {company_profile.get('name', 'BHAGAVATHI KRISHNA EXPORTS')}.\n\nThank you,\nExport Documentation Team"
    
    body_html = body_text.replace('\n', '<br>')
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <body style="font-family: Arial, sans-serif; color: #1e293b; padding: 20px;">
        <h2 style="color: #1e3a8a;">{company_profile.get('name', 'BHAGAVATHI KRISHNA EXPORTS')}</h2>
        <p style="font-size: 14px; color: #475569; line-height: 1.5;">{body_html}</p>
        <hr style="border: none; border-top: 1px solid #cbd5e1; margin: 20px 0;">
        <p style="font-size: 12px; color: #64748b;">This is an official document email sent from BKNR ERP System.</p>
    </body>
    </html>
    """
    
    from app.utils.email_service import send_email
    try:
        send_email(
            to_email=to_email,
            subject=subject_str,
            html=html_content,
            text=body_text,
            attachment_bytes=pdf_bytes,
            attachment_name=filename,
            attachment_type="application/pdf"
        )
        
        write_audit(
            db, doc_type, row.id, comp_code, "EMAIL_SENT", "NONE",
            f"Sent {cfg['title']} PDF ({filename}) via email to {to_email}", request.session.get("email"),
        )
        db.commit()
        return {"success": True, "message": f"Email successfully sent with attached PDF ({filename}) to {to_email}"}
    except Exception as e:
        logger.error(f"Failed to send document email with PDF: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")
