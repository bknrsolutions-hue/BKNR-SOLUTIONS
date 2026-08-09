import re
import json
import logging
import textwrap
from io import BytesIO
from pathlib import Path
from typing import Any
from decimal import Decimal
from datetime import date, datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pydantic import BaseModel, model_validator
from fastapi import APIRouter, Request, Depends, HTTPException
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, text, or_

from app.database import get_db
from app.database.models.invoices import (
    ProformaInvoice,
    ExportShipment,
    ExportComplianceTracker,
    CommercialInvoice,
    PackingList,
    ContainerStuffing,
    ShippingBill,
    BillOfLading,
    HealthCertificate,
    ExportDocumentFile,
    ExportDocumentApproval,
    ExportRequiredDocument,
)
from app.database.models.users import Company, User
from app.database.models.enterprise_finance import ProductionCostAllocation
from app.database.models.processing import AuditLog
from app.services.posting_engine import PostingEngineService
from app.services.bill_accounting import ensure_bill_accounting_schema
from app.services.cache import invalidate_company_cache

templates = Jinja2Templates(directory="app/templates")


def write_audit(db: Session, table: str, rec_id: int, company_id: str, action: str, old: str, new: str, email: str):
    audit = AuditLog(
        table_name=table,
        record_id=rec_id,
        company_id=company_id,
        field_name=action,
        old_value=old,
        new_value=new,
        edited_by=email,
        edited_at=datetime.utcnow()
    )
    db.add(audit)

def jinja_from_json(val):
    if not val:
        return []
    if isinstance(val, (list, dict)):
        return val
    try:
        return json.loads(val)
    except Exception:
        return []

templates.env.filters["from_json"] = jinja_from_json
logger = logging.getLogger(__name__)

EXPORT_PDF_DIR = Path("uploads/export_documents_private")
_EXPORT_SCHEMA_READY = False


def repost_invoice_cogs(db: Session, company_id: str, invoice: CommercialInvoice, email: str) -> float:
    """Value packing-list batches and keep one COGS voucher per invoice."""
    if invoice.cogs_journal_id:
        PostingEngineService.reverse_voucher(
            db, company_id, invoice.cogs_journal_id,
            "Packing list valuation revised", email or "SYSTEM",
        )
        invoice.cogs_journal_id = None
    rows = db.query(PackingList).filter(
        PackingList.company_id == company_id,
        PackingList.invoice_no == invoice.invoice_no,
        PackingList.is_cancelled != True,
    ).all()
    total_cogs = 0.0
    missing_batches = []
    for row in rows:
        batch = str(row.batch_no or "").strip()
        if not batch:
            missing_batches.append(row.packing_no)
            continue
        allocation = db.query(ProductionCostAllocation).filter(
            ProductionCostAllocation.company_id == company_id,
            ProductionCostAllocation.batch_number == batch,
            ProductionCostAllocation.status == "FG_TRANSFERRED",
            ProductionCostAllocation.is_cancelled != True,
        ).first()
        if not allocation or float(allocation.cost_per_kg or 0) <= 0:
            missing_batches.append(batch)
            continue
        total_cogs += float(row.net_weight or 0) * float(allocation.cost_per_kg or 0)
    total_cogs = round(total_cogs, 2)
    if missing_batches:
        raise ValueError("Complete FG cost allocation for packing batches: " + ", ".join(sorted(set(missing_batches))))
    if total_cogs <= 0:
        return 0.0
    voucher = PostingEngineService.create_voucher(
        db, company_id, "Journal", invoice.invoice_date,
        f"Cost of goods sold for export invoice {invoice.invoice_no}",
        [
            {"ledger_name": "Cost of Goods Sold A/c", "group_name": "Direct Expenses", "group_type": "EXPENSE", "debit_amount": total_cogs, "credit_amount": 0.0, "remarks": invoice.invoice_no},
            {"ledger_name": "Finished Goods Inventory A/c", "group_name": "Stock-in-hand", "group_type": "ASSET", "debit_amount": 0.0, "credit_amount": total_cogs, "remarks": invoice.invoice_no},
        ],
        reference_no=invoice.invoice_no, created_by=email or "SYSTEM",
    )
    invoice.cogs_journal_id = voucher.id
    return total_cogs

EXPORT_SUPPORT_DOCUMENT_TYPES = [
    # Order and buyer approval stage
    {"code": "PROFORMA_INVOICE", "label": "Proforma Invoice (PI)", "stage": "Order & Contract"},
    {"code": "BUYER_PO", "label": "Buyer Purchase Order", "stage": "Order & Contract"},
    {"code": "SALES_CONTRACT", "label": "Sales / Export Contract", "stage": "Order & Contract"},
    {"code": "LC_COPY", "label": "Letter of Credit (LC) Copy", "stage": "Order & Contract"},
    {"code": "LC_AMENDMENT", "label": "LC Amendment", "stage": "Order & Contract"},
    {"code": "ADVANCE_PAYMENT_PROOF", "label": "Advance Payment / SWIFT Proof", "stage": "Order & Contract"},
    {"code": "BUYER_APPROVAL", "label": "Buyer Approval / Email Copy", "stage": "Order & Contract"},
    {"code": "PRODUCT_SPECIFICATION", "label": "Product Specification / MSDS", "stage": "Order & Contract"},
    {"code": "LABEL_ARTWORK_APPROVAL", "label": "Label / Artwork Approval", "stage": "Order & Contract"},
    {"code": "IMPORT_PERMIT", "label": "Buyer Country Import Permit", "stage": "Order & Contract"},

    # Seafood production, quality and traceability stage
    {"code": "BATCH_TRACEABILITY", "label": "Batch & Lot Traceability Record", "stage": "Seafood Quality"},
    {"code": "FARM_CATCH_DECLARATION", "label": "Farm / Catch Declaration", "stage": "Seafood Quality"},
    {"code": "CATCH_CERTIFICATE", "label": "Catch Certificate", "stage": "Seafood Quality"},
    {"code": "HACCP_CHECKLIST", "label": "HACCP / Processing Checklist", "stage": "Seafood Quality"},
    {"code": "QC_INSPECTION_REPORT", "label": "QC Inspection Report", "stage": "Seafood Quality"},
    {"code": "LAB_TEST_REPORT", "label": "Laboratory Test Report", "stage": "Seafood Quality"},
    {"code": "MICROBIOLOGY_REPORT", "label": "Microbiology Report", "stage": "Seafood Quality"},
    {"code": "ANTIBIOTIC_RESIDUE_REPORT", "label": "Antibiotic Residue Test Report", "stage": "Seafood Quality"},
    {"code": "HEAVY_METAL_REPORT", "label": "Heavy Metal Test Report", "stage": "Seafood Quality"},
    {"code": "WATER_ICE_TEST_REPORT", "label": "Water / Ice Quality Report", "stage": "Seafood Quality"},
    {"code": "TEMPERATURE_LOG", "label": "Cold-chain Temperature Log", "stage": "Seafood Quality"},
    {"code": "WEIGHT_PACKING_VERIFICATION", "label": "Weight & Packing Verification", "stage": "Seafood Quality"},

    # Statutory and certification stage
    {"code": "EIA_INSPECTION_REPORT", "label": "EIA Inspection Report", "stage": "Certificates"},
    {"code": "HEALTH_CERTIFICATE_COPY", "label": "Health Certificate", "stage": "Certificates"},
    {"code": "PHYTO_CERTIFICATE", "label": "Phytosanitary Certificate", "stage": "Certificates"},
    {"code": "VETERINARY_CERTIFICATE", "label": "Veterinary Certificate", "stage": "Certificates"},
    {"code": "COO", "label": "Certificate of Origin", "stage": "Certificates"},
    {"code": "FUMIGATION_CERTIFICATE", "label": "Fumigation Certificate", "stage": "Certificates"},
    {"code": "HALAL_CERTIFICATE", "label": "Halal Certificate", "stage": "Certificates"},
    {"code": "ANIMAL_QUARANTINE_NOC", "label": "Animal Quarantine / NOC", "stage": "Certificates"},

    # Commercial, customs and logistics stage
    {"code": "COMMERCIAL_INVOICE", "label": "Commercial Invoice", "stage": "Shipping & Customs"},
    {"code": "PACKING_LIST", "label": "Packing List", "stage": "Shipping & Customs"},
    {"code": "CONTAINER_STUFFING_REPORT", "label": "Container Stuffing Report", "stage": "Shipping & Customs"},
    {"code": "CONTAINER_SEAL_REPORT", "label": "Container Seal / Inspection Report", "stage": "Shipping & Customs"},
    {"code": "VGM_DECLARATION", "label": "VGM Declaration", "stage": "Shipping & Customs"},
    {"code": "SHIPPING_BILL", "label": "Shipping Bill", "stage": "Shipping & Customs"},
    {"code": "CUSTOMS_LEO_COPY", "label": "Customs Let Export Order (LEO)", "stage": "Shipping & Customs"},
    {"code": "BILL_OF_LADING_DRAFT", "label": "Bill of Lading Draft", "stage": "Shipping & Customs"},
    {"code": "BL_COPY", "label": "Final Bill of Lading / AWB", "stage": "Shipping & Customs"},
    {"code": "INSURANCE_CERTIFICATE", "label": "Marine Insurance Certificate", "stage": "Shipping & Customs"},
    {"code": "FREIGHT_INVOICE", "label": "Freight Invoice", "stage": "Shipping & Customs"},
    {"code": "CHA_INVOICE", "label": "CHA / Customs Broker Invoice", "stage": "Shipping & Customs"},
    {"code": "PORT_TERMINAL_RECEIPT", "label": "Port / Terminal Receipt", "stage": "Shipping & Customs"},

    # Bank submission, payment and closure stage
    {"code": "BANK_SUBMISSION_SET", "label": "Bank Document Submission Set", "stage": "Bank & Payment"},
    {"code": "BILL_OF_EXCHANGE", "label": "Bill of Exchange", "stage": "Bank & Payment"},
    {"code": "NEGOTIATION_COLLECTION_PROOF", "label": "Negotiation / Collection Proof", "stage": "Bank & Payment"},
    {"code": "PAYMENT_SWIFT_COPY", "label": "Payment SWIFT Copy", "stage": "Bank & Payment"},
    {"code": "PAYMENT_RECEIPT", "label": "Payment Receipt / Bank Credit Advice", "stage": "Bank & Payment"},
    {"code": "FIRC", "label": "Foreign Inward Remittance Certificate (FIRC)", "stage": "Bank & Payment"},
    {"code": "EBRC", "label": "Electronic Bank Realisation Certificate (e-BRC)", "stage": "Bank & Payment"},
    {"code": "CREDIT_DEBIT_NOTE", "label": "Credit / Debit Note", "stage": "Bank & Payment"},
    {"code": "DUTY_DRAWBACK_PROOF", "label": "Duty Drawback Credit Proof", "stage": "Bank & Payment"},
    {"code": "RODTEP_CREDIT_PROOF", "label": "RoDTEP Credit Proof", "stage": "Bank & Payment"},
    {"code": "EXPORT_CLOSURE_CONFIRMATION", "label": "Export File Closure Confirmation", "stage": "Bank & Payment"},
]

EXPORT_GENERATE_DOCUMENTS = {
    "SALES_CONTRACT", "PRODUCT_SPECIFICATION",
    "BATCH_TRACEABILITY", "HACCP_CHECKLIST", "TEMPERATURE_LOG",
    "WEIGHT_PACKING_VERIFICATION", "COMMERCIAL_INVOICE", "PACKING_LIST",
    "CONTAINER_STUFFING_REPORT", "VGM_DECLARATION", "BILL_OF_EXCHANGE",
    "BANK_SUBMISSION_SET", "CREDIT_DEBIT_NOTE", "EXPORT_CLOSURE_CONFIRMATION",
}
EXPORT_HYBRID_DOCUMENTS = {
    "SALES_CONTRACT", "PRODUCT_SPECIFICATION",
    "CATCH_CERTIFICATE", "HEALTH_CERTIFICATE_COPY", "COMMERCIAL_INVOICE",
    "PACKING_LIST", "CONTAINER_STUFFING_REPORT", "VGM_DECLARATION",
    "BILL_OF_EXCHANGE", "BANK_SUBMISSION_SET", "CREDIT_DEBIT_NOTE",
}


def export_document_mode(document_code: str) -> str:
    if document_code == "PROFORMA_INVOICE":
        return "IMPORT_FINAL_PDF"
    if document_code in EXPORT_HYBRID_DOCUMENTS:
        return "GENERATE_AND_IMPORT_FINAL"
    if document_code in EXPORT_GENERATE_DOCUMENTS:
        return "GENERATE"
    return "IMPORT_PDF"

EXPORT_REQUIREMENT_STAGE_FIELDS = {
    "Order & Contract": [
        {"name": "buyer_name", "label": "Buyer", "type": "select", "lookup": "buyers"},
        {"name": "buyer_agent", "label": "Buyer Agent", "type": "select", "lookup": "buyer_agents"},
        {"name": "destination_country", "label": "Destination Country", "type": "select", "lookup": "countries"},
        {"name": "buyer_reference", "label": "Buyer Reference", "type": "text"},
        {"name": "contract_date", "label": "Contract / Approval Date", "type": "date"},
        {"name": "validity_date", "label": "Validity Date", "type": "date"},
        {"name": "incoterm", "label": "Incoterm", "type": "select", "options": ["FOB", "CFR", "CIF", "EXW", "FCA", "CPT", "CIP", "DDP"]},
        {"name": "payment_terms", "label": "Payment Terms", "type": "text"},
        {"name": "product_description", "label": "Product / Specification", "type": "textarea"},
    ],
    "Seafood Quality": [
        {"name": "batch_no", "label": "Batch Number", "type": "text"},
        {"name": "lot_no", "label": "Lot Number", "type": "text"},
        {"name": "species", "label": "Species", "type": "select", "lookup": "species", "multiple": True},
        {"name": "variety", "label": "Variety", "type": "select", "lookup": "varieties", "multiple": True},
        {"name": "grade", "label": "Grade / Size", "type": "select", "lookup": "grades", "multiple": True},
        {"name": "brand", "label": "Brand", "type": "select", "lookup": "brands"},
        {"name": "glaze", "label": "Glaze", "type": "select", "lookup": "glazes", "multiple": True},
        {"name": "freezer", "label": "Freezer", "type": "select", "lookup": "freezers", "multiple": True},
        {"name": "packing_style", "label": "Packing Style", "type": "select", "lookup": "packing_styles", "multiple": True},
        {"name": "lab_name", "label": "Laboratory / Inspector", "type": "text"},
        {"name": "sample_date", "label": "Sample Date", "type": "date"},
        {"name": "result", "label": "Test / Inspection Result", "type": "select", "options": ["PASS", "FAIL", "CONDITIONAL", "NA"]},
        {"name": "temperature", "label": "Temperature", "type": "text"},
    ],
    "Certificates": [
        {"name": "certificate_no", "label": "Certificate Number", "type": "text"},
        {"name": "authority", "label": "Issuing Authority", "type": "text"},
        {"name": "factory_approval_no", "label": "Factory Approval Number", "type": "text"},
        {"name": "destination_country", "label": "Destination Country", "type": "select", "lookup": "countries"},
        {"name": "species", "label": "Species / Product", "type": "select", "lookup": "species", "multiple": True},
        {"name": "health_marks", "label": "Health Marks / Endorsement", "type": "textarea"},
    ],
    "Shipping & Customs": [
        {"name": "invoice_no", "label": "Invoice Number", "type": "text"},
        {"name": "container_no", "label": "Container Number", "type": "text"},
        {"name": "seal_no", "label": "Seal Number", "type": "text"},
        {"name": "shipping_line", "label": "Shipping Line / CHA", "type": "select", "lookup": "shipping_vendors"},
        {"name": "freezer", "label": "Freezer", "type": "select", "lookup": "freezers", "multiple": True},
        {"name": "packing_style", "label": "Packing Style", "type": "select", "lookup": "packing_styles", "multiple": True},
        {"name": "vessel_voyage", "label": "Vessel / Voyage", "type": "text"},
        {"name": "port_of_loading", "label": "Port of Loading", "type": "text"},
        {"name": "port_of_discharge", "label": "Port of Discharge", "type": "text"},
        {"name": "etd", "label": "ETD", "type": "date"},
        {"name": "eta", "label": "ETA", "type": "date"},
        {"name": "gross_weight", "label": "Gross Weight", "type": "number"},
        {"name": "net_weight", "label": "Net Weight", "type": "number"},
    ],
    "Bank & Payment": [
        {"name": "bank_account", "label": "Company Bank Account", "type": "select", "lookup": "bank_accounts"},
        {"name": "swift_reference", "label": "SWIFT / Bank Reference", "type": "text"},
        {"name": "invoice_no", "label": "Invoice Number", "type": "text"},
        {"name": "submission_date", "label": "Submission Date", "type": "date"},
        {"name": "receipt_date", "label": "Receipt / Credit Date", "type": "date"},
        {"name": "bill_reference", "label": "Bill / e-BRC Reference", "type": "text"},
        {"name": "realisation_status", "label": "Realisation Status", "type": "select", "options": ["PENDING", "PARTIAL", "REALISED", "CLOSED"]},
    ],
    "Custom": [
        {"name": "custom_reference_1", "label": "Additional Reference 1", "type": "text"},
        {"name": "custom_reference_2", "label": "Additional Reference 2", "type": "text"},
        {"name": "description", "label": "Document Details", "type": "textarea"},
    ],
}

EXPORT_REQUIREMENT_COMMON_FIELDS = [
    {"name": "document_no", "label": "Document Number / Reference", "type": "text", "required": True},
    {"name": "document_date", "label": "Document Date", "type": "date", "required": True},
    {"name": "expiry_date", "label": "Expiry / Valid Until", "type": "date"},
    {"name": "issuer_name", "label": "Issuer / Organisation", "type": "text"},
    {"name": "reference_no", "label": "Secondary Reference", "type": "text"},
    {"name": "currency", "label": "Currency", "type": "select", "options": ["USD", "EUR", "GBP", "AED", "JPY", "INR"]},
    {"name": "amount", "label": "Amount / Value", "type": "number"},
    {"name": "status_note", "label": "Status / Notes", "type": "textarea"},
]


def invalidate_export_cache(company_id: str | None):
    if company_id:
        invalidate_company_cache(company_id, "export_documents")


def _dt(value):
    if value is None:
        return None
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "document")).strip("_")[:120]


def export_doc_config():
    return {
        "proforma_invoice": {
            "model": ProformaInvoice, "no": "pi_no", "date": "pi_date",
            "title": "Proforma Invoice", "template": "export_documents/print_document_pdf.html",
            "fields": [
                ("PI No", "pi_no"), ("PI Date", "pi_date"), ("Valid Until", "validity_date"),
                ("Buyer", "buyer_name"), ("Buyer Address", "buyer_address"),
                ("Country", "country"), ("Buyer PO", "po_number"), ("Currency", "currency"),
                ("Incoterm", "incoterm"), ("Payment Terms", "payment_terms"),
                ("Port of Loading", "port_of_loading"), ("Port of Discharge", "port_of_discharge"),
                ("Product Description", "product_description"), ("Quantity", "quantity"),
                ("Unit", "unit"), ("Unit Price", "unit_price"), ("Total Amount", "total_amount"),
                ("Status", "status"), ("Approval Status", "approval_status"),
                ("Approved By", "approved_by"), ("Approval Remarks", "approval_remarks"),
                ("Remarks", "remarks"),
            ],
        },
        "export_shipment": {
            "model": ExportShipment, "no": "shipment_no", "date": "created_at",
            "title": "Export Shipment File", "template": "export_documents/print_document_pdf.html",
            "fields": [
                ("Shipment No", "shipment_no"), ("PO Number", "po_number"), ("Invoice No", "invoice_no"),
                ("Container No", "container_no"), ("Buyer", "buyer_name"), ("Country", "country"),
                ("ETD", "etd"), ("ETA", "eta"), ("Completion Date", "completion_date"),
                ("Status", "status"), ("Approval Status", "approval_status"),
            ],
        },
        "commercial_invoice": {
            "model": CommercialInvoice, "no": "invoice_no", "date": "invoice_date",
            "title": "Commercial Invoice", "template": "export_documents/print_document_pdf.html",
            "fields": [
                ("Invoice No", "invoice_no"), ("Invoice Date", "invoice_date"), ("Shipment No", "shipment_no"),
                ("PO Number", "po_number"), ("Buyer", "buyer_name"), ("Buyer Address", "buyer_address"),
                ("Consignee", "consignee_name"), ("Notify Party", "notify_party"), ("Country", "country"),
                ("Currency", "currency"), ("Total Amount", "total_amount"), ("Exchange Rate", "exchange_rate"),
                ("Invoice Value INR", "invoice_value_inr"), ("Payment Terms", "payment_terms"),
                ("Shipment Terms", "shipment_terms"), ("Port of Loading", "port_of_loading"),
                ("Port of Discharge", "port_of_discharge"), ("Final Destination", "final_destination"),
                ("Shipment Type", "shipment_type"), ("Total Master Cartons", "total_mc"),
                ("Total Net Weight", "total_net_weight"), ("Total Gross Weight", "total_gross_weight"),
                ("Payment Status", "payment_status"), ("Approval Status", "approval_status"),
            ],
        },
        "packing_list": {
            "model": PackingList, "no": "packing_no", "date": "created_at",
            "title": "Packing List", "template": "export_documents/print_document_pdf.html",
            "fields": [
                ("Packing No", "packing_no"), ("Invoice No", "invoice_no"), ("PO Number", "po_number"),
                ("Container No", "container_no"), ("Buyer", "buyer_name"), ("Product", "product_name"),
                ("Grade", "grade"), ("Batch No", "batch_no"), ("Lot No", "lot_no"), ("Glaze", "glaze"),
                ("Freezing Type", "freezing_type"), ("HS Code", "hs_code"), ("Packing Style", "packing_style"),
                ("Inner Pack", "inner_pack"), ("Outer Pack", "outer_pack"), ("Master Cartons", "master_cartons"),
                ("Net Weight", "net_weight"), ("Gross Weight", "gross_weight"), ("Pallet Count", "pallet_count"),
                ("Manufacturing Date", "manufacturing_date"), ("Expiry Date", "expiry_date"),
            ],
        },
        "container_stuffing": {
            "model": ContainerStuffing, "no": "container_no", "date": "stuffing_date",
            "title": "Container Stuffing Report", "template": "export_documents/print_document_pdf.html",
            "fields": [
                ("Container No", "container_no"), ("Invoice No", "invoice_no"), ("PO Number", "po_number"),
                ("Buyer", "buyer_name"), ("Seal No", "seal_no"), ("Shipping Line", "shipping_line"),
                ("Stuffing Date", "stuffing_date"), ("Stuffing Location", "stuffing_location"),
                ("Container Type", "container_type"), ("Container Size", "container_size"),
                ("Set Temperature", "temperature"), ("Vehicle No", "vehicle_no"),
                ("Container Condition", "container_condition"), ("Temperature Before Loading", "temperature_before_loading"),
                ("Temperature After Loading", "temperature_after_loading"), ("Driver", "driver_name"),
                ("Loading Supervisor", "loading_supervisor"), ("Approval Status", "approval_status"), ("Remarks", "remarks"),
            ],
        },
        "shipping_bill": {
            "model": ShippingBill, "no": "shipping_bill_no", "date": "shipping_bill_date",
            "title": "Shipping Bill Summary", "template": "export_documents/print_document_pdf.html",
            "fields": [
                ("Shipping Bill No", "shipping_bill_no"), ("Shipping Bill Date", "shipping_bill_date"),
                ("Invoice No", "invoice_no"), ("Container No", "container_no"), ("PO Number", "po_number"),
                ("Buyer", "buyer_name"), ("FOB Value INR", "shipping_bill_value"),
                ("Drawback Amount", "drawback_amount"), ("Scheme", "scheme"), ("Customs Status", "customs_status"),
                ("Port", "port"), ("CHA Name", "cha_name"), ("CHA Bill No", "cha_bill_no"), ("Vessel", "vessel_name"),
                ("Voyage No", "voyage_no"), ("ETD", "etd"), ("ETA", "eta"),
                ("Approval Status", "approval_status"), ("Remarks", "remarks"),
            ],
        },
        "bill_of_lading": {
            "model": BillOfLading, "no": "bl_no", "date": "bl_date",
            "title": "Bill of Lading", "template": "export_documents/print_document_pdf.html",
            "fields": [
                ("B/L No", "bl_no"), ("B/L Date", "bl_date"), ("On Board Date", "onboard_date"),
                ("Invoice No", "invoice_no"), ("Container No", "container_no"), ("PO Number", "po_number"),
                ("Buyer", "buyer_name"), ("Shipping Line", "shipping_line"), ("Seal No", "seal_no"),
                ("Freight Terms", "freight_terms"), ("Original B/L Count", "no_of_original_bl"),
                ("Marks & Numbers", "marks_and_numbers"), ("Packages Description", "packages_description"),
                ("Place of Receipt", "place_of_receipt"), ("Place of Delivery", "place_of_delivery"),
                ("Gross Weight", "gross_weight"), ("Net Weight", "net_weight"),
                ("Approval Status", "approval_status"),
            ],
        },
        "health_certificate": {
            "model": HealthCertificate, "no": "certificate_no", "date": "issue_date",
            "title": "Health Certificate", "template": "export_documents/print_document_pdf.html",
            "fields": [
                ("Certificate No", "certificate_no"), ("Issue Date", "issue_date"), ("Authority", "authority"),
                ("Factory Approval No", "factory_approval_no"), ("Invoice No", "invoice_no"),
                ("Container No", "container_no"), ("PO Number", "po_number"), ("Buyer", "buyer_name"),
                ("Country", "country"), ("Species", "species"), ("Temperature Verified", "temperature_verified"),
                ("Issued By", "issued_by"), ("Status", "status"), ("Remarks", "remarks"),
            ],
        },
    }


def get_export_record_or_404(db: Session, request: Request, doc_type: str, record_id: int):
    cfg = export_doc_config().get(doc_type)
    if not cfg:
        raise HTTPException(status_code=404, detail="Unsupported document type")
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Unauthorized")
    row = db.query(cfg["model"]).filter(cfg["model"].id == record_id, cfg["model"].company_id == comp_code).first()
    if not row:
        raise HTTPException(status_code=404, detail="Record not found")
    return cfg, row, comp_code


def store_export_pdf(
    db: Session,
    company_id: str,
    module_name: str,
    record_id: int,
    document_no: str,
    document_kind: str,
    file_name: str,
    content: bytes,
    uploaded_by: str | None,
    remarks: str | None = None,
) -> ExportDocumentFile:
    EXPORT_PDF_DIR.mkdir(parents=True, exist_ok=True)
    for old in db.query(ExportDocumentFile).filter(
        ExportDocumentFile.company_id == company_id,
        ExportDocumentFile.module_name == module_name,
        ExportDocumentFile.record_id == record_id,
        ExportDocumentFile.document_kind == document_kind,
        ExportDocumentFile.is_current == True,
    ).all():
        old.is_current = False
    version_no = (
        db.query(func.coalesce(func.max(ExportDocumentFile.version_no), 0))
        .filter(
            ExportDocumentFile.company_id == company_id,
            ExportDocumentFile.module_name == module_name,
            ExportDocumentFile.record_id == record_id,
            ExportDocumentFile.document_kind == document_kind,
        )
        .scalar()
        + 1
    )
    final_name = f"{safe_filename(module_name)}_{safe_filename(document_no)}_v{version_no}_{safe_filename(file_name)}"
    disk_path = EXPORT_PDF_DIR / final_name
    disk_path.write_bytes(content)
    file_row = ExportDocumentFile(
        company_id=company_id,
        module_name=module_name,
        record_id=record_id,
        document_no=document_no,
        document_kind=document_kind,
        file_name=final_name,
        file_path=None,
        content_type="application/pdf",
        file_bytes=content,
        file_size=len(content),
        version_no=version_no,
        uploaded_by=uploaded_by,
        remarks=remarks,
    )
    db.add(file_row)
    db.flush()
    file_row.file_path = f"/export_documents/files/{file_row.id}/download"
    return file_row


def get_export_company_profile(db: Session, company_id: str) -> dict:
    company_code_clean = str(company_id or "").strip().upper()
    company = db.query(Company).filter(
        func.upper(func.trim(Company.company_code)) == company_code_clean
    ).first()
    
    bank_info = {
        "bank_name": "HDFC BANK LIMITED",
        "account_number": "50200084920194",
        "ifsc_code": "HDFC0001234",
        "swift_code": "HDFCINBBXXX",
        "branch": "APSEZ VISAKHAPATNAM BRANCH, ANDHRA PRADESH",
        "currency_code": "USD",
        "account_type": "EXPORT CURRENT",
    }
    try:
        from app.database.models.enterprise_finance import BankMaster
        bank = db.query(BankMaster).filter(
            func.upper(func.trim(BankMaster.company_id)) == company_code_clean,
            BankMaster.is_active != False
        ).order_by(BankMaster.is_default.desc(), BankMaster.is_export_account.desc(), BankMaster.id.asc()).first()

        if bank:
            bank_info = {
                "bank_name": bank.bank_name or bank_info["bank_name"],
                "account_number": bank.account_number or bank_info["account_number"],
                "ifsc_code": bank.ifsc_code or bank_info["ifsc_code"],
                "swift_code": bank.swift_code or bank_info["swift_code"],
                "branch": bank.branch or bank_info["branch"],
                "currency_code": bank.currency_code or bank_info["currency_code"],
                "account_type": bank.account_type or bank_info["account_type"],
            }
    except Exception as exc:
        logger.warning("Error fetching bank details for company %s: %s", company_id, exc)

    return {
        "name": company.company_name if company else (company_id or "BHAGAVATHI KRISHNA EXPORTS"),
        "address": company.address if company else "Survey No 142/2, APSEZ, Atchutapuram, Visakhapatnam - 531011, AP, India",
        "email": company.email if company else "export@bknrexports.com",
        "phone": getattr(company, "phone", "") or "+91 891 2748899",
        "code": (
            company.mpeda_registration_code
            if company and company.mpeda_registration_code
            else "BKNR-EXP-01"
        ),
        "tenant_code": company_id,
        "bank": bank_info,
    }


def ensure_export_document_schema(db: Session = Depends(get_db)) -> None:
    """Keep export routes compatible while pending migrations are deployed."""
    global _EXPORT_SCHEMA_READY
    if _EXPORT_SCHEMA_READY:
        return
    ProformaInvoice.__table__.create(bind=db.get_bind(), checkfirst=True)
    db.execute(text("ALTER TABLE proforma_invoices ADD COLUMN IF NOT EXISTS approval_status VARCHAR NOT NULL DEFAULT 'PENDING'"))
    db.execute(text("ALTER TABLE proforma_invoices ADD COLUMN IF NOT EXISTS approved_by VARCHAR"))
    db.execute(text("ALTER TABLE proforma_invoices ADD COLUMN IF NOT EXISTS approved_at TIMESTAMP"))
    db.execute(text("ALTER TABLE proforma_invoices ADD COLUMN IF NOT EXISTS approval_remarks TEXT"))
    ExportRequiredDocument.__table__.create(bind=db.get_bind(), checkfirst=True)
    ExportDocumentApproval.__table__.create(bind=db.get_bind(), checkfirst=True)
    db.commit()
    _EXPORT_SCHEMA_READY = True


def is_supporting_document_admin(request: Request) -> bool:
    role = str(request.session.get("role") or "").strip().lower()
    email = str(request.session.get("email") or "").strip().lower()
    return role in {"admin", "super_admin"} or email == "bknr.solutions@gmail.com"


def set_document_path(row, path: str):
    if hasattr(row, "document_path"):
        row.document_path = path


def get_invoice_packing_rows(db: Session, row) -> list[PackingList]:
    if not isinstance(row, PackingList):
        return []
    return db.query(PackingList).filter(
        PackingList.company_id == row.company_id,
        PackingList.invoice_no == row.invoice_no,
        PackingList.is_cancelled != True,
    ).order_by(PackingList.invoice_item_no, PackingList.id).all()


def update_company_bank_master(db: Session, comp_code: str, payload: Any):
    if not any([getattr(payload, "bank_name", None), getattr(payload, "account_number", None), getattr(payload, "ifsc_code", None), getattr(payload, "swift_code", None), getattr(payload, "branch", None)]):
        return
    company_code_clean = (comp_code or "").strip().upper()
    from app.database.models.enterprise_finance import BankMaster
    bank_rec = db.query(BankMaster).filter(
        func.upper(func.trim(BankMaster.company_id)) == company_code_clean,
        BankMaster.is_active != False
    ).order_by(BankMaster.is_default.desc(), BankMaster.is_export_account.desc(), BankMaster.id.asc()).first()

    if bank_rec:
        if payload.bank_name: bank_rec.bank_name = payload.bank_name.strip()
        if payload.account_number: bank_rec.account_number = payload.account_number.strip()
        if payload.ifsc_code: bank_rec.ifsc_code = payload.ifsc_code.strip()
        if payload.swift_code: bank_rec.swift_code = payload.swift_code.strip()
        if payload.branch: bank_rec.branch = payload.branch.strip()
    else:
        bank_rec = BankMaster(
            company_id=comp_code,
            bank_name=payload.bank_name.strip() if payload.bank_name else "HDFC BANK LIMITED",
            account_number=payload.account_number.strip() if payload.account_number else "50200084920194",
            ifsc_code=payload.ifsc_code.strip() if payload.ifsc_code else "HDFC0001234",
            swift_code=payload.swift_code.strip() if payload.swift_code else "HDFCINBBXXX",
            branch=payload.branch.strip() if payload.branch else "APSEZ VISAKHAPATNAM BRANCH",
            is_default=True,
            is_export_account=True,
            is_active=True,
        )
        db.add(bank_rec)


def proforma_payload_values(payload: Any) -> dict:
    values = payload.model_dump()
    for bank_field in ["bank_name", "account_number", "ifsc_code", "swift_code", "branch"]:
        values.pop(bank_field, None)
    values["pi_no"] = payload.pi_no.strip()
    values["buyer_name"] = payload.buyer_name.strip()
    values["total_amount"] = (payload.quantity * payload.unit_price).quantize(Decimal("0.01"))
    return values


class SupportingDocumentApprovalSchema(BaseModel):
    decision: str
    remarks: str | None = None


class ExportShipmentSchema(BaseModel):
    shipment_no: str
    po_number: str
    invoice_no: str = None
    container_no: str = None
    buyer_name: str
    country: str
    etd: date = None
    eta: date = None

    @model_validator(mode="after")
    def eta_must_be_after_etd(self):
        if self.etd and self.eta and self.eta < self.etd:
            raise ValueError("ETA must be on or after ETD")
        return self


class ProformaInvoiceSchema(BaseModel):
    pi_no: str
    pi_date: date
    validity_date: date = None
    po_number: str = None
    buyer_name: str
    buyer_address: str
    country: str
    currency: str = "USD"
    incoterm: str
    payment_terms: str
    port_of_loading: str = None
    port_of_discharge: str = None
    product_description: str
    quantity: Decimal
    unit: str = "KG"
    unit_price: Decimal
    status: str = "DRAFT"
    remarks: str = None
    brand: str = None
    packing_style: str = None
    freezer: str = None
    count_glaze: str = None
    weight_glaze: str = None
    species: str = None
    variety: str = None
    grade: str = None
    no_of_pieces: str = None
    no_of_mc: int = 0
    items_json: str = None
    bank_name: str = None
    account_number: str = None
    ifsc_code: str = None
    swift_code: str = None
    branch: str = None

    @model_validator(mode="after")
    def validate_pi(self):
        if self.quantity is not None and self.quantity <= 0:
            raise ValueError("quantity must be greater than zero")
        if self.validity_date and self.pi_date and self.validity_date < self.pi_date:
            raise ValueError("validity_date must be on or after pi_date")
        return self


class CommercialInvoiceSchema(BaseModel):
    shipment_no: str
    invoice_no: str
    po_number: str
    container_no: str = None
    buyer_name: str
    invoice_date: date
    buyer_address: str
    consignee_name: str = None
    notify_party: str = None
    country: str
    currency: str = "USD"
    exchange_rate: Decimal = Decimal("83.50")
    total_amount: Decimal
    payment_terms: str
    shipment_terms: str

    @model_validator(mode="after")
    def validate_amounts(self):
        if self.exchange_rate is not None and self.exchange_rate <= 0:
            raise ValueError("exchange_rate must be greater than zero")
        if self.total_amount is not None and self.total_amount <= 0:
            raise ValueError("total_amount must be greater than zero")
        return self


class PackingListSchema(BaseModel):
    packing_no: str
    invoice_no: str
    po_number: str = None
    container_no: str = None
    buyer_name: str = None
    product_name: str
    grade: str
    batch_no: str = None
    lot_no: str = None
    glaze: str = None
    freezing_type: str = None
    hs_code: str = None
    manufacturing_date: date = None
    expiry_date: date = None
    packing_style: str
    inner_pack: str = None
    outer_pack: str = None
    master_cartons: int = 0
    net_weight: float = 0.0
    gross_weight: float = 0.0
    pallet_count: int = 0
    inventory_batch_id: str = None
    stock_entry_no: str = None
    invoice_item_no: int = None

    @model_validator(mode="after")
    def gross_must_be_gte_net(self):
        if self.gross_weight is not None and self.net_weight is not None and self.gross_weight < self.net_weight:
            raise ValueError("gross_weight must be >= net_weight")
        return self


class PackingListBulkLineSchema(BaseModel):
    product_name: str
    grade: str
    batch_no: str = None
    lot_no: str = None
    glaze: str = None
    freezing_type: str = None
    hs_code: str = None
    manufacturing_date: date = None
    expiry_date: date = None
    packing_style: str
    inner_pack: str = None
    outer_pack: str = None
    master_cartons: int = 0
    net_weight: float = 0.0
    gross_weight: float = 0.0
    pallet_count: int = 0
    inventory_batch_id: str = None
    stock_entry_no: str = None


class PackingListBulkSchema(BaseModel):
    packing_no: str
    invoice_no: str
    po_number: str = None
    container_no: str = None
    buyer_name: str = None
    items: list[PackingListBulkLineSchema]


class ContainerStuffingSchema(BaseModel):
    container_no: str
    invoice_no: str = None
    po_number: str = None
    buyer_name: str = None
    seal_no: str
    shipping_line: str = None
    stuffing_date: date
    stuffing_location: str = None
    container_type: str = "Reefer"
    container_size: str = "40FT"
    temperature: float
    vehicle_no: str
    loading_supervisor: str


class ShippingBillSchema(BaseModel):
    shipping_bill_no: str
    shipping_bill_date: date
    invoice_no: str
    container_no: str = None
    po_number: str = None
    buyer_name: str = None
    shipping_bill_value: float = 0.0
    drawback_amount: float = 0.0
    scheme: str = "NONE"
    customs_status: str = "LEO"
    port: str
    cha_name: str
    vessel_name: str
    voyage_no: str
    etd: date
    eta: date

    @model_validator(mode="after")
    def validate_shipping_bill(self):
        if self.etd and self.eta and self.eta < self.etd:
            raise ValueError("ETA must be on or after ETD")
        if self.drawback_amount is not None and self.drawback_amount < 0:
            raise ValueError("drawback_amount must be non-negative")
        return self


class BillOfLadingSchema(BaseModel):
    bl_no: str
    bl_date: date
    invoice_no: str
    container_no: str
    po_number: str = None
    buyer_name: str = None
    shipping_line: str
    seal_no: str
    freight_terms: str = "PREPAID"
    no_of_original_bl: int = 3
    gross_weight: float = 0.0
    net_weight: float = 0.0

    @model_validator(mode="after")
    def validate_bl(self):
        if self.no_of_original_bl is not None and self.no_of_original_bl < 1:
            raise ValueError("no_of_original_bl must be at least 1")
        return self


class HealthCertificateSchema(BaseModel):
    certificate_no: str
    issue_date: date
    authority: str = "EIA"
    invoice_no: str
    container_no: str
    po_number: str = None
    buyer_name: str = None
    country: str = None
    species: str = None
    temperature_verified: bool = True
    issued_by: str = None


def build_document_payload(cfg, row, packing_rows: list[PackingList] | None = None):
    configured_labels = {attr: label for label, attr in cfg["fields"]}
    ordered_attrs = [attr for _, attr in cfg["fields"]]
    for column in row.__table__.columns:
        if column.name not in ordered_attrs:
            ordered_attrs.append(column.name)

    grouped = {}
    for attr in ordered_attrs:
        value = getattr(row, attr, None)
        grouped.setdefault("Additional Details", []).append({
            "name": attr,
            "label": configured_labels.get(attr, str(attr).replace("_", " ").title()),
            "value": str(value) if value is not None else "—",
        })

    sections = [{"title": "Document Particulars", "fields": grouped["Additional Details"]}]
    sections[0]["rows"] = pack_export_print_rows(sections[0]["fields"])
    return {
        "title": cfg["title"],
        "document_no": getattr(row, cfg["no"], ""),
        "document_date": getattr(row, cfg["date"], None),
        "sections": sections,
    }


def pack_export_print_rows(fields: list[dict]) -> list[list[dict]]:
    rows = []
    current = []
    used = 0
    for field in fields:
        item = {**field, "span": 2}
        if used + 2 > 6:
            rows.append(current)
            current, used = [], 0
        current.append(item)
        used += 2
    if current:
        rows.append(current)
    return rows


def process_items_with_spans(raw_items, default_desc="", default_brand="", default_pack=""):
    if not raw_items:
        return []
    norm_items = []
    for it in raw_items:
        desc_val = (it.get("product_description") or it.get("item_name") or default_desc or "Seafood Export Product").strip()
        brand_val = (it.get("brand") or default_brand or "").strip()
        pack_val = (it.get("packing_style") or default_pack or "").strip()
        grade_val = (it.get("grade") or "").strip()
        mc = int(it.get("no_of_mc") or it.get("mc") or 0)
        qty = float(it.get("quantity") or it.get("quantity_kg") or 0.0)
        price = float(it.get("unit_price") or it.get("rate_per_kg") or 0.0)
        amt = float(it.get("total_amount") or it.get("amount") or (qty * price))
        is_unmerged = bool(it.get("_isUnmerged") or it.get("is_unmerged"))
        norm_items.append({
            "product_description": desc_val,
            "brand": brand_val,
            "packing_style": pack_val,
            "grade": grade_val,
            "no_of_mc": mc,
            "quantity": qty,
            "unit_price": price,
            "total_amount": amt,
            "is_unmerged": is_unmerged,
            "desc_span": 1,
            "brand_span": 1,
            "pack_span": 1,
        })

    i = 0
    n = len(norm_items)
    while i < n:
        count = 1
        curr = norm_items[i]
        if not curr.get("is_unmerged"):
            while i + count < n:
                nxt = norm_items[i + count]
                if (
                    not nxt.get("is_unmerged")
                    and curr["product_description"] == nxt["product_description"]
                    and curr["brand"] == nxt["brand"]
                    and curr["packing_style"] == nxt["packing_style"]
                ):
                    count += 1
                else:
                    break

        norm_items[i]["desc_span"] = count
        norm_items[i]["brand_span"] = count
        norm_items[i]["pack_span"] = count
        for k in range(1, count):
            norm_items[i + k]["desc_span"] = 0
            norm_items[i + k]["brand_span"] = 0
            norm_items[i + k]["pack_span"] = 0
        i += count

    return norm_items


def render_document_pdf(
    cfg,
    row,
    company_id: str,
    doc_type: str,
    company_profile: dict | None = None,
    packing_rows: list[PackingList] | None = None,
) -> bytes:
    payload = build_document_payload(cfg, row, packing_rows)
    raw_items = []
    if hasattr(row, "items_json") and row.items_json:
        try:
            raw_items = json.loads(row.items_json)
        except Exception:
            raw_items = []
    if not raw_items and hasattr(row, "product_description"):
        raw_items = [{
            "product_description": getattr(row, "product_description", ""),
            "brand": getattr(row, "brand", ""),
            "packing_style": getattr(row, "packing_style", ""),
            "grade": getattr(row, "grade", ""),
            "no_of_mc": getattr(row, "no_of_mc", 0),
            "quantity": getattr(row, "quantity", 0),
            "unit_price": getattr(row, "unit_price", 0),
            "total_amount": getattr(row, "total_amount", 0),
        }]
    processed_items = process_items_with_spans(raw_items)
    try:
        html = templates.env.get_template("export_documents/print_document_pdf.html").render(
            **payload,
            company_id=company_id,
            company=company_profile or {"name": company_id, "address": "", "email": "", "code": company_id},
            record=row,
            items=processed_items,
            doc_type=doc_type,
            generated_at=datetime.utcnow(),
        )
        from xhtml2pdf import pisa
        output = BytesIO()
        result = pisa.CreatePDF(BytesIO(html.encode("utf-8")), dest=output, encoding="utf-8")
        if not result.err:
            return output.getvalue()
    except Exception as exc:
        logger.warning("HTML PDF rendering failed for %s: %s", doc_type, exc)
    return make_simple_pdf(cfg["title"], str(getattr(row, cfg["no"], row.id)), ["Document export copy"], company_id)


def make_simple_pdf(title: str, document_no: str, lines: list[str], company_name: str = "COMPANY") -> bytes:
    output = BytesIO()
    output.write(b"%PDF-1.4\n1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] >>\nendobj\nxref\n0 4\n0000000000 65535 f \n0000000010 00000 n \n0000000060 00000 n \n0000000117 00000 n \ntrailer\n<< /Size 4 /Root 1 0 R >>\nstartxref\n185\n%%EOF")
    return output.getvalue()


def document_register_workbook(db: Session, company_id: str, doc_type: str | None = None) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Register"
    sheet.append(["Company", company_id])
    sheet.append(["Document Register"])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# 📋 REQUIREMENT FIELD VALUE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def requirement_field_values(values: list) -> list:
    """Deduplicate and strip empty strings from a multi-select value list."""
    seen = []
    for v in values:
        stripped = str(v).strip() if v is not None else ""
        if stripped and stripped not in seen:
            seen.append(stripped)
    return seen


def requirement_display_value(values: list) -> str:
    """Join a list of values into a comma-separated display string."""
    return ", ".join(str(v).strip() for v in values if v is not None and str(v).strip())


# ─────────────────────────────────────────────────────────────────────────────
# 📧 EMAIL APPROVAL STATUS REFRESH
# ─────────────────────────────────────────────────────────────────────────────

def refresh_email_approval_status(file_row, approvals: list) -> None:
    """Recompute approval_status on a file row based on the approval records.

    Rules:
    - If every approval is APPROVED  → APPROVED
    - If any approval is REJECTED    → REJECTED
    - Otherwise                      → PENDING
    """
    if not approvals:
        file_row.approval_status = "PENDING"
        return

    statuses = [str(a.decision or "").upper().strip() for a in approvals]

    if any(s == "REJECTED" for s in statuses):
        file_row.approval_status = "REJECTED"
    elif all(s == "APPROVED" for s in statuses):
        file_row.approval_status = "APPROVED"
    else:
        file_row.approval_status = "PENDING"

