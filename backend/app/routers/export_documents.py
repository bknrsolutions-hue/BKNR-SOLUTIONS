from fastapi import APIRouter, Request, Depends, Body, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field, model_validator
from decimal import Decimal
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, and_, or_, text
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
import re
import json
import logging
import textwrap
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.utils.download_security import require_download_grant
from app.utils.data_management_audit import log_data_management_action
from app.database.models.invoices import (
    ProformaInvoice,
    ExportShipment,
    ExportComplianceTracker,
    CommercialInvoice,
    PackingList,
    ContainerStuffing,
    ShippingBill,
    BillOfLading,
    HealthCertificate,
    ExportDocumentFile,
    ExportDocumentApproval,
    ExportRequiredDocument,
)
from app.database.models.users import Company, User
from app.database.models.criteria import (
    buyers, buyer_agents, countries, species, varieties, grades, brands,
    glazes, freezers, packing_styles, shipping_vendors,
)
from app.database.models.enterprise_finance import BankMaster, ProductionCostAllocation
from app.database.models.inventory_management import pending_orders, sales_dispatch
from app.database.models.processing import AuditLog  # Audit trails
from app.services.cache import cache_get_or_set, invalidate_company_cache
from app.services.bill_accounting import (
    cancel_linked_bill_voucher,
    ensure_bill_accounting_schema,
    post_export_sales_invoice,
)
from app.services.posting_engine import PostingEngineService

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")
logger = logging.getLogger(__name__)

EXPORT_PDF_DIR = Path("uploads/export_documents_private")
_EXPORT_SCHEMA_READY = False


def repost_invoice_cogs(db: Session, company_id: str, invoice: CommercialInvoice, email: str) -> float:
    """Value packing-list batches and keep one COGS voucher per invoice."""
    if invoice.cogs_journal_id:
        PostingEngineService.reverse_voucher(
            db, company_id, invoice.cogs_journal_id,
            "Packing list valuation revised", email or "SYSTEM",
        )
        invoice.cogs_journal_id = None
    rows = db.query(PackingList).filter(
        PackingList.company_id == company_id,
        PackingList.invoice_no == invoice.invoice_no,
        PackingList.is_cancelled != True,
    ).all()
    total_cogs = 0.0
    missing_batches = []
    for row in rows:
        batch = str(row.batch_no or "").strip()
        if not batch:
            missing_batches.append(row.packing_no)
            continue
        allocation = db.query(ProductionCostAllocation).filter(
            ProductionCostAllocation.company_id == company_id,
            ProductionCostAllocation.batch_number == batch,
            ProductionCostAllocation.status == "FG_TRANSFERRED",
            ProductionCostAllocation.is_cancelled != True,
        ).first()
        if not allocation or float(allocation.cost_per_kg or 0) <= 0:
            missing_batches.append(batch)
            continue
        total_cogs += float(row.net_weight or 0) * float(allocation.cost_per_kg or 0)
    total_cogs = round(total_cogs, 2)
    if missing_batches:
        raise ValueError("Complete FG cost allocation for packing batches: " + ", ".join(sorted(set(missing_batches))))
    if total_cogs <= 0:
        return 0.0
    voucher = PostingEngineService.create_voucher(
        db, company_id, "Journal", invoice.invoice_date,
        f"Cost of goods sold for export invoice {invoice.invoice_no}",
        [
            {"ledger_name": "Cost of Goods Sold A/c", "group_name": "Direct Expenses", "group_type": "EXPENSE", "debit_amount": total_cogs, "credit_amount": 0.0, "remarks": invoice.invoice_no},
            {"ledger_name": "Finished Goods Inventory A/c", "group_name": "Stock-in-hand", "group_type": "ASSET", "debit_amount": 0.0, "credit_amount": total_cogs, "remarks": invoice.invoice_no},
        ],
        reference_no=invoice.invoice_no, created_by=email or "SYSTEM",
    )
    invoice.cogs_journal_id = voucher.id
    return total_cogs

EXPORT_SUPPORT_DOCUMENT_TYPES = [
    # Order and buyer approval stage
    {"code": "PROFORMA_INVOICE", "label": "Proforma Invoice (PI)", "stage": "Order & Contract"},
    {"code": "BUYER_PO", "label": "Buyer Purchase Order", "stage": "Order & Contract"},
    {"code": "SALES_CONTRACT", "label": "Sales / Export Contract", "stage": "Order & Contract"},
    {"code": "LC_COPY", "label": "Letter of Credit (LC) Copy", "stage": "Order & Contract"},
    {"code": "LC_AMENDMENT", "label": "LC Amendment", "stage": "Order & Contract"},
    {"code": "ADVANCE_PAYMENT_PROOF", "label": "Advance Payment / SWIFT Proof", "stage": "Order & Contract"},
    {"code": "BUYER_APPROVAL", "label": "Buyer Approval / Email Copy", "stage": "Order & Contract"},
    {"code": "PRODUCT_SPECIFICATION", "label": "Product Specification / MSDS", "stage": "Order & Contract"},
    {"code": "LABEL_ARTWORK_APPROVAL", "label": "Label / Artwork Approval", "stage": "Order & Contract"},
    {"code": "IMPORT_PERMIT", "label": "Buyer Country Import Permit", "stage": "Order & Contract"},

    # Seafood production, quality and traceability stage
    {"code": "BATCH_TRACEABILITY", "label": "Batch & Lot Traceability Record", "stage": "Seafood Quality"},
    {"code": "FARM_CATCH_DECLARATION", "label": "Farm / Catch Declaration", "stage": "Seafood Quality"},
    {"code": "CATCH_CERTIFICATE", "label": "Catch Certificate", "stage": "Seafood Quality"},
    {"code": "HACCP_CHECKLIST", "label": "HACCP / Processing Checklist", "stage": "Seafood Quality"},
    {"code": "QC_INSPECTION_REPORT", "label": "QC Inspection Report", "stage": "Seafood Quality"},
    {"code": "LAB_TEST_REPORT", "label": "Laboratory Test Report", "stage": "Seafood Quality"},
    {"code": "MICROBIOLOGY_REPORT", "label": "Microbiology Report", "stage": "Seafood Quality"},
    {"code": "ANTIBIOTIC_RESIDUE_REPORT", "label": "Antibiotic Residue Test Report", "stage": "Seafood Quality"},
    {"code": "HEAVY_METAL_REPORT", "label": "Heavy Metal Test Report", "stage": "Seafood Quality"},
    {"code": "WATER_ICE_TEST_REPORT", "label": "Water / Ice Quality Report", "stage": "Seafood Quality"},
    {"code": "TEMPERATURE_LOG", "label": "Cold-chain Temperature Log", "stage": "Seafood Quality"},
    {"code": "WEIGHT_PACKING_VERIFICATION", "label": "Weight & Packing Verification", "stage": "Seafood Quality"},

    # Statutory and certification stage
    {"code": "EIA_INSPECTION_REPORT", "label": "EIA Inspection Report", "stage": "Certificates"},
    {"code": "HEALTH_CERTIFICATE_COPY", "label": "Health Certificate", "stage": "Certificates"},
    {"code": "PHYTO_CERTIFICATE", "label": "Phytosanitary Certificate", "stage": "Certificates"},
    {"code": "VETERINARY_CERTIFICATE", "label": "Veterinary Certificate", "stage": "Certificates"},
    {"code": "COO", "label": "Certificate of Origin", "stage": "Certificates"},
    {"code": "FUMIGATION_CERTIFICATE", "label": "Fumigation Certificate", "stage": "Certificates"},
    {"code": "HALAL_CERTIFICATE", "label": "Halal Certificate", "stage": "Certificates"},
    {"code": "ANIMAL_QUARANTINE_NOC", "label": "Animal Quarantine / NOC", "stage": "Certificates"},

    # Commercial, customs and logistics stage
    {"code": "COMMERCIAL_INVOICE", "label": "Commercial Invoice", "stage": "Shipping & Customs"},
    {"code": "PACKING_LIST", "label": "Packing List", "stage": "Shipping & Customs"},
    {"code": "CONTAINER_STUFFING_REPORT", "label": "Container Stuffing Report", "stage": "Shipping & Customs"},
    {"code": "CONTAINER_SEAL_REPORT", "label": "Container Seal / Inspection Report", "stage": "Shipping & Customs"},
    {"code": "VGM_DECLARATION", "label": "VGM Declaration", "stage": "Shipping & Customs"},
    {"code": "SHIPPING_BILL", "label": "Shipping Bill", "stage": "Shipping & Customs"},
    {"code": "CUSTOMS_LEO_COPY", "label": "Customs Let Export Order (LEO)", "stage": "Shipping & Customs"},
    {"code": "BILL_OF_LADING_DRAFT", "label": "Bill of Lading Draft", "stage": "Shipping & Customs"},
    {"code": "BL_COPY", "label": "Final Bill of Lading / AWB", "stage": "Shipping & Customs"},
    {"code": "INSURANCE_CERTIFICATE", "label": "Marine Insurance Certificate", "stage": "Shipping & Customs"},
    {"code": "FREIGHT_INVOICE", "label": "Freight Invoice", "stage": "Shipping & Customs"},
    {"code": "CHA_INVOICE", "label": "CHA / Customs Broker Invoice", "stage": "Shipping & Customs"},
    {"code": "PORT_TERMINAL_RECEIPT", "label": "Port / Terminal Receipt", "stage": "Shipping & Customs"},

    # Bank submission, payment and closure stage
    {"code": "BANK_SUBMISSION_SET", "label": "Bank Document Submission Set", "stage": "Bank & Payment"},
    {"code": "BILL_OF_EXCHANGE", "label": "Bill of Exchange", "stage": "Bank & Payment"},
    {"code": "NEGOTIATION_COLLECTION_PROOF", "label": "Negotiation / Collection Proof", "stage": "Bank & Payment"},
    {"code": "PAYMENT_SWIFT_COPY", "label": "Payment SWIFT Copy", "stage": "Bank & Payment"},
    {"code": "PAYMENT_RECEIPT", "label": "Payment Receipt / Bank Credit Advice", "stage": "Bank & Payment"},
    {"code": "FIRC", "label": "Foreign Inward Remittance Certificate (FIRC)", "stage": "Bank & Payment"},
    {"code": "EBRC", "label": "Electronic Bank Realisation Certificate (e-BRC)", "stage": "Bank & Payment"},
    {"code": "CREDIT_DEBIT_NOTE", "label": "Credit / Debit Note", "stage": "Bank & Payment"},
    {"code": "DUTY_DRAWBACK_PROOF", "label": "Duty Drawback Credit Proof", "stage": "Bank & Payment"},
    {"code": "RODTEP_CREDIT_PROOF", "label": "RoDTEP Credit Proof", "stage": "Bank & Payment"},
    {"code": "EXPORT_CLOSURE_CONFIRMATION", "label": "Export File Closure Confirmation", "stage": "Bank & Payment"},
]

# Document handling follows the same separation used by enterprise output
# systems: structured data is independent from the rendered/imported PDF.
EXPORT_GENERATE_DOCUMENTS = {
    "SALES_CONTRACT", "PRODUCT_SPECIFICATION",
    "BATCH_TRACEABILITY", "HACCP_CHECKLIST", "TEMPERATURE_LOG",
    "WEIGHT_PACKING_VERIFICATION", "COMMERCIAL_INVOICE", "PACKING_LIST",
    "CONTAINER_STUFFING_REPORT", "VGM_DECLARATION", "BILL_OF_EXCHANGE",
    "BANK_SUBMISSION_SET", "CREDIT_DEBIT_NOTE", "EXPORT_CLOSURE_CONFIRMATION",
}
EXPORT_HYBRID_DOCUMENTS = {
    "SALES_CONTRACT", "PRODUCT_SPECIFICATION",
    "CATCH_CERTIFICATE", "HEALTH_CERTIFICATE_COPY", "COMMERCIAL_INVOICE",
    "PACKING_LIST", "CONTAINER_STUFFING_REPORT", "VGM_DECLARATION",
    "BILL_OF_EXCHANGE", "BANK_SUBMISSION_SET", "CREDIT_DEBIT_NOTE",
}


def export_document_mode(document_code: str) -> str:
    if document_code == "PROFORMA_INVOICE":
        return "IMPORT_FINAL_PDF"
    if document_code in EXPORT_HYBRID_DOCUMENTS:
        return "GENERATE_AND_IMPORT_FINAL"
    if document_code in EXPORT_GENERATE_DOCUMENTS:
        return "GENERATE"
    return "IMPORT_PDF"

EXPORT_REQUIREMENT_STAGE_FIELDS = {
    "Order & Contract": [
        {"name": "buyer_name", "label": "Buyer", "type": "select", "lookup": "buyers"},
        {"name": "buyer_agent", "label": "Buyer Agent", "type": "select", "lookup": "buyer_agents"},
        {"name": "destination_country", "label": "Destination Country", "type": "select", "lookup": "countries"},
        {"name": "buyer_reference", "label": "Buyer Reference", "type": "text"},
        {"name": "contract_date", "label": "Contract / Approval Date", "type": "date"},
        {"name": "validity_date", "label": "Validity Date", "type": "date"},
        {"name": "incoterm", "label": "Incoterm", "type": "select", "options": ["FOB", "CFR", "CIF", "EXW", "FCA", "CPT", "CIP", "DDP"]},
        {"name": "payment_terms", "label": "Payment Terms", "type": "text"},
        {"name": "product_description", "label": "Product / Specification", "type": "textarea"},
    ],
    "Seafood Quality": [
        {"name": "batch_no", "label": "Batch Number", "type": "text"},
        {"name": "lot_no", "label": "Lot Number", "type": "text"},
        {"name": "species", "label": "Species", "type": "select", "lookup": "species", "multiple": True},
        {"name": "variety", "label": "Variety", "type": "select", "lookup": "varieties", "multiple": True},
        {"name": "grade", "label": "Grade / Size", "type": "select", "lookup": "grades", "multiple": True},
        {"name": "brand", "label": "Brand", "type": "select", "lookup": "brands"},
        {"name": "glaze", "label": "Glaze", "type": "select", "lookup": "glazes", "multiple": True},
        {"name": "freezer", "label": "Freezer", "type": "select", "lookup": "freezers", "multiple": True},
        {"name": "packing_style", "label": "Packing Style", "type": "select", "lookup": "packing_styles", "multiple": True},
        {"name": "lab_name", "label": "Laboratory / Inspector", "type": "text"},
        {"name": "sample_date", "label": "Sample Date", "type": "date"},
        {"name": "result", "label": "Test / Inspection Result", "type": "select", "options": ["PASS", "FAIL", "CONDITIONAL", "NA"]},
        {"name": "temperature", "label": "Temperature", "type": "text"},
    ],
    "Certificates": [
        {"name": "certificate_no", "label": "Certificate Number", "type": "text"},
        {"name": "authority", "label": "Issuing Authority", "type": "text"},
        {"name": "factory_approval_no", "label": "Factory Approval Number", "type": "text"},
        {"name": "destination_country", "label": "Destination Country", "type": "select", "lookup": "countries"},
        {"name": "species", "label": "Species / Product", "type": "select", "lookup": "species", "multiple": True},
        {"name": "health_marks", "label": "Health Marks / Endorsement", "type": "textarea"},
    ],
    "Shipping & Customs": [
        {"name": "invoice_no", "label": "Invoice Number", "type": "text"},
        {"name": "container_no", "label": "Container Number", "type": "text"},
        {"name": "seal_no", "label": "Seal Number", "type": "text"},
        {"name": "shipping_line", "label": "Shipping Line / CHA", "type": "select", "lookup": "shipping_vendors"},
        {"name": "freezer", "label": "Freezer", "type": "select", "lookup": "freezers", "multiple": True},
        {"name": "packing_style", "label": "Packing Style", "type": "select", "lookup": "packing_styles", "multiple": True},
        {"name": "vessel_voyage", "label": "Vessel / Voyage", "type": "text"},
        {"name": "port_of_loading", "label": "Port of Loading", "type": "text"},
        {"name": "port_of_discharge", "label": "Port of Discharge", "type": "text"},
        {"name": "etd", "label": "ETD", "type": "date"},
        {"name": "eta", "label": "ETA", "type": "date"},
        {"name": "gross_weight", "label": "Gross Weight", "type": "number"},
        {"name": "net_weight", "label": "Net Weight", "type": "number"},
    ],
    "Bank & Payment": [
        {"name": "bank_account", "label": "Company Bank Account", "type": "select", "lookup": "bank_accounts"},
        {"name": "swift_reference", "label": "SWIFT / Bank Reference", "type": "text"},
        {"name": "invoice_no", "label": "Invoice Number", "type": "text"},
        {"name": "submission_date", "label": "Submission Date", "type": "date"},
        {"name": "receipt_date", "label": "Receipt / Credit Date", "type": "date"},
        {"name": "bill_reference", "label": "Bill / e-BRC Reference", "type": "text"},
        {"name": "realisation_status", "label": "Realisation Status", "type": "select", "options": ["PENDING", "PARTIAL", "REALISED", "CLOSED"]},
    ],
    "Custom": [
        {"name": "custom_reference_1", "label": "Additional Reference 1", "type": "text"},
        {"name": "custom_reference_2", "label": "Additional Reference 2", "type": "text"},
        {"name": "description", "label": "Document Details", "type": "textarea"},
    ],
}

EXPORT_REQUIREMENT_COMMON_FIELDS = [
    {"name": "document_no", "label": "Document Number / Reference", "type": "text", "required": True},
    {"name": "document_date", "label": "Document Date", "type": "date", "required": True},
    {"name": "expiry_date", "label": "Expiry / Valid Until", "type": "date"},
    {"name": "issuer_name", "label": "Issuer / Organisation", "type": "text"},
    {"name": "reference_no", "label": "Secondary Reference", "type": "text"},
    {"name": "currency", "label": "Currency", "type": "select", "options": ["USD", "EUR", "GBP", "AED", "JPY", "INR"]},
    {"name": "amount", "label": "Amount / Value", "type": "number"},
    {"name": "status_note", "label": "Status / Notes", "type": "textarea"},
]


def invalidate_export_cache(company_id: str | None):
    if company_id:
        invalidate_company_cache(company_id, "export_documents")


def _dt(value):
    if value is None:
        return None
    return value.isoformat() if hasattr(value, "isoformat") else str(value)


def safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(value or "document")).strip("_")[:120]


def export_doc_config():
    return {
        "proforma_invoice": {
            "model": ProformaInvoice, "no": "pi_no", "date": "pi_date",
            "title": "Proforma Invoice", "template": "export_documents/print_document.html",
            "fields": [
                ("PI No", "pi_no"), ("PI Date", "pi_date"), ("Valid Until", "validity_date"),
                ("Buyer", "buyer_name"), ("Buyer Address", "buyer_address"),
                ("Country", "country"), ("Buyer PO", "po_number"), ("Currency", "currency"),
                ("Incoterm", "incoterm"), ("Payment Terms", "payment_terms"),
                ("Port of Loading", "port_of_loading"), ("Port of Discharge", "port_of_discharge"),
                ("Product Description", "product_description"), ("Quantity", "quantity"),
                ("Unit", "unit"), ("Unit Price", "unit_price"), ("Total Amount", "total_amount"),
                ("Status", "status"), ("Approval Status", "approval_status"),
                ("Approved By", "approved_by"), ("Approval Remarks", "approval_remarks"),
                ("Remarks", "remarks"),
            ],
        },
        "export_shipment": {
            "model": ExportShipment, "no": "shipment_no", "date": "created_at",
            "title": "Export Shipment File", "template": "export_documents/print_document.html",
            "fields": [
                ("Shipment No", "shipment_no"), ("PO Number", "po_number"), ("Invoice No", "invoice_no"),
                ("Container No", "container_no"), ("Buyer", "buyer_name"), ("Country", "country"),
                ("ETD", "etd"), ("ETA", "eta"), ("Completion Date", "completion_date"),
                ("Status", "status"), ("Approval Status", "approval_status"),
            ],
        },
        "commercial_invoice": {
            "model": CommercialInvoice, "no": "invoice_no", "date": "invoice_date",
            "title": "Commercial Invoice", "template": "export_documents/print_document.html",
            "fields": [
                ("Invoice No", "invoice_no"), ("Invoice Date", "invoice_date"), ("Shipment No", "shipment_no"),
                ("PO Number", "po_number"), ("Buyer", "buyer_name"), ("Buyer Address", "buyer_address"),
                ("Consignee", "consignee_name"), ("Notify Party", "notify_party"), ("Country", "country"),
                ("Currency", "currency"), ("Total Amount", "total_amount"), ("Exchange Rate", "exchange_rate"),
                ("Invoice Value INR", "invoice_value_inr"), ("Payment Terms", "payment_terms"),
                ("Shipment Terms", "shipment_terms"), ("Port of Loading", "port_of_loading"),
                ("Port of Discharge", "port_of_discharge"), ("Final Destination", "final_destination"),
                ("Shipment Type", "shipment_type"), ("Total Master Cartons", "total_mc"),
                ("Total Net Weight", "total_net_weight"), ("Total Gross Weight", "total_gross_weight"),
                ("Payment Status", "payment_status"), ("Approval Status", "approval_status"),
            ],
        },
        "packing_list": {
            "model": PackingList, "no": "packing_no", "date": "created_at",
            "title": "Packing List", "template": "export_documents/print_document.html",
            "fields": [
                ("Packing No", "packing_no"), ("Invoice No", "invoice_no"), ("PO Number", "po_number"),
                ("Container No", "container_no"), ("Buyer", "buyer_name"), ("Product", "product_name"),
                ("Grade", "grade"), ("Batch No", "batch_no"), ("Lot No", "lot_no"), ("Glaze", "glaze"),
                ("Freezing Type", "freezing_type"), ("HS Code", "hs_code"), ("Packing Style", "packing_style"),
                ("Inner Pack", "inner_pack"), ("Outer Pack", "outer_pack"), ("Master Cartons", "master_cartons"),
                ("Net Weight", "net_weight"), ("Gross Weight", "gross_weight"), ("Pallet Count", "pallet_count"),
                ("Manufacturing Date", "manufacturing_date"), ("Expiry Date", "expiry_date"),
            ],
        },
        "container_stuffing": {
            "model": ContainerStuffing, "no": "container_no", "date": "stuffing_date",
            "title": "Container Stuffing Report", "template": "export_documents/print_document.html",
            "fields": [
                ("Container No", "container_no"), ("Invoice No", "invoice_no"), ("PO Number", "po_number"),
                ("Buyer", "buyer_name"), ("Seal No", "seal_no"), ("Shipping Line", "shipping_line"),
                ("Stuffing Date", "stuffing_date"), ("Stuffing Location", "stuffing_location"),
                ("Container Type", "container_type"), ("Container Size", "container_size"),
                ("Set Temperature", "temperature"), ("Vehicle No", "vehicle_no"),
                ("Container Condition", "container_condition"), ("Temperature Before Loading", "temperature_before_loading"),
                ("Temperature After Loading", "temperature_after_loading"), ("Driver", "driver_name"),
                ("Loading Supervisor", "loading_supervisor"), ("Approval Status", "approval_status"), ("Remarks", "remarks"),
            ],
        },
        "shipping_bill": {
            "model": ShippingBill, "no": "shipping_bill_no", "date": "shipping_bill_date",
            "title": "Shipping Bill Summary", "template": "export_documents/print_document.html",
            "fields": [
                ("Shipping Bill No", "shipping_bill_no"), ("Shipping Bill Date", "shipping_bill_date"),
                ("Invoice No", "invoice_no"), ("Container No", "container_no"), ("PO Number", "po_number"),
                ("Buyer", "buyer_name"), ("FOB Value INR", "shipping_bill_value"),
                ("Drawback Amount", "drawback_amount"), ("Scheme", "scheme"), ("Customs Status", "customs_status"),
                ("Port", "port"), ("CHA Name", "cha_name"), ("CHA Bill No", "cha_bill_no"), ("Vessel", "vessel_name"),
                ("Voyage No", "voyage_no"), ("ETD", "etd"), ("ETA", "eta"),
                ("Approval Status", "approval_status"), ("Remarks", "remarks"),
            ],
        },
        "bill_of_lading": {
            "model": BillOfLading, "no": "bl_no", "date": "bl_date",
            "title": "Bill of Lading", "template": "export_documents/print_document.html",
            "fields": [
                ("B/L No", "bl_no"), ("B/L Date", "bl_date"), ("On Board Date", "onboard_date"),
                ("Invoice No", "invoice_no"), ("Container No", "container_no"), ("PO Number", "po_number"),
                ("Buyer", "buyer_name"), ("Shipping Line", "shipping_line"), ("Seal No", "seal_no"),
                ("Freight Terms", "freight_terms"), ("Original B/L Count", "no_of_original_bl"),
                ("Marks & Numbers", "marks_and_numbers"), ("Packages Description", "packages_description"),
                ("Place of Receipt", "place_of_receipt"), ("Place of Delivery", "place_of_delivery"),
                ("Gross Weight", "gross_weight"), ("Net Weight", "net_weight"),
                ("Approval Status", "approval_status"),
            ],
        },
        "health_certificate": {
            "model": HealthCertificate, "no": "certificate_no", "date": "issue_date",
            "title": "Health Certificate", "template": "export_documents/print_document.html",
            "fields": [
                ("Certificate No", "certificate_no"), ("Issue Date", "issue_date"), ("Authority", "authority"),
                ("Factory Approval No", "factory_approval_no"), ("Invoice No", "invoice_no"),
                ("Container No", "container_no"), ("PO Number", "po_number"), ("Buyer", "buyer_name"),
                ("Country", "country"), ("Species", "species"), ("Temperature Verified", "temperature_verified"),
                ("Issued By", "issued_by"), ("Status", "status"), ("Remarks", "remarks"),
            ],
        },
    }


def get_export_record_or_404(db: Session, request: Request, doc_type: str, record_id: int):
    cfg = export_doc_config().get(doc_type)
    if not cfg:
        raise HTTPException(status_code=404, detail="Unsupported document type")
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Unauthorized")
    row = db.query(cfg["model"]).filter(cfg["model"].id == record_id, cfg["model"].company_id == comp_code).first()
    if not row:
        raise HTTPException(status_code=404, detail="Record not found")
    return cfg, row, comp_code


def store_export_pdf(
    db: Session,
    company_id: str,
    module_name: str,
    record_id: int,
    document_no: str,
    document_kind: str,
    file_name: str,
    content: bytes,
    uploaded_by: str | None,
    remarks: str | None = None,
) -> ExportDocumentFile:
    EXPORT_PDF_DIR.mkdir(parents=True, exist_ok=True)
    for old in db.query(ExportDocumentFile).filter(
        ExportDocumentFile.company_id == company_id,
        ExportDocumentFile.module_name == module_name,
        ExportDocumentFile.record_id == record_id,
        ExportDocumentFile.document_kind == document_kind,
        ExportDocumentFile.is_current == True,
    ).all():
        old.is_current = False
    version_no = (
        db.query(func.coalesce(func.max(ExportDocumentFile.version_no), 0))
        .filter(
            ExportDocumentFile.company_id == company_id,
            ExportDocumentFile.module_name == module_name,
            ExportDocumentFile.record_id == record_id,
            ExportDocumentFile.document_kind == document_kind,
        )
        .scalar()
        + 1
    )
    final_name = f"{safe_filename(module_name)}_{safe_filename(document_no)}_v{version_no}_{safe_filename(file_name)}"
    disk_path = EXPORT_PDF_DIR / final_name
    disk_path.write_bytes(content)
    file_row = ExportDocumentFile(
        company_id=company_id,
        module_name=module_name,
        record_id=record_id,
        document_no=document_no,
        document_kind=document_kind,
        file_name=final_name,
        file_path=None,
        content_type="application/pdf",
        file_bytes=content,
        file_size=len(content),
        version_no=version_no,
        uploaded_by=uploaded_by,
        remarks=remarks,
    )
    db.add(file_row)
    db.flush()
    file_row.file_path = f"/export_documents/files/{file_row.id}/download"
    return file_row


def set_document_path(row, path: str):
    if hasattr(row, "document_path"):
        row.document_path = path


EXPORT_FIELD_LABELS = {
    "id": "Internal Reference ID",
    "company_id": "Company Code",
    "pi_no": "PI Number",
    "po_number": "Buyer PO Number",
    "bl_no": "B/L Number",
    "no_of_original_bl": "Number of Original B/L",
    "etd": "ETD",
    "eta": "ETA",
    "hs_code": "HS Code",
    "total_mc": "Total Master Cartons",
    "gstr1_updated": "GSTR-1 Updated",
    "journal_id": "Sales Journal ID",
    "cogs_journal_id": "COGS Journal ID",
    "customer_ledger_id": "Customer Ledger ID",
    "sales_ledger_id": "Sales Ledger ID",
    "document_path": "Stored Document Reference",
    "photo_path": "Stuffing Photo Reference",
}

EXPORT_DOCUMENT_PREFIXES = {
    ProformaInvoice: "PI",
    ExportShipment: "ES",
    CommercialInvoice: "CI",
    PackingList: "PL",
    ContainerStuffing: "CS",
    ShippingBill: "SB",
    BillOfLading: "BL",
    HealthCertificate: "HC",
}

EXPORT_INTERNAL_PATH_FIELDS = {"document_path", "photo_path"}
EXPORT_REFERENCE_FIELDS = {
    "id", "company_id", "pi_no", "shipment_no", "invoice_no", "packing_no",
    "shipping_bill_no", "bl_no", "certificate_no", "po_number", "container_no",
    "seal_no", "cha_bill_no", "stock_entry_no", "inventory_batch_id",
    "invoice_item_no",
}
EXPORT_PARTY_FIELDS = {
    "buyer_name", "buyer_address", "consignee_name", "notify_party", "country",
    "company_name", "authority", "issued_by", "cha_name", "driver_name",
    "loading_supervisor",
}
EXPORT_FINANCIAL_FIELDS = {
    "currency", "exchange_rate", "quantity", "unit", "unit_price", "total_amount",
    "invoice_value_inr", "shipping_bill_value", "drawback_amount", "scheme",
    "payment_terms", "payment_status", "incoterm", "shipment_terms",
}
EXPORT_CARGO_FIELDS = {
    "product_description", "product_name", "grade", "batch_no", "lot_no", "glaze",
    "freezing_type", "hs_code", "packing_style", "inner_pack", "outer_pack",
    "master_cartons", "total_mc", "net_weight", "gross_weight", "total_net_weight",
    "total_gross_weight", "pallet_count", "species", "marks_and_numbers",
    "packages_description", "manufacturing_date", "expiry_date",
}
EXPORT_LOGISTICS_FIELDS = {
    "port_of_loading", "port_of_discharge", "final_destination", "shipment_type",
    "shipping_line", "stuffing_date", "stuffing_location", "container_type",
    "container_size", "container_condition", "start_time", "end_time",
    "temperature_before_loading", "temperature_after_loading", "temperature",
    "vehicle_no", "port", "vessel_name", "voyage_no", "etd", "eta",
    "completion_date", "bl_date", "onboard_date", "place_of_receipt",
    "place_of_delivery", "freight_terms", "no_of_original_bl",
    "temperature_verified", "shipping_bill_date", "issue_date", "pi_date",
    "validity_date", "invoice_date",
}
EXPORT_CONTROL_FIELDS = {
    "status", "customs_status", "approval_status", "approved_by", "approved_at",
    "approval_remarks", "remarks", "is_completed", "is_cancelled",
}
EXPORT_AUDIT_FIELDS = {
    "created_by", "updated_by", "created_at", "updated_at", "journal_id",
    "cogs_journal_id", "gstr1_updated", "customer_ledger_id", "sales_ledger_id",
    "document_path", "photo_path",
}
EXPORT_FULL_WIDTH_PRINT_FIELDS = {
    "buyer_address", "notify_party", "product_description", "remarks",
    "approval_remarks", "marks_and_numbers", "packages_description",
}
EXPORT_MEDIUM_WIDTH_PRINT_FIELDS = {
    "buyer_name", "consignee_name", "authority", "issued_by", "cha_name",
    "payment_terms", "shipment_terms", "port_of_loading", "port_of_discharge",
    "final_destination", "stuffing_location", "shipping_line",
}


def humanize_export_field(field_name: str) -> str:
    if field_name in EXPORT_FIELD_LABELS:
        return EXPORT_FIELD_LABELS[field_name]
    return str(field_name).replace("_", " ").title()


def export_display_value(value):
    if value is None or value == "":
        return "—"
    if isinstance(value, bool):
        return "YES" if value else "NO"
    if isinstance(value, datetime):
        return value.strftime("%d-%b-%Y %I:%M %p")
    if isinstance(value, date):
        return value.strftime("%d-%b-%Y")
    if isinstance(value, Decimal):
        return f"{value:,.2f}"
    if isinstance(value, float):
        return f"{value:,.2f}"
    return value


def export_field_section(field_name: str) -> str:
    if field_name in EXPORT_REFERENCE_FIELDS:
        return "Document References"
    if field_name in EXPORT_PARTY_FIELDS:
        return "Buyer, Parties & Authorities"
    if field_name in EXPORT_FINANCIAL_FIELDS:
        return "Commercial & Financial Details"
    if field_name in EXPORT_CARGO_FIELDS:
        return "Cargo, Packing & Traceability"
    if field_name in EXPORT_LOGISTICS_FIELDS:
        return "Shipping, Dates & Logistics"
    if field_name in EXPORT_CONTROL_FIELDS:
        return "Status, Approval & Remarks"
    if field_name in EXPORT_AUDIT_FIELDS:
        return "Document Control & Audit Trail"
    return "Additional Document Details"


def export_print_field_span(field_name: str, value) -> int:
    """Return a six-column print-grid span based on field content/type."""
    text_value = str(value or "")
    if field_name in EXPORT_FULL_WIDTH_PRINT_FIELDS or len(text_value) > 72:
        return 6
    if field_name in EXPORT_MEDIUM_WIDTH_PRINT_FIELDS or len(text_value) > 30:
        return 3
    return 2


def pack_export_print_rows(fields: list[dict]) -> list[list[dict]]:
    """Pack homogeneous 3-up, 2-up or full-width fields without broken gaps."""
    rows = []
    current = []
    used = 0
    current_span = None
    for field in fields:
        item = {**field, "span": export_print_field_span(field["name"], field["value"])}
        span = item["span"]
        if current and (current_span != span or used + span > 6):
            if used < 6:
                current.append({"empty": True, "span": 6 - used})
            rows.append(current)
            current, used = [], 0
        current.append(item)
        current_span = span
        used += span
        if used == 6:
            rows.append(current)
            current, used, current_span = [], 0, None
    if current:
        if used < 6:
            current.append({"empty": True, "span": 6 - used})
        rows.append(current)
    return rows


def get_export_company_profile(db: Session, company_id: str) -> dict:
    company = db.query(Company).filter(Company.company_code == company_id).first()
    return {
        "name": company.company_name if company else company_id,
        "address": company.address if company else "",
        "email": company.email if company else "",
        "code": (
            company.mpeda_registration_code
            if company and company.mpeda_registration_code
            else "MPEDA CODE NOT REGISTERED"
        ),
        "tenant_code": company_id,
    }


PACKING_SCHEDULE_FIELDS = {
    "packing_no", "invoice_item_no", "inventory_batch_id", "stock_entry_no",
    "product_name", "grade", "batch_no", "lot_no", "glaze", "freezing_type",
    "hs_code", "manufacturing_date", "expiry_date", "packing_style",
    "inner_pack", "outer_pack", "master_cartons", "net_weight", "gross_weight",
    "pallet_count",
}


def get_invoice_packing_rows(db: Session, row) -> list[PackingList]:
    if not isinstance(row, PackingList):
        return []
    return db.query(PackingList).filter(
        PackingList.company_id == row.company_id,
        PackingList.invoice_no == row.invoice_no,
        PackingList.is_cancelled != True,
    ).order_by(PackingList.invoice_item_no, PackingList.id).all()


def build_document_payload(cfg, row, packing_rows: list[PackingList] | None = None):
    configured_labels = {attr: label for label, attr in cfg["fields"]}
    ordered_attrs = [attr for _, attr in cfg["fields"]]
    for column in row.__table__.columns:
        if column.name not in ordered_attrs:
            ordered_attrs.append(column.name)

    grouped = {}
    for attr in ordered_attrs:
        # Accounting links, storage references and creation metadata are useful
        # inside ERP screens, but must not appear on customer-facing documents.
        if attr in EXPORT_AUDIT_FIELDS:
            continue
        value = getattr(row, attr, None)
        if attr in EXPORT_INTERNAL_PATH_FIELDS:
            value = "AVAILABLE IN SYSTEM" if value else "NOT ATTACHED"
        grouped.setdefault(export_field_section(attr), []).append({
            "name": attr,
            "label": configured_labels.get(attr, humanize_export_field(attr)),
            "value": export_display_value(value),
        })

    section_order = [
        "Document References",
        "Buyer, Parties & Authorities",
        "Commercial & Financial Details",
        "Cargo, Packing & Traceability",
        "Shipping, Dates & Logistics",
        "Status, Approval & Remarks",
        "Additional Document Details",
        "Document Control & Audit Trail",
    ]
    sections = [
        {"title": title, "fields": grouped[title]}
        for title in section_order
        if grouped.get(title)
    ]
    if isinstance(row, PackingList):
        for section in sections:
            section["fields"] = [
                field for field in section["fields"]
                if field["name"] not in PACKING_SCHEDULE_FIELDS
            ]
        sections = [section for section in sections if section["fields"]]
    for section in sections:
        section["rows"] = pack_export_print_rows(section["fields"])
    field_count = sum(len(section["fields"]) for section in sections)
    payload = {
        "title": cfg["title"],
        "document_no": getattr(row, cfg["no"], ""),
        "document_date": getattr(row, cfg["date"], None),
        "fields": [
            (configured_labels.get(attr, humanize_export_field(attr)), export_display_value(getattr(row, attr, None)))
            for attr in ordered_attrs
            if attr not in EXPORT_AUDIT_FIELDS
        ],
        "sections": sections,
        "field_count": field_count,
        "document_reference": (
            f"{EXPORT_DOCUMENT_PREFIXES.get(cfg['model'], 'DOC')}-{int(getattr(row, 'id', 0) or 0):06d}"
        ),
    }
    if isinstance(row, CommercialInvoice):
        payload["line_items"] = [
            {
                "product": item.product_name,
                "grade": item.grade,
                "packing_style": item.packing_style,
                "batch_no": item.batch_no,
                "master_cartons": item.master_cartons or 0,
                "net_weight": item.net_weight or 0,
                "gross_weight": item.gross_weight or 0,
            }
            for item in row.packing_items
            if not item.is_cancelled
        ]
    if isinstance(row, PackingList):
        source_rows = packing_rows or [row]
        payload["packing_line_items"] = [
            {
                "packing_no": item.packing_no,
                "invoice_item_no": item.invoice_item_no,
                "inventory_batch_id": item.inventory_batch_id,
                "stock_entry_no": item.stock_entry_no,
                "product_name": item.product_name,
                "grade": item.grade,
                "batch_no": item.batch_no,
                "lot_no": item.lot_no,
                "glaze": item.glaze,
                "freezing_type": item.freezing_type,
                "hs_code": item.hs_code,
                "manufacturing_date": export_display_value(item.manufacturing_date),
                "expiry_date": export_display_value(item.expiry_date),
                "packing_style": item.packing_style,
                "inner_pack": item.inner_pack,
                "outer_pack": item.outer_pack,
                "master_cartons": int(item.master_cartons or 0),
                "net_weight": float(item.net_weight or 0),
                "gross_weight": float(item.gross_weight or 0),
                "pallet_count": int(item.pallet_count or 0),
            }
            for item in source_rows
        ]
    line_item_count = len(payload.get("line_items", []))
    line_item_count += len(payload.get("packing_line_items", []))
    estimated_rows = (
        sum(len(section["rows"]) + 1 for section in sections)
        + line_item_count
        + 16
    )
    payload["line_item_count"] = line_item_count
    payload["fit_scale"] = min(1.0, round(40 / max(40, estimated_rows), 3))
    payload["fit_font_size"] = max(3.2, min(6.8, round(7.0 * payload["fit_scale"], 2)))
    return payload


def render_document_pdf(
    cfg,
    row,
    company_id: str,
    doc_type: str,
    company_profile: dict | None = None,
    packing_rows: list[PackingList] | None = None,
) -> bytes:
    payload = build_document_payload(cfg, row, packing_rows)
    html = templates.env.get_template("export_documents/print_document_pdf.html").render(
        **payload,
        company_id=company_id,
        company=company_profile or {"name": company_id, "address": "", "email": "", "code": company_id},
        record=row,
        doc_type=doc_type,
        generated_at=datetime.utcnow(),
    )
    try:
        from xhtml2pdf import pisa

        output = BytesIO()
        result = pisa.CreatePDF(BytesIO(html.encode("utf-8")), dest=output, encoding="utf-8")
        if not result.err:
            return output.getvalue()
    except Exception as exc:
        logger.warning("HTML PDF rendering failed for %s: %s", doc_type, exc)

    lines = []
    for section in payload["sections"]:
        lines.append(f"## {section['title']}")
        lines.extend(f"{field['label']}: {field['value']}" for field in section["fields"])
    if payload.get("packing_line_items"):
        lines.append("## Consolidated Packing Schedule")
        for index, item in enumerate(payload["packing_line_items"], start=1):
            lines.append(
                f"{index}. {item['packing_no']} / Item {item['invoice_item_no'] or '—'} | "
                f"{item['product_name']} / HS {item['hs_code'] or '—'} | "
                f"Grade {item['grade']} / {item['glaze'] or '—'} / {item['freezing_type'] or '—'} | "
                f"Batch {item['batch_no'] or '—'} / Lot {item['lot_no'] or '—'} | "
                f"Mfg {item['manufacturing_date']} / Exp {item['expiry_date']} | "
                f"{item['packing_style']} / Inner {item['inner_pack'] or '—'} / Outer {item['outer_pack'] or '—'} | "
                f"Inventory {item['inventory_batch_id'] or '—'} / Stock {item['stock_entry_no'] or '—'} | "
                f"MC {item['master_cartons']} / Pallet {item['pallet_count']} | "
                f"NW {item['net_weight']:,.2f} / GW {item['gross_weight']:,.2f}"
            )
    return make_simple_pdf(
        payload["title"],
        payload["document_no"],
        lines,
        (company_profile or {}).get("name") or company_id,
    )


def render_requirement_pdf(
    definition: dict,
    details: dict,
    company_id: str,
    reference,
    company_profile: dict | None = None,
) -> bytes:
    """Render a controlled PDF from a generic requirement data model."""
    document_no = str(details.get("document_no") or getattr(reference, "shipment_no", None) or getattr(reference, "pi_no", reference.id))
    fields = [
        (field["label"], requirement_display_value(details.get(field["name"])))
        for field in definition["fields"]
        if requirement_field_values(details.get(field["name"]))
    ]
    requirement_section = {"title": "Document Particulars", "fields": [
        {"name": field["name"], "label": field["label"], "value": export_display_value(details.get(field["name"]))}
        for field in definition["fields"]
    ]}
    requirement_section["rows"] = pack_export_print_rows(requirement_section["fields"])
    html = templates.env.get_template("export_documents/print_document_pdf.html").render(
        title=definition["label"], document_no=document_no,
        document_date=details.get("document_date"), fields=fields,
        sections=[requirement_section],
        document_reference=f"DOC-{int(getattr(reference, 'id', 0) or 0):06d}",
        line_items=[], company_id=company_id,
        company=company_profile or {"name": company_id, "address": "", "email": "", "code": company_id},
        record=reference,
        doc_type=definition["code"].lower(), generated_at=datetime.utcnow(),
    )
    try:
        from xhtml2pdf import pisa

        output = BytesIO()
        result = pisa.CreatePDF(BytesIO(html.encode("utf-8")), dest=output, encoding="utf-8")
        if not result.err:
            return output.getvalue()
    except Exception as exc:
        logger.warning("Requirement PDF rendering failed for %s: %s", definition["code"], exc)
    return make_simple_pdf(
        definition["label"],
        document_no,
        [f"{label}: {value}" for label, value in fields],
        (company_profile or {}).get("name") or company_id,
    )


def style_register_sheet(sheet, title: str, company_id: str, fields: list[tuple[str, str]], rows: list) -> None:
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(fields)))
    sheet.cell(1, 1, title).font = Font(size=16, bold=True, color="FFFFFF")
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor="123B5D")
    sheet.cell(1, 1).alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(fields)))
    sheet.cell(2, 1, f"Company: {company_id} | Generated: {datetime.utcnow():%d-%b-%Y %H:%M UTC}")
    sheet.cell(2, 1).alignment = Alignment(horizontal="center")
    for column, (label, _) in enumerate(fields, start=1):
        cell = sheet.cell(4, column, label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="176B87")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_number, row in enumerate(rows, start=5):
        for column, (_, attr) in enumerate(fields, start=1):
            value = getattr(row, attr, None)
            sheet.cell(row_number, column, _dt(value) if hasattr(value, "isoformat") else value)
    for column, (label, attr) in enumerate(fields, start=1):
        max_len = max([len(str(label))] + [len(str(getattr(row, attr, "") or "")) for row in rows[:500]])
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max_len + 2, 12), 42)
    sheet.auto_filter.ref = f"A4:{get_column_letter(max(1, len(fields)))}{max(4, len(rows) + 4)}"


def document_register_workbook(db: Session, company_id: str, doc_type: str | None = None) -> bytes:
    configs = export_doc_config()
    selected = {doc_type: configs[doc_type]} if doc_type in configs else configs
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    company_profile = get_export_company_profile(db, company_id)
    for key, cfg in selected.items():
        rows = db.query(cfg["model"]).filter(
            cfg["model"].company_id == company_id
        ).order_by(desc(cfg["model"].id)).all()
        sheet = workbook.create_sheet(title=cfg["title"][:31])
        style_register_sheet(
            sheet,
            cfg["title"],
            company_profile["code"],
            cfg["fields"],
            rows,
        )
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def pdf_escape(value) -> str:
    return str(value if value is not None else "").replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def make_simple_pdf(title: str, document_no: str, lines: list[str], company_name: str = "COMPANY") -> bytes:
    """Dependency-free corporate single-page portrait PDF fallback."""
    wrapped_lines = []
    for line in lines:
        if line.startswith("## "):
            wrapped_lines.append(line)
            continue
        chunks = textwrap.wrap(str(line), width=29, break_long_words=False, break_on_hyphens=False) or ["—"]
        wrapped_lines.extend(chunks)
    column_count = 3
    rows_per_column = max(1, (len(wrapped_lines) + column_count - 1) // column_count)
    y_step = min(13.0, max(3.8, 585 / rows_per_column))
    font_size = min(7.2, max(3.8, y_step * 0.62))
    columns = [
        wrapped_lines[index * rows_per_column:(index + 1) * rows_per_column]
        for index in range(column_count)
    ]
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [5 0 R] /Count 1 >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>",
    ]
    commands = [
        "0.071 0.231 0.365 rg",
        "0 785 595 57 re f",
        "1 1 1 rg",
        "BT /F2 16 Tf 34 815 Td",
        f"({pdf_escape(str(company_name).upper())}) Tj",
        "/F1 7 Tf 0 -15 Td (CONTROLLED CORPORATE DOCUMENT) Tj ET",
        "0.071 0.231 0.365 rg",
        "BT /F2 13 Tf 330 815 Td",
        f"({pdf_escape(title.upper())}) Tj",
        "/F1 9 Tf 0 -17 Td",
        f"(Document No: {pdf_escape(document_no)}) Tj ET",
        "0.70 0.78 0.82 RG 34 763 527 1 re f",
    ]
    column_x = [38, 213, 388]
    for column_index, page_lines in enumerate(columns):
        y = 747
        for line in page_lines:
            if line.startswith("## "):
                commands.extend([
                    "0.918 0.961 0.973 rg",
                    f"{column_x[column_index]} {y - 2} 164 {max(9, y_step)} re f",
                    "0.071 0.231 0.365 rg",
                    f"BT /F2 {max(4.2, font_size)} Tf {column_x[column_index] + 4} {y + 1} Td ({pdf_escape(line[3:].upper())}) Tj ET",
                ])
                y -= y_step + 3
            else:
                commands.extend([
                    "0.09 0.13 0.20 rg",
                    f"BT /F1 {font_size} Tf {column_x[column_index] + 4} {y} Td ({pdf_escape(line)}) Tj ET",
                ])
                y -= y_step
    commands.extend([
        "0.20 0.28 0.34 RG",
        "38 83 160 1 re f",
        "213 83 160 1 re f",
        "388 83 164 1 re f",
        "0.09 0.13 0.20 rg",
        "BT /F2 6 Tf 38 69 Td (PREPARED BY) Tj ET",
        "BT /F2 6 Tf 213 69 Td (VERIFIED BY) Tj ET",
        "BT /F2 6 Tf 388 69 Td (AUTHORISED SIGNATORY) Tj ET",
        "0.35 0.43 0.49 rg",
        "BT /F1 5 Tf 38 59 Td (Export Documentation) Tj ET",
        "BT /F1 5 Tf 213 59 Td (Commercial / Logistics / Compliance) Tj ET",
        "BT /F1 5 Tf 388 59 Td (Company Seal) Tj ET",
        "0.70 0.78 0.82 RG 34 30 527 1 re f",
        "0.35 0.43 0.49 rg",
        f"BT /F1 6 Tf 34 18 Td (CONTROLLED DOCUMENT COPY - {pdf_escape(title)} - {pdf_escape(document_no)}) Tj ET",
        "BT /F1 6 Tf 500 18 Td (Page 1 of 1) Tj ET",
    ])
    stream = "\n".join(commands).encode("latin-1", "replace")
    objects.append(
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
        b"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents 6 0 R >>"
    )
    objects.append(b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream")

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{idx} 0 obj\n".encode())
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")
    xref = len(pdf)
    pdf.extend(f"xref\n0 {len(objects)+1}\n0000000000 65535 f \n".encode())
    for off in offsets[1:]:
        pdf.extend(f"{off:010d} 00000 n \n".encode())
    pdf.extend(f"trailer << /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode())
    return bytes(pdf)


def ensure_export_document_schema(db: Session = Depends(get_db)) -> None:
    """Keep export routes compatible while pending migrations are deployed."""
    global _EXPORT_SCHEMA_READY
    if _EXPORT_SCHEMA_READY:
        return
    ProformaInvoice.__table__.create(bind=db.get_bind(), checkfirst=True)
    # Some installations already had the PI table before email/document approval
    # was introduced. Keep that existing data and add only the missing columns.
    db.execute(text(
        "ALTER TABLE proforma_invoices ADD COLUMN IF NOT EXISTS "
        "approval_status VARCHAR NOT NULL DEFAULT 'PENDING'"
    ))
    db.execute(text("ALTER TABLE proforma_invoices ADD COLUMN IF NOT EXISTS approved_by VARCHAR"))
    db.execute(text("ALTER TABLE proforma_invoices ADD COLUMN IF NOT EXISTS approved_at TIMESTAMP"))
    db.execute(text("ALTER TABLE proforma_invoices ADD COLUMN IF NOT EXISTS approval_remarks TEXT"))
    db.execute(text(
        "CREATE INDEX IF NOT EXISTS ix_proforma_invoices_approval_status "
        "ON proforma_invoices (approval_status)"
    ))
    ExportRequiredDocument.__table__.create(bind=db.get_bind(), checkfirst=True)
    ExportDocumentApproval.__table__.create(bind=db.get_bind(), checkfirst=True)
    db.execute(text(
        "ALTER TABLE export_document_files ADD COLUMN IF NOT EXISTS "
        "approval_status VARCHAR NOT NULL DEFAULT 'PENDING'"
    ))
    db.execute(text("ALTER TABLE export_document_files ADD COLUMN IF NOT EXISTS approved_by VARCHAR"))
    db.execute(text("ALTER TABLE export_document_files ADD COLUMN IF NOT EXISTS approved_at TIMESTAMP"))
    db.execute(text("ALTER TABLE export_document_files ADD COLUMN IF NOT EXISTS approval_remarks TEXT"))
    db.execute(text("ALTER TABLE export_document_files ADD COLUMN IF NOT EXISTS document_date DATE"))
    db.execute(text("ALTER TABLE export_document_files ADD COLUMN IF NOT EXISTS expiry_date DATE"))
    db.execute(text("ALTER TABLE export_document_files ADD COLUMN IF NOT EXISTS issuer_name VARCHAR"))
    db.execute(text("ALTER TABLE export_document_files ADD COLUMN IF NOT EXISTS reference_no VARCHAR"))
    db.execute(text("ALTER TABLE export_document_files ADD COLUMN IF NOT EXISTS currency VARCHAR"))
    db.execute(text("ALTER TABLE export_document_files ADD COLUMN IF NOT EXISTS amount NUMERIC(18, 2)"))
    db.execute(text("ALTER TABLE export_document_files ADD COLUMN IF NOT EXISTS details_json TEXT"))
    db.execute(text(
        "CREATE INDEX IF NOT EXISTS ix_export_document_files_approval_status "
        "ON export_document_files (approval_status)"
    ))
    for table_name in (
        "commercial_invoices", "packing_lists", "container_stuffing",
        "shipping_bills", "bill_of_ladings", "health_certificates",
    ):
        db.execute(text(
            f"ALTER TABLE {table_name} ADD COLUMN IF NOT EXISTS "
            "is_cancelled BOOLEAN NOT NULL DEFAULT FALSE"
        ))
    db.execute(text(
        "ALTER TABLE export_compliance_tracker "
        "ADD COLUMN IF NOT EXISTS company_id VARCHAR"
    ))
    db.execute(text("""
        UPDATE export_compliance_tracker ect
           SET company_id = es.company_id
          FROM export_shipments es
         WHERE es.shipment_no = ect.shipment_no
           AND ect.company_id IS NULL
    """))
    db.commit()
    _EXPORT_SCHEMA_READY = True


# Applied to every route registered below, including direct API calls.
router.dependencies.append(Depends(ensure_export_document_schema))


def build_export_dashboard_context(db: Session, comp_code: str) -> dict:
    ensure_bill_accounting_schema(db)
    active_shipments = db.query(ExportShipment).filter(
        ExportShipment.company_id == comp_code,
        ExportShipment.is_cancelled != True,
    )
    packing_filter = (
        PackingList.company_id == comp_code,
        PackingList.is_cancelled != True,
    )
    compliance_rows = db.query(ExportComplianceTracker).filter(
        ExportComplianceTracker.company_id == comp_code,
    ).all()
    packing_documents = db.query(func.count(func.distinct(PackingList.packing_no))).filter(
        *packing_filter,
    ).scalar() or 0
    packing_lines = db.query(PackingList).filter(*packing_filter).count()
    stats = {
        "proforma_invoices": db.query(ProformaInvoice).filter(
            ProformaInvoice.company_id == comp_code,
            ProformaInvoice.is_cancelled != True,
        ).count(),
        "shipments": active_shipments.count(),
        "invoices": db.query(CommercialInvoice).filter(
            CommercialInvoice.company_id == comp_code,
            CommercialInvoice.is_cancelled != True,
        ).count(),
        "packing_lists": int(packing_documents),
        "packing_lines": packing_lines,
        "stuffing": db.query(ContainerStuffing).filter(
            ContainerStuffing.company_id == comp_code,
            ContainerStuffing.is_cancelled != True,
        ).count(),
        "shipping_bills": db.query(ShippingBill).filter(
            ShippingBill.company_id == comp_code,
            ShippingBill.is_cancelled != True,
        ).count(),
        "bill_of_lading": db.query(BillOfLading).filter(
            BillOfLading.company_id == comp_code,
            BillOfLading.is_cancelled != True,
        ).count(),
        "health_certificates": db.query(HealthCertificate).filter(
            HealthCertificate.company_id == comp_code,
            HealthCertificate.is_cancelled != True,
        ).count(),
        "compliance": len(compliance_rows),
        "pending_approvals": db.query(ExportDocumentFile).filter(
            ExportDocumentFile.company_id == comp_code,
            ExportDocumentFile.is_current == True,
            ExportDocumentFile.approval_status == "PENDING",
        ).count(),
    }
    recent_shipments = [
        {
            "id": row.id,
            "shipment_no": row.shipment_no,
            "invoice_no": row.invoice_no,
            "buyer_name": row.buyer_name,
            "country": row.country,
            "status": row.status or "OPEN",
            "etd": _dt(row.etd),
            "eta": _dt(row.eta),
        }
        for row in active_shipments.order_by(desc(ExportShipment.id)).limit(10).all()
    ]
    recent_invoices = [
        {
            "invoice_no": row.invoice_no,
            "shipment_no": row.shipment_no,
            "buyer_name": row.buyer_name,
            "invoice_date": _dt(row.invoice_date),
            "currency": row.currency or "USD",
            "total_amount": float(row.total_amount or 0),
            "payment_status": row.payment_status or "PENDING",
            "approval_status": row.approval_status or "PENDING",
        }
        for row in db.query(CommercialInvoice)
        .filter(
            CommercialInvoice.company_id == comp_code,
            CommercialInvoice.is_cancelled != True,
        )
        .order_by(desc(CommercialInvoice.id))
        .limit(10)
        .all()
    ]
    shipment_status = [
        {"label": status or "OPEN", "value": int(count or 0)}
        for status, count in db.query(ExportShipment.status, func.count(ExportShipment.id))
        .filter(
            ExportShipment.company_id == comp_code,
            ExportShipment.is_cancelled != True,
        )
        .group_by(ExportShipment.status)
        .order_by(desc(func.count(ExportShipment.id)))
        .all()
    ]
    currency_totals = [
        {
            "label": currency or "USD",
            "value": float(total or 0),
            "count": int(count or 0),
        }
        for currency, total, count in db.query(
            CommercialInvoice.currency,
            func.sum(CommercialInvoice.total_amount),
            func.count(CommercialInvoice.id),
        )
        .filter(
            CommercialInvoice.company_id == comp_code,
            CommercialInvoice.is_cancelled != True,
        )
        .group_by(CommercialInvoice.currency)
        .order_by(desc(func.sum(CommercialInvoice.total_amount)))
        .all()
    ]
    compliance_fields = (
        ("Commercial Invoice", "invoice_pending"),
        ("Packing List", "packing_list_pending"),
        ("Health Certificate", "health_cert_pending"),
        ("Shipping Bill", "shipping_bill_pending"),
        ("Bill of Lading", "bl_pending"),
        ("Payment Closure", "payment_pending"),
    )
    document_completion = []
    for label, field_name in compliance_fields:
        pending = sum(1 for row in compliance_rows if bool(getattr(row, field_name, False)))
        document_completion.append({
            "label": label,
            "complete": max(0, len(compliance_rows) - pending),
            "pending": pending,
            "total": len(compliance_rows),
        })
    return {
        "stats": stats,
        "shipment_status": shipment_status,
        "document_completion": document_completion,
        "currency_totals": currency_totals,
        "recent_shipments": recent_shipments,
        "recent_invoices": recent_invoices,
        "company_id": comp_code,
    }


@router.get("/dashboard/data")
def export_documents_dashboard_data(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Authentication required")
    return cache_get_or_set(
        f"bknr:export_documents:{comp_code}:dashboard:v2",
        lambda: build_export_dashboard_context(db, comp_code),
        ttl=45,
    )


@router.get("/dashboard", response_class=HTMLResponse)
def export_documents_dashboard(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return RedirectResponse("/auth/login")

    context = cache_get_or_set(
        f"bknr:export_documents:{comp_code}:dashboard:v2",
        lambda: build_export_dashboard_context(db, comp_code),
        ttl=45,
    )
    return templates.TemplateResponse(
        request=request,
        name="export_documents/dashboard.html",
        context=context,
    )


class RequiredDocumentItemSchema(BaseModel):
    code: str
    label: str | None = None


class RequiredDocumentSelectionSchema(BaseModel):
    po_number: str
    documents: list[RequiredDocumentItemSchema]


class SupportingDocumentApprovalSchema(BaseModel):
    decision: str
    remarks: str | None = None


class RequirementGenerateSchema(BaseModel):
    shipment_id: int
    details: dict = Field(default_factory=dict)
    approver_emails: list[str] = Field(default_factory=list)


def is_supporting_document_admin(request: Request) -> bool:
    role = str(request.session.get("role") or "").strip().lower()
    email = str(request.session.get("email") or "").strip().lower()
    return role in {"admin", "super_admin"} or email == "bknr.solutions@gmail.com"


def export_requirement_definition(db: Session, company_id: str, document_kind: str) -> dict | None:
    code = re.sub(r"[^A-Z0-9_]+", "_", document_kind.strip().upper()).strip("_")[:80]
    canonical = next((item for item in EXPORT_SUPPORT_DOCUMENT_TYPES if item["code"] == code), None)
    if canonical:
        definition = dict(canonical)
    else:
        custom = db.query(ExportRequiredDocument).filter(
            ExportRequiredDocument.company_id == company_id,
            ExportRequiredDocument.document_kind == code,
        ).order_by(desc(ExportRequiredDocument.id)).first()
        if not custom:
            return None
        definition = {"code": code, "label": custom.document_label, "stage": "Custom"}
    definition["fields"] = EXPORT_REQUIREMENT_COMMON_FIELDS + EXPORT_REQUIREMENT_STAGE_FIELDS.get(
        definition["stage"], EXPORT_REQUIREMENT_STAGE_FIELDS["Custom"],
    )
    if code == "PROFORMA_INVOICE":
        definition["fields"] = [
            {
                **field,
                **({"label": "PI Valid Until"} if field["name"] == "expiry_date" else {}),
            }
            for field in definition["fields"]
        ]
    definition["page_id"] = f"export_requirement_{code}"
    definition["page_url"] = f"/page/export_requirement_{code}"
    definition["template_url"] = f"/export_documents/requirement/{code}/entry"
    definition["document_mode"] = export_document_mode(code)
    definition["workspace_url"] = "/page/proforma_invoice" if code == "PROFORMA_INVOICE" else None
    return definition


def requirement_field_values(value) -> list[str]:
    """Normalize single and multi-select form values for validation/rendering."""
    raw_values = value if isinstance(value, list) else [value]
    return list(dict.fromkeys(str(item).strip() for item in raw_values if item is not None and str(item).strip()))


def requirement_display_value(value):
    values = requirement_field_values(value)
    return ", ".join(values) if isinstance(value, list) else (values[0] if values else None)


def export_company_email_options(db: Session, company_code: str, session_email: str | None) -> list[dict]:
    company = db.query(Company).filter(Company.company_code == company_code).first()
    users = []
    if company:
        users = db.query(User).filter(
            User.company_id == company.id,
            or_(User.is_active == True, User.is_active.is_(None)),
        ).order_by(User.name, User.email).all()
    options = []
    seen = set()
    for user in users:
        email = (user.email or "").strip().lower()
        if email and email not in seen:
            seen.add(email)
            options.append({"email": email, "name": user.name or email, "designation": user.designation or ""})
    current = (session_email or "").strip().lower()
    if current and current not in seen:
        options.insert(0, {"email": current, "name": "Current Session User", "designation": ""})
    return options


def export_requirement_lookup_options(db: Session, company_code: str) -> dict[str, list[str]]:
    """Return normalized, company-scoped master values used by requirement forms."""
    def values(model, column) -> list[str]:
        rows = db.query(column).filter(model.company_id == company_code).distinct().all()
        return sorted({str(value).strip() for (value,) in rows if value and str(value).strip()}, key=str.casefold)

    bank_rows = db.query(BankMaster).filter(
        BankMaster.company_id == company_code,
        BankMaster.is_active == True,
    ).order_by(BankMaster.bank_name, BankMaster.account_number).all()
    bank_accounts = [
        f"{row.bank_name} · {row.account_number} · {row.currency_code or 'INR'}"
        for row in bank_rows
    ]
    return {
        "buyers": values(buyers, buyers.buyer_name),
        "buyer_agents": values(buyer_agents, buyer_agents.agent_name),
        "countries": values(countries, countries.country_name),
        "species": values(species, species.species_name),
        "varieties": values(varieties, varieties.variety_name),
        "grades": values(grades, grades.grade_name),
        "brands": values(brands, brands.brand_name),
        "glazes": values(glazes, glazes.glaze_name),
        "freezers": values(freezers, freezers.freezer_name),
        "packing_styles": values(packing_styles, packing_styles.packing_style),
        "shipping_vendors": values(shipping_vendors, shipping_vendors.vendor_name),
        "bank_accounts": bank_accounts,
    }


def serialize_email_approval(row: ExportDocumentApproval, current_email: str) -> dict:
    return {
        "id": row.id,
        "email": row.approver_email,
        "decision": row.decision or "PENDING",
        "remarks": row.remarks,
        "assigned_by": row.assigned_by,
        "assigned_at": _dt(row.assigned_at),
        "decided_at": _dt(row.decided_at),
        "is_current_user": row.approver_email.lower() == current_email.lower(),
    }


def refresh_email_approval_status(file_row: ExportDocumentFile, approvals: list[ExportDocumentApproval]) -> None:
    decisions = [row.decision or "PENDING" for row in approvals]
    if decisions and all(decision == "APPROVED" for decision in decisions):
        file_row.approval_status = "APPROVED"
    elif "REJECTED" in decisions:
        file_row.approval_status = "REJECTED"
    else:
        file_row.approval_status = "PENDING"
    decided = [row for row in approvals if row.decision != "PENDING"]
    file_row.approved_by = ", ".join(row.approver_email for row in decided) or None
    file_row.approved_at = max((row.decided_at for row in decided if row.decided_at), default=None)


def serialize_requirement_file(row: ExportDocumentFile, approvals: list[ExportDocumentApproval], current_email: str) -> dict:
    try:
        details = json.loads(row.details_json or "{}")
    except (TypeError, ValueError):
        details = {}
    approval_rows = [serialize_email_approval(item, current_email) for item in approvals]
    pending = [item["email"] for item in approval_rows if item["decision"] == "PENDING"]
    return {
        "id": row.id,
        "shipment_id": row.record_id,
        "document_kind": row.document_kind,
        "document_no": row.document_no,
        "document_date": _dt(row.document_date),
        "expiry_date": _dt(row.expiry_date),
        "issuer_name": row.issuer_name,
        "reference_no": row.reference_no,
        "currency": row.currency,
        "amount": str(row.amount) if row.amount is not None else None,
        "details": details,
        "file_name": row.file_name,
        "version_no": row.version_no,
        "uploaded_by": row.uploaded_by,
        "uploaded_at": _dt(row.uploaded_at),
        "remarks": row.remarks,
        "approval_status": row.approval_status or "PENDING",
        "approval_remarks": row.approval_remarks,
        "approvals": approval_rows,
        "pending_approvers": pending,
        "approval_progress": f"{len([item for item in approval_rows if item['decision'] == 'APPROVED'])}/{len(approval_rows)}",
        "can_current_user_approve": any(item["is_current_user"] and item["decision"] != "APPROVED" for item in approval_rows),
        "download_url": f"/export_documents/files/{row.id}/download",
    }


@router.get("/requirement-pages/catalog")
def export_requirement_catalog(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    current_email = (request.session.get("email") or "").strip().lower()
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    definitions = [export_requirement_definition(db, comp_code, item["code"]) for item in EXPORT_SUPPORT_DOCUMENT_TYPES]
    custom_codes = db.query(ExportRequiredDocument.document_kind).filter(
        ExportRequiredDocument.company_id == comp_code,
    ).distinct().all()
    canonical_codes = {item["code"] for item in EXPORT_SUPPORT_DOCUMENT_TYPES}
    definitions.extend(
        definition for (code,) in custom_codes
        if code not in canonical_codes
        if (definition := export_requirement_definition(db, comp_code, code))
    )
    pending_counts = dict(
        db.query(ExportDocumentFile.document_kind, func.count(ExportDocumentApproval.id))
        .join(ExportDocumentApproval, ExportDocumentApproval.file_id == ExportDocumentFile.id)
        .filter(
            ExportDocumentFile.company_id == comp_code,
            ExportDocumentFile.is_current == True,
            ExportDocumentApproval.company_id == comp_code,
            func.lower(ExportDocumentApproval.approver_email) == current_email,
            ExportDocumentApproval.decision == "PENDING",
        )
        .group_by(ExportDocumentFile.document_kind)
        .all()
    ) if current_email else {}
    for definition in definitions:
        definition["pending_for_me"] = int(pending_counts.get(definition["code"], 0))
    return {"success": True, "document_types": definitions}


@router.get("/requirement-pages/entry", response_class=HTMLResponse)
def export_requirement_catalog_entry(request: Request, db: Session = Depends(get_db)):
    context = export_requirement_catalog(request, db)
    if isinstance(context, JSONResponse):
        return RedirectResponse("/auth/login")
    return templates.TemplateResponse(
        request=request,
        name="export_documents/requirement_forms.html",
        context={"request": request, **context},
    )


@router.get("/requirement/{document_kind}/data")
def export_requirement_page_data(document_kind: str, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    current_email = (request.session.get("email") or "").strip().lower()
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    definition = export_requirement_definition(db, comp_code, document_kind)
    if not definition:
        return JSONResponse({"success": False, "message": "Document type not found"}, status_code=404)
    is_pre_po = definition["code"] == "PROFORMA_INVOICE"
    if is_pre_po:
        references = db.query(ProformaInvoice).filter(
            ProformaInvoice.company_id == comp_code,
            ProformaInvoice.is_cancelled != True,
        ).order_by(desc(ProformaInvoice.pi_date), desc(ProformaInvoice.id)).all()
        module_name = "export_proforma_requirement"
    else:
        references = db.query(ExportShipment).filter(
            ExportShipment.company_id == comp_code,
            ExportShipment.is_cancelled != True,
        ).order_by(desc(ExportShipment.id)).all()
        module_name = "export_supporting"
    reference_by_id = {row.id: row for row in references}
    all_files = db.query(ExportDocumentFile).filter(
        ExportDocumentFile.company_id == comp_code,
        ExportDocumentFile.module_name == module_name,
        ExportDocumentFile.document_kind == definition["code"],
    ).order_by(desc(ExportDocumentFile.version_no), desc(ExportDocumentFile.id)).all()
    files = [row for row in all_files if row.is_current]
    versions_by_reference = {}
    for version in all_files:
        try:
            version_details = json.loads(version.details_json or "{}")
        except (TypeError, ValueError):
            version_details = {}
        versions_by_reference.setdefault(version.record_id, []).append({
            "id": version.id,
            "version_no": version.version_no,
            "file_name": version.file_name,
            "file_origin": version_details.get("_file_origin") or ("GENERATED" if "generated" in (version.remarks or "").lower() else "IMPORTED"),
            "uploaded_by": version.uploaded_by,
            "uploaded_at": _dt(version.uploaded_at),
            "is_current": bool(version.is_current),
            "download_url": f"/export_documents/files/{version.id}/download",
        })
    file_ids = [row.id for row in files]
    approval_rows = db.query(ExportDocumentApproval).filter(
        ExportDocumentApproval.company_id == comp_code,
        ExportDocumentApproval.file_id.in_(file_ids),
    ).order_by(ExportDocumentApproval.id).all() if file_ids else []
    approval_map = {}
    for approval in approval_rows:
        approval_map.setdefault(approval.file_id, []).append(approval)
    entries = []
    for file_row in files:
        payload = serialize_requirement_file(file_row, approval_map.get(file_row.id, []), current_email)
        payload["versions"] = versions_by_reference.get(file_row.record_id, [])
        payload["file_origin"] = payload["details"].get("_file_origin") or "IMPORTED"
        reference = reference_by_id.get(file_row.record_id)
        if is_pre_po:
            payload.update({
                "po_number": reference.po_number if reference else None,
                "shipment_no": reference.pi_no if reference else None,
                "buyer_name": reference.buyer_name if reference else None,
                "country": reference.country if reference else None,
            })
        else:
            payload.update({
                "po_number": reference.po_number if reference else None,
                "shipment_no": reference.shipment_no if reference else None,
                "buyer_name": reference.buyer_name if reference else None,
                "country": reference.country if reference else None,
            })
        entries.append(payload)
    required_pos = {po for (po,) in db.query(ExportRequiredDocument.po_number).filter(
        ExportRequiredDocument.company_id == comp_code,
        ExportRequiredDocument.document_kind == definition["code"],
    ).all()}
    return {
        "success": True,
        "definition": {
            **definition,
            "pre_po_allowed": is_pre_po,
            "reference_label": "Proforma Invoice Number / Buyer" if is_pre_po else "PO Number / Shipment",
            "create_reference_url": "/page/proforma_invoice" if is_pre_po else None,
        },
        "current_email": current_email,
        "email_options": export_company_email_options(db, comp_code, current_email),
        "lookup_options": export_requirement_lookup_options(db, comp_code),
        "po_options": [{
            "shipment_id": row.id,
            "shipment_no": row.pi_no if is_pre_po else row.shipment_no,
            "po_number": (row.po_number or "PRE-PO") if is_pre_po else row.po_number,
            "buyer_name": row.buyer_name,
            "country": row.country,
            "document_date": row.pi_date.isoformat() if is_pre_po and row.pi_date else None,
            "validity_date": row.validity_date.isoformat() if is_pre_po and row.validity_date else None,
            "is_required": False if is_pre_po else row.po_number in required_pos,
        } for row in references],
        "entries": entries,
    }


@router.get("/requirement/{document_kind}/entry", response_class=HTMLResponse)
def export_requirement_page_entry(document_kind: str, request: Request, db: Session = Depends(get_db)):
    context = export_requirement_page_data(document_kind, request, db)
    if isinstance(context, JSONResponse):
        return context
    return templates.TemplateResponse(
        request=request,
        name="export_documents/requirement_document.html",
        context={"request": request, **context},
    )


@router.post("/requirement/{document_kind}/generate")
def export_requirement_page_generate(
    document_kind: str,
    payload: RequirementGenerateSchema,
    request: Request,
    db: Session = Depends(get_db),
):
    comp_code = request.session.get("company_code")
    email = (request.session.get("email") or "").strip().lower()
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    definition = export_requirement_definition(db, comp_code, document_kind)
    if not definition:
        return JSONResponse({"success": False, "message": "Document type not found"}, status_code=404)
    if definition["document_mode"] not in {"GENERATE", "GENERATE_AND_IMPORT_FINAL"}:
        return JSONResponse({"success": False, "message": "This document must be imported as an original PDF"}, status_code=400)

    is_pre_po = definition["code"] == "PROFORMA_INVOICE"
    model = ProformaInvoice if is_pre_po else ExportShipment
    reference = db.query(model).filter(
        model.id == payload.shipment_id,
        model.company_id == comp_code,
        model.is_cancelled != True,
    ).first()
    if not reference:
        return JSONResponse({"success": False, "message": "Select a valid PI reference" if is_pre_po else "Select a valid PO / shipment"}, status_code=404)

    details = payload.details if isinstance(payload.details, dict) else {}
    selected_emails = list(dict.fromkeys(str(value).strip().lower() for value in payload.approver_emails if str(value).strip()))
    missing_fields = [
        field["label"] for field in definition["fields"]
        if field.get("required") and not requirement_field_values(details.get(field["name"]))
    ]
    if missing_fields:
        return JSONResponse({"success": False, "message": f"Required fields missing: {', '.join(missing_fields)}"}, status_code=400)
    lookup_options = export_requirement_lookup_options(db, comp_code)
    invalid_lookups = [
        field["label"] for field in definition["fields"]
        if field.get("lookup") and any(
            value not in set(lookup_options.get(field["lookup"], []))
            for value in requirement_field_values(details.get(field["name"]))
        )
    ]
    if invalid_lookups:
        return JSONResponse({"success": False, "message": f"Select valid master values for: {', '.join(invalid_lookups)}"}, status_code=400)
    valid_emails = {item["email"] for item in export_company_email_options(db, comp_code, email)}
    if not selected_emails or any(value not in valid_emails for value in selected_emails):
        return JSONResponse({"success": False, "message": "Select at least one valid approval email"}, status_code=400)

    try:
        parsed_document_date = date.fromisoformat(details["document_date"]) if details.get("document_date") else None
        parsed_expiry_date = date.fromisoformat(details["expiry_date"]) if details.get("expiry_date") else None
        parsed_amount = Decimal(str(details["amount"])) if details.get("amount") not in (None, "") else None
    except (ValueError, ArithmeticError):
        return JSONResponse({"success": False, "message": "Enter valid document dates and amount"}, status_code=400)
    if parsed_document_date and parsed_expiry_date and parsed_expiry_date < parsed_document_date:
        return JSONResponse({"success": False, "message": "Expiry date cannot be before document date"}, status_code=400)
    if parsed_amount is not None and (not parsed_amount.is_finite() or parsed_amount < 0):
        return JSONResponse({"success": False, "message": "Amount must be a valid non-negative number"}, status_code=400)

    document_no = str(details.get("document_no") or getattr(reference, "pi_no", None) or getattr(reference, "shipment_no", reference.id)).strip()[:160]
    pdf_bytes = render_requirement_pdf(
        definition,
        details,
        comp_code,
        reference,
        get_export_company_profile(db, comp_code),
    )
    module_name = "export_proforma_requirement" if is_pre_po else "export_supporting"
    file_row = store_export_pdf(
        db=db, company_id=comp_code, module_name=module_name, record_id=reference.id,
        document_no=document_no, document_kind=definition["code"],
        file_name=f"{safe_filename(document_no)}.pdf", content=pdf_bytes,
        uploaded_by=email, remarks="System generated controlled PDF",
    )
    file_row.document_date = parsed_document_date
    file_row.expiry_date = parsed_expiry_date
    file_row.amount = parsed_amount
    file_row.issuer_name = str(details.get("issuer_name") or "")[:255] or None
    file_row.reference_no = str(details.get("reference_no") or "")[:255] or None
    file_row.currency = str(details.get("currency") or "")[:10] or None
    allowed_fields = {field["name"] for field in definition["fields"]}
    file_row.details_json = json.dumps({
        **{key: value for key, value in details.items() if key in allowed_fields},
        "_file_origin": "GENERATED",
    }, ensure_ascii=False)
    file_row.approval_status = "PENDING"
    db.flush()
    for approver_email in selected_emails:
        db.add(ExportDocumentApproval(
            company_id=comp_code, file_id=file_row.id, approver_email=approver_email,
            decision="PENDING", assigned_by=email,
        ))
    write_audit(
        db, "export_supporting", file_row.id, comp_code, "PDF_GENERATE_SAVE", "NONE",
        f"TYPE={definition['code']} | REF={document_no} | VERSION={file_row.version_no}", email,
    )
    db.commit()
    invalidate_export_cache(comp_code)
    return {
        "success": True,
        "message": f"{definition['label']} PDF generated and saved as version {file_row.version_no}",
        "file_id": file_row.id,
        "download_url": f"/export_documents/files/{file_row.id}/download",
    }


@router.post("/requirement/{document_kind}/upload")
async def export_requirement_page_upload(
    document_kind: str,
    request: Request,
    shipment_id: int = Form(...),
    details_json: str = Form("{}"),
    approver_emails: str = Form("[]"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    comp_code = request.session.get("company_code")
    email = (request.session.get("email") or "").strip().lower()
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    definition = export_requirement_definition(db, comp_code, document_kind)
    if not definition:
        return JSONResponse({"success": False, "message": "Document type not found"}, status_code=404)
    is_pre_po = definition["code"] == "PROFORMA_INVOICE"
    if is_pre_po:
        reference = db.query(ProformaInvoice).filter(
            ProformaInvoice.id == shipment_id,
            ProformaInvoice.company_id == comp_code,
            ProformaInvoice.is_cancelled != True,
        ).first()
        module_name = "export_proforma_requirement"
        reference_no = reference.pi_no if reference else None
        po_number = reference.po_number if reference else None
    else:
        reference = db.query(ExportShipment).filter(
            ExportShipment.id == shipment_id,
            ExportShipment.company_id == comp_code,
            ExportShipment.is_cancelled != True,
        ).first()
        module_name = "export_supporting"
        reference_no = reference.shipment_no if reference else None
        po_number = reference.po_number if reference else None
    if not reference:
        return JSONResponse({"success": False, "message": "Select a valid PI reference" if is_pre_po else "Select a valid PO / shipment"}, status_code=404)
    try:
        details = json.loads(details_json or "{}")
        raw_emails = json.loads(approver_emails or "[]")
    except (TypeError, ValueError, json.JSONDecodeError):
        return JSONResponse({"success": False, "message": "Invalid form details or approver list"}, status_code=400)
    if not isinstance(details, dict) or not isinstance(raw_emails, list):
        return JSONResponse({"success": False, "message": "Invalid form details or approver list"}, status_code=400)
    selected_emails = list(dict.fromkeys(
        str(value).strip().lower() for value in raw_emails if str(value).strip()
    ))
    missing_fields = [
        field["label"] for field in definition["fields"]
        if field.get("required") and not requirement_field_values(details.get(field["name"]))
    ]
    if missing_fields:
        return JSONResponse({"success": False, "message": f"Required fields missing: {', '.join(missing_fields)}"}, status_code=400)
    lookup_options = export_requirement_lookup_options(db, comp_code)
    invalid_lookups = []
    for field in definition["fields"]:
        lookup_key = field.get("lookup")
        selected_values = requirement_field_values(details.get(field["name"]))
        if lookup_key and any(value not in set(lookup_options.get(lookup_key, [])) for value in selected_values):
            invalid_lookups.append(field["label"])
    if invalid_lookups:
        return JSONResponse({
            "success": False,
            "message": f"Select valid master values for: {', '.join(invalid_lookups)}",
        }, status_code=400)
    valid_emails = {item["email"] for item in export_company_email_options(db, comp_code, email)}
    if not selected_emails:
        return JSONResponse({"success": False, "message": "Select at least one approval email"}, status_code=400)
    if any(value not in valid_emails for value in selected_emails):
        return JSONResponse({"success": False, "message": "One or more approver emails are invalid"}, status_code=400)
    if file.content_type != "application/pdf" and not (file.filename or "").lower().endswith(".pdf"):
        return JSONResponse({"success": False, "message": "Only PDF files are allowed"}, status_code=400)
    content = await file.read()
    if not content or len(content) > 25 * 1024 * 1024 or not content.startswith(b"%PDF-"):
        return JSONResponse({"success": False, "message": "Invalid PDF or file exceeds 25 MB"}, status_code=400)
    parsed_dates = {}
    for key in ("document_date", "expiry_date"):
        value = details.get(key)
        if value:
            try:
                parsed_dates[key] = date.fromisoformat(value)
            except ValueError:
                return JSONResponse({"success": False, "message": f"Invalid {key.replace('_', ' ')}"}, status_code=400)
    if parsed_dates.get("document_date") and parsed_dates.get("expiry_date") and parsed_dates["expiry_date"] < parsed_dates["document_date"]:
        return JSONResponse({"success": False, "message": "Expiry date cannot be before document date"}, status_code=400)
    amount = details.get("amount")
    try:
        parsed_amount = Decimal(str(amount)) if amount not in (None, "") else None
    except Exception:
        return JSONResponse({"success": False, "message": "Amount must be numeric"}, status_code=400)
    if parsed_amount is not None and (not parsed_amount.is_finite() or parsed_amount < 0):
        return JSONResponse({"success": False, "message": "Amount must be a valid non-negative number"}, status_code=400)
    document_no = str(details.get("document_no") or reference_no).strip()[:160]
    file_row = store_export_pdf(
        db=db, company_id=comp_code, module_name=module_name, record_id=reference.id,
        document_no=document_no, document_kind=definition["code"],
        file_name=file.filename or f"{document_no}.pdf", content=content,
        uploaded_by=email, remarks=str(details.get("status_note") or "")[:1000] or None,
    )
    file_row.document_date = parsed_dates.get("document_date")
    file_row.expiry_date = parsed_dates.get("expiry_date")
    file_row.issuer_name = str(details.get("issuer_name") or "")[:255] or None
    file_row.reference_no = str(details.get("reference_no") or "")[:255] or None
    file_row.currency = str(details.get("currency") or "")[:10] or None
    file_row.amount = parsed_amount
    allowed_fields = {field["name"] for field in definition["fields"]}
    cleaned_details = {
        **{key: value for key, value in details.items() if key in allowed_fields},
        "_file_origin": "IMPORTED_FINAL" if definition["document_mode"] in {"GENERATE_AND_IMPORT_FINAL", "IMPORT_FINAL_PDF"} else "IMPORTED_ORIGINAL",
    }
    file_row.details_json = json.dumps(cleaned_details, ensure_ascii=False)
    file_row.approval_status = "PENDING"
    db.flush()
    for approver_email in selected_emails:
        db.add(ExportDocumentApproval(
            company_id=comp_code, file_id=file_row.id, approver_email=approver_email,
            decision="PENDING", assigned_by=email,
        ))
    write_audit(
        db, "export_supporting", file_row.id, comp_code, "DETAILS_FILE_UPLOAD", "NONE",
        f"PO={po_number or 'PRE-PO'} | REF={reference_no} | TYPE={definition['code']} | APPROVERS={', '.join(selected_emails)}", email,
    )
    db.commit()
    invalidate_export_cache(comp_code)
    return {"success": True, "message": f"{definition['label']} uploaded and sent to {len(selected_emails)} approver(s)"}


@router.post("/requirement/{document_kind}/files/{file_id}/approval")
def export_requirement_email_approval(
    document_kind: str,
    file_id: int,
    payload: SupportingDocumentApprovalSchema,
    request: Request,
    db: Session = Depends(get_db),
):
    comp_code = request.session.get("company_code")
    email = (request.session.get("email") or "").strip().lower()
    if not comp_code or not email:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    definition = export_requirement_definition(db, comp_code, document_kind)
    file_row = db.query(ExportDocumentFile).filter(
        ExportDocumentFile.id == file_id,
        ExportDocumentFile.company_id == comp_code,
        ExportDocumentFile.document_kind == (definition or {}).get("code"),
        ExportDocumentFile.is_current == True,
    ).first()
    if not file_row:
        return JSONResponse({"success": False, "message": "Current document not found"}, status_code=404)
    assignment = db.query(ExportDocumentApproval).filter(
        ExportDocumentApproval.file_id == file_id,
        ExportDocumentApproval.company_id == comp_code,
        func.lower(ExportDocumentApproval.approver_email) == email,
    ).first()
    if not assignment:
        return JSONResponse({"success": False, "message": "This document was not assigned to your email"}, status_code=403)
    decision = payload.decision.strip().upper()
    remarks = (payload.remarks or "").strip()[:500]
    if decision not in {"APPROVED", "REJECTED"}:
        return JSONResponse({"success": False, "message": "Decision must be APPROVED or REJECTED"}, status_code=400)
    if decision == "REJECTED" and not remarks:
        return JSONResponse({"success": False, "message": "Rejection remarks are required"}, status_code=400)
    old_decision = assignment.decision or "PENDING"
    assignment.decision = decision
    assignment.remarks = remarks or None
    assignment.decided_at = datetime.utcnow()
    approvals = db.query(ExportDocumentApproval).filter(
        ExportDocumentApproval.file_id == file_id,
        ExportDocumentApproval.company_id == comp_code,
    ).all()
    refresh_email_approval_status(file_row, approvals)
    file_row.approval_remarks = "; ".join(
        f"{row.approver_email}: {row.decision}" for row in approvals
    )
    write_audit(
        db, "export_supporting", file_id, comp_code, "EMAIL_APPROVAL",
        f"{email}: {old_decision}", f"{email}: {decision}{f' | {remarks}' if remarks else ''}", email,
    )
    db.commit()
    invalidate_export_cache(comp_code)
    return {"success": True, "message": f"Your decision was saved. Overall status: {file_row.approval_status}"}


@router.get("/supporting_documents/data")
def export_supporting_documents_data(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    current_email = (request.session.get("email") or "").strip().lower()
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    shipments = (
        db.query(ExportShipment)
        .filter(ExportShipment.company_id == comp_code, ExportShipment.is_cancelled != True)
        .order_by(desc(ExportShipment.id))
        .all()
    )
    shipment_by_id = {row.id: row for row in shipments}
    po_options_by_number = {}

    def add_po_option(po_number, buyer_name=None, country=None, source=None):
        value = str(po_number or "").strip()
        if not value or value.upper() in {"N/A", "NA", "-", "NONE", "NULL", "PRE-PO"}:
            return
        key = value.upper()
        existing = po_options_by_number.get(key, {})
        po_options_by_number[key] = {
            "po_number": existing.get("po_number") or value,
            "shipment_id": existing.get("shipment_id"),
            "shipment_no": existing.get("shipment_no"),
            "buyer_name": existing.get("buyer_name") or str(buyer_name or "").strip(),
            "country": existing.get("country") or str(country or "").strip(),
            "source": existing.get("source") or source,
        }

    pending_po_rows = db.query(
        pending_orders.po_number, pending_orders.buyer, pending_orders.country,
    ).filter(
        pending_orders.company_id == comp_code,
        pending_orders.po_number.isnot(None),
        pending_orders.po_number != "",
    ).all()
    for po_number, buyer_name, country in pending_po_rows:
        add_po_option(po_number, buyer_name, country, "PENDING_ORDERS")

    sales_po_rows = db.query(
        sales_dispatch.po_number, sales_dispatch.buyer_name, sales_dispatch.country,
    ).filter(
        sales_dispatch.company_id == comp_code,
        sales_dispatch.po_number.isnot(None),
        sales_dispatch.po_number != "",
    ).all()
    for po_number, buyer_name, country in sales_po_rows:
        add_po_option(po_number, buyer_name, country, "SALES")

    for row in shipments:
        po_options_by_number[str(row.po_number or "").strip().upper()] = {
            "po_number": row.po_number,
            "shipment_id": row.id,
            "shipment_no": row.shipment_no,
            "buyer_name": row.buyer_name,
            "country": row.country,
            "source": "EXPORT_SHIPMENT",
        }

    requirements = (
        db.query(ExportRequiredDocument)
        .filter(ExportRequiredDocument.company_id == comp_code)
        .order_by(ExportRequiredDocument.po_number, ExportRequiredDocument.document_label)
        .all()
    )
    file_rows = (
        db.query(
            ExportDocumentFile.id,
            ExportDocumentFile.record_id,
            ExportDocumentFile.document_kind,
            ExportDocumentFile.document_no,
            ExportDocumentFile.file_name,
            ExportDocumentFile.version_no,
            ExportDocumentFile.is_current,
            ExportDocumentFile.uploaded_at,
            ExportDocumentFile.remarks,
            ExportDocumentFile.approval_status,
            ExportDocumentFile.approved_by,
            ExportDocumentFile.approved_at,
            ExportDocumentFile.approval_remarks,
        )
        .filter(
            ExportDocumentFile.company_id == comp_code,
            ExportDocumentFile.module_name == "export_supporting",
        )
        .order_by(desc(ExportDocumentFile.uploaded_at), desc(ExportDocumentFile.id))
        .all()
    )
    supporting_file_ids = [row.id for row in file_rows]
    email_approval_rows = db.query(ExportDocumentApproval).filter(
        ExportDocumentApproval.company_id == comp_code,
        ExportDocumentApproval.file_id.in_(supporting_file_ids),
    ).order_by(ExportDocumentApproval.id).all() if supporting_file_ids else []
    email_approvals_by_file = {}
    for approval in email_approval_rows:
        email_approvals_by_file.setdefault(approval.file_id, []).append(approval)

    def approval_summary(file_row):
        approvals = email_approvals_by_file.get(file_row.id, []) if file_row else []
        serialized = [serialize_email_approval(item, current_email) for item in approvals]
        return {
            "approvals": serialized,
            "pending_approvers": [item["email"] for item in serialized if item["decision"] == "PENDING"],
            "approval_progress": f"{len([item for item in serialized if item['decision'] == 'APPROVED'])}/{len(serialized)}" if serialized else None,
            "can_current_user_approve": any(item["is_current_user"] and item["decision"] != "APPROVED" for item in serialized),
        }

    canonical_labels = {item["code"]: item["label"] for item in EXPORT_SUPPORT_DOCUMENT_TYPES}
    latest_files = {}
    for file_row in file_rows:
        if not file_row.is_current:
            continue
        shipment = shipment_by_id.get(file_row.record_id)
        if not shipment:
            continue
        latest_files.setdefault((shipment.po_number, file_row.document_kind), (file_row, shipment))

    checklist = []
    required_keys = set()
    for requirement in requirements:
        key = (requirement.po_number, requirement.document_kind)
        required_keys.add(key)
        matched = latest_files.get(key)
        file_row, shipment = matched if matched else (None, None)
        po_meta = po_options_by_number.get(str(requirement.po_number or "").strip().upper(), {})
        checklist.append({
            "requirement_id": requirement.id,
            "po_number": requirement.po_number,
            "shipment_id": shipment.id if shipment else po_meta.get("shipment_id"),
            "shipment_no": shipment.shipment_no if shipment else po_meta.get("shipment_no"),
            "buyer_name": shipment.buyer_name if shipment else po_meta.get("buyer_name"),
            "document_kind": requirement.document_kind,
            "document_label": requirement.document_label,
            "required": True,
            "status": "UPLOADED" if file_row else "PENDING",
            "file_id": file_row.id if file_row else None,
            "file_name": file_row.file_name if file_row else None,
            "document_no": file_row.document_no if file_row else None,
            "version_no": file_row.version_no if file_row else None,
            "uploaded_at": _dt(file_row.uploaded_at) if file_row else None,
            "remarks": file_row.remarks if file_row else None,
            "approval_status": (file_row.approval_status or "PENDING") if file_row else None,
            "approved_by": file_row.approved_by if file_row else None,
            "approved_at": _dt(file_row.approved_at) if file_row else None,
            "approval_remarks": file_row.approval_remarks if file_row else None,
            "download_url": f"/export_documents/files/{file_row.id}/download" if file_row else None,
            "page_url": f"/page/export_requirement_{requirement.document_kind}",
            **approval_summary(file_row),
        })

    for key, (file_row, shipment) in latest_files.items():
        if key in required_keys:
            continue
        checklist.append({
            "requirement_id": None,
            "po_number": shipment.po_number,
            "shipment_id": shipment.id,
            "shipment_no": shipment.shipment_no,
            "buyer_name": shipment.buyer_name,
            "document_kind": file_row.document_kind,
            "document_label": canonical_labels.get(file_row.document_kind, file_row.document_kind.replace("_", " ").title()),
            "required": False,
            "status": "UPLOADED",
            "file_id": file_row.id,
            "file_name": file_row.file_name,
            "document_no": file_row.document_no,
            "version_no": file_row.version_no,
            "uploaded_at": _dt(file_row.uploaded_at),
            "remarks": file_row.remarks,
            "approval_status": file_row.approval_status or "PENDING",
            "approved_by": file_row.approved_by,
            "approved_at": _dt(file_row.approved_at),
            "approval_remarks": file_row.approval_remarks,
            "download_url": f"/export_documents/files/{file_row.id}/download",
            "page_url": f"/page/export_requirement_{file_row.document_kind}",
            **approval_summary(file_row),
        })

    checklist.sort(key=lambda row: (row["po_number"] or "", row["status"] != "PENDING", row["document_label"] or ""))
    requirements_by_po = {}
    for row in requirements:
        requirements_by_po.setdefault(row.po_number, []).append({
            "code": row.document_kind,
            "label": row.document_label,
        })

    po_groups = []
    checklist_by_po = {}
    for row in checklist:
        checklist_by_po.setdefault(row["po_number"], []).append(row)
    for po in sorted(po_options_by_number.values(), key=lambda item: item["po_number"].upper()):
        rows = checklist_by_po.get(po["po_number"], [])
        required_rows = [row for row in rows if row["required"]]
        uploaded_rows = [row for row in required_rows if row["status"] == "UPLOADED"]
        approved_rows = [row for row in uploaded_rows if row["approval_status"] == "APPROVED"]
        rejected_rows = [row for row in uploaded_rows if row["approval_status"] == "REJECTED"]
        pending_upload = len(required_rows) - len(uploaded_rows)
        pending_approval = len(uploaded_rows) - len(approved_rows) - len(rejected_rows)
        if not required_rows:
            po_status = "NOT_CONFIGURED"
        elif len(approved_rows) == len(required_rows):
            po_status = "COMPLETE"
        elif rejected_rows:
            po_status = "REJECTED"
        else:
            po_status = "PENDING"
        po_groups.append({
            **po,
            "rows": rows,
            "required_count": len(required_rows),
            "uploaded_count": len(uploaded_rows),
            "approved_count": len(approved_rows),
            "pending_upload_count": pending_upload,
            "pending_approval_count": pending_approval,
            "rejected_count": len(rejected_rows),
            "extra_count": len([row for row in rows if not row["required"]]),
            "status": po_status,
        })

    audit_rows = (
        db.query(AuditLog)
        .filter(AuditLog.company_id == comp_code, AuditLog.table_name == "export_supporting")
        .order_by(desc(AuditLog.edited_at), desc(AuditLog.id))
        .limit(100)
        .all()
    )
    audit_logs = [{
        "id": row.id,
        "record_id": row.record_id,
        "action": row.field_name,
        "old_value": row.old_value,
        "new_value": row.new_value,
        "edited_by": row.edited_by,
        "edited_at": _dt(row.edited_at),
    } for row in audit_rows]
    document_groups = []
    for document_type in EXPORT_SUPPORT_DOCUMENT_TYPES:
        if not document_groups or document_groups[-1]["stage"] != document_type["stage"]:
            document_groups.append({"stage": document_type["stage"], "items": []})
        document_groups[-1]["items"].append(document_type)

    completed_po_numbers = {
        group["po_number"].strip().upper()
        for group in po_groups
        if group["required_count"] > 0 and group["status"] == "COMPLETE"
    }
    selectable_po_options = [
        po for po in sorted(po_options_by_number.values(), key=lambda item: item["po_number"].upper())
        if po["po_number"].strip().upper() not in completed_po_numbers
    ]

    return {
        "success": True,
        "company_id": comp_code,
        "po_options": selectable_po_options,
        "document_types": EXPORT_SUPPORT_DOCUMENT_TYPES,
        "document_groups": document_groups,
        "requirements_by_po": requirements_by_po,
        "checklist": checklist,
        "po_groups": po_groups,
        "audit_logs": audit_logs,
        "is_admin": is_supporting_document_admin(request),
    }


@router.post("/supporting_documents/requirements")
def save_export_required_documents(
    payload: RequiredDocumentSelectionSchema,
    request: Request,
    db: Session = Depends(get_db),
):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    po_number = payload.po_number.strip()
    shipment = db.query(ExportShipment).filter(
        ExportShipment.company_id == comp_code,
        func.upper(func.trim(ExportShipment.po_number)) == po_number.upper(),
        ExportShipment.is_cancelled != True,
    ).first()
    if not shipment:
        pending_source = db.query(pending_orders).filter(
            pending_orders.company_id == comp_code,
            func.upper(func.trim(pending_orders.po_number)) == po_number.upper(),
        ).first()
        sales_source = db.query(sales_dispatch).filter(
            sales_dispatch.company_id == comp_code,
            func.upper(func.trim(sales_dispatch.po_number)) == po_number.upper(),
        ).first()
        source = pending_source or sales_source
        if not source:
            return JSONResponse({
                "success": False,
                "message": "Select a valid PO number from Pending Orders or Sales",
            }, status_code=404)
        buyer_name = getattr(source, "buyer", None) or getattr(source, "buyer_name", None) or "Buyer Not Available"
        country = getattr(source, "country", None) or "Country Not Available"
        shipment_base = re.sub(r"[^A-Z0-9]+", "-", po_number.upper()).strip("-")[:50] or "PO"
        shipment_no = f"AUTO-{shipment_base}"
        suffix = 1
        while db.query(ExportShipment.id).filter(
            ExportShipment.company_id == comp_code,
            ExportShipment.shipment_no == shipment_no,
        ).first():
            suffix += 1
            shipment_no = f"AUTO-{shipment_base}-{suffix}"
        shipment = ExportShipment(
            company_id=comp_code,
            shipment_no=shipment_no,
            po_number=po_number,
            buyer_name=str(buyer_name).strip(),
            country=str(country).strip(),
            status="OPEN",
            approval_status="PENDING",
            created_by=email or "SYSTEM",
        )
        db.add(shipment)
        db.flush()

    canonical_labels = {item["code"]: item["label"] for item in EXPORT_SUPPORT_DOCUMENT_TYPES}
    selected = {}
    for document in payload.documents:
        code = re.sub(r"[^A-Z0-9_]+", "_", document.code.strip().upper()).strip("_")[:80]
        if not code:
            continue
        label = (document.label or canonical_labels.get(code) or code.replace("_", " ").title()).strip()[:160]
        selected[code] = label

    existing_rows = db.query(ExportRequiredDocument).filter(
        ExportRequiredDocument.company_id == comp_code,
        ExportRequiredDocument.po_number == po_number,
    ).all()
    existing_by_code = {row.document_kind: row for row in existing_rows}
    old_codes = sorted(existing_by_code)

    for code, row in existing_by_code.items():
        if code not in selected:
            db.delete(row)
    for code, label in selected.items():
        row = existing_by_code.get(code)
        if row:
            row.document_label = label
            row.updated_at = datetime.utcnow()
        else:
            db.add(ExportRequiredDocument(
                company_id=comp_code,
                po_number=po_number,
                document_kind=code,
                document_label=label,
                created_by=email,
            ))

    write_audit(
        db, "export_supporting", shipment.id, comp_code,
        "REQUIREMENTS_UPDATE", ", ".join(old_codes) or "NONE",
        ", ".join(sorted(selected)) or "NONE", email,
    )
    db.commit()
    invalidate_export_cache(comp_code)
    return {
        "success": True,
        "message": f"Required document list saved for PO {po_number}",
        "count": len(selected),
    }


@router.post("/supporting_documents/files/{file_id}/approval")
def decide_supporting_document_approval(
    file_id: int,
    payload: SupportingDocumentApprovalSchema,
    request: Request,
    db: Session = Depends(get_db),
):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    decision = payload.decision.strip().upper()
    if decision not in {"APPROVED", "REJECTED"}:
        return JSONResponse({"success": False, "message": "Decision must be APPROVED or REJECTED"}, status_code=400)
    remarks = (payload.remarks or "").strip()[:500]
    if decision == "REJECTED" and not remarks:
        return JSONResponse({"success": False, "message": "Rejection remarks are required"}, status_code=400)

    file_row = db.query(ExportDocumentFile).filter(
        ExportDocumentFile.id == file_id,
        ExportDocumentFile.company_id == comp_code,
        ExportDocumentFile.module_name == "export_supporting",
        ExportDocumentFile.is_current == True,
    ).first()
    if not file_row:
        return JSONResponse({"success": False, "message": "Current supporting document not found"}, status_code=404)

    assignments = db.query(ExportDocumentApproval).filter(
        ExportDocumentApproval.company_id == comp_code,
        ExportDocumentApproval.file_id == file_id,
    ).all()
    if assignments:
        assignment = next((row for row in assignments if row.approver_email.lower() == (email or "").lower()), None)
        if not assignment:
            return JSONResponse({"success": False, "message": "This document was not assigned to your email"}, status_code=403)
        old_status = assignment.decision or "PENDING"
        assignment.decision = decision
        assignment.remarks = remarks or None
        assignment.decided_at = datetime.utcnow()
        refresh_email_approval_status(file_row, assignments)
        file_row.approval_remarks = "; ".join(f"{row.approver_email}: {row.decision}" for row in assignments)
        write_audit(
            db, "export_supporting", file_row.id, comp_code, "EMAIL_APPROVAL",
            f"{email}: {old_status}", f"{email}: {decision}{f' | {remarks}' if remarks else ''}", email,
        )
        db.commit()
        invalidate_export_cache(comp_code)
        return {"success": True, "message": f"Decision saved. Overall status: {file_row.approval_status}"}

    if not is_supporting_document_admin(request):
        return JSONResponse({"success": False, "message": "Admin approval is required"}, status_code=403)

    old_status = file_row.approval_status or "PENDING"
    file_row.approval_status = decision
    file_row.approved_by = email
    file_row.approved_at = datetime.utcnow()
    file_row.approval_remarks = remarks or None
    write_audit(
        db, "export_supporting", file_row.id, comp_code,
        f"DOCUMENT_{decision}", old_status,
        f"{decision}{f' | {remarks}' if remarks else ''}", email,
    )
    db.commit()
    invalidate_export_cache(comp_code)
    return {"success": True, "message": f"Document {decision.lower()} successfully"}


@router.get("/supporting_documents/entry", response_class=HTMLResponse)
def export_supporting_documents_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return RedirectResponse("/auth/login")
    context = export_supporting_documents_data(request, db)
    return templates.TemplateResponse(
        request=request,
        name="export_documents/supporting_documents.html",
        context={"request": request, **context},
    )


@router.post("/supporting_documents/upload")
async def export_supporting_documents_upload(
    request: Request,
    shipment_id: int = Form(...),
    document_kind: str = Form(...),
    document_no: str = Form(None),
    remarks: str = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    return JSONResponse({
        "success": False,
        "message": "Use the document-specific entry page to enter details, upload PDF and select approval emails",
        "page_url": f"/page/export_requirement_{document_kind}",
    }, status_code=400)

# Pydantic schemas for data validation
class ExportShipmentSchema(BaseModel):
    shipment_no: str
    po_number: str
    invoice_no: str = None
    container_no: str = None
    buyer_name: str
    country: str
    etd: date = None
    eta: date = None

    @model_validator(mode="after")
    def validate_dates(self):
        if self.etd and self.eta and self.eta < self.etd:
            raise ValueError("ETA cannot be before ETD")
        return self


class ProformaInvoiceSchema(BaseModel):
    pi_no: str
    pi_date: date
    validity_date: date = None
    po_number: str = None
    buyer_name: str
    buyer_address: str
    country: str
    currency: str = "USD"
    incoterm: str
    payment_terms: str
    port_of_loading: str = None
    port_of_discharge: str = None
    product_description: str
    quantity: Decimal
    unit: str = "KG"
    unit_price: Decimal
    status: str = "DRAFT"
    remarks: str = None
    brand: str = None
    packing_style: str = None
    freezer: str = None
    count_glaze: str = None
    weight_glaze: str = None
    species: str = None
    variety: str = None
    grade: str = None
    no_of_pieces: str = None
    no_of_mc: int = 0
    items_json: str = None

    @model_validator(mode="after")
    def validate_proforma(self):
        if self.validity_date and self.validity_date < self.pi_date:
            raise ValueError("Validity date cannot be before PI date")
        if self.quantity <= 0 or self.unit_price < 0:
            raise ValueError("Quantity must be greater than zero and unit price cannot be negative")
        if self.status not in {"DRAFT", "SENT", "ACCEPTED", "EXPIRED"}:
            raise ValueError("Invalid proforma invoice status")
        return self


class CommercialInvoiceSchema(BaseModel):
    shipment_no: str
    invoice_no: str
    po_number: str
    container_no: str = None
    buyer_name: str
    invoice_date: date
    buyer_address: str
    consignee_name: str = None
    notify_party: str = None
    country: str
    currency: str = "USD"
    exchange_rate: Decimal = Decimal("83.50")
    total_amount: Decimal
    payment_terms: str
    shipment_terms: str

    @model_validator(mode="after")
    def validate_amounts(self):
        if self.exchange_rate <= 0 or self.total_amount <= 0:
            raise ValueError("Exchange rate and invoice amount must be greater than zero")
        return self

class PackingListSchema(BaseModel):
    packing_no: str
    invoice_no: str
    po_number: str = None
    container_no: str = None
    buyer_name: str = None
    product_name: str
    grade: str
    batch_no: str = None
    lot_no: str = None
    glaze: str = None
    freezing_type: str = None
    hs_code: str = None
    manufacturing_date: date = None
    expiry_date: date = None
    packing_style: str
    inner_pack: str = None
    outer_pack: str = None
    master_cartons: int = 0
    net_weight: float = 0.0
    gross_weight: float = 0.0
    pallet_count: int = 0
    inventory_batch_id: str = None
    stock_entry_no: str = None
    invoice_item_no: int = None

    @model_validator(mode="after")
    def validate_quantities(self):
        if self.master_cartons < 0 or self.pallet_count < 0 or self.net_weight < 0 or self.gross_weight < 0:
            raise ValueError("Cartons, pallets and weights cannot be negative")
        if self.gross_weight < self.net_weight:
            raise ValueError("Gross weight cannot be less than net weight")
        if self.manufacturing_date and self.expiry_date and self.expiry_date < self.manufacturing_date:
            raise ValueError("Expiry date cannot be before manufacturing date")
        return self


class PackingListBulkLineSchema(BaseModel):
    product_name: str
    grade: str
    batch_no: str = None
    lot_no: str = None
    glaze: str = None
    freezing_type: str = None
    hs_code: str = None
    manufacturing_date: date = None
    expiry_date: date = None
    packing_style: str
    inner_pack: str = None
    outer_pack: str = None
    master_cartons: int = 0
    net_weight: float = 0.0
    gross_weight: float = 0.0
    pallet_count: int = 0
    inventory_batch_id: str = None
    stock_entry_no: str = None

    @model_validator(mode="after")
    def validate_line(self):
        if self.master_cartons < 0 or self.pallet_count < 0 or self.net_weight < 0 or self.gross_weight < 0:
            raise ValueError("Cartons, pallets and weights cannot be negative")
        if self.gross_weight < self.net_weight:
            raise ValueError("Gross weight cannot be less than net weight")
        if self.manufacturing_date and self.expiry_date and self.expiry_date < self.manufacturing_date:
            raise ValueError("Expiry date cannot be before manufacturing date")
        return self


class PackingListBulkSchema(BaseModel):
    packing_no: str
    invoice_no: str
    po_number: str = None
    container_no: str = None
    buyer_name: str = None
    items: list[PackingListBulkLineSchema] = Field(..., min_length=1, max_length=100)

class ContainerStuffingSchema(BaseModel):
    container_no: str
    invoice_no: str = None
    po_number: str = None
    buyer_name: str = None
    seal_no: str
    shipping_line: str = None
    stuffing_date: date
    stuffing_location: str = None
    container_type: str = "Reefer"
    container_size: str = "40FT"
    temperature: float
    vehicle_no: str
    loading_supervisor: str

class ShippingBillSchema(BaseModel):
    shipping_bill_no: str
    shipping_bill_date: date
    invoice_no: str
    container_no: str = None
    po_number: str = None
    buyer_name: str = None
    shipping_bill_value: float = 0.0
    drawback_amount: float = 0.0
    scheme: str = "NONE"
    customs_status: str = "LEO"
    port: str
    cha_name: str
    vessel_name: str
    voyage_no: str
    etd: date
    eta: date

    @model_validator(mode="after")
    def validate_shipping_bill(self):
        if self.shipping_bill_value < 0 or self.drawback_amount < 0:
            raise ValueError("Shipping bill and drawback values cannot be negative")
        if self.eta < self.etd:
            raise ValueError("ETA cannot be before ETD")
        return self

class BillOfLadingSchema(BaseModel):
    bl_no: str
    bl_date: date
    invoice_no: str
    container_no: str
    po_number: str = None
    buyer_name: str = None
    shipping_line: str
    seal_no: str
    freight_terms: str = "PREPAID"
    no_of_original_bl: int = 3
    gross_weight: float = 0.0
    net_weight: float = 0.0

    @model_validator(mode="after")
    def validate_bl(self):
        if self.no_of_original_bl <= 0:
            raise ValueError("Original B/L count must be greater than zero")
        if self.net_weight < 0 or self.gross_weight < 0 or self.gross_weight < self.net_weight:
            raise ValueError("B/L weights are invalid")
        return self

class HealthCertificateSchema(BaseModel):
    certificate_no: str
    issue_date: date
    authority: str = "EIA"
    invoice_no: str
    container_no: str
    po_number: str = None
    buyer_name: str = None
    country: str = None
    species: str = None
    temperature_verified: bool = True
    issued_by: str = None


# Helper function to audit actions
def write_audit(db: Session, table: str, rec_id: int, company_id: str, action: str, old: str, new: str, email: str):
    audit = AuditLog(
        table_name=table,
        record_id=rec_id,
        company_id=company_id,
        field_name=action,
        old_value=old,
        new_value=new,
        edited_by=email,
        edited_at=datetime.utcnow()
    )
    db.add(audit)
    invalidate_export_cache(company_id)


def require_company_invoice(db: Session, company_id: str, invoice_no: str) -> CommercialInvoice:
    invoice = db.query(CommercialInvoice).filter(
        CommercialInvoice.company_id == company_id,
        CommercialInvoice.invoice_no == invoice_no,
        CommercialInvoice.is_cancelled != True,
    ).first()
    if not invoice:
        raise ValueError("Select a valid commercial invoice for this company")
    return invoice


def refresh_compliance(db: Session, company_id: str, shipment_no: str) -> None:
    tracker = db.query(ExportComplianceTracker).filter(
        ExportComplianceTracker.company_id == company_id,
        ExportComplianceTracker.shipment_no == shipment_no,
    ).first()
    if not tracker:
        return
    invoice = db.query(CommercialInvoice).filter(
        CommercialInvoice.company_id == company_id,
        CommercialInvoice.shipment_no == shipment_no,
        CommercialInvoice.is_cancelled != True,
    ).first()
    tracker.invoice_pending = invoice is None
    if not invoice:
        tracker.packing_list_pending = True
        tracker.health_cert_pending = True
        tracker.shipping_bill_pending = True
        tracker.bl_pending = True
        return
    common = (lambda model: db.query(model).filter(
        model.company_id == company_id,
        model.invoice_no == invoice.invoice_no,
        model.is_cancelled != True,
    ).first() is None)
    tracker.packing_list_pending = common(PackingList)
    tracker.health_cert_pending = common(HealthCertificate)
    tracker.shipping_bill_pending = common(ShippingBill)
    tracker.bl_pending = common(BillOfLading)


def apply_invoice_container_defaults(db: Session, company_id: str, invoices: list[CommercialInvoice]) -> list[CommercialInvoice]:
    """Populate missing invoice container numbers from stuffing/shipment links for form defaults."""
    for invoice in invoices:
        if invoice.container_no:
            continue
        stuffing = db.query(ContainerStuffing).filter(
            ContainerStuffing.company_id == company_id,
            ContainerStuffing.invoice_no == invoice.invoice_no,
            ContainerStuffing.is_cancelled != True,
        ).order_by(desc(ContainerStuffing.id)).first()
        if stuffing:
            invoice.container_no = stuffing.container_no
            continue
        shipment = db.query(ExportShipment).filter(
            ExportShipment.company_id == company_id,
            ExportShipment.shipment_no == invoice.shipment_no,
            ExportShipment.is_cancelled != True,
        ).first()
        if shipment and shipment.container_no:
            invoice.container_no = shipment.container_no
    return invoices


# ============================================================
# PROFORMA INVOICES
# ============================================================
def serialize_proforma(row: ProformaInvoice) -> dict:
    return {
        "id": row.id,
        "pi_no": row.pi_no,
        "pi_date": _dt(row.pi_date),
        "validity_date": _dt(row.validity_date),
        "po_number": row.po_number,
        "buyer_name": row.buyer_name,
        "buyer_address": row.buyer_address,
        "country": row.country,
        "currency": row.currency,
        "incoterm": row.incoterm,
        "payment_terms": row.payment_terms,
        "port_of_loading": row.port_of_loading,
        "port_of_discharge": row.port_of_discharge,
        "product_description": row.product_description,
        "quantity": str(row.quantity or 0),
        "unit": row.unit,
        "unit_price": str(row.unit_price or 0),
        "total_amount": str(row.total_amount or 0),
        "status": row.status,
        "approval_status": row.approval_status or "PENDING",
        "approved_by": row.approved_by,
        "approved_at": _dt(row.approved_at),
        "approval_remarks": row.approval_remarks,
        "remarks": row.remarks,
        "brand": getattr(row, "brand", "") or "",
        "packing_style": getattr(row, "packing_style", "") or "",
        "freezer": getattr(row, "freezer", "") or "",
        "count_glaze": getattr(row, "count_glaze", "") or "",
        "weight_glaze": getattr(row, "weight_glaze", "") or "",
        "species": getattr(row, "species", "") or "",
        "variety": getattr(row, "variety", "") or "",
        "grade": getattr(row, "grade", "") or "",
        "no_of_pieces": getattr(row, "no_of_pieces", "") or "",
        "no_of_mc": getattr(row, "no_of_mc", 0) or 0,
        "items_json": getattr(row, "items_json", "") or "",
        "created_by": row.created_by,
        "created_at": _dt(row.created_at),
    }


@router.get("/proforma_invoice/entry", response_class=HTMLResponse)
def proforma_invoice_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        return RedirectResponse("/", status_code=302)
    ensure_bill_accounting_schema(db)
    history = db.query(ProformaInvoice).filter(
        ProformaInvoice.company_id == comp_code,
        ProformaInvoice.is_cancelled != True,
    ).order_by(desc(ProformaInvoice.pi_date), desc(ProformaInvoice.id)).all()
    return templates.TemplateResponse(
        request=request,
        name="export_documents/proforma_invoice.html",
        context={"history": history, "company_id": comp_code},
    )


@router.get("/proforma_invoice/data")
def proforma_invoice_data(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Unauthorized")
    ensure_bill_accounting_schema(db)
    rows = db.query(ProformaInvoice).filter(
        ProformaInvoice.company_id == comp_code,
        ProformaInvoice.is_cancelled != True,
    ).order_by(desc(ProformaInvoice.pi_date), desc(ProformaInvoice.id)).all()
    audit_rows = db.query(AuditLog).filter(
        AuditLog.company_id == comp_code,
        AuditLog.table_name == "proforma_invoices",
    ).order_by(desc(AuditLog.edited_at), desc(AuditLog.id)).limit(100).all()
    buyer_rows = db.query(buyers).filter(
        buyers.company_id == comp_code,
    ).order_by(buyers.buyer_name).all()
    country_rows = db.query(countries).filter(
        countries.company_id == comp_code,
    ).order_by(countries.country_name).all()
    current_year = date.today().year
    existing_numbers = [
        row.pi_no for row in db.query(ProformaInvoice.pi_no).filter(
            ProformaInvoice.company_id == comp_code,
            ProformaInvoice.pi_no.ilike(f"PI-{current_year}-%"),
        ).all()
    ]
    sequence_values = []
    for number in existing_numbers:
        match = re.search(r"(\d+)$", str(number or ""))
        if match:
            sequence_values.append(int(match.group(1)))
    next_pi_no = f"PI-{current_year}-{max(sequence_values, default=0) + 1:04d}"
    return {
        "success": True,
        "can_approve": is_supporting_document_admin(request),
        "next_pi_no": next_pi_no,
        "buyers": [{
            "name": row.buyer_name,
            "address": row.address or "",
            "country": row.country or "",
            "currency": row.currency_code or "USD",
            "payment_terms": f"{int(row.payment_terms_days or 0)} Days" if row.payment_terms_days else "",
            "contact_person": row.contact_person or "",
            "email": row.buyer_email or "",
            "iec_code": row.iec_code or "",
        } for row in buyer_rows],
        "countries": [row.country_name for row in country_rows],
        "brands": [b.brand_name for b in db.query(brands).filter(brands.company_id == comp_code).order_by(brands.brand_name).all()],
        "packing_styles": [p.packing_style for p in db.query(packing_styles).filter(packing_styles.company_id == comp_code).order_by(packing_styles.packing_style).all()],
        "freezers": [f.freezer_name for f in db.query(freezers).filter(freezers.company_id == comp_code).order_by(freezers.freezer_name).all()],
        "glazes": [g.glaze_name for g in db.query(glazes).filter(glazes.company_id == comp_code).order_by(glazes.glaze_name).all()],
        "species": [s.species_name for s in db.query(species).filter(species.company_id == comp_code).order_by(species.species_name).all()],
        "varieties": [v.variety_name for v in db.query(varieties).filter(varieties.company_id == comp_code).order_by(varieties.variety_name).all()],
        "grades": [g.grade_name for g in db.query(grades).filter(grades.company_id == comp_code).order_by(grades.grade_name).all()],
        "rows": [serialize_proforma(row) for row in rows],
        "audit_logs": [{
            "id": audit.id,
            "record_id": audit.record_id,
            "action": audit.field_name,
            "old_value": audit.old_value,
            "new_value": audit.new_value,
            "edited_by": audit.edited_by,
            "edited_at": _dt(audit.edited_at),
        } for audit in audit_rows],
    }


def proforma_payload_values(payload: ProformaInvoiceSchema) -> dict:
    values = payload.model_dump()
    values["pi_no"] = payload.pi_no.strip()
    values["buyer_name"] = payload.buyer_name.strip()
    values["total_amount"] = (payload.quantity * payload.unit_price).quantize(Decimal("0.01"))
    return values


@router.post("/proforma_invoice/save")
def proforma_invoice_save(request: Request, payload: ProformaInvoiceSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    exists = db.query(ProformaInvoice).filter(
        ProformaInvoice.company_id == comp_code,
        ProformaInvoice.pi_no == payload.pi_no.strip(),
    ).first()
    if exists:
        return JSONResponse({"success": False, "message": "PI number already exists"}, status_code=400)
    entry = ProformaInvoice(company_id=comp_code, created_by=email, **proforma_payload_values(payload))
    db.add(entry)
    db.flush()
    write_audit(db, "proforma_invoices", entry.id, comp_code, "CREATE", "NONE", f"PI {entry.pi_no}", email)
    db.commit()
    invalidate_export_cache(comp_code)
    return {
        "success": True,
        "message": "Proforma invoice created successfully",
        "record_id": entry.id,
        "print_url": f"/export_documents/proforma_invoice/print/{entry.id}",
        "pdf_url": f"/export_documents/proforma_invoice/pdf/{entry.id}",
    }


@router.put("/proforma_invoice/{record_id}")
def proforma_invoice_update(record_id: int, request: Request, payload: ProformaInvoiceSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    entry = db.query(ProformaInvoice).filter(
        ProformaInvoice.id == record_id,
        ProformaInvoice.company_id == comp_code,
        ProformaInvoice.is_cancelled != True,
    ).first()
    if not entry:
        return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)
    duplicate = db.query(ProformaInvoice).filter(
        ProformaInvoice.company_id == comp_code,
        ProformaInvoice.pi_no == payload.pi_no.strip(),
        ProformaInvoice.id != record_id,
    ).first()
    if duplicate:
        return JSONResponse({"success": False, "message": "PI number already exists"}, status_code=400)
    old_status = entry.status
    old_approval = entry.approval_status or "PENDING"
    for field, value in proforma_payload_values(payload).items():
        setattr(entry, field, value)
    entry.updated_by = email
    entry.approval_status = "PENDING"
    entry.approved_by = None
    entry.approved_at = None
    entry.approval_remarks = None
    write_audit(db, "proforma_invoices", entry.id, comp_code, "UPDATE", old_status, entry.status, email)
    if old_approval != "PENDING":
        write_audit(db, "proforma_invoices", entry.id, comp_code, "APPROVAL_RESET", old_approval, "PENDING", email)
    db.commit()
    invalidate_export_cache(comp_code)
    return {"success": True, "message": "Proforma invoice updated successfully"}


@router.post("/proforma_invoice/cancel/{record_id}")
def proforma_invoice_cancel(record_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    entry = db.query(ProformaInvoice).filter(
        ProformaInvoice.id == record_id,
        ProformaInvoice.company_id == comp_code,
        ProformaInvoice.is_cancelled != True,
    ).first()
    if not entry:
        return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)
    old_status = entry.status
    entry.is_cancelled = True
    entry.status = "CANCELLED"
    entry.updated_by = email
    write_audit(db, "proforma_invoices", entry.id, comp_code, "CANCEL", old_status, "CANCELLED", email)
    db.commit()
    invalidate_export_cache(comp_code)
    return {"success": True, "message": "Proforma invoice cancelled successfully"}


@router.post("/proforma_invoice/{record_id}/approval")
def proforma_invoice_approval(
    record_id: int,
    payload: SupportingDocumentApprovalSchema,
    request: Request,
    db: Session = Depends(get_db),
):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    if not is_supporting_document_admin(request):
        return JSONResponse({"success": False, "message": "Admin approval is required"}, status_code=403)
    decision = payload.decision.strip().upper()
    remarks = (payload.remarks or "").strip()[:500]
    if decision not in {"APPROVED", "REJECTED"}:
        return JSONResponse({"success": False, "message": "Decision must be APPROVED or REJECTED"}, status_code=400)
    if decision == "REJECTED" and not remarks:
        return JSONResponse({"success": False, "message": "Rejection remarks are required"}, status_code=400)
    entry = db.query(ProformaInvoice).filter(
        ProformaInvoice.id == record_id,
        ProformaInvoice.company_id == comp_code,
        ProformaInvoice.is_cancelled != True,
    ).first()
    if not entry:
        return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)
    old_status = entry.approval_status or "PENDING"
    entry.approval_status = decision
    entry.approved_by = email
    entry.approved_at = datetime.utcnow()
    entry.approval_remarks = remarks or None
    write_audit(
        db, "proforma_invoices", entry.id, comp_code, f"DOCUMENT_{decision}",
        old_status, f"{decision}{f' | {remarks}' if remarks else ''}", email,
    )
    db.commit()
    invalidate_export_cache(comp_code)
    return {"success": True, "message": f"Proforma invoice {decision.lower()} successfully"}


# ============================================================
# 1. EXPORT SHIPMENTS
# ============================================================
@router.get("/export_shipment/entry", response_class=HTMLResponse)
def export_shipment_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code: return RedirectResponse("/", status_code=302)
    history = db.query(ExportShipment).filter(ExportShipment.company_id == comp_code, ExportShipment.is_cancelled != True).order_by(desc(ExportShipment.created_at)).all()
    return templates.TemplateResponse(request=request, name="export_documents/export_shipment.html", context={"history": history, "company_id": comp_code})

@router.post("/export_shipment/save")
def export_shipment_save(request: Request, payload: ExportShipmentSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code: return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    
    exists = db.query(ExportShipment).filter(ExportShipment.company_id == comp_code, ExportShipment.shipment_no == payload.shipment_no).first()
    if exists: return JSONResponse({"success": False, "message": "Shipment Number already registered"}, status_code=400)
    
    entry = ExportShipment(company_id=comp_code, created_by=email, **payload.dict())
    db.add(entry)
    db.flush()
    
    # Create companion Compliance checklist automatically
    compliance = ExportComplianceTracker(
        company_id=comp_code,
        shipment_no=entry.shipment_no,
        invoice_pending=True,
        packing_list_pending=True,
        health_cert_pending=True,
        shipping_bill_pending=True,
        bl_pending=True,
        payment_pending=True
    )
    db.add(compliance)
    
    write_audit(db, "export_shipments", entry.id, comp_code, "CREATE", "NONE", f"Shipment Registered: {payload.shipment_no}", email)
    db.commit()
    return {"success": True, "message": "Export shipment registered successfully"}

@router.post("/export_shipment/delete/{log_id}")
def export_shipment_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(ExportShipment).filter(ExportShipment.id == log_id, ExportShipment.company_id == comp_code).first()
    if entry:
        if entry.invoice_no:
            return JSONResponse({"success": False, "message": "Cancel linked export documents before cancelling this shipment"}, status_code=400)
        write_audit(db, "export_shipments", entry.id, comp_code, "CANCEL", entry.status, "CANCELLED", email)
        entry.is_cancelled = True
        entry.status = "CANCELLED"
        db.commit()
        return {"success": True, "message": "Export shipment deleted successfully"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# 2. COMMERCIAL INVOICE
# ============================================================
@router.get("/commercial_invoice/entry", response_class=HTMLResponse)
def commercial_invoice_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code: return RedirectResponse("/", status_code=302)
    history = db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.is_cancelled != True).order_by(desc(CommercialInvoice.invoice_date)).all()
    shipments = db.query(ExportShipment).filter(ExportShipment.company_id == comp_code, ExportShipment.is_cancelled != True).all()
    return templates.TemplateResponse(request=request, name="export_documents/commercial_invoice.html", context={"history": history, "shipments": shipments, "company_id": comp_code})

@router.post("/commercial_invoice/save")
def commercial_invoice_save(request: Request, payload: CommercialInvoiceSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code: return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    ensure_bill_accounting_schema(db)
    
    exists = db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.invoice_no == payload.invoice_no).first()
    if exists: return JSONResponse({"success": False, "message": "Commercial Invoice No already registered"}, status_code=400)
    shipment = db.query(ExportShipment).filter(
        ExportShipment.company_id == comp_code,
        ExportShipment.shipment_no == payload.shipment_no,
        ExportShipment.is_cancelled != True,
    ).first()
    if not shipment:
        return JSONResponse({"success": False, "message": "Select a valid export shipment for this company"}, status_code=400)
    if shipment.invoice_no:
        return JSONResponse({"success": False, "message": "This shipment already has a commercial invoice"}, status_code=400)
    
    try:
        inr_value = payload.total_amount * payload.exchange_rate
        entry = CommercialInvoice(
            company_id=comp_code,
            invoice_value_inr=inr_value,
            created_by=email,
            **payload.dict()
        )
        db.add(entry)
        db.flush()

        voucher = post_export_sales_invoice(
            db,
            comp_code,
            payload.invoice_date,
            payload.invoice_no,
            payload.buyer_name,
            inr_value,
            email,
        )
        entry.journal_id = voucher.id
        entry.status = "POSTED"

        # Update ExportShipment and ExportComplianceTracker status
        shipment.invoice_no = entry.invoice_no
        refresh_compliance(db, comp_code, shipment.shipment_no)

        write_audit(db, "commercial_invoices", entry.id, comp_code, "CREATE", "NONE", f"Invoice Registered: {payload.invoice_no}", email)
        db.commit()
        invalidate_export_cache(comp_code)
        return {
            "success": True,
            "message": "Commercial invoice registered and posted to accounts",
            "record_id": entry.id,
            "print_url": f"/export_documents/commercial_invoice/print/{entry.id}",
            "pdf_url": f"/export_documents/commercial_invoice/pdf/{entry.id}",
        }
    except Exception as exc:
        db.rollback()
        logger.exception("Commercial invoice accounting post failed")
        return JSONResponse({"success": False, "message": str(exc)}, status_code=400)

@router.post("/commercial_invoice/cancel/{log_id}")
@router.post("/commercial_invoice/delete/{log_id}")
def commercial_invoice_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    ensure_bill_accounting_schema(db)
    entry = db.query(CommercialInvoice).filter(CommercialInvoice.id == log_id, CommercialInvoice.company_id == comp_code).first()
    if entry:
        for model in (PackingList, ContainerStuffing, ShippingBill, BillOfLading, HealthCertificate):
            if db.query(model).filter(model.company_id == comp_code, model.invoice_no == entry.invoice_no, model.is_cancelled != True).first():
                return JSONResponse({"success": False, "message": "Cancel linked export documents before cancelling this invoice"}, status_code=400)
        cancel_linked_bill_voucher(db, comp_code, entry.journal_id, email)
        cancel_linked_bill_voucher(db, comp_code, entry.cogs_journal_id, email)
        write_audit(db, "commercial_invoices", entry.id, comp_code, "CANCEL", entry.status, "CANCELLED", email)
        entry.is_cancelled = True
        entry.status = "CANCELLED"
        shipment = db.query(ExportShipment).filter(ExportShipment.company_id == comp_code, ExportShipment.shipment_no == entry.shipment_no).first()
        if shipment and shipment.invoice_no == entry.invoice_no:
            shipment.invoice_no = None
        refresh_compliance(db, comp_code, entry.shipment_no)
        db.commit()
        invalidate_export_cache(comp_code)
        return {"success": True, "message": "Commercial invoice cancelled successfully"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# 3. PACKING LIST
# ============================================================
@router.get("/packing_list/entry", response_class=HTMLResponse)
def packing_list_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code: return RedirectResponse("/", status_code=302)
    history = db.query(PackingList).filter(PackingList.company_id == comp_code, PackingList.is_cancelled != True).order_by(desc(PackingList.created_at)).all()
    invoices = apply_invoice_container_defaults(db, comp_code, db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.is_cancelled != True).all())
    return templates.TemplateResponse(request=request, name="export_documents/packing_list.html", context={"history": history, "invoices": invoices, "company_id": comp_code})

@router.post("/packing_list/save")
def packing_list_save(request: Request, payload: PackingListSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code: return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    try:
        invoice = require_company_invoice(db, comp_code, payload.invoice_no)
    except ValueError as exc:
        return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
    entry = PackingList(company_id=comp_code, created_by=email, **payload.dict())
    db.add(entry)
    db.flush()
    cogs_value = repost_invoice_cogs(db, comp_code, invoice, email)
    
    # Update ExportComplianceTracker packing list pending status
    refresh_compliance(db, comp_code, invoice.shipment_no)
            
    write_audit(db, "packing_lists", entry.id, comp_code, "CREATE", "NONE", f"Packing Item: {payload.packing_no}", email)
    db.commit()
    return {"success": True, "message": "Packing list and COGS accounting recorded successfully", "cogs_value": cogs_value}


@router.post("/packing_list/save-bulk")
def packing_list_save_bulk(request: Request, payload: PackingListBulkSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    try:
        invoice = require_company_invoice(db, comp_code, payload.invoice_no)
        created_entries = []
        common = {
            "company_id": comp_code,
            "packing_no": payload.packing_no.strip(),
            "invoice_no": payload.invoice_no,
            "po_number": payload.po_number,
            "container_no": payload.container_no,
            "buyer_name": payload.buyer_name,
            "created_by": email,
        }
        if not common["packing_no"]:
            raise ValueError("Packing document number is required")
        for index, item in enumerate(payload.items, start=1):
            entry = PackingList(
                **common,
                invoice_item_no=index,
                **item.model_dump(),
            )
            db.add(entry)
            db.flush()
            created_entries.append(entry)
            write_audit(
                db,
                "packing_lists",
                entry.id,
                comp_code,
                "CREATE",
                "NONE",
                f"Packing {payload.packing_no} · Line {index} · {item.product_name} · {item.grade}",
                email,
            )
        cogs_value = repost_invoice_cogs(db, comp_code, invoice, email)
        refresh_compliance(db, comp_code, invoice.shipment_no)
        db.commit()
        invalidate_export_cache(comp_code)
        return {
            "success": True,
            "message": f"Packing list saved successfully with {len(created_entries)} line items",
            "record_ids": [entry.id for entry in created_entries],
            "cogs_value": cogs_value,
        }
    except Exception as exc:
        db.rollback()
        logger.exception("Bulk packing list save failed")
        return JSONResponse({"success": False, "message": str(exc)}, status_code=400)

@router.post("/packing_list/delete/{log_id}")
def packing_list_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(PackingList).filter(PackingList.id == log_id, PackingList.company_id == comp_code).first()
    if entry:
        invoice = require_company_invoice(db, comp_code, entry.invoice_no)
        write_audit(db, "packing_lists", entry.id, comp_code, "CANCEL", "ACTIVE", "CANCELLED", email)
        entry.is_cancelled = True
        repost_invoice_cogs(db, comp_code, invoice, email)
        refresh_compliance(db, comp_code, invoice.shipment_no)
        db.commit()
        return {"success": True, "message": "Packing list entry deleted successfully"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# 4. CONTAINER STUFFING
# ============================================================
@router.get("/container_stuffing/entry", response_class=HTMLResponse)
def container_stuffing_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code: return RedirectResponse("/", status_code=302)
    history = db.query(ContainerStuffing).filter(ContainerStuffing.company_id == comp_code, ContainerStuffing.is_cancelled != True).order_by(desc(ContainerStuffing.stuffing_date)).all()
    invoices = apply_invoice_container_defaults(db, comp_code, db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.is_cancelled != True).all())
    return templates.TemplateResponse(request=request, name="export_documents/container_stuffing.html", context={"history": history, "invoices": invoices, "company_id": comp_code})

@router.post("/container_stuffing/save")
def container_stuffing_save(request: Request, payload: ContainerStuffingSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code: return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    invoice = None
    if payload.invoice_no:
        try:
            invoice = require_company_invoice(db, comp_code, payload.invoice_no)
        except ValueError as exc:
            return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
    
    exists = db.query(ContainerStuffing).filter(ContainerStuffing.company_id == comp_code, ContainerStuffing.container_no == payload.container_no).first()
    if exists: return JSONResponse({"success": False, "message": "Container stuffing already logged"}, status_code=400)
    
    entry = ContainerStuffing(company_id=comp_code, created_by=email, **payload.dict())
    db.add(entry)
    db.flush()
    
    # Update ExportShipment container No
    if payload.invoice_no:
        if invoice:
            invoice.container_no = entry.container_no
            shipment = db.query(ExportShipment).filter(ExportShipment.company_id == comp_code, ExportShipment.shipment_no == invoice.shipment_no).first()
            if shipment:
                shipment.container_no = entry.container_no
                
    write_audit(db, "container_stuffing", entry.id, comp_code, "CREATE", "NONE", f"Container Stuffing: {payload.container_no}", email)
    db.commit()
    return {"success": True, "message": "Container stuffing log recorded successfully"}

@router.post("/container_stuffing/delete/{log_id}")
def container_stuffing_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(ContainerStuffing).filter(ContainerStuffing.id == log_id, ContainerStuffing.company_id == comp_code).first()
    if entry:
        write_audit(db, "container_stuffing", entry.id, comp_code, "CANCEL", "ACTIVE", "CANCELLED", email)
        entry.is_cancelled = True
        db.commit()
        return {"success": True, "message": "Container stuffing record deleted successfully"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# 5. SHIPPING BILL
# ============================================================
@router.get("/shipping_bill/entry", response_class=HTMLResponse)
def shipping_bill_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code: return RedirectResponse("/", status_code=302)
    history = db.query(ShippingBill).filter(ShippingBill.company_id == comp_code, ShippingBill.is_cancelled != True).order_by(desc(ShippingBill.shipping_bill_date)).all()
    invoices = apply_invoice_container_defaults(db, comp_code, db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.is_cancelled != True).all())
    return templates.TemplateResponse(request=request, name="export_documents/shipping_bill.html", context={"history": history, "invoices": invoices, "company_id": comp_code})

@router.post("/shipping_bill/save")
def shipping_bill_save(request: Request, payload: ShippingBillSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code: return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    try:
        invoice = require_company_invoice(db, comp_code, payload.invoice_no)
    except ValueError as exc:
        return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
    
    exists = db.query(ShippingBill).filter(ShippingBill.company_id == comp_code, ShippingBill.shipping_bill_no == payload.shipping_bill_no).first()
    if exists: return JSONResponse({"success": False, "message": "Shipping Bill Number already registered"}, status_code=400)
    
    entry = ShippingBill(company_id=comp_code, created_by=email, **payload.dict())
    db.add(entry)
    db.flush()
    
    # Update ExportComplianceTracker shipping bill pending status
    refresh_compliance(db, comp_code, invoice.shipment_no)
            
    write_audit(db, "shipping_bills", entry.id, comp_code, "CREATE", "NONE", f"Shipping Bill: {payload.shipping_bill_no}", email)
    db.commit()
    return {"success": True, "message": "Shipping Bill successfully registered"}

@router.post("/shipping_bill/delete/{log_id}")
def shipping_bill_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(ShippingBill).filter(ShippingBill.id == log_id, ShippingBill.company_id == comp_code).first()
    if entry:
        invoice = require_company_invoice(db, comp_code, entry.invoice_no)
        write_audit(db, "shipping_bills", entry.id, comp_code, "CANCEL", "ACTIVE", "CANCELLED", email)
        entry.is_cancelled = True
        refresh_compliance(db, comp_code, invoice.shipment_no)
        db.commit()
        return {"success": True, "message": "Shipping bill record removed successfully"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# 6. BILL OF LADING
# ============================================================
@router.get("/bill_of_lading/entry", response_class=HTMLResponse)
def bill_of_lading_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code: return RedirectResponse("/", status_code=302)
    history = db.query(BillOfLading).filter(BillOfLading.company_id == comp_code, BillOfLading.is_cancelled != True).order_by(desc(BillOfLading.bl_date)).all()
    invoices = apply_invoice_container_defaults(db, comp_code, db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.is_cancelled != True).all())
    return templates.TemplateResponse(request=request, name="export_documents/bill_of_lading.html", context={"history": history, "invoices": invoices, "company_id": comp_code})

@router.post("/bill_of_lading/save")
def bill_of_lading_save(request: Request, payload: BillOfLadingSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code: return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    try:
        invoice = require_company_invoice(db, comp_code, payload.invoice_no)
    except ValueError as exc:
        return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
    
    exists = db.query(BillOfLading).filter(BillOfLading.company_id == comp_code, BillOfLading.bl_no == payload.bl_no).first()
    if exists: return JSONResponse({"success": False, "message": "BL Number already registered"}, status_code=400)
    
    entry = BillOfLading(company_id=comp_code, created_by=email, **payload.dict())
    db.add(entry)
    db.flush()
    
    # Update ExportComplianceTracker BL pending status
    refresh_compliance(db, comp_code, invoice.shipment_no)
            
    write_audit(db, "bill_of_ladings", entry.id, comp_code, "CREATE", "NONE", f"BL Entry: {payload.bl_no}", email)
    db.commit()
    return {"success": True, "message": "Bill of Lading recorded successfully"}

@router.post("/bill_of_lading/delete/{log_id}")
def bill_of_lading_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(BillOfLading).filter(BillOfLading.id == log_id, BillOfLading.company_id == comp_code).first()
    if entry:
        invoice = require_company_invoice(db, comp_code, entry.invoice_no)
        write_audit(db, "bill_of_ladings", entry.id, comp_code, "CANCEL", "ACTIVE", "CANCELLED", email)
        entry.is_cancelled = True
        refresh_compliance(db, comp_code, invoice.shipment_no)
        db.commit()
        return {"success": True, "message": "Bill of lading entry removed"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# 7. HEALTH CERTIFICATE
# ============================================================
@router.get("/health_certificate/entry", response_class=HTMLResponse)
def health_certificate_entry(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code: return RedirectResponse("/", status_code=302)
    history = db.query(HealthCertificate).filter(HealthCertificate.company_id == comp_code, HealthCertificate.is_cancelled != True).order_by(desc(HealthCertificate.issue_date)).all()
    invoices = apply_invoice_container_defaults(db, comp_code, db.query(CommercialInvoice).filter(CommercialInvoice.company_id == comp_code, CommercialInvoice.is_cancelled != True).all())
    return templates.TemplateResponse(request=request, name="export_documents/health_certificate.html", context={"history": history, "invoices": invoices, "company_id": comp_code})

@router.post("/health_certificate/save")
def health_certificate_save(request: Request, payload: HealthCertificateSchema, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    if not comp_code: return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    try:
        invoice = require_company_invoice(db, comp_code, payload.invoice_no)
    except ValueError as exc:
        return JSONResponse({"success": False, "message": str(exc)}, status_code=400)
    
    exists = db.query(HealthCertificate).filter(HealthCertificate.company_id == comp_code, HealthCertificate.certificate_no == payload.certificate_no).first()
    if exists: return JSONResponse({"success": False, "message": "Certificate Number already exists"}, status_code=400)
    
    entry = HealthCertificate(company_id=comp_code, created_by=email, **payload.dict())
    db.add(entry)
    db.flush()
    
    # Update ExportComplianceTracker Health Certificate pending status
    refresh_compliance(db, comp_code, invoice.shipment_no)
            
    write_audit(db, "health_certificates", entry.id, comp_code, "CREATE", "NONE", f"Health Cert: {payload.certificate_no}", email)
    db.commit()
    return {"success": True, "message": "Health Certificate recorded successfully"}

@router.post("/health_certificate/delete/{log_id}")
def health_certificate_delete(log_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    entry = db.query(HealthCertificate).filter(HealthCertificate.id == log_id, HealthCertificate.company_id == comp_code).first()
    if entry:
        invoice = require_company_invoice(db, comp_code, entry.invoice_no)
        write_audit(db, "health_certificates", entry.id, comp_code, "CANCEL", "ACTIVE", "CANCELLED", email)
        entry.is_cancelled = True
        refresh_compliance(db, comp_code, invoice.shipment_no)
        db.commit()
        return {"success": True, "message": "Health certificate deleted successfully"}
    return JSONResponse({"success": False, "message": "Record not found"}, status_code=404)


# ============================================================
# PRINT / PDF / UPLOADED COPY STORAGE
# ============================================================
@router.get("/registers.xlsx")
def export_all_document_registers(request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Unauthorized")
    content = document_register_workbook(db, comp_code)
    log_data_management_action(comp_code, "REGISTER", "All Export Registers", "Success", "Downloaded complete export-document workbook")
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Export_Document_Registers_{safe_filename(comp_code)}.xlsx"'},
    )


@router.get("/{doc_type}/register.xlsx")
def export_document_register(doc_type: str, request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Unauthorized")
    if doc_type not in export_doc_config():
        raise HTTPException(status_code=404, detail="Unsupported document type")
    content = document_register_workbook(db, comp_code, doc_type)
    label = doc_type.replace("_", " ").title()
    log_data_management_action(comp_code, "REGISTER", f"{label} Register", "Success", f"Downloaded {label} tenant register")
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_filename(doc_type)}_Register.xlsx"'},
    )


@router.get("/shipment/{shipment_id}/dossier.zip")
def export_shipment_dossier(shipment_id: int, request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Unauthorized")
    shipment = db.query(ExportShipment).filter(
        ExportShipment.id == shipment_id,
        ExportShipment.company_id == comp_code,
    ).first()
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")

    records = [("export_shipment", shipment)]
    invoice = None
    if shipment.invoice_no:
        invoice = db.query(CommercialInvoice).filter(
            CommercialInvoice.company_id == comp_code,
            CommercialInvoice.invoice_no == shipment.invoice_no,
        ).first()
    if invoice:
        records.append(("commercial_invoice", invoice))
        linked_models = (
            ("packing_list", PackingList),
            ("container_stuffing", ContainerStuffing),
            ("shipping_bill", ShippingBill),
            ("bill_of_lading", BillOfLading),
            ("health_certificate", HealthCertificate),
        )
        for doc_type, model in linked_models:
            linked_rows = db.query(model).filter(
                model.company_id == comp_code,
                model.invoice_no == invoice.invoice_no,
            ).order_by(model.id).all()
            records.extend((doc_type, row) for row in linked_rows)

    output = BytesIO()
    company_profile = get_export_company_profile(db, comp_code)
    manifest_rows = []
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for index, (doc_type, row) in enumerate(records, start=1):
            cfg = export_doc_config()[doc_type]
            document_no = str(getattr(row, cfg["no"], row.id))
            file_name = f"{index:02d}_{safe_filename(cfg['title'])}_{safe_filename(document_no)}.pdf"
            archive.writestr(
                file_name,
                render_document_pdf(
                    cfg,
                    row,
                    comp_code,
                    doc_type,
                    company_profile,
                    get_invoice_packing_rows(db, row),
                ),
            )
            manifest_rows.append((cfg["title"], document_no, "SYSTEM GENERATED", file_name))

        supporting = db.query(ExportDocumentFile).filter(
            ExportDocumentFile.company_id == comp_code,
            ExportDocumentFile.module_name == "export_supporting",
            ExportDocumentFile.record_id == shipment.id,
            ExportDocumentFile.is_current == True,
        ).order_by(ExportDocumentFile.document_kind).all()
        for file_row in supporting:
            file_name = f"Supporting/{safe_filename(file_row.document_kind)}_{safe_filename(file_row.file_name)}"
            archive.writestr(file_name, file_row.file_bytes)
            manifest_rows.append((file_row.document_kind, file_row.document_no or "", "UPLOADED COPY", file_name))

        present_types = {doc_type for doc_type, _ in records}
        for required_type in (
            "export_shipment", "commercial_invoice", "packing_list", "container_stuffing",
            "shipping_bill", "bill_of_lading", "health_certificate",
        ):
            if required_type not in present_types:
                manifest_rows.append((export_doc_config()[required_type]["title"], "", "MISSING", ""))

        manifest = openpyxl.Workbook()
        sheet = manifest.active
        sheet.title = "Shipment Dossier"
        sheet.append(["BKNR EXPORT SHIPMENT DOSSIER"])
        sheet.append(["Shipment No", shipment.shipment_no])
        sheet.append(["Company", comp_code])
        sheet.append(["Generated UTC", datetime.utcnow().strftime("%d-%b-%Y %H:%M")])
        sheet.append([])
        sheet.append(["Document Type", "Document No", "Source", "File"])
        for row in manifest_rows:
            sheet.append(row)
        for cell in sheet[6]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="176B87")
        for column, width in zip("ABCD", (28, 24, 20, 64)):
            sheet.column_dimensions[column].width = width
        manifest_output = BytesIO()
        manifest.save(manifest_output)
        archive.writestr("00_Dossier_Manifest.xlsx", manifest_output.getvalue())

    write_audit(
        db, "export_shipments", shipment.id, comp_code, "DOSSIER_EXPORT", "NONE",
        f"Exported {len(manifest_rows)} documents", request.session.get("email"),
    )
    db.commit()
    log_data_management_action(comp_code, "DOSSIER", f"Shipment {shipment.shipment_no}", "Success", "Downloaded complete shipment dossier")
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="Shipment_{safe_filename(shipment.shipment_no)}_Dossier.zip"'},
    )


@router.get("/{doc_type}/print/{record_id}", response_class=HTMLResponse)
def export_document_print(doc_type: str, record_id: int, request: Request, db: Session = Depends(get_db)):
    cfg, row, comp_code = get_export_record_or_404(db, request, doc_type, record_id)
    payload = build_document_payload(cfg, row, get_invoice_packing_rows(db, row))
    company = get_export_company_profile(db, comp_code)
    return templates.TemplateResponse(
        request=request,
        name=cfg["template"],
        context={
            **payload,
            "company_id": comp_code,
            "company": company,
            "record": row,
            "doc_type": doc_type,
            "generated_at": datetime.utcnow(),
        },
    )


@router.get("/{doc_type}/pdf/{record_id}")
def export_document_pdf(doc_type: str, record_id: int, request: Request, db: Session = Depends(get_db)):
    cfg, row, comp_code = get_export_record_or_404(db, request, doc_type, record_id)
    pdf_bytes = render_document_pdf(
        cfg,
        row,
        comp_code,
        doc_type,
        get_export_company_profile(db, comp_code),
        get_invoice_packing_rows(db, row),
    )
    document_no = str(getattr(row, cfg["no"], record_id))
    file_row = store_export_pdf(
        db=db,
        company_id=comp_code,
        module_name=doc_type,
        record_id=row.id,
        document_no=document_no,
        document_kind="GENERATED_PDF",
        file_name=f"{safe_filename(document_no)}.pdf",
        content=pdf_bytes,
        uploaded_by=request.session.get("email"),
        remarks="System generated international format PDF",
    )
    # Register PDFs are working drafts. Controlled approval starts only when
    # the final copy is saved through Document Center with selected emails.
    file_row.approval_status = "DRAFT"
    set_document_path(row, file_row.file_path)
    write_audit(db, doc_type, row.id, comp_code, "PDF_GENERATE", "NONE", file_row.file_path, request.session.get("email"))
    db.commit()
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={safe_filename(document_no)}.pdf"},
    )


@router.post("/{doc_type}/upload_pdf/{record_id}")
async def export_document_upload_pdf(
    doc_type: str,
    record_id: int,
    request: Request,
    file: UploadFile = File(...),
    document_kind: str = Form("SIGNED_COPY"),
    remarks: str = Form(None),
    db: Session = Depends(get_db),
):
    cfg, row, comp_code = get_export_record_or_404(db, request, doc_type, record_id)
    if file.content_type != "application/pdf" and not file.filename.lower().endswith(".pdf"):
        return JSONResponse({"success": False, "message": "Only PDF files are allowed"}, status_code=400)
    content = await file.read()
    if not content:
        return JSONResponse({"success": False, "message": "Empty PDF file"}, status_code=400)
    if len(content) > 25 * 1024 * 1024:
        return JSONResponse({"success": False, "message": "PDF size cannot exceed 25 MB"}, status_code=400)
    if not content.startswith(b"%PDF-"):
        return JSONResponse({"success": False, "message": "Invalid PDF file"}, status_code=400)
    document_no = str(getattr(row, cfg["no"], record_id))
    file_row = store_export_pdf(
        db=db,
        company_id=comp_code,
        module_name=doc_type,
        record_id=row.id,
        document_no=document_no,
        document_kind=document_kind,
        file_name=file.filename or f"{document_no}.pdf",
        content=content,
        uploaded_by=request.session.get("email"),
        remarks=remarks,
    )
    # Direct register uploads remain drafts; the Document Center is the single
    # controlled import/approval workflow and prevents duplicate final copies.
    file_row.approval_status = "DRAFT"
    set_document_path(row, file_row.file_path)
    write_audit(db, doc_type, row.id, comp_code, "PDF_UPLOAD", "NONE", file_row.file_path, request.session.get("email"))
    db.commit()
    return {
        "success": True,
        "message": "Draft PDF saved. Import the final copy in Document Center for selected-email approval.",
        "file_id": file_row.id,
        "file_path": file_row.file_path,
        "approval_status": "DRAFT",
    }


@router.get("/files/{file_id}/download")
def export_document_file_download(file_id: int, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Unauthorized")
    file_row = db.query(ExportDocumentFile).filter(ExportDocumentFile.id == file_id, ExportDocumentFile.company_id == comp_code).first()
    if not file_row:
        raise HTTPException(status_code=404, detail="File not found")
    return StreamingResponse(
        BytesIO(file_row.file_bytes),
        media_type=file_row.content_type or "application/pdf",
        headers={"Content-Disposition": f"inline; filename={safe_filename(file_row.file_name)}"},
    )
