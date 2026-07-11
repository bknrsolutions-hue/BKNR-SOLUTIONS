from app.utils.timezone import ist_now
# ============================================================
# PEELING REPORT ROUTER – FINAL (SPECIES + FY LOCK + DB SYNC)
# ============================================================

from fastapi import APIRouter, Request, Depends, Body, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func
import datetime as dt
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.services.floor_balance_sync import refresh_floor_balance
from app.utils.global_filters import get_global_filters
from app.utils.cancel_math import signed_number

from app.database import get_db
from app.database.models.processing import Peeling, AuditLog
from app.database.models.criteria import contractors, varieties as Varieties
from app.services.bill_accounting import (
    cancel_linked_bill_voucher,
    ensure_bill_accounting_schema,
    post_contractor_source_charge,
)

router = APIRouter(
    prefix="/peeling_report",
    tags=["PEELING REPORT"]
)

templates = Jinja2Templates(directory="app/templates")

# --- Helper: Get Financial Year (April to March) ---
def get_fin_year(date_val):
    if not date_val: return None
    return date_val.year if date_val.month >= 4 else date_val.year - 1


def contractor_gst_percent(db: Session, company_id: str, contractor_name: str) -> float:
    row = db.query(contractors).filter(
        contractors.company_id == company_id,
        contractors.contractor_name == contractor_name,
    ).first()
    return float(row.gst_percent or 0) if row else 0.0


def repost_peeling_accounts(db: Session, row: Peeling, company_id: str, email: str):
    if row.journal_id:
        cancel_linked_bill_voucher(db, company_id, row.journal_id, email)
        row.journal_id = None
    if row.is_cancelled or float(row.amount or 0) <= 0:
        return
    voucher = post_contractor_source_charge(
        db=db,
        company_id=company_id,
        voucher_date=row.date,
        reference_no=f"PEL-{row.id}",
        contractor_name=row.contractor_name,
        charge_type="Peeling",
        taxable_amount=row.amount,
        gst_percent=contractor_gst_percent(db, company_id, row.contractor_name),
        created_by=email,
        quantity=row.peeled_qty,
        rate=row.rate,
    )
    row.journal_id = voucher.id


# ------------------------------------------------------------
# 1. MAIN REPORT PAGE (GET) - FY FILTERED & AUTO REFRESH
# ------------------------------------------------------------
@router.get("", response_class=HTMLResponse)
async def peeling_report(request: Request, db: Session = Depends(get_db)):
    production_for, location = get_global_filters(request)
    comp_code = request.session.get("company_code")
    role = request.session.get("role")
    
    # Financial Year from Query Params (e.g., ?fy=2026)
    selected_fy = request.query_params.get("fy")

    if not comp_code:
        return RedirectResponse("/", status_code=302)

    # 1. Current System FY & Variety Criteria
    current_system_fy = get_fin_year(ist_now().date())
    if selected_fy is None:
        selected_fy = str(current_system_fy)
    var_list = db.query(Varieties).filter(Varieties.company_id == comp_code).all()
    yield_map = {v.variety_name: float(v.peeling_yield or 0) for v in var_list}

    # 2. Fetch Rows based on Selected FY (Universal Filters Applied Here)
    rows = []
    if selected_fy:
        try:
            fy_start_int = int(selected_fy)
            start_date = dt.date(fy_start_int, 4, 1)
            end_date = dt.date(fy_start_int + 1, 3, 31)
            
            # 🟢 UPDATED: Filter rows specifically using global parameters and FY range
            query = db.query(Peeling).filter(
                Peeling.company_id == comp_code,
                Peeling.date >= start_date,
                Peeling.date <= end_date
            )

            if production_for:
                query = query.filter(Peeling.production_for == production_for)

            if location:
                query = query.filter(Peeling.peeling_at == location)

            rows = query.order_by(Peeling.date.desc(), Peeling.time.desc()).all()
        except ValueError:
            rows = []

    # 3. Auto-Refresh Logic (Only if viewing CURRENT FY)
    needs_commit = False
    if selected_fy and int(selected_fy) == current_system_fy:
        for r in rows:
            fresh_target = yield_map.get(r.variety_name, 0.0)
            
            # Sync target_yield_percent if different
            if float(r.target_yield_percent or 0) != fresh_target:
                r.target_yield_percent = fresh_target
                needs_commit = True
            
            # Always recalculate display fields for consistency
            h_qty = float(r.hlso_qty or 0)
            p_qty = float(r.peeled_qty or 0)
            target_y = float(r.target_yield_percent or 0)

            r.yield_percent = round((p_qty / h_qty * 100), 2) if h_qty > 0 else 0
            r.diff_percent = round(r.yield_percent - target_y, 2)
            
            if target_y > 0:
                expected_hlso = p_qty / (target_y / 100)
                r.diff_qty = round(expected_hlso - h_qty, 2)
            else:
                r.diff_qty = 0.0

    if needs_commit:
        db.commit()

    # Unique filter options for the UI
    def get_unique(field):
        return sorted({getattr(r, field) for r in rows if getattr(r, field)})

    from app.utils.report_permissions import check_report_permission
    return templates.TemplateResponse(
        request=request,
        name="reports/peeling_report.html",
        context={
            "rows": rows,
            "selected_fy": selected_fy,
            "batches": get_unique("batch_number"),
            "contractors": get_unique("contractor_name"),
            "varieties_dropdown": sorted(list(yield_map.keys())),
            "locations": get_unique("peeling_at"),
            "production_for_list": get_unique("production_for"),
            "is_admin": role == "admin",
            "can_edit": check_report_permission(request, "report_edit"),
            "can_delete": check_report_permission(request, "report_delete"),
            "can_print": check_report_permission(request, "report_print"),
            "can_export": check_report_permission(request, "report_export"),
            "datetime": datetime
        }
    )


# ------------------------------------------------------------
# 2. INLINE UPDATE (POST) - DON'T TOUCH GLOBAL FILTERS
# ------------------------------------------------------------
@router.post("/update")
async def update_peeling(
    request: Request, 
    payload: dict = Body(...), 
    db: Session = Depends(get_db)
):
    comp_code = request.session.get("company_code")
    user_email = request.session.get("email")
    ensure_bill_accounting_schema(db)
    
    from app.utils.report_permissions import enforce_report_permission
    enforce_report_permission(request, "report_edit")

    row = db.query(Peeling).filter(Peeling.id == payload.get("id"), Peeling.company_id == comp_code).first()
    if not row:
        raise HTTPException(status_code=404, detail="Record not found")

    current_fy = get_fin_year(dt.date.today())
    record_fy = get_fin_year(row.date)

    # Basic fields update & Audit logging
    updatable_fields = [
        "batch_number", "contractor_name", "variety_name", "hlso_count",
        "hlso_qty", "peeled_qty", "rate", "peeling_at", "production_for"
    ]

    for field in updatable_fields:
        if field in payload:
            new_val = payload[field]
            old_val = getattr(row, field)

            if field in ["hlso_qty", "peeled_qty", "rate", "hlso_count"]:
                try: new_val = float(new_val or 0)
                except: new_val = 0.0

            if str(old_val) != str(new_val):
                db.add(AuditLog(
                    table_name="peeling", record_id=row.id, company_id=comp_code,
                    field_name=field, old_value=str(old_val), new_value=str(new_val),
                    edited_by=user_email, edited_at=dt.datetime.now(dt.timezone.utc)
                ))
                setattr(row, field, new_val)

    # --- RECALCULATION & TARGET SYNC ---
    if record_fy == current_fy:
        var_obj = db.query(Varieties).filter(
            Varieties.company_id == comp_code, 
            Varieties.variety_name == row.variety_name
        ).first()
        row.target_yield_percent = float(var_obj.peeling_yield or 0) if var_obj else 0.0

    # Final Calculations
    h_qty = float(row.hlso_qty or 0)
    p_qty = float(row.peeled_qty or 0)
    rt = float(row.rate or 0)
    target_y = float(row.target_yield_percent or 0)

    row.yield_percent = round((p_qty / h_qty * 100), 2) if h_qty > 0 else 0
    row.amount = round(p_qty * rt, 2)
    row.diff_percent = round(row.yield_percent - target_y, 2)

    if target_y > 0:
        row.diff_qty = round((p_qty / (target_y / 100)) - h_qty, 2)
    else:
        row.diff_qty = 0.0

    repost_peeling_accounts(db, row, comp_code, user_email)
    db.commit()
    refresh_floor_balance(db, comp_code)
    return {"status": "success", "target_yield": row.target_yield_percent}


# ------------------------------------------------------------
# 3. EXPORT REGION (PDF & EXCEL SPECIFIC ROUTINGS)
# ------------------------------------------------------------
@router.get("/export_pdf")
async def peeling_export_pdf(
    request: Request, 
    ids: str = Query(None), 
    db: Session = Depends(get_db)
):
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Session expired")

    from app.utils.report_permissions import enforce_report_permission
    enforce_report_permission(request, "report_export")

    q = db.query(Peeling).filter(
        Peeling.company_id == comp_code
    )
    
    # 1. 🟢 Apply global filter constraints to the PDF generation scope
    production_for, location = get_global_filters(request)

    if production_for:
        q = q.filter(Peeling.production_for == production_for)

    if location:
        q = q.filter(Peeling.peeling_at == location)

    if ids:
        id_list = [int(i) for i in ids.split(",") if i.isdigit()]
        q = q.filter(Peeling.id.in_(id_list))
        
    rows = q.order_by(Peeling.date.asc()).all()
    
    return templates.TemplateResponse(
        request=request,
        name="reports/peeling_report_pdf.html", 
        context={
            "rows": rows,
            "print_date": ist_now(),
            "company_code": comp_code
        }
    )


@router.get("/export_excel")
def peeling_export_excel(
    request: Request, 
    ids: str = Query(None), 
    db: Session = Depends(get_db)
):
    comp_code = request.session.get("company_code")
    q = db.query(Peeling).filter(
        Peeling.company_id == comp_code
    )
    
    # 2. 🟢 Apply global filter constraints to the Excel workbook data scope
    production_for, location = get_global_filters(request)

    if production_for:
        q = q.filter(Peeling.production_for == production_for)

    if location:
        q = q.filter(Peeling.peeling_at == location)

    if ids: 
        q = q.filter(Peeling.id.in_([int(i) for i in ids.split(",") if i.isdigit()]))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Peeling Ledger"
    
    ws.views.sheetView[0].showGridLines = True
    
    # Styles Setup
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, bold=False)
    total_font = Font(name=font_family, size=10, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )
    total_border = Border(
        top=Side(style='thin', color='000000'), 
        bottom=Side(style='double', color='000000')
    )

    # Header Row
    headers = [
        "Date", "Batch No", "Contractor", "Variety", "HL-Count", 
        "HLSO Qty (Kg)", "Peeled Qty (Kg)", "Target Yield %", 
        "Actual Yield %", "Diff %", "Diff Qty (Kg)", "Rate (₹)", "Amount (₹)"
    ]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data Rows Insertion
    start_row = 2
    for r in q.order_by(Peeling.date.asc()).all():
        ws.append([
            str(r.date.strftime('%d-%m-%Y') if r.date else ''),
            r.batch_number or '',
            r.contractor_name or '',
            r.variety_name or '',
            r.hlso_count or 0,
            r.hlso_qty or 0.0,
            r.peeled_qty or 0.0,
            r.target_yield_percent or 0.0,
            r.yield_percent or 0.0,
            r.diff_percent or 0.0,
            r.diff_qty or 0.0,
            r.rate or 0.0,
            r.amount or 0.0
        ])
        
    end_row = ws.max_row
    
    # Formatting Cell Values & Alignments
    for row in range(start_row, end_row + 1):
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            
            if col in [1, 2]:
                cell.alignment = Alignment(horizontal="center")
            elif col in [3, 4]:
                cell.alignment = Alignment(horizontal="left")
            elif col == 5:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            elif col in [6, 7, 11, 12, 13]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
            elif col in [8, 9, 10]:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '0.00"%"'

    # Grand Totals Row Compilation
    tot_row = end_row + 1
    ws.cell(row=tot_row, column=1, value="GRAND TOTALS").font = total_font
    ws.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=4)
    ws.cell(row=tot_row, column=1).alignment = Alignment(horizontal="right", bold=True)
    
    ws.cell(row=tot_row, column=5, value=f"=SUM(E{start_row}:E{end_row})").number_format = '#,##0'
    ws.cell(row=tot_row, column=6, value=f"=SUM(F{start_row}:F{end_row})").number_format = '#,##0.00'
    ws.cell(row=tot_row, column=7, value=f"=SUM(G{start_row}:G{end_row})").number_format = '#,##0.00'
    
    ws.cell(row=tot_row, column=9, value=f"=IF(F{tot_row}>0,(G{tot_row}/F{tot_row})*100,0)").number_format = '0.00"%"'
    ws.cell(row=tot_row, column=11, value=f"=SUM(K{start_row}:K{end_row})").number_format = '#,##0.00'
    ws.cell(row=tot_row, column=13, value=f"=SUM(M{start_row}:M{end_row})").number_format = '#,##0.00'
    
    for col in range(1, 14):
        t_cell = ws.cell(row=tot_row, column=col)
        t_cell.font = total_font
        t_cell.border = total_border
        if col >= 5:
            t_cell.alignment = Alignment(horizontal="right")

    # Column Width Optimization Setup
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": "attachment; filename=PEELING_REPORT.xlsx"}
    )


# ------------------------------------------------------------
# 4. CONTRACTOR BILLS - DON'T TOUCH GLOBAL FILTERS
# ------------------------------------------------------------
@router.get("/contractor_monthly_bill")
def peeling_monthly_bill(
    request: Request, 
    month: str, 
    contractor: str, 
    ids: str = None, 
    db: Session = Depends(get_db)
):
    comp_code = request.session.get("company_code")
    email = request.session.get("email")
    ensure_bill_accounting_schema(db)
    from app.utils.report_permissions import enforce_report_permission
    enforce_report_permission(request, "report_print")
    query_bill = db.query(Peeling).filter(
        Peeling.company_id == comp_code,
        Peeling.contractor_name == contractor
    )
    
    if ids: 
        query_bill = query_bill.filter(Peeling.id.in_([int(i) for i in ids.split(",") if i.isdigit()]))
    else: 
        query_bill = query_bill.filter(func.to_char(Peeling.date, "YYYY-MM") == month)
        
    rows = query_bill.order_by(Peeling.date.asc()).all()

    # --- Calculations ---
    t_hlso = sum(signed_number(r, r.hlso_qty) for r in rows)
    t_peeled = sum(signed_number(r, r.peeled_qty) for r in rows)
    
    avg_yield = (t_peeled / t_hlso * 100) if t_hlso > 0 else 0

    data = {
        "request": request, 
        "rows": rows, 
        "contractor_name": contractor, 
        "month_year": month,
        "total_hlso": round(t_hlso, 2),
        "total_peeled": round(t_peeled, 2),
        "avg_yield": round(avg_yield, 2),
        "grand_total": round(sum(signed_number(r, r.amount) for r in rows), 2),
        "bill_date": ist_now()
    }
    return templates.TemplateResponse(name="reports/peeling_monthly_bill.html", request=request, context=data)


# ------------------------------------------------------------
# 5. AUDIT HISTORY & TRANSACTION DELETION - DON'T TOUCH
# ------------------------------------------------------------
@router.get("/audit_all")
async def get_all_peeling_audit(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    logs = (
        db.query(AuditLog, Peeling.batch_number)
        .join(Peeling, AuditLog.record_id == Peeling.id)
        .filter(AuditLog.table_name == "peeling", AuditLog.company_id == comp_code)
        .order_by(AuditLog.edited_at.desc()).limit(100).all()
    )
    return [{
        "timestamp": l.AuditLog.edited_at.strftime("%d-%m-%Y %H:%M:%S"),
        "user": l.AuditLog.edited_by.split('@')[0] if l.AuditLog.edited_by else "System",
        "email": l.AuditLog.edited_by if l.AuditLog.edited_by else "System",
        "batch": f"Batch: {l.batch_number}" if l.batch_number else f"ID Ref: {l.AuditLog.record_id}",
        "action": f"Changed {l.AuditLog.field_name.replace('_', ' ').title()}" if l.AuditLog.field_name != "DELETE" else "Deleted Record",
        "details": f"{l.AuditLog.old_value} ➔ {l.AuditLog.new_value}"
    } for l in logs]


@router.post("/delete")
async def delete_peeling(
    request: Request, 
    payload: dict = Body(...), 
    db: Session = Depends(get_db)
):
    comp_code = request.session.get("company_code")
    from app.utils.report_permissions import enforce_report_permission
    enforce_report_permission(request, "report_delete")
    row = db.query(Peeling).filter(Peeling.id == payload.get("id"), Peeling.company_id == comp_code).first()
    if row:
        db.add(AuditLog(
            table_name="peeling", record_id=row.id, company_id=comp_code, 
            field_name="is_cancelled", old_value="False", new_value="True", 
            edited_by=request.session.get("email"), edited_at=dt.datetime.now(dt.timezone.utc)
        ))
        row.is_cancelled = True
        row.status = "Cancelled"
        cancel_linked_bill_voucher(db, comp_code, row.journal_id, email)
        db.commit()
        refresh_floor_balance(db, comp_code)
        return {"status": "success"}
    return {"status": "error"}
