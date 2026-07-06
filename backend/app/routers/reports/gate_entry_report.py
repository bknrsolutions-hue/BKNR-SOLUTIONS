# ============================================================================
# GATE ENTRY REPORT ROUTER – STOCK STYLE (FY LOCK + META SYNC)
# ============================================================================

from fastapi import APIRouter, Request, Depends, Body, HTTPException, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func  # 👈 func.trim() వాడటానికి ఇంపోర్ట్ చేశాను
from datetime import datetime, date
from app.utils.timezone import ist_now
import datetime as dt
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.services.pdf_renderer import render_pdf_from_html
from app.utils.global_filters import get_global_filters

from app.database import get_db
from app.database.models.processing import GateEntry, AuditLog
from app.services.cache import cache_get_or_set

router = APIRouter(
    prefix="/gate_entry",
    tags=["GATE ENTRY REPORT"]
)

templates = Jinja2Templates(directory="app/templates")

# --- Helper: Get Financial Year (April to March) ---
def get_fin_year(date_val):
    if not date_val: return None
    return date_val.year if date_val.month >= 4 else date_val.year - 1


def row_to_dict(row):
    return {k: v for k, v in row.__dict__.items() if not k.startswith("_")}

# ============================================================================
# 1. MAIN REPORT (GET) - WITH FY LOCK & AUTO META-DATA
# ============================================================================
@router.get("", response_class=HTMLResponse)
async def gate_entry_report(
    request: Request,
    fy: str = Query(None), # Financial Year Filter
    db: Session = Depends(get_db)
):
    production_for, location = get_global_filters(request)
    
    # 🟢 TEMPORARY DEBUG PRINT: టెర్మినల్ అవుట్‌పుట్ చెక్ చేయడానికి యాడ్ చేశాను
    print("GLOBAL FILTERS =", production_for, location)
    
    company_id = request.session.get("company_code")
    role = request.session.get("role")
    if not company_id:
        return RedirectResponse("/", status_code=302)

    def build_report_context():
        # 🟢 META DROPDOWNS SYNC WITH func.trim()
        meta_q = db.query(GateEntry).filter(
            GateEntry.company_id == company_id,
            GateEntry.is_cancelled != True
        )
        
        if production_for:
            meta_q = meta_q.filter(func.trim(GateEntry.production_for) == func.trim(production_for))
        if location:
            meta_q = meta_q.filter(func.trim(GateEntry.receiving_center) == func.trim(location))
            
        meta_base = meta_q.all()
        
        def extract_global_unique(field_attr):
            return sorted(list({getattr(r, field_attr) for r in meta_base if getattr(r, field_attr)}))

        base_context = {
            "suppliers_list": extract_global_unique("supplier_name"),
            "factories_list": extract_global_unique("receiving_center"),
            "locations_list": extract_global_unique("purchasing_location"),
            "vehicles_list": extract_global_unique("vehicle_number"),
            "production_for_list": extract_global_unique("production_for"),
            "selected_production_for": production_for,
            "selected_location": location,
        }

        if not fy:
            return {**base_context, "rows": [], "selected_fy": None}

        selected_fy = int(fy)
        start_date = dt.date(selected_fy, 4, 1)
        end_date = dt.date(selected_fy + 1, 3, 31)

        query = db.query(GateEntry).filter(
            GateEntry.company_id == company_id,
            GateEntry.date >= start_date,
            GateEntry.date <= end_date,
            GateEntry.is_cancelled != True
        )

        if production_for:
            query = query.filter(func.trim(GateEntry.production_for) == func.trim(production_for))
        if location:
            query = query.filter(func.trim(GateEntry.receiving_center) == func.trim(location))

        rows = [row_to_dict(row) for row in query.order_by(GateEntry.date.desc(), GateEntry.time.desc()).all()]
        return {**base_context, "rows": rows, "selected_fy": str(selected_fy)}

    cache_key = f"bknr:processing_reports:{company_id}:gate_report:{fy or 'NONE'}:{production_for or 'ALL'}:{location or 'ALL'}"
    context = cache_get_or_set(cache_key, build_report_context, ttl=75)
    context = dict(context)
    
    from app.utils.report_permissions import check_report_permission
    context.update({
        "can_edit": check_report_permission(request, "report_edit"),
        "can_delete": check_report_permission(request, "report_delete"),
        "can_print": check_report_permission(request, "report_print"),
        "can_export": check_report_permission(request, "report_export"),
        "is_admin": role == "admin", 
        "today_date": ist_now()
    })

    if not fy:
        return templates.TemplateResponse(
            request=request,
            name="reports/gate_entry_report.html",
            context=context
        )

    return templates.TemplateResponse(
        request=request,
        name="reports/gate_entry_report.html",
        context=context
    )

# ============================================================================
# 2. DYNAMIC REAL-TIME PDF METADATA FILTER EXPORT
# ============================================================================
@router.get("/export_pdf")
async def gate_export_pdf(
    request: Request, 
    fy: str = Query(None),
    supplier: str = Query(None),
    factory: str = Query(None),
    db: Session = Depends(get_db)
):
    from app.utils.report_permissions import enforce_report_permission
    enforce_report_permission(request, "report_export")
    company_id = request.session.get("company_code")
    company_name = request.session.get("company_name", "BKNR ENTERPRISES")
    if not fy: 
        raise HTTPException(status_code=400, detail="Financial Year parameter missing")
        
    selected_fy = int(fy)
    start_date = date(selected_fy, 4, 1)
    end_date = date(selected_fy + 1, 3, 31)

    query = db.query(GateEntry).filter(
        GateEntry.company_id == company_id,
        GateEntry.date >= start_date,
        GateEntry.date <= end_date,
        GateEntry.is_cancelled != True
    )
    
    # 🟢 GLOBAL FILTERS LOCK FOR PDF WITH func.trim()
    production_for, location = get_global_filters(request)

    if production_for:
        query = query.filter(func.trim(GateEntry.production_for) == func.trim(production_for))
    if location:
        query = query.filter(func.trim(GateEntry.receiving_center) == func.trim(location))
    
    if supplier and supplier.strip() != "":
        query = query.filter(GateEntry.supplier_name == supplier)
    if factory and factory.strip() != "":
        query = query.filter(GateEntry.receiving_center == factory)

    rows = query.order_by(GateEntry.date.asc()).all()

    html_content = templates.get_template("reports/gate_entry_print.html").render({
        "request": request,
        "company_name": company_name,
        "rows": rows,
        "printed_on": ist_now(),
        "auto": 0  
    })
    
    pdf = render_pdf_from_html(html_content)
    return StreamingResponse(
        BytesIO(pdf), 
        media_type="application/pdf", 
        headers={"Content-Disposition": f"attachment; filename=GATE_ENTRY_FY{fy}.pdf"}
    )

# ============================================================================
# 3. DYNAMIC EXCEL ENGINE (CORPORATE DESIGN SHEET BUILDER)
# ============================================================================
@router.get("/export_excel")
async def gate_export_excel(
    request: Request, 
    fy: str = Query(None),
    supplier: str = Query(None),
    factory: str = Query(None),
    db: Session = Depends(get_db)
):
    from app.utils.report_permissions import enforce_report_permission
    enforce_report_permission(request, "report_export")
    company_id = request.session.get("company_code")
    company_name = request.session.get("company_name", "BKNR ENTERPRISES")
    if not fy: 
        raise HTTPException(status_code=400, detail="Financial Year parameter missing")
        
    selected_fy = int(fy)
    start_date = date(selected_fy, 4, 1)
    end_date = date(selected_fy + 1, 3, 31)

    query = db.query(GateEntry).filter(
        GateEntry.company_id == company_id,
        GateEntry.date >= start_date,
        GateEntry.date <= end_date,
        GateEntry.is_cancelled != True
    )
    
    # 🟢 GLOBAL FILTERS LOCK FOR EXCEL WITH func.trim()
    production_for, location = get_global_filters(request)

    if production_for:
        query = query.filter(func.trim(GateEntry.production_for) == func.trim(production_for))
    if location:
        query = query.filter(func.trim(GateEntry.receiving_center) == func.trim(location))
    
    if supplier and supplier.strip() != "":
        query = query.filter(GateEntry.supplier_name == supplier)
    if factory and factory.strip() != "":
        query = query.filter(GateEntry.receiving_center == factory)

    rows = query.order_by(GateEntry.date.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Gate Entry Report"
    ws.views.sheetView[0].showGridLines = True

    ws.merge_cells("A1:M1")
    ws["A1"] = company_name.upper()
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="003366")
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws.merge_cells("A2:M2")
    ws["A2"] = f"GATE ENTRY SHEET | FY: {fy}-{int(fy)+1}"
    ws["A2"].font = Font(name="Arial", size=10, bold=True, color="475569")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append([]) 

    headers = ["SL", "Date", "Time", "Batch Number", "Challan No", "Gate Pass", "Factory", "Supplier Name", "Location", "Vehicle No", "Material Box", "Empty Box", "Ice Box"]
    ws.append(headers)
    
    header_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    start_data_row = 5
    for idx, r in enumerate(rows, 1):
        dt_str = r.date.strftime("%d-%m-%Y") if isinstance(r.date, (date, datetime)) else str(r.date)
        tm_str = r.time.strftime("%H:%M") if r.time else ""
        
        ws.append([
            idx, dt_str, tm_str, r.batch_number, r.challan_number or "", r.gate_pass_number or "",
            r.receiving_center or "", r.supplier_name or "", r.purchasing_location or "", r.vehicle_number or "",
            r.no_of_material_boxes or 0, r.no_of_empty_boxes or 0, r.no_of_ice_boxes or 0
        ])
        
        current_row = start_data_row + idx - 1
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.font = Font(name="Arial", size=9)
            cell.border = thin_border
            if col_num in [8, 9]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_num >= 11:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = center_align

    last_data_row = start_data_row + len(rows) - 1
    total_row_idx = last_data_row + 1
    
    ws.append(["GRAND TOTAL SUMMARY COUNT", "", "", "", "", "", "", "", "", "", f"=SUM(K{start_data_row}:K{last_data_row})", f"=SUM(L{start_data_row}:L{last_data_row})", f"=SUM(M{start_data_row}:M{last_data_row})"])
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=10)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_idx, column=col_num)
        cell.font = Font(name="Arial", size=9, bold=True, color="003366")
        cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        cell.border = thin_border
        if col_num >= 11:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            ws.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="right")

    for col in ws.columns:
        max_len = 0
        first_cell = col[0]
        if hasattr(first_cell, 'column_letter'):
            col_letter = first_cell.column_letter
        else:
            col_letter = get_column_letter(first_cell.column)
            
        for cell in col:
            if cell.row > 3 and cell.row < total_row_idx and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=GATE_ENTRY_FY{fy}.xlsx"}
    )

# ============================================================
# 4. ACTION CONTROLLERS SYSTEM
# ============================================================
@router.post("/update")
async def update_gate_entry(
    request: Request, 
    payload: dict = Body(...), 
    db: Session = Depends(get_db)
):
    from app.utils.report_permissions import enforce_report_permission
    enforce_report_permission(request, "report_edit")
    company_id = request.session.get("company_code")
    user_email = request.session.get("email")

    row = db.query(GateEntry).filter(
        GateEntry.id == payload.get("id"), 
        GateEntry.company_id == company_id
    ).first()
    
    if not row: 
        raise HTTPException(status_code=404, detail="Record not found")

    fields = [
        "batch_number", "challan_number", "gate_pass_number", "receiving_center",
        "supplier_name", "purchasing_location", "vehicle_number", "production_for",
        "no_of_material_boxes", "no_of_empty_boxes", "no_of_ice_boxes"
    ]

    for f in fields:
        if f in payload:
            old_val = str(getattr(row, f))
            new_val = payload[f]
            
            if f in ["no_of_material_boxes", "no_of_empty_boxes", "no_of_ice_boxes"]:
                try: new_val = float(new_val or 0)
                except: new_val = 0.0

            if old_val != str(new_val):
                db.add(AuditLog(
                    table_name="gate_entry", 
                    record_id=row.id, 
                    company_id=company_id,
                    field_name=f, 
                    old_value=old_val, 
                    new_value=str(new_val),
                    edited_by=user_email, 
                    edited_at=dt.datetime.now(dt.timezone.utc)
                ))
                setattr(row, f, new_val)

    db.commit()
    return {"status": "success"}

@router.get("/audit")
async def get_gate_audit(request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    logs = (
        db.query(AuditLog, GateEntry.batch_number)
        .join(GateEntry, AuditLog.record_id == GateEntry.id)
        .filter(AuditLog.table_name == "gate_entry", AuditLog.company_id == comp_code)
        .order_by(AuditLog.edited_at.desc()).limit(100).all()
    )
    return [{
        "timestamp": l.AuditLog.edited_at.strftime("%d-%m-%Y %H:%M:%S"),
        "user": l.AuditLog.edited_by.split('@')[0] if l.AuditLog.edited_by else "System",
        "email": l.AuditLog.edited_by if l.AuditLog.edited_by else "System",
        "batch": f"Batch: {l.batch_number}" if l.batch_number else f"ID Ref: {l.AuditLog.record_id}",
        "action": f"Changed {l.AuditLog.field_name.replace('_', ' ').title()}" if l.AuditLog.field_name != "DELETE" else "Deleted Record",
        "details": f"{l.AuditLog.old_value} ➔ {l.AuditLog.new_value}"
    } for l in logs]

@router.post("/delete")
async def delete_gate_row(request: Request, payload: dict = Body(...), db: Session = Depends(get_db)):
    from app.utils.report_permissions import enforce_report_permission
    enforce_report_permission(request, "report_delete")
    company_id = request.session.get("company_code")
    row = db.query(GateEntry).filter(GateEntry.id == payload.get("id"), GateEntry.company_id == company_id).first()
    if row:
        db.add(AuditLog(
            table_name="gate_entry", record_id=row.id, company_id=company_id, 
            field_name="is_cancelled", old_value="False", new_value="True", 
            edited_by=request.session.get("email"), edited_at=dt.datetime.now(dt.timezone.utc)
        ))
        row.is_cancelled = True
        db.commit()
        return {"status": "success"}
    return {"status": "error"}
