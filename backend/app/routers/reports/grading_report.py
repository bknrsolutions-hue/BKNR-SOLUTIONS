# ============================================================================
# GRADING SUMMARY REPORT – FULL ROUTER (LOCKED, UPDATED & FY FILTER)
# ============================================================================

from fastapi import APIRouter, Request, Depends, Query, HTTPException, Body
from fastapi.responses import (
    HTMLResponse,
    RedirectResponse,
    JSONResponse,
    StreamingResponse
)
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import desc, update, func
from collections import defaultdict
from io import BytesIO
import json
import datetime as dt
from datetime import datetime
from app.utils.global_filters import get_global_filters

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.database import get_db
from app.database.models.processing import Grading, DeHeading, AuditLog
from app.database.models.criteria import HOSO_HLSO_Yields

router = APIRouter(
    prefix="/grading_report",
    tags=["GRADING SUMMARY REPORT"]
)

templates = Jinja2Templates(directory="app/templates")

# --- Helper: Get Financial Year (April to March) ---
def get_fin_year(date_val):
    if not date_val: return None
    return date_val.year if date_val.month >= 4 else date_val.year - 1

# ============================================================================
# PERMISSION CHECK
# ============================================================================
def allow_grading(request: Request):
    role = request.session.get("role", "admin")
    permissions_str = request.session.get("permissions", "")
    allowed_routes = [p.strip() for p in permissions_str.split(",") if p.strip()]
    if role != "admin" and "grading_report" not in allowed_routes:
        raise HTTPException(status_code=403, detail="Access Denied")

# ============================================================================
# MAIN REPORT VIEW WITH FINANCIAL YEAR FILTER
# ============================================================================
@router.get("", response_class=HTMLResponse)
def grading_report(
    request: Request,
    fy: str = Query(None), # Financial Year Filter (Default None, no auto-selection)
    db: Session = Depends(get_db),
    _ = Depends(allow_grading)
):
    production_for, location = get_global_filters(request)
    company_id = request.session.get("company_code")
    if not company_id:
        return RedirectResponse("/auth/login", status_code=303)

    # 1. Base Queries setup for unique filter generation (Universal Filter Core Applied)
    grading_base_query = db.query(Grading).filter(
        Grading.company_id == company_id,
        Grading.is_cancelled != True
    )
    
    if production_for:
        grading_base_query = grading_base_query.filter(Grading.production_for == production_for)

    if location:
        grading_base_query = grading_base_query.filter(Grading.peeling_at == location)
    
    rows = []
    detailed_rows = [] 
    selected_year = None

    # CRITICAL CHANGE: Only process and fetch data if 'fy' parameter is explicitly provided
    if fy:
        deheading_base_query = db.query(DeHeading).filter(
            DeHeading.company_id == company_id,
            DeHeading.is_cancelled != True
        )

        if production_for:
            deheading_base_query = deheading_base_query.filter(DeHeading.production_for == production_for)

        if location:
            deheading_base_query = deheading_base_query.filter(DeHeading.peeling_at == location)

        # 2. Date Boundaries based on Selected Financial Year (April 1st to March 31st)
        selected_year = int(fy)
        start_date = dt.date(selected_year, 4, 1)
        end_date = dt.date(selected_year + 1, 3, 31)

        # Apply date filters to target execution data
        grading_rows = grading_base_query.filter(Grading.date >= start_date, Grading.date <= end_date).all()
        deheading_rows = deheading_base_query.filter(DeHeading.date >= start_date, DeHeading.date <= end_date).all()

        # 🟢 Extracting detailed rows for JS to use in Card View Edits
        for raw_row in grading_rows:
            d = {k: getattr(raw_row, k) for k in raw_row.__dict__ if not k.startswith('_')}
            if 'date' in d and d['date']: d['date'] = str(d['date'])
            if 'time' in d and d['time']: d['time'] = str(d['time'])
            detailed_rows.append(d)

        # 3. Yield Map Preparation
        yield_map = {
            (r.species, str(r.hoso_count)): float(r.hlso_yield_pct or 0) / 100
            for r in db.query(HOSO_HLSO_Yields)
            .filter(HOSO_HLSO_Yields.company_id == company_id)
            .all()
        }

        # 4. De-Heading Data Mapping (Actual HOSO Source) filtered by FY
        deheading_hoso_map = defaultdict(float)
        for r in deheading_rows:
            deheading_hoso_map[(r.batch_number, r.species, str(r.hoso_count))] += float(r.hoso_qty or 0)

        # 5. Grading Raw Data Grouping
        grouped = defaultdict(list)
        for r in grading_rows:
            grouped[(r.batch_number, r.species, str(r.hoso_count), r.variety_name)].append(r)

        idx = 1
        summary_group = defaultdict(list)

        # 6. Summary Calculation Logic
        for (batch, species, hoso_count, variety), items in grouped.items():
            graded_qty_sum = sum(float(i.quantity or 0) for i in items)
            base = sum(float(i.graded_count or 0) * float(i.quantity or 0) for i in items)
            
            yield_factor = yield_map.get((species, hoso_count), 0)

            # Actual HOSO Logic
            if variety == "HOSO":
                actual_hoso_qty = graded_qty_sum
            elif variety == "HLSO":
                actual_hoso_qty = deheading_hoso_map.get((batch, species, hoso_count), 0)
            else:
                actual_hoso_qty = 0

            # Workout & Yield Calculations
            workout = (base / graded_qty_sum) if graded_qty_sum > 0 else 0
            if variety == "HLSO":
                workout = workout * 2.2 * yield_factor

            yield_pct = (graded_qty_sum / actual_hoso_qty * 100) if actual_hoso_qty > 0 else 0
            grading_hoso_qty = (graded_qty_sum / yield_factor) if variety == "HLSO" and yield_factor > 0 else graded_qty_sum
            
            diff_kg = grading_hoso_qty - actual_hoso_qty if variety == "HLSO" else 0
            diff_pct = (diff_kg / actual_hoso_qty * 100) if actual_hoso_qty > 0 else 0

            summary_group[(batch, species)].append({
                "batch": batch,
                "species": species,
                "hoso_count": hoso_count,
                "variety": variety,
                "hoso_qty": round(actual_hoso_qty, 2),
                "graded_qty": round(graded_qty_sum, 2),
                "workout_count": round(workout, 2),
                "yield_pct": round(yield_pct, 2),
                "grading_hoso_qty": round(grading_hoso_qty, 2),
                "weight_diff_kg": round(diff_kg, 2),
                "weight_diff_pct": round(diff_pct, 2),
                "fy_year": str(selected_year)
            })

        # 7. Construct Rows with Subtotals
        for (batch, species), items in summary_group.items():
            sh = sg = sw = sgh = sdiff = 0
            for r in items:
                r["id"] = idx
                rows.append(r)
                idx += 1
                sh += r["hoso_qty"]
                sg += r["graded_qty"]
                sw += r["workout_count"]
                sgh += r["grading_hoso_qty"]
                sdiff += r["weight_diff_kg"]

            rows.append({
                "id": "", "batch": batch, "species": species, "hoso_count": "", "variety": "SUB TOTAL",
                "hoso_qty": round(sh, 2), "graded_qty": round(sg, 2), "workout_count": round(sw, 2),
                "yield_pct": round((sg / sh * 100), 2) if sh > 0 else 0,
                "grading_hoso_qty": round(sgh, 2), "weight_diff_kg": round(sdiff, 2), "weight_diff_pct": 0,
                "fy_year": str(selected_year),
                "is_subtotal": True
            })

    # 8. Dynamic Unique Option Lists generation for Search Columns (All time base)
    all_grading_records = grading_base_query.all()
    
    def get_unique_options(field_attr):
        return sorted(list({str(getattr(r, field_attr)) for r in all_grading_records if getattr(r, field_attr)}))
        
    # Generate unique list of Financial Years from database entries
    unique_fy_years = sorted(list({str(get_fin_year(r.date)) for r in all_grading_records if r.date}), reverse=True)

    from app.utils.report_permissions import check_report_permission
    return templates.TemplateResponse(
        request=request,
        name="reports/grading_report.html",
        context={
            "rows": rows,
            "detailed_rows": detailed_rows, 
            "selected_fy": fy,
            "fy_years": unique_fy_years,
            "batches": get_unique_options("batch_number"),
            "species_list": get_unique_options("species"),
            "varieties": get_unique_options("variety_name"),
            "counts": get_unique_options("hoso_count"),
            "datetime": datetime,
            "can_edit": check_report_permission(request, "report_edit"),
            "can_delete": check_report_permission(request, "report_delete"),
            "can_print": check_report_permission(request, "report_print"),
            "can_export": check_report_permission(request, "report_export"),
        }
    )

# ============================================================
# DETAILED VIEW API – (ALL COLUMNS INCLUDED)
# ============================================================
@router.get("/details")
def grading_details(
    request: Request,
    source: str = Query(...),
    batch: str = Query(None),
    species: str = Query(None),
    hoso_count: str = Query(None),
    variety: str = Query(None),
    db: Session = Depends(get_db)
):
    company_id = request.session.get("company_code")
    if not company_id: return []

    if source == "all":
        data = db.query(Grading).filter(
            Grading.company_id == company_id,
            Grading.is_cancelled != True
        ).order_by(desc(Grading.date), desc(Grading.time)).all()
    elif source == "hoso":
        data = db.query(DeHeading).filter(
            DeHeading.company_id == company_id, 
            DeHeading.batch_number == batch,
            DeHeading.species == species, 
            DeHeading.hoso_count == hoso_count,
            DeHeading.is_cancelled != True
        ).all()
    else:
        data = db.query(Grading).filter(
            Grading.company_id == company_id, 
            Grading.batch_number == batch,
            Grading.species == species, 
            Grading.hoso_count == hoso_count, 
            Grading.variety_name == variety,
            Grading.is_cancelled != True
        ).all()

    result = []
    for r in data:
        d = {k: v for k, v in r.__dict__.items() if not k.startswith('_')}
        if 'date' in d and d['date']: d['date'] = str(d['date'])
        if 'time' in d and d['time']: d['time'] = str(d['time'])
        result.append(d)
    return result

# ============================================================
# UPDATE RECORD (INLINE EDIT)
# ============================================================
@router.post("/update")
async def update_grading(request: Request, db: Session = Depends(get_db)):
    from app.utils.report_permissions import enforce_report_permission
    enforce_report_permission(request, "report_edit")
    company_id = request.session.get("company_code")
    user_email = request.session.get("email")
    data = await request.json()
    record_id = data.get("id")

    if not record_id:
        return JSONResponse({"status": "error", "message": "Missing ID"}, status_code=400)

    original = db.query(Grading).filter(Grading.id == record_id, Grading.company_id == company_id).first()
    if not original:
        return JSONResponse({"status": "error", "message": "Record not found"}, status_code=404)

    # Tracking changes for Audit Log
    for field, new_val in data.items():
        if field in ("id", "company_id"): continue
        old_val = getattr(original, field, None)
        
        if str(old_val) != str(new_val):
            log = AuditLog(
                table_name="grading",
                record_id=record_id,
                field_name=field,
                old_value=str(old_val),
                new_value=str(new_val),
                edited_by=user_email,
                edited_at=datetime.utcnow(),
                company_id=company_id
            )
            db.add(log)
            setattr(original, field, new_val)

    db.commit()
    return {"status": "success"}

# ============================================================
# DELETE RECORD
# ============================================================
@router.post("/delete")
async def delete_grading(request: Request, db: Session = Depends(get_db)):
    from app.utils.report_permissions import enforce_report_permission
    enforce_report_permission(request, "report_delete")
    company_id = request.session.get("company_code")
    user_email = request.session.get("email")
    data = await request.json()
    record_id = data.get("id")

    row = db.query(Grading).filter(Grading.id == record_id, Grading.company_id == company_id).first()
    if row:
        db.add(AuditLog(
            table_name="grading", record_id=row.id, company_id=company_id,
            field_name="DELETE", old_value="Record", new_value="Deleted",
            edited_by=user_email, edited_at=datetime.utcnow()
        ))
        db.delete(row)
        db.commit()
        return {"status": "success"}
    
    return {"status": "error", "message": "Not found"}

# ============================================================
# AUDIT LOGS (FETCH)
# ============================================================
@router.get("/audit")
def get_grading_audits(request: Request, db: Session = Depends(get_db)):
    company_id = request.session.get("company_code")
    logs = (
        db.query(AuditLog, Grading.batch_number)
        .join(Grading, AuditLog.record_id == Grading.id)
        .filter(AuditLog.table_name == "grading", AuditLog.company_id == company_id)
        .order_by(desc(AuditLog.edited_at)).limit(100).all()
    )
    return [{
        "timestamp": l.AuditLog.edited_at.strftime("%d-%m-%Y %H:%M:%S"),
        "user": l.AuditLog.edited_by.split('@')[0] if l.AuditLog.edited_by else "System",
        "email": l.AuditLog.edited_by if l.AuditLog.edited_by else "System",
        "batch": f"Batch: {l.batch_number}" if l.batch_number else f"ID Ref: {l.AuditLog.record_id}",
        "action": f"Changed {l.AuditLog.field_name.replace('_', ' ').title()}" if l.AuditLog.field_name != "DELETE" else "Deleted Record",
        "details": f"{l.AuditLog.old_value} ➔ {l.AuditLog.new_value}"
    } for l in logs]

# ============================================================
# EXPORT EXCEL (WITH UNIVERSAL FILTERS LAYER)
# ============================================================
@router.get("/export_excel")
def export_excel(request: Request, db: Session = Depends(get_db)):
    from app.utils.report_permissions import enforce_report_permission
    enforce_report_permission(request, "report_export")
    company_id = request.session.get("company_code")
    
    # 🟢 Mapped and forced global criteria evaluation inside spreadsheet compiler
    production_for, location = get_global_filters(request)

    query = db.query(Grading).filter(
        Grading.company_id == company_id,
        Grading.is_cancelled != True
    )

    if production_for:
        query = query.filter(Grading.production_for == production_for)

    if location:
        query = query.filter(Grading.peeling_at == location)

    rows = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Grading Report"

    headers = ["Date", "Time", "Batch", "Species", "Variety", "Count", "Quantity", "Graded Count"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([str(r.date), str(r.time), r.batch_number, r.species, r.variety_name, r.hoso_count, r.quantity, r.graded_count])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": "attachment; filename=Grading_Detailed_Report.xlsx"}
    )