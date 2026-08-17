from app.utils.timezone import ist_now
from fastapi import APIRouter, Request, Depends, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text, or_
from app.database import get_db
from app.utils.download_security import issue_download_grant, require_download_grant
from app.utils.data_management_audit import DATA_MANAGEMENT_HISTORY_FILE, log_data_management_action

import os
import json
import random
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import pandas as pd
import numpy as np
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import math
from datetime import date, time, datetime, timedelta
from zoneinfo import ZoneInfo
import re
from sqlalchemy.types import Date, Time, DateTime, Integer, BigInteger, SmallInteger, Float, Numeric, Boolean, String, Text

# =====================================================
# 🟢 1. ALL MODELS IMPORTS
# =====================================================
# Users & Company for OTP validation
from app.database.models.users import User, Company

# Processing
from app.database.models.processing import GateEntry, RawMaterialPurchasing, DeHeading, Grading, Peeling, Soaking, Production, AuditLog as ProcessingAuditLog
from app.database.models.reprocess import Reprocess
# Inventory Management
from app.database.models.inventory_management import stock_entry, pending_orders, sales_dispatch, cold_storage_holding, cold_storage
# Bills
from app.database.models.bills import ElectricityLog, DieselLog, PurchaseInvoice, ContainerLog, QATestingLog, OtherExpense
# General Stock
from app.database.models.general_stock import GeneralStock, GeneralStoreItems
# Payments
from app.database.models.payments import (
    CustomerReceivable,
    VendorPayment,
    BankTransaction,
    ExpenseVoucher,
    JournalEntry,
    JournalEntryLine,
    PaymentReceipt,
    BuyerAgingSummary,
    ERPAlertEngine,
)

from app.database.models.enterprise_finance import LedgerMaster
# Masters
from app.database.models.criteria import brands, purposes, production_at, production_for, glazes, grades, varieties, countries, buyers, buyer_agents, packing_styles, production_types, chemicals, contractors, suppliers, peeling_rates, species, purchasing_locations, vehicle_numbers, coldstore_locations, freezers, grade_to_hoso, HOSO_HLSO_Yields, peeling_at, shipping_vendors, vendors, hsn_codes
# HRMS
from app.database.models.attendance import EmployeeRegistration, DailyAttendance, EmployeeIncrement, EmployeeStatutoryMaster, EmployeeSalaryAdvance

router = APIRouter()
SENDER_EMAIL = os.getenv("SMTP_EMAIL", "bknr.solutions@gmail.com")
SENDER_NAME = os.getenv("EMAIL_SENDER_NAME", "SVBK")
if not SENDER_NAME or "bknr" in SENDER_NAME.lower():
    SENDER_NAME = "SVBK"
SUPPORT_EMAIL = os.getenv("SUPPORT_EMAIL", "bknr.solutions@gmail.com")

# =====================================================
# 🟢 2. GLOBAL DICTIONARY FOR DYNAMIC IMPORT
# =====================================================
ALL_MODELS = {
    "GateEntry": GateEntry, "RawMaterialPurchasing": RawMaterialPurchasing, "DeHeading": DeHeading,
    "Grading": Grading, "Peeling": Peeling, "Soaking": Soaking, "Production": Production, "Reprocess": Reprocess,
    "StockEntry": stock_entry, "PendingOrders": pending_orders, "SalesDispatch": sales_dispatch,
    "ColdStorageHolding": cold_storage_holding, "ColdStorageMaster": cold_storage,
    "PurchaseInvoice": PurchaseInvoice, "ContainerLog": ContainerLog, "ElectricityLog": ElectricityLog,
    "DieselLog": DieselLog, "QATestingLog": QATestingLog, "OtherExpense": OtherExpense,
    "GeneralStock": GeneralStock, "GeneralStoreItems": GeneralStoreItems,
    "CustomerReceivable": CustomerReceivable, "VendorPayment": VendorPayment, "BankTransaction": BankTransaction,
    "ExpenseVoucher": ExpenseVoucher, "JournalEntry": JournalEntry, "LedgerMaster": LedgerMaster,
    "PaymentReceipt": PaymentReceipt, "ERPAlertEngine": ERPAlertEngine,
    "Brands": brands, "Purposes": purposes, "ProductionAt": production_at, "ProductionFor": production_for,
    "Glazes": glazes, "Grades": grades, "Varieties": varieties, "Countries": countries, "Buyers": buyers,
    "BuyerAgents": buyer_agents, "PackingStyles": packing_styles, "ProductionTypes": production_types,
    "Chemicals": chemicals, "Contractors": contractors, "Suppliers": suppliers, "Species": species,
    "HSNCodes": hsn_codes, "EmployeeRegistration": EmployeeRegistration, "DailyAttendance": DailyAttendance,
    "EmployeeIncrement": EmployeeIncrement, "EmployeeStatutoryMaster": EmployeeStatutoryMaster,
    "EmployeeSalaryAdvance": EmployeeSalaryAdvance
}

REGISTER_GROUPS = {
    "processing": {
        "gate-entry": ("Gate Entry Register", GateEntry, "GateEntry"),
        "raw-material-purchasing": ("Raw Material Purchasing Register", RawMaterialPurchasing, "RawMaterial"),
        "de-heading": ("De-Heading Register", DeHeading, "DeHeading"),
        "grading": ("Grading Register", Grading, "Grading"),
        "peeling": ("Peeling Register", Peeling, "Peeling"),
        "soaking": ("Soaking Register", Soaking, "Soaking"),
        "production": ("Production Register", Production, "Production"),
        "reprocess": ("Reprocess Register", Reprocess, "Reprocess"),
    },
    "inventory": {
        "stock-entry": ("Stock Entry Register", stock_entry, "StockEntry"),
        "pending-orders": ("Pending Orders Register", pending_orders, "PendingOrders"),
        "sales-dispatch": ("Sales Dispatch Register", sales_dispatch, "SalesDispatch"),
        "cold-storage-holding": ("Cold Storage Holding Register", cold_storage_holding, "ColdStorageHolding"),
        "cold-storage-master": ("Cold Storage Master Register", cold_storage, "ColdStorageMaster"),
    },
    "accounts": {
        "ledger-master": ("Ledger Master Register", LedgerMaster, "LedgerMaster"),
        "journal-entries": ("Journal Entries Register", JournalEntry, "JournalEntry"),
        "customer-receivables": ("Customer Receivables Register", CustomerReceivable, "CustomerReceivable"),
        "vendor-payments": ("Vendor Payments Register", VendorPayment, "VendorPayment"),
        "bank-transactions": ("Bank Transactions Register", BankTransaction, "BankTransaction"),
        "expense-vouchers": ("Expense Vouchers Register", ExpenseVoucher, "ExpenseVoucher"),
        "payment-receipts": ("Payment Receipts Register", PaymentReceipt, "PaymentReceipt"),
        "purchase-invoices": ("Purchase Invoices Register", PurchaseInvoice, "PurchaseInvoice"),
        "container-costs": ("Container & Logistics Register", ContainerLog, "ContainerLog"),
    },
    "hrms": {
        "employee-registration": ("Employee Registration Register", EmployeeRegistration, "EmployeeRegistration"),
        "daily-attendance": ("Daily Attendance Register", DailyAttendance, "DailyAttendance"),
        "employee-increments": ("Employee Increment Register", EmployeeIncrement, "EmployeeIncrement"),
        "statutory-master": ("Employee Statutory Register", EmployeeStatutoryMaster, "StatutoryMaster"),
        "salary-advances": ("Salary Advance Register", EmployeeSalaryAdvance, "SalaryAdvance"),
    },
}


# =====================================================
# 🟢 3. HISTORY LOGGING HELPER
# =====================================================
HISTORY_LOG_FILE = str(DATA_MANAGEMENT_HISTORY_FILE)
log_data_action = log_data_management_action

def build_security_otp_email(otp: str, action: str, module: str, company_code: str):
    action_label = str(action or "").replace("_", " ").title()
    module_label = str(module or "").replace("_", " ").title()
    html = f"""
    <!doctype html>
    <html>
    <body style="margin:0;padding:0;background:#eef6ff;font-family:Arial,Helvetica,sans-serif;color:#1f2937;">
      <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#eef6ff;padding:24px 12px;">
        <tr>
          <td align="center">
            <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="max-width:560px;background:#ffffff;border:1px solid #dbeafe;border-radius:12px;overflow:hidden;">
              <tr>
                <td style="padding:18px 22px;background:#f8fbff;border-bottom:1px solid #e5eefb;">
                  <div style="font-size:18px;font-weight:800;color:#1d4ed8;">SVBK</div>
                  <div style="font-size:12px;color:#64748b;margin-top:4px;">Data management security verification</div>
                </td>
              </tr>
              <tr>
                <td style="padding:24px 22px;">
                  <h2 style="margin:0 0 12px;font-size:20px;color:#0f172a;">Security OTP required</h2>
                  <p style="margin:0 0 18px;color:#475569;font-size:14px;line-height:1.6;">An OTP was requested for a protected data-management action.</p>
                  <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="border-collapse:collapse;font-size:14px;margin-bottom:18px;">
                    <tr><td style="padding:8px;border-bottom:1px solid #e5eefb;color:#64748b;width:36%;">Action</td><td style="padding:8px;border-bottom:1px solid #e5eefb;color:#0f172a;font-weight:700;">{action_label}</td></tr>
                    <tr><td style="padding:8px;border-bottom:1px solid #e5eefb;color:#64748b;">Module</td><td style="padding:8px;border-bottom:1px solid #e5eefb;color:#0f172a;">{module_label}</td></tr>
                    <tr><td style="padding:8px;border-bottom:1px solid #e5eefb;color:#64748b;">Company</td><td style="padding:8px;border-bottom:1px solid #e5eefb;color:#0f172a;">{company_code}</td></tr>
                  </table>
                  <div style="padding:18px;background:#f0f7ff;border:1px solid #bfdbfe;border-radius:10px;text-align:center;">
                    <div style="font-size:12px;color:#64748b;text-transform:uppercase;letter-spacing:.08em;">Verification Code</div>
                    <div style="font-size:32px;font-weight:800;color:#1d4ed8;letter-spacing:6px;margin-top:6px;">{otp}</div>
                  </div>
                  <p style="margin:14px 0 0;color:#64748b;font-size:13px;line-height:1.6;">Share this code only with an authorized admin. If you did not request this action, contact support immediately.</p>
                </td>
              </tr>
              <tr>
                <td style="padding:16px 22px;background:#f8fbff;border-top:1px solid #e5eefb;color:#64748b;font-size:12px;line-height:1.6;">
                  Sent by <strong>{SENDER_NAME}</strong> from {SENDER_EMAIL}<br>
                  For support, contact {SUPPORT_EMAIL}. This is an automated email.
                </td>
              </tr>
            </table>
          </td>
        </tr>
      </table>
    </body>
    </html>
    """
    text = f"SVBK security OTP: {otp}\nAction: {action_label}\nModule: {module_label}\nCompany: {company_code}\nSupport: {SUPPORT_EMAIL}"
    return html, text


# =====================================================
# 🟢 4. COMMON HELPER FUNCTIONS
# =====================================================
def get_comp_code(request: Request):
    comp_code = request.session.get("company_code")
    if not comp_code:
        raise HTTPException(status_code=401, detail="Session expired or unauthorized")
    return str(comp_code)

def export_sheet(writer, data, sheet_name):
    if data:
        df = pd.DataFrame([{k: v for k, v in vars(row).items() if k != "_sa_instance_state"} for row in data])
        def excel_safe_value(value):
            if isinstance(value, datetime) and value.tzinfo is not None and value.utcoffset() is not None:
                return value.astimezone(ZoneInfo("Asia/Kolkata")).replace(tzinfo=None)
            if isinstance(value, (dict, list, tuple, set)):
                return json.dumps(value, default=str, ensure_ascii=False)
            return value
        df = df.map(excel_safe_value)
        df.insert(0, "Sl No", range(1, len(df) + 1))
    else:
        df = pd.DataFrame()
    acronym_labels = {
        "id": "ID", "po": "PO", "grn": "GRN", "gst": "GST", "hsn": "HSN",
        "esi": "ESI", "pf": "PF", "uan": "UAN", "utr": "UTR", "ifsc": "IFSC",
        "pan": "PAN", "tds": "TDS", "qty": "Qty", "rm": "RM",
    }
    df.columns = [
        " ".join(acronym_labels.get(part.lower(), part.capitalize()) for part in str(column).split("_"))
        for column in df.columns
    ]
    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

def style_department_register_workbook(writer, company_code: str):
    for sheet in writer.book.worksheets:
        sheet.insert_rows(1, amount=3)
        max_column = max(1, sheet.max_column)
        max_row = max(4, sheet.max_row)
        title = re.sub(r"(?<!^)(?=[A-Z])", " ", sheet.title).strip()
        if not title.lower().endswith("register"):
            title = f"{title} Register"

        sheet.freeze_panes = "A5"
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
        title_cell = sheet.cell(1, 1, title)
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="123B5D")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 28

        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)
        meta_cell = sheet.cell(
            2, 1,
            f"Company: {company_code} | Generated: {datetime.utcnow():%d-%b-%Y %H:%M UTC}",
        )
        meta_cell.alignment = Alignment(horizontal="center", vertical="center")
        meta_cell.font = Font(size=10, color="475569")

        for column in range(1, max_column + 1):
            header = sheet.cell(4, column)
            header.font = Font(bold=True, color="FFFFFF")
            header.fill = PatternFill("solid", fgColor="176B87")
            header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[4].height = 24

        for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=False)

        for column in range(1, max_column + 1):
            values = [
                len(str(sheet.cell(row, column).value or ""))
                for row in range(4, min(sheet.max_row, 504) + 1)
            ]
            sheet.column_dimensions[get_column_letter(column)].width = min(max(max(values, default=10) + 2, 12), 42)
        sheet.auto_filter.ref = f"A4:{get_column_letter(max_column)}{max_row}"

def register_rows(db: Session, model, company_code: str):
    query = db.query(model)
    if hasattr(model, "company_id"):
        query = query.filter(model.company_id == company_code)
    return query.all()

def generate_export_response(comp_code, module_name, export_logic, db):
    export_dir = "exports"
    os.makedirs(export_dir, exist_ok=True)
    filename = f"SVBK_{module_name}_{comp_code}_{ist_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(export_dir, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        export_logic(writer, db, comp_code)
        style_department_register_workbook(writer, comp_code)

    return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# =====================================================
# 🟢 5. OTP SECURITY LOGIC
# =====================================================
OTP_STORE = {}

class OTPRequest(BaseModel):
    action: str
    module: str

class OTPVerify(BaseModel):
    action: str
    otp: str

@router.post("/data-management/generate-otp")
async def generate_otp(payload: OTPRequest, request: Request, db: Session = Depends(get_db)):
    comp_code = request.session.get("company_code")
    if not comp_code: return {"success": False, "error": "Session Expired"}

    try:
        company = db.query(Company).filter(Company.company_code == comp_code).first()
        if not company: return {"success": False, "error": "Company not found."}

        all_users = db.query(User).filter(User.company_id == company.id).all()
        authorized_emails = []

        for u in all_users:
            perms = u.permissions or ""
            has_attr_access = getattr(u, 'data_management_access', False)
            if u.role == 'admin' or "data_management" in perms or has_attr_access:
                if u.email:
                    authorized_emails.append(u.email)
    except Exception as e:
        print(f"Error fetching users: {e}")
        authorized_emails = []

    if not authorized_emails:
        return {"success": False, "error": f"No authorized data managers found for {comp_code}."}

    otp = str(random.randint(100000, 999999))
    session_id = request.session.get("session_id", comp_code) 
    OTP_STORE[f"{session_id}_{payload.action}"] = otp

    SMTP_SERVER = "smtp.gmail.com"
    SMTP_PORT = 587
    sender_password = os.getenv("SMTP_PASSWORD")
    if not sender_password:
        return {"success": False, "error": "Security email delivery is not configured."}
    sent_count = 0

    try:
        html_body, text_body = build_security_otp_email(otp, payload.action, payload.module, comp_code)
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SENDER_EMAIL, sender_password)
        for email_id in authorized_emails:
            try:
                msg = MIMEMultipart()
                msg['From'] = f"{SENDER_NAME} <{SENDER_EMAIL}>"
                msg['To'] = email_id
                msg['Subject'] = f"SVBK - Security OTP for {str(payload.action).upper()}"
                msg.attach(MIMEText(text_body, 'plain'))
                msg.attach(MIMEText(html_body, 'html'))
                server.send_message(msg)
                sent_count += 1
            except: pass
        server.quit()
    except Exception:
        return {"success": False, "error": "Security email delivery failed. Please try again."}

    return {"success": True, "message": f"OTP sent to {sent_count} admins."}

@router.post("/data-management/verify-otp")
async def verify_otp(payload: OTPVerify, request: Request):
    comp_code = request.session.get("company_code")
    session_id = request.session.get("session_id", comp_code)
    key = f"{session_id}_{payload.action}"
    
    if OTP_STORE.get(key) == payload.otp:
        del OTP_STORE[key]
        response = {"success": True, "message": "OTP Verified!"}
        if payload.action in {"download", "export"}:
            response["download_token"] = issue_download_grant(request)
        return response
    return {"success": False, "error": "Invalid OTP."}


# =====================================================
# 🟢 6. PAGE RENDER & TEMPLATE DOWNLOAD
# =====================================================
@router.get("/data-management", response_class=HTMLResponse)
async def data_management(request: Request, db: Session = Depends(get_db)):
    templates = request.app.state.templates
    return templates.TemplateResponse(
        request=request, 
        name="admin/data_management.html", 
        context={"request": request}
    )
    
@router.get("/data-management/template/blank")
async def download_blank_template(request: Request, table: str = ""):
    """Download a blank Excel sheet for one table or all tables."""
    require_download_grant(request)
    from fastapi import Query as FastQuery
    template_dir = "templates_excel"
    os.makedirs(template_dir, exist_ok=True)

    def build_blank_df(model_class):
        columns = [c.name for c in model_class.__table__.columns if c.name != 'id']
        # Build a sample hint row so the user knows what format each column expects
        hint_row = {}
        for c in model_class.__table__.columns:
            if c.name == 'id': continue
            t = str(c.type).upper()
            if 'DATE' in t:   hint_row[c.name] = 'YYYY-MM-DD'
            elif 'TIME' in t: hint_row[c.name] = 'HH:MM:SS'
            elif 'FLOAT' in t or 'NUMERIC' in t: hint_row[c.name] = '0.00'
            elif 'INT' in t:  hint_row[c.name] = '0'
            elif 'BOOL' in t: hint_row[c.name] = 'True/False'
            else:              hint_row[c.name] = ''
        df = pd.DataFrame([hint_row], columns=columns)
        return df

    # Single table
    if table and table in ALL_MODELS:
        model_class = ALL_MODELS[table]
        df = build_blank_df(model_class)
        safe_sheet = table[:31]
        filename = f"BKNR_BlankSheet_{table}_{ist_now().strftime('%Y%m%d')}.xlsx"
        file_path = os.path.join(template_dir, filename)
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # INFO sheet
            pd.DataFrame({
                "Info": [f"SVBK Blank Import Sheet — {table}"],
                "Note": ["Row 1 shows format hints. Delete row 1 before importing actual data."]
            }).to_excel(writer, sheet_name="INFO", index=False)
            df.to_excel(writer, sheet_name=safe_sheet, index=False)
        return FileResponse(file_path, filename=filename,
                            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # All tables (one workbook, each table = one sheet)
    filename = f"BKNR_AllBlankSheets_{ist_now().strftime('%Y%m%d')}.xlsx"
    file_path = os.path.join(template_dir, filename)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        pd.DataFrame({
            "Info": ["SVBK — All Tables Blank Import Sheets"],
            "Note": ["Each sheet = one DB table. Row 1 = format hints. Delete Row 1 before importing real data."]
        }).to_excel(writer, sheet_name="INFO", index=False)
        for tbl_name, model_class in ALL_MODELS.items():
            try:
                df = build_blank_df(model_class)
                df.to_excel(writer, sheet_name=tbl_name[:31], index=False)
            except Exception:
                pass
    return FileResponse(file_path, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Kept for backward compatibility
@router.get("/data-management/template/master")
async def download_master_template(request: Request):
    return await download_blank_template(request=request, table="GateEntry")


# =====================================================
# 🟢 7. SEGREGATED EXPORT ROUTES (7 MODULES)
# =====================================================
@router.get("/export/processing", operation_id="export_processing_get")
@router.post("/export/processing", operation_id="export_processing_post")
async def export_processing(request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    def logic(writer, db, cc):
        export_sheet(writer, db.query(GateEntry).filter(GateEntry.company_id == cc).all(), "GateEntry")
        export_sheet(writer, db.query(RawMaterialPurchasing).filter(RawMaterialPurchasing.company_id == cc).all(), "RawMaterial")
        export_sheet(writer, db.query(DeHeading).filter(DeHeading.company_id == cc).all(), "DeHeading")
        export_sheet(writer, db.query(Grading).filter(Grading.company_id == cc).all(), "Grading")
        export_sheet(writer, db.query(Peeling).filter(Peeling.company_id == cc).all(), "Peeling")
        export_sheet(writer, db.query(Soaking).filter(Soaking.company_id == cc).all(), "Soaking")
        export_sheet(writer, db.query(Production).filter(Production.company_id == cc).all(), "Production")
        export_sheet(writer, db.query(Reprocess).filter(Reprocess.company_id == cc).all(), "Reprocess")
        log_data_action(cc, "EXPORT", "Processing Module", "Success", "Exported Processing records")
    return generate_export_response(get_comp_code(request), "Processing", logic, db)

@router.get("/export/inventory", operation_id="export_inventory_get")
@router.post("/export/inventory", operation_id="export_inventory_post")
async def export_inventory(request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    def logic(writer, db, cc):
        export_sheet(writer, db.query(stock_entry).filter(stock_entry.company_id == cc).all(), "StockEntry")
        export_sheet(writer, db.query(pending_orders).filter(pending_orders.company_id == cc).all(), "PendingOrders")
        export_sheet(writer, db.query(sales_dispatch).filter(sales_dispatch.company_id == cc).all(), "SalesDispatch")
        export_sheet(writer, db.query(cold_storage_holding).filter(cold_storage_holding.company_id == cc).all(), "ColdStorageHolding")
        export_sheet(writer, db.query(cold_storage).filter(cold_storage.company_id == cc).all(), "ColdStorageMaster")
        log_data_action(cc, "EXPORT", "Inventory Module", "Success", "Exported Inventory records")
    return generate_export_response(get_comp_code(request), "Inventory", logic, db)

@router.get("/export/bills", operation_id="export_bills_get")
@router.post("/export/bills", operation_id="export_bills_post")
async def export_bills(request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    def logic(writer, db, cc):
        export_sheet(writer, db.query(PurchaseInvoice).filter(PurchaseInvoice.company_id == cc).all(), "PurchaseInvoice")
        export_sheet(writer, db.query(ContainerLog).filter(ContainerLog.company_id == cc).all(), "ContainerLog")
        export_sheet(writer, db.query(ElectricityLog).all(), "ElectricityLog") 
        export_sheet(writer, db.query(DieselLog).all(), "DieselLog") 
        export_sheet(writer, db.query(QATestingLog).all(), "QATestingLog") 
        export_sheet(writer, db.query(OtherExpense).all(), "OtherExpense")
        log_data_action(cc, "EXPORT", "Bills Module", "Success", "Exported Commercial Bills")
    return generate_export_response(get_comp_code(request), "Bills", logic, db)

@router.get("/export/general-stock", operation_id="export_general_stock_get")
@router.post("/export/general-stock", operation_id="export_general_stock_post")
async def export_general_stock(request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    def logic(writer, db, cc):
        # 🌟 FIX: Removed extract_company_id, using 'cc' directly
        export_sheet(writer, db.query(GeneralStock).filter(GeneralStock.company_id == cc).all(), "GeneralStock")
        export_sheet(writer, db.query(GeneralStoreItems).filter(GeneralStoreItems.company_id == cc).all(), "GeneralStoreItems")
        log_data_action(cc, "EXPORT", "General Stock Module", "Success", "Exported General Stock records")
    return generate_export_response(get_comp_code(request), "GeneralStock", logic, db)

@router.get("/export/payments", operation_id="export_payments_get")
@router.post("/export/payments", operation_id="export_payments_post")
async def export_payments(request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    def logic(writer, db, cc):
        export_sheet(writer, db.query(CustomerReceivable).filter(CustomerReceivable.company_id == cc).all(), "CustomerReceivable")
        export_sheet(writer, db.query(VendorPayment).filter(VendorPayment.company_id == cc).all(), "VendorPayment")
        export_sheet(writer, db.query(BankTransaction).filter(BankTransaction.company_id == cc).all(), "BankTransaction")
        export_sheet(writer, db.query(ExpenseVoucher).filter(ExpenseVoucher.company_id == cc).all(), "ExpenseVoucher")
        export_sheet(writer, db.query(JournalEntry).filter(JournalEntry.company_id == cc).all(), "JournalEntry")
        export_sheet(writer, db.query(LedgerMaster).filter(LedgerMaster.company_id == cc).all(), "LedgerMaster")
        export_sheet(writer, db.query(PaymentReceipt).filter(PaymentReceipt.company_id == cc).all(), "PaymentReceipt")
        export_sheet(writer, db.query(ERPAlertEngine).filter(ERPAlertEngine.company_id == cc).all(), "ERPAlertEngine")
        log_data_action(cc, "EXPORT", "Finance/Payments Module", "Success", "Exported Financial records")
    return generate_export_response(get_comp_code(request), "Payments", logic, db)

@router.get("/export/accounts", operation_id="export_accounts_get")
@router.post("/export/accounts", operation_id="export_accounts_post")
async def export_accounts(request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    def logic(writer, db, cc):
        for _, model, sheet_name in REGISTER_GROUPS["accounts"].values():
            export_sheet(writer, register_rows(db, model, cc), sheet_name)
        log_data_action(cc, "EXPORT", "Accounts Module", "Success", "Exported tenant accounts and ledger registers")
    return generate_export_response(get_comp_code(request), "Accounts", logic, db)

@router.get("/export/masters", operation_id="export_masters_get")
@router.post("/export/masters", operation_id="export_masters_post")
async def export_masters(request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    def logic(writer, db, cc):
        export_sheet(writer, db.query(brands).filter(brands.company_id == cc).all(), "Brands")
        export_sheet(writer, db.query(purposes).filter(purposes.company_id == cc).all(), "Purposes")
        export_sheet(writer, db.query(glazes).filter(glazes.company_id == cc).all(), "Glazes")
        export_sheet(writer, db.query(grades).filter(grades.company_id == cc).all(), "Grades")
        export_sheet(writer, db.query(varieties).filter(varieties.company_id == cc).all(), "Varieties")
        export_sheet(writer, db.query(countries).filter(countries.company_id == cc).all(), "Countries")
        export_sheet(writer, db.query(buyers).filter(buyers.company_id == cc).all(), "Buyers")
        export_sheet(writer, db.query(contractors).filter(contractors.company_id == cc).all(), "Contractors")
        export_sheet(writer, db.query(suppliers).filter(suppliers.company_id == cc).all(), "Suppliers")
        export_sheet(writer, db.query(species).filter(species.company_id == cc).all(), "Species")
        export_sheet(writer, db.query(hsn_codes).filter(hsn_codes.company_id == cc).all(), "HSNCodes")
        log_data_action(cc, "EXPORT", "Masters Module", "Success", "Exported System Masters")
    return generate_export_response(get_comp_code(request), "Masters", logic, db)

@router.get("/export/hrms", operation_id="export_hrms_get")
@router.post("/export/hrms", operation_id="export_hrms_post")
async def export_hrms(request: Request, db: Session = Depends(get_db)):
    require_download_grant(request)
    def logic(writer, db, cc):
        export_sheet(writer, db.query(EmployeeRegistration).filter(EmployeeRegistration.company_id == cc).all(), "EmployeeReg")
        export_sheet(writer, db.query(DailyAttendance).filter(DailyAttendance.company_id == cc).all(), "DailyAttendance")
        export_sheet(writer, db.query(EmployeeIncrement).filter(EmployeeIncrement.company_id == cc).all(), "EmployeeIncrement")
        export_sheet(writer, db.query(EmployeeStatutoryMaster).filter(EmployeeStatutoryMaster.company_id == cc).all(), "StatutoryMaster")
        export_sheet(writer, db.query(EmployeeSalaryAdvance).filter(EmployeeSalaryAdvance.company_id == cc).all(), "SalaryAdvance")
        log_data_action(cc, "EXPORT", "HRMS Module", "Success", "Exported HRMS records")
    return generate_export_response(get_comp_code(request), "HRMS", logic, db)

@router.get("/data-management/register/{module}/{register_key}.xlsx")
async def download_module_register(
    module: str,
    register_key: str,
    request: Request,
    db: Session = Depends(get_db),
):
    require_download_grant(request)
    company_code = get_comp_code(request)
    group = REGISTER_GROUPS.get(str(module or "").strip().lower())
    register = group.get(str(register_key or "").strip().lower()) if group else None
    if not register:
        raise HTTPException(status_code=404, detail="Register not found")
    label, model, sheet_name = register

    def logic(writer, db, cc):
        rows = register_rows(db, model, cc)
        export_sheet(writer, rows, sheet_name)
        log_data_action(cc, "REGISTER", label, "Success", f"Downloaded {len(rows)} tenant records")

    safe_name = re.sub(r"[^A-Za-z0-9]+", "", sheet_name) or "Register"
    return generate_export_response(company_code, safe_name, logic, db)


# =====================================================
# 🟢 8. DYNAMIC MAPPING IMPORT LOGIC
# =====================================================
class ImportMappingPayload(BaseModel):
    filename: str
    table_name: str
    sheet_name: str
    mapping: dict

class ClearTablePayload(BaseModel):
    table_name: str

@router.get("/data-management/db-schema")
async def get_db_schema():
    schema = {}
    for table_name, model in ALL_MODELS.items():
        columns = [c.name for c in model.__table__.columns if c.name != 'id']
        schema[table_name] = columns
    return {"success": True, "tables": schema}

@router.post("/data-management/inspect-file")
async def inspect_excel_file(excel_file: UploadFile = File(...)):
    try:
        upload_dir = "uploads"
        os.makedirs(upload_dir, exist_ok=True)
        filename = f"import_temp_{ist_now().strftime('%Y%m%d%H%M%S')}_{excel_file.filename}"
        filepath = os.path.join(upload_dir, filename)

        with open(filepath, "wb") as buffer:
            buffer.write(await excel_file.read())

        xl = pd.ExcelFile(filepath)
        sheet_data = {}
        for sheet in xl.sheet_names:
            df = xl.parse(sheet, nrows=0) 
            sheet_data[sheet] = df.columns.tolist()

        return {"success": True, "filename": filename, "sheets": sheet_data}
    except Exception as e:
        return {"success": False, "error": str(e)}

def coerce_cell_value(val, col_obj):
    if val is None or pd.isna(val):
        return None

    col_type = getattr(col_obj, 'type', None)
    if col_type is None:
        if isinstance(val, float) and math.isnan(val):
            return None
        return val

    # 1. TIME Column
    if isinstance(col_type, Time):
        if isinstance(val, time) and not isinstance(val, datetime):
            return val
        if isinstance(val, (datetime, pd.Timestamp)):
            return val.time()
        if isinstance(val, timedelta):
            total_sec = int(val.total_seconds()) % 86400
            h = total_sec // 3600
            m = (total_sec % 3600) // 60
            s = total_sec % 60
            micro = val.microseconds
            return time(h, m, s, micro)
        if isinstance(val, str):
            val_str = val.strip()
            if not val_str or val_str.lower() in ('none', 'nan', 'nat', 'null', '-'):
                return None
            for fmt in ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p", "%H:%M:%S.%f"):
                try:
                    return datetime.strptime(val_str, fmt).time()
                except ValueError:
                    pass
            try:
                dt = pd.to_datetime(val_str)
                return dt.time()
            except Exception:
                return None
        return None

    # 2. DATE Column
    if isinstance(col_type, Date) and not isinstance(col_type, DateTime):
        if isinstance(val, date) and not isinstance(val, datetime):
            return val
        if isinstance(val, (datetime, pd.Timestamp)):
            return val.date()
        if isinstance(val, str):
            val_str = val.strip()
            if not val_str or val_str.lower() in ('none', 'nan', 'nat', 'null', '-'):
                return None
            try:
                return pd.to_datetime(val_str).date()
            except Exception:
                return None
        return None

    # 3. DATETIME Column
    if isinstance(col_type, DateTime):
        if isinstance(val, pd.Timestamp):
            return val.to_pydatetime()
        if isinstance(val, datetime):
            return val
        if isinstance(val, date):
            return datetime.combine(val, time.min)
        if isinstance(val, str):
            val_str = val.strip()
            if not val_str or val_str.lower() in ('none', 'nan', 'nat', 'null', '-'):
                return None
            try:
                return pd.to_datetime(val_str).to_pydatetime()
            except Exception:
                return None
        return None

    # 4. BOOLEAN Column
    if isinstance(col_type, Boolean):
        if isinstance(val, bool):
            return val
        if isinstance(val, (int, float)):
            if math.isnan(val):
                return None
            return bool(val)
        if isinstance(val, str):
            val_str = val.strip().lower()
            if not val_str or val_str in ('none', 'nan', 'null', '-'):
                return None
            return val_str in ('true', '1', 'yes', 't', 'y')
        return bool(val)

    # 5. INTEGER Column
    if isinstance(col_type, (Integer, BigInteger, SmallInteger)):
        if isinstance(val, (int, float)):
            if math.isnan(val):
                return None
            return int(val)
        if isinstance(val, str):
            val_str = val.strip()
            if not val_str or val_str.lower() in ('none', 'nan', 'null', '-'):
                return None
            try:
                return int(float(val_str))
            except ValueError:
                return None

    # 6. FLOAT / NUMERIC Column
    if isinstance(col_type, (Float, Numeric)):
        if isinstance(val, (int, float)):
            if math.isnan(val):
                return None
            return float(val)
        if isinstance(val, str):
            val_str = val.strip()
            if not val_str or val_str.lower() in ('none', 'nan', 'null', '-'):
                return None
            try:
                return float(val_str)
            except ValueError:
                return None

    # 7. STRING / TEXT Column
    if isinstance(col_type, (String, Text)):
        if isinstance(val, float) and math.isnan(val):
            return None
        val_str = str(val).strip()
        if val_str.lower() in ('nan', 'none', 'null'):
            return None
        return val_str

    if isinstance(val, float) and math.isnan(val):
        return None
    return val


def ensure_date_obj(val):
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.date()
    if isinstance(val, str) and val.strip():
        try:
            return pd.to_datetime(val.strip()).date()
        except Exception:
            pass
    return ist_now().date()


def sync_imported_data_to_modules(db: Session, comp_code: str, table_name: str, records: list):
    """
    Ensures that imported records trigger corresponding accounts flow
    (journals, vouchers, contractor charges) and floor balance stock updates.
    """
    if not records:
        return

    import logging
    logger = logging.getLogger(__name__)

    try:
        from app.services.bill_accounting import post_contractor_source_charge, post_vendor_bill, purchase_stock_account, resolve_posting_ledger
        from app.services.floor_balance_sync import refresh_floor_balance
    except Exception as e:
        logger.error(f"Error importing sync modules: {e}")
        return

    batches_to_refresh = set()

    for rec in records:
        b_num = getattr(rec, "batch_number", None)
        if b_num:
            batches_to_refresh.add(str(b_num).strip())

        # 1. DE-HEADING
        if table_name == "DeHeading":
            rec_id = getattr(rec, "id", None)
            contractor = getattr(rec, "contractor", None)
            amount = float(getattr(rec, "amount", 0) or 0)
            rec_date = ensure_date_obj(getattr(rec, "date", None))
            hoso_qty = float(getattr(rec, "hoso_qty", 0) or 0)
            rate = float(getattr(rec, "rate_per_kg", 0) or 0)
            email = getattr(rec, "email", "SYSTEM") or "SYSTEM"

            if rec_id and contractor and amount > 0:
                try:
                    voucher = post_contractor_source_charge(
                        db, comp_code, rec_date, f"DEH-{rec_id}",
                        contractor, "De-Heading", amount, 0, email, hoso_qty, rate
                    )
                    if voucher and hasattr(rec, "journal_id"):
                        rec.journal_id = getattr(voucher, "id", None)
                except Exception as e:
                    logger.error(f"Error posting contractor charge for imported DeHeading {rec_id}: {e}")

        # 2. PEELING
        elif table_name == "Peeling":
            rec_id = getattr(rec, "id", None)
            contractor = getattr(rec, "contractor_name", None)
            amount = float(getattr(rec, "amount", 0) or 0)
            rec_date = ensure_date_obj(getattr(rec, "date", None))
            peeled_qty = float(getattr(rec, "peeled_qty", 0) or 0)
            rate = float(getattr(rec, "rate_per_kg", 0) or 0)
            email = getattr(rec, "email", "SYSTEM") or "SYSTEM"

            if rec_id and contractor and amount > 0:
                try:
                    voucher = post_contractor_source_charge(
                        db, comp_code, rec_date, f"PEL-{rec_id}",
                        contractor, "Peeling", amount, 0, email, peeled_qty, rate
                    )
                    if voucher and hasattr(rec, "journal_id"):
                        rec.journal_id = getattr(voucher, "id", None)
                except Exception as e:
                    logger.error(f"Error posting contractor charge for imported Peeling {rec_id}: {e}")

        # 3. DAILY ATTENDANCE (Contractor Labour)
        elif table_name == "DailyAttendance":
            rec_id = getattr(rec, "id", None)
            contractor = getattr(rec, "contractor_name", None)
            amount = float(getattr(rec, "calculated_amount", 0) or getattr(rec, "approved_duty_credit", 0) or 0)
            duty_date = ensure_date_obj(getattr(rec, "duty_date", None))
            email = getattr(rec, "email", "SYSTEM") or "SYSTEM"

            if rec_id and contractor and amount > 0:
                try:
                    voucher = post_contractor_source_charge(
                        db, comp_code, duty_date, f"ATT-{rec_id}",
                        contractor, "Contract Labour", amount, 0, email
                    )
                    if voucher and hasattr(rec, "journal_id"):
                        rec.journal_id = getattr(voucher, "id", None)
                except Exception as e:
                    logger.error(f"Error posting contractor charge for imported DailyAttendance {rec_id}: {e}")

        # 4. PURCHASE INVOICE
        elif table_name == "PurchaseInvoice":
            rec_id = getattr(rec, "id", None)
            inv_no = getattr(rec, "invoice_no", None)
            vendor_id = getattr(rec, "vendor_id", None)
            grand_total = float(getattr(rec, "grand_total", 0) or 0)
            tax_amount = float(getattr(rec, "tax_amount", 0) or 0)
            taxable_val = float(getattr(rec, "qty", 0) or 0) * float(getattr(rec, "base_price", 0) or 0)
            inv_date = ensure_date_obj(getattr(rec, "invoice_date", None))
            email = getattr(rec, "email", "SYSTEM") or "SYSTEM"

            if inv_no and grand_total > 0 and not getattr(rec, "journal_id", None):
                try:
                    from app.database.models.criteria import vendors
                    v_row = db.query(vendors).filter(vendors.id == vendor_id, vendors.company_id == comp_code).first() if vendor_id else None
                    v_name = v_row.name if v_row else f"Vendor {vendor_id or 'Import'}"
                    p_name = getattr(rec, "product_name", "Purchase Product") or "Purchase Product"
                    default_post = purchase_stock_account(p_name)
                    post_ledger = resolve_posting_ledger(
                        db, comp_code, None,
                        default_post["ledger_name"], default_post["group_name"],
                        default_post["group_type"], default_post["parent_group_name"]
                    )
                    voucher = post_vendor_bill(
                        db, comp_code, inv_date, str(inv_no).upper(), v_name,
                        post_ledger["ledger_name"], taxable_val, tax_amount, grand_total,
                        f"Imported Purchase invoice {inv_no}", email,
                        expense_group_name=post_ledger["group_name"],
                        expense_group_type=post_ledger["group_type"],
                        expense_parent_group_name=post_ledger["parent_group_name"],
                        voucher_type="Purchase"
                    )
                    if voucher:
                        rec.journal_id = voucher.id
                        rec.status = "POSTED"
                except Exception as e:
                    logger.error(f"Error posting vendor bill for imported PurchaseInvoice {rec_id}: {e}")

    try:
        db.commit()
    except Exception as e:
        logger.error(f"Error committing synced journal_ids: {e}")

    # 5. REFRESH FLOOR BALANCE
    if table_name in ["DeHeading", "Peeling", "RawMaterialPurchasing", "Grading", "Reprocess", "Soaking", "Production"]:
        try:
            if batches_to_refresh:
                for b in batches_to_refresh:
                    refresh_floor_balance(db, comp_code, batch_number=b)
            else:
                refresh_floor_balance(db, comp_code)
        except Exception as e:
            logger.error(f"Error refreshing floor balance for {table_name}: {e}")


@router.post("/data-management/execute-import")
async def execute_dynamic_import(payload: ImportMappingPayload, request: Request, db: Session = Depends(get_db)):
    comp_code = get_comp_code(request)
    filepath = os.path.join("uploads", payload.filename)

    if not os.path.exists(filepath):
        return {"success": False, "error": "File expired or not found. Please re-upload."}

    try:
        ModelClass = ALL_MODELS.get(payload.table_name)
        if not ModelClass:
            return {"success": False, "error": "Invalid database table selected."}

        df = pd.read_excel(filepath, sheet_name=payload.sheet_name)
        df = df.where(pd.notnull(df), None)

        model_cols = {c.name: c for c in ModelClass.__table__.columns}
        records_to_insert = []
        for index, row in df.iterrows():
            record_data = {}
            for db_col, excel_col in payload.mapping.items():
                if excel_col and excel_col in df.columns:
                    val = row[excel_col]
                    col_obj = model_cols.get(db_col)
                    coerced = coerce_cell_value(val, col_obj)
                    if coerced is not None:
                        record_data[db_col] = coerced

            if hasattr(ModelClass, "company_id"):
                record_data["company_id"] = comp_code

            records_to_insert.append(ModelClass(**record_data))

        db.add_all(records_to_insert)
        db.commit()

        # 🌟 Sync imported data with accounts flow (vouchers/journals) & floor balance stock
        sync_imported_data_to_modules(db, comp_code, payload.table_name, records_to_insert)

        if os.path.exists(filepath):
            os.remove(filepath)

        msg = f"Successfully dynamically mapped and imported {len(records_to_insert)} records."
        log_data_action(comp_code, "IMPORT", payload.table_name, "Success", msg)

        return {"success": True, "rows_imported": len(records_to_insert), "table": payload.table_name}
    except Exception as e:
        db.rollback()
        log_data_action(comp_code, "IMPORT", payload.table_name, "Failed", str(e))
        return {"success": False, "error": str(e)}


# =====================================================
# 🟢 9. UNDO & CLEAR TABLE (EMERGENCY CLEANUP)
# =====================================================
@router.post("/data-management/undo-import")
async def undo_last_import(payload: ClearTablePayload, request: Request, db: Session = Depends(get_db)):
    comp_code = get_comp_code(request)
    ModelClass = ALL_MODELS.get(payload.table_name)
    if not ModelClass:
        return {"success": False, "error": "Invalid database table selected."}

    logs = []
    if os.path.exists(HISTORY_LOG_FILE):
        try:
            with open(HISTORY_LOG_FILE, "r") as f:
                logs = json.load(f)
        except Exception:
            pass

    import_logs = [
        l for l in logs 
        if l.get("company_code") == comp_code 
        and l.get("type") == "IMPORT" 
        and l.get("module") == payload.table_name 
        and l.get("status") == "Success"
    ]

    if not import_logs:
        return {"success": False, "error": "No recent successful import found in history to undo."}

    last_import = import_logs[-1] 
    details = last_import.get("details", "")

    # Regex to find number in "imported X records"
    match = re.search(r'imported (\d+) records', details.lower())
    if not match:
        return {"success": False, "error": "Could not determine how many rows to delete from history log."}
    
    rows_to_delete = int(match.group(1))
    if rows_to_delete <= 0:
        return {"success": False, "error": "Zero rows were imported in the last session."}

    try:
        query = db.query(ModelClass.id)
        db_columns = [c.name for c in ModelClass.__table__.columns]
        
        # 🌟 FIX: Directly use VNBK2162 (comp_code) string for safe match
        if "company_id" in db_columns:
            query = query.filter(ModelClass.company_id == comp_code)
        
        top_ids = query.order_by(ModelClass.id.desc()).limit(rows_to_delete).all()
        ids_to_delete = [row[0] for row in top_ids]
        
        if not ids_to_delete:
            return {"success": False, "error": "No database records found to delete."}

        deleted_count = db.query(ModelClass).filter(ModelClass.id.in_(ids_to_delete)).delete(synchronize_session=False)
        db.commit()

        msg = f"Reverted last import. Deleted {deleted_count} records."
        log_data_action(comp_code, "UNDO IMPORT", payload.table_name, "Success", msg)

        return {"success": True, "rows_deleted": deleted_count, "table": payload.table_name}

    except Exception as e:
        db.rollback()
        log_data_action(comp_code, "UNDO IMPORT", payload.table_name, "Failed", str(e))
        return {"success": False, "error": str(e)}

@router.post("/data-management/clear-table")
async def clear_table_data(payload: ClearTablePayload, request: Request, db: Session = Depends(get_db)):
    comp_code = get_comp_code(request)
    ModelClass = ALL_MODELS.get(payload.table_name)

    if not ModelClass:
        return {"success": False, "error": "Invalid database table selected."}

    try:
        db_columns = [c.name for c in ModelClass.__table__.columns]
        query = db.query(ModelClass)

        # 🌟 FIX: Directly use VNBK2162 (comp_code) string for safe match
        if "company_id" in db_columns:
            query = query.filter(ModelClass.company_id == comp_code)
        else:
            return {"success": False, "error": "Action denied. Cannot bulk delete global shared tables."}

        deleted_count = query.delete(synchronize_session=False)
        db.commit()

        msg = f"Permanently deleted all {deleted_count} records from table."
        log_data_action(comp_code, "CLEAR", payload.table_name, "Success", msg)

        return {"success": True, "rows_deleted": deleted_count, "table": payload.table_name}

    except Exception as e:
        db.rollback()
        log_data_action(comp_code, "CLEAR", payload.table_name, "Failed", str(e))
        return {"success": False, "error": str(e)}


@router.get("/data-management/history")
async def import_history(request: Request):
    comp_code = request.session.get("company_code")
    logs = []
    if os.path.exists(HISTORY_LOG_FILE):
        try:
            with open(HISTORY_LOG_FILE, "r") as f:
                all_logs = json.load(f)
                logs = [l for l in all_logs if str(l.get("company_code") or "") == str(comp_code or "")]
        except Exception as e:
            pass
            
    logs.reverse() 
    return {"success": True, "history": logs}
